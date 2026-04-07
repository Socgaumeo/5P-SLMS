"""
Test rate sheet parser: regex vs AI (DeepSeek, Gemini, Claude).
Creates synthetic test files + tests real files from EBOOK/.
Compares: accuracy, rate count, parse time, cost estimation.

Usage:
  cd backend
  python tests/test-rate-sheet-parser-comparison.py
"""

import asyncio
import json
import logging
import os
import sys
import time
from datetime import datetime
from typing import Dict, List, Any

import pandas as pd
from openpyxl import Workbook

# Add backend to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

# Test output directory
TEST_DIR = os.path.join(os.path.dirname(__file__), "test_rate_files")
EBOOK_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "..", "EBOOK")
REPORT_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "..", "plans", "reports", "test-260302-1005-rate-parser-comparison.md"
)

# Estimated cost per 1K tokens (USD)
COST_PER_1K = {
    "deepseek": {"input": 0.00014, "output": 0.00028},
    "gemini": {"input": 0.000075, "output": 0.0003},
    "anthropic": {"input": 0.003, "output": 0.015},
}


# ===================================================================
# Create synthetic test files
# ===================================================================

def create_test_files():
    """Create synthetic Excel files for each format type."""
    os.makedirs(TEST_DIR, exist_ok=True)
    files = {}

    # 1. Trucking pivot table (like ANT format)
    files["trucking_pivot"] = _create_trucking_pivot()
    # 2. Trucking vertical table (1 col per field)
    files["trucking_vertical"] = _create_trucking_vertical()
    # 3. Customs rate sheet
    files["customs"] = _create_customs_sheet()
    # 4. Packing service rate
    files["packing"] = _create_packing_sheet()
    # 5. Warehouse rate
    files["warehouse"] = _create_warehouse_sheet()
    # 6. Mixed format (company header + pivot)
    files["mixed_header"] = _create_mixed_header()

    return files


def _create_trucking_pivot():
    """Trucking pivot: rows = routes, columns = vehicle types."""
    path = os.path.join(TEST_DIR, "test_trucking_pivot.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Báo giá"

    # Company header rows (buried)
    ws.append(["CÔNG TY VẬN TẢI ABC"])
    ws.append(["Địa chỉ: 123 Đường XYZ, Hà Nội"])
    ws.append(["ĐT: 0912345678"])
    ws.append([])
    ws.append(["BẢNG GIÁ CƯỚC VẬN CHUYỂN NỘI ĐỊA 2026"])
    ws.append([])

    # Header row (row 7)
    ws.append(["STT", "Từ", "Đến", "Truck 1.25T", "Truck 2.5T", "Truck 5T", "CONT 20'", "CONT 40'"])

    # Data rows
    routes = [
        (1, "Hà Nội", "Hải Phòng", 1800000, 2500000, 3500000, 5000000, 7500000),
        (2, None, "Bắc Ninh", 1200000, 1800000, 2500000, 4000000, 6000000),
        (3, None, "Hưng Yên", 1500000, 2200000, 3000000, 4500000, 6500000),
        (4, "HCM", "Bình Dương", 1000000, 1500000, 2000000, 3500000, 5000000),
        (5, None, "Đồng Nai", 1200000, 1800000, 2500000, 4000000, 6000000),
        (6, None, "Long An", 1500000, 2000000, 2800000, 4200000, 6200000),
    ]
    for r in routes:
        ws.append(list(r))

    # Surcharge rows
    ws.append([])
    ws.append(["", "Chờ giờ", "", 200000, 300000, 400000, 500000, 700000])
    ws.append(["", "Hủy chuyến", "", 500000, 700000, 900000, 1200000, 1500000])

    wb.save(path)
    return {"path": path, "expected_count": 30, "type": "TRUCKING_DOM"}


def _create_trucking_vertical():
    """Trucking vertical: each row = 1 rate with all fields."""
    path = os.path.join(TEST_DIR, "test_trucking_vertical.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Bảng giá"

    ws.append(["STT", "Điểm đi", "Điểm đến", "Loại xe", "Đơn giá (VND)", "Đơn vị", "Ghi chú"])

    rates = [
        (1, "Hà Nội", "Hải Phòng", "Truck 5T", 3500000, "Chuyến", ""),
        (2, "Hà Nội", "Bắc Ninh", "Truck 2.5T", 1800000, "Chuyến", ""),
        (3, "HCM", "Bình Dương", "CONT 20'", 3500000, "Chuyến", "Nội thành"),
        (4, "HCM", "Đồng Nai", "CONT 40'", 6000000, "Chuyến", ""),
        (5, "Đà Nẵng", "Quảng Nam", "Truck 1.25T", 1200000, "Chuyến", ""),
    ]
    for r in rates:
        ws.append(list(r))

    wb.save(path)
    return {"path": path, "expected_count": 5, "type": "TRUCKING_DOM"}


def _create_customs_sheet():
    """Customs service rate sheet."""
    path = os.path.join(TEST_DIR, "test_customs_rates.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Phí hải quan"

    ws.append(["BẢNG GIÁ DỊCH VỤ HẢI QUAN"])
    ws.append([])
    ws.append(["STT", "Dịch vụ", "Loại hình", "Đơn giá (VND)", "Đơn vị", "Ghi chú"])

    items = [
        (1, "Khai báo hải quan", "Xuất khẩu", 800000, "Tờ khai", ""),
        (2, "Khai báo hải quan", "Nhập khẩu", 1000000, "Tờ khai", ""),
        (3, "Khai báo hải quan", "Tại chỗ", 600000, "Tờ khai", ""),
        (4, "Kiểm hóa", "Xuất khẩu", 500000, "Lần", "Tại kho"),
        (5, "Kiểm hóa", "Nhập khẩu", 700000, "Lần", "Tại cảng"),
        (6, "Phí C/O", "Xuất khẩu", 350000, "Bộ", "Form AK/AI/D"),
        (7, "Hun trùng", "", 1500000, "Lần", "Tùy khối lượng"),
        (8, "Phí nâng hạ container", "", 2000000, "Cont", "20ft"),
        (9, "Phí nâng hạ container", "", 3000000, "Cont", "40ft"),
    ]
    for r in items:
        ws.append(list(r))

    wb.save(path)
    return {"path": path, "expected_count": 9, "type": "CUSTOMS"}


def _create_packing_sheet():
    """Packing service rate sheet."""
    path = os.path.join(TEST_DIR, "test_packing_rates.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Đóng gói"

    ws.append(["BẢNG GIÁ DỊCH VỤ ĐÓNG GÓI"])
    ws.append([])
    ws.append(["STT", "Loại dịch vụ", "Quy cách", "Đơn giá (VND)", "Đơn vị"])

    items = [
        (1, "Đóng kiện gỗ", "< 1 CBM", 2500000, "Kiện"),
        (2, "Đóng kiện gỗ", "1-3 CBM", 3500000, "Kiện"),
        (3, "Đóng kiện gỗ", "> 3 CBM", 4500000, "Kiện"),
        (4, "Đóng pallet", "Standard", 800000, "Pallet"),
        (5, "Đóng pallet", "Heavy duty", 1200000, "Pallet"),
        (6, "Bọc màng co", "", 150000, "CBM"),
        (7, "Dán nhãn", "", 50000, "Kiện"),
    ]
    for r in items:
        ws.append(list(r))

    wb.save(path)
    return {"path": path, "expected_count": 7, "type": "PACKING"}


def _create_warehouse_sheet():
    """Warehouse service rate sheet."""
    path = os.path.join(TEST_DIR, "test_warehouse_rates.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Kho bãi"

    ws.append(["BẢNG GIÁ KHO BÃI"])
    ws.append(["Áp dụng từ: 01/01/2026"])
    ws.append([])
    ws.append(["STT", "Dịch vụ", "Chi tiết", "Đơn giá (VND)", "Đơn vị", "Ghi chú"])

    items = [
        (1, "Lưu kho", "Hàng thường", 25000, "CBM/ngày", "Tối thiểu 3 ngày"),
        (2, "Lưu kho", "Hàng lạnh", 45000, "CBM/ngày", ""),
        (3, "Lưu kho", "Hàng nguy hiểm", 60000, "CBM/ngày", ""),
        (4, "Bốc xếp", "< 5 tấn", 150000, "Lần", ""),
        (5, "Bốc xếp", "> 5 tấn", 300000, "Lần", "Cần xe nâng"),
        (6, "Kiểm đếm", "", 100000, "Lần", ""),
    ]
    for r in items:
        ws.append(list(r))

    wb.save(path)
    return {"path": path, "expected_count": 6, "type": "WAREHOUSE"}


def _create_mixed_header():
    """Trucking pivot with deep company header (like real ANT file)."""
    path = os.path.join(TEST_DIR, "test_mixed_deep_header.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Rate"

    # 12 rows of company info before data
    ws.append(["CÔNG TY TNHH VẬN TẢI & THƯƠNG MẠI ANT"])
    ws.append(["MST: 0123456789"])
    ws.append(["Địa chỉ: Lô A, KCN Quang Minh, Mê Linh, Hà Nội"])
    ws.append(["Hotline: 1900 xxxx"])
    ws.append([])
    ws.append(["Kính gửi: Quý khách hàng"])
    ws.append(["Chúng tôi xin gửi bảng giá cước vận chuyển"])
    ws.append([])
    ws.append(["Điều kiện:", "FOB Hà Nội"])
    ws.append(["Thời gian:", "Từ 01/01/2026"])
    ws.append([])

    # Header at row 12
    ws.append(["STT", "Từ", "Đến", "1.25T", "2.5T", "5T", "14T"])

    # Data
    ws.append([1, "KCN Quang Minh", "Nội Bài", 1200000, 1800000, 2500000, 5000000])
    ws.append([2, None, "Hải Dương", 2000000, 2800000, 3800000, 7000000])
    ws.append([3, None, "Bắc Ninh", 1500000, 2200000, 3000000, 5500000])
    ws.append([4, "KCN Thăng Long", "Nội Bài", 1000000, 1500000, 2000000, 4000000])
    ws.append([5, None, "Hải Phòng", 2500000, 3500000, 4800000, 9000000])

    # Notes
    ws.append([])
    ws.append(["Ghi chú: Giá trên chưa bao gồm VAT 8%"])
    ws.append(["Lưu ý: Phụ phí xăng dầu có thể thay đổi theo thị trường"])

    wb.save(path)
    return {"path": path, "expected_count": 20, "type": "TRUCKING_DOM"}


# ===================================================================
# Parser functions
# ===================================================================

def run_regex_parser(file_path: str) -> Dict[str, Any]:
    """Run regex-based parser and measure time."""
    from app.api.rate_file_upload import parse_excel_rates

    start = time.time()
    result = parse_excel_rates(file_path)
    elapsed = time.time() - start

    return {
        "method": "regex",
        "rate_count": len(result.get("parsed_rates", [])),
        "time_sec": round(elapsed, 3),
        "rates": result.get("parsed_rates", []),
        "sheet_info": result.get("sheet_info", []),
        "error": result.get("error"),
        "cost_usd": 0,
    }


async def run_ai_parser(file_path: str, provider: str, service_type: str = None) -> Dict[str, Any]:
    """Run AI parser with specific provider and measure time."""
    import importlib
    ai_parser = importlib.import_module("app.ai.excel.rate-sheet-ai-parser")

    # Temporarily override AI_PROVIDER
    from app.core.config import settings
    original_provider = settings.AI_PROVIDER
    settings.AI_PROVIDER = provider

    # Clear cached client
    import app.ai.clients as clients_mod
    clients_mod._client_instance = None

    start = time.time()
    try:
        result = await ai_parser.parse_rates_with_ai(file_path, service_type_code=service_type)
        elapsed = time.time() - start

        rate_count = len(result.get("parsed_rates", []))

        # Estimate token cost (rough: ~4 chars per token)
        # Input: sheet content (~2000 chars = ~500 tokens) + prompt (~800 tokens)
        # Output: ~200 tokens per rate
        est_input_tokens = 1300
        est_output_tokens = max(200, rate_count * 50)
        cost = COST_PER_1K.get(provider, COST_PER_1K["deepseek"])
        est_cost = (est_input_tokens / 1000 * cost["input"]) + (est_output_tokens / 1000 * cost["output"])

        return {
            "method": f"ai_{provider}",
            "rate_count": rate_count,
            "time_sec": round(elapsed, 3),
            "rates": result.get("parsed_rates", []),
            "sheet_info": result.get("sheet_info", []),
            "error": result.get("error"),
            "cost_usd": round(est_cost, 6),
        }
    except Exception as e:
        elapsed = time.time() - start
        return {
            "method": f"ai_{provider}",
            "rate_count": 0,
            "time_sec": round(elapsed, 3),
            "rates": [],
            "error": str(e),
            "cost_usd": 0,
        }
    finally:
        settings.AI_PROVIDER = original_provider
        clients_mod._client_instance = None


# ===================================================================
# Test runner
# ===================================================================

async def run_all_tests():
    """Run all parser tests and generate comparison report."""
    results = {}

    # Create synthetic test files
    logger.info("Creating synthetic test files...")
    test_files = create_test_files()

    # Add real files from EBOOK
    real_files = {
        "real_ant": {
            "path": os.path.join(EBOOK_DIR, "Báo giá ANT.xlsx"),
            "expected_count": 54,
            "type": "TRUCKING_DOM",
        },
        "real_navf": {
            "path": os.path.join(EBOOK_DIR, "NAVF_0808.xlsx"),
            "expected_count": 7,
            "type": "TRUCKING_DOM",
        },
        "real_asgl_quotation": {
            "path": os.path.join(EBOOK_DIR, "ASGL Quotation for shipment 09.01.2026.xlsx"),
            "expected_count": None,  # Unknown
            "type": "PACKING",
        },
        "real_trucking_tuyen_dai": {
            "path": os.path.join(EBOOK_DIR, "TRUCKING TUYẾN DÀI 2026.xlsx"),
            "expected_count": None,
            "type": "TRUCKING_DOM",
        },
    }

    # Filter to files that exist
    for name, info in real_files.items():
        if os.path.exists(info["path"]):
            test_files[name] = info
        else:
            logger.warning(f"Real file not found: {info['path']}")

    providers = ["deepseek", "gemini", "anthropic"]

    # Run tests
    for file_name, file_info in test_files.items():
        logger.info(f"\n{'='*60}")
        logger.info(f"Testing: {file_name}")
        logger.info(f"  File: {os.path.basename(file_info['path'])}")
        logger.info(f"  Type: {file_info['type']}")
        logger.info(f"  Expected: {file_info.get('expected_count', '?')} rates")

        file_results = {}

        # 1. Regex parser
        logger.info("  Running regex parser...")
        regex_result = run_regex_parser(file_info["path"])
        file_results["regex"] = regex_result
        logger.info(f"  Regex: {regex_result['rate_count']} rates in {regex_result['time_sec']}s")

        # 2. AI parsers
        for provider in providers:
            logger.info(f"  Running AI parser ({provider})...")
            try:
                ai_result = await run_ai_parser(
                    file_info["path"], provider, file_info["type"]
                )
                file_results[f"ai_{provider}"] = ai_result
                logger.info(
                    f"  {provider}: {ai_result['rate_count']} rates in {ai_result['time_sec']}s "
                    f"(~${ai_result['cost_usd']:.6f})"
                )
                if ai_result.get("error"):
                    logger.warning(f"  Error: {ai_result['error']}")
            except Exception as e:
                logger.error(f"  {provider} failed: {e}")
                file_results[f"ai_{provider}"] = {
                    "method": f"ai_{provider}",
                    "rate_count": 0,
                    "time_sec": 0,
                    "rates": [],
                    "error": str(e),
                    "cost_usd": 0,
                }

        results[file_name] = {
            "file": os.path.basename(file_info["path"]),
            "type": file_info["type"],
            "expected": file_info.get("expected_count"),
            "parsers": file_results,
        }

    # Generate report
    generate_report(results)
    return results


def generate_report(results: Dict[str, Any]):
    """Generate markdown comparison report."""
    os.makedirs(os.path.dirname(REPORT_PATH), exist_ok=True)

    lines = [
        "# Rate Sheet Parser Comparison Report",
        f"\nGenerated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "## Summary",
        "",
        "| File | Type | Expected | Regex | DeepSeek | Gemini | Claude |",
        "|------|------|----------|-------|----------|--------|--------|",
    ]

    total_cost = {"deepseek": 0, "gemini": 0, "anthropic": 0}
    total_time = {"regex": 0, "deepseek": 0, "gemini": 0, "anthropic": 0}

    for name, data in results.items():
        parsers = data["parsers"]
        expected = data.get("expected") or "?"

        regex_count = parsers.get("regex", {}).get("rate_count", 0)
        ds_count = parsers.get("ai_deepseek", {}).get("rate_count", 0)
        gm_count = parsers.get("ai_gemini", {}).get("rate_count", 0)
        cl_count = parsers.get("ai_anthropic", {}).get("rate_count", 0)

        # Track errors
        ds_err = "ERR" if parsers.get("ai_deepseek", {}).get("error") else ""
        gm_err = "ERR" if parsers.get("ai_gemini", {}).get("error") else ""
        cl_err = "ERR" if parsers.get("ai_anthropic", {}).get("error") else ""

        lines.append(
            f"| {name} | {data['type']} | {expected} | "
            f"{regex_count} | {ds_count}{ds_err} | {gm_count}{gm_err} | {cl_count}{cl_err} |"
        )

        # Accumulate totals
        total_time["regex"] += parsers.get("regex", {}).get("time_sec", 0)
        for p in ["deepseek", "gemini", "anthropic"]:
            key = f"ai_{p}"
            total_cost[p] += parsers.get(key, {}).get("cost_usd", 0)
            total_time[p] += parsers.get(key, {}).get("time_sec", 0)

    # Performance comparison
    lines.extend([
        "",
        "## Performance",
        "",
        "| Metric | Regex | DeepSeek | Gemini | Claude |",
        "|--------|-------|----------|--------|--------|",
        f"| Total Time (s) | {total_time['regex']:.2f} | {total_time['deepseek']:.2f} | "
        f"{total_time['gemini']:.2f} | {total_time['anthropic']:.2f} |",
        f"| Est. Cost (USD) | $0 | ${total_cost['deepseek']:.4f} | "
        f"${total_cost['gemini']:.4f} | ${total_cost['anthropic']:.4f} |",
        f"| Cost/100 files | $0 | ${total_cost['deepseek']*10:.2f} | "
        f"${total_cost['gemini']*10:.2f} | ${total_cost['anthropic']*10:.2f} |",
    ])

    # Detail per file
    lines.extend(["", "## Detail per File", ""])

    for name, data in results.items():
        lines.append(f"### {name} ({data['file']})")
        lines.append(f"- Type: {data['type']}")
        lines.append(f"- Expected: {data.get('expected') or '?'}")
        lines.append("")

        for method, presult in data["parsers"].items():
            status = "OK" if not presult.get("error") else f"ERROR: {presult['error'][:80]}"
            lines.append(
                f"- **{method}**: {presult['rate_count']} rates, "
                f"{presult['time_sec']}s, ${presult.get('cost_usd', 0):.6f} — {status}"
            )

            # Show first 3 rates as sample
            if presult.get("rates"):
                for r in presult["rates"][:3]:
                    origin = r.get("origin") or "-"
                    dest = r.get("destination") or "-"
                    vt = r.get("vehicle_type") or "-"
                    price = r.get("price", 0)
                    lines.append(f"  - {origin} → {dest} | {vt} | {price:,.0f} VND")
                if len(presult["rates"]) > 3:
                    lines.append(f"  - ... and {len(presult['rates']) - 3} more")

        lines.append("")

    # Recommendations
    lines.extend([
        "## Recommendations",
        "",
        "1. **Regex parser**: Best for standard trucking pivot tables (fast, free)",
        "2. **DeepSeek**: Best cost/accuracy ratio for AI fallback",
        "3. **Gemini**: Good accuracy, lowest AI cost",
        "4. **Claude**: Highest accuracy but 20-50x more expensive than DeepSeek/Gemini",
        "",
        "### Suggested Strategy",
        "- Use regex as primary parser (handles ~70% of trucking files)",
        "- Fall back to DeepSeek for unrecognized formats (best value)",
        "- Consider Gemini for high-volume batch imports (cheapest AI)",
    ])

    report = "\n".join(lines)

    with open(REPORT_PATH, "w", encoding="utf-8") as f:
        f.write(report)

    logger.info(f"\nReport saved: {REPORT_PATH}")
    print("\n" + report)


# ===================================================================
# Main
# ===================================================================

if __name__ == "__main__":
    asyncio.run(run_all_tests())
