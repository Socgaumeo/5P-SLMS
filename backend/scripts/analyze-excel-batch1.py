#!/usr/bin/env python3
"""Analyze structure of Excel files - Batch 1 (12 files)."""

import openpyxl
import os
import sys

FILES = [
    "/Users/bear1108/Documents/Tháng 3/DAINESE/BẢNG KÊ PHÍ CO DAINESE T3.2026. 5P.xlsx",
    "/Users/bear1108/Documents/Tháng 3/DAINESE/Bảng kê nhập tháng 3.2026.sea.air.xlsx",
    "/Users/bear1108/Documents/Tháng 3/DAINESE/Bảng kê tháng 3.2026. tc.nhap cpn.xlsx",
    "/Users/bear1108/Documents/Tháng 3/DAINESE/Copy of (DAINESE-5PVN) BẢNG KÊ TT T3.2026 bs2.xlsx",
    "/Users/bear1108/Documents/Tháng 3/DAINESE/Copy of Bảng kê xuất tháng 3.2026 final.xlsx",
    "/Users/bear1108/Documents/Tháng 3/DONSUNG/BẢNG KÊ DỊCH VỤ KHO T3.2026 DONGSUNGrev.xlsx",
    "/Users/bear1108/Documents/Tháng 3/GANG THÉP TN/BangKe_GangThep_T3_2026_v10.xlsx",
    "/Users/bear1108/Documents/Tháng 3/GANG THÉP TN/Debit_GangThep_CPN_T3_2026.xlsx",
    "/Users/bear1108/Documents/Tháng 3/GLOREX/Debit 5PVN_GLOREX 3.2026.QUỐC TẾ.xlsx",
    "/Users/bear1108/Documents/Tháng 3/GLOREX/Debit 5PVN_GLOREX T3.2026 TẠI CHỖ.xlsx",
    "/Users/bear1108/Documents/Tháng 3/GLOREX/Debit_TCH_5PVN_GLOBAL_T3_2026_full (4) - Copy.xlsx",
    "/Users/bear1108/Documents/Tháng 3/GLOREX/Debit_TCH_5PVN_GLOREX_T3_2026_full (4).xlsx",
]


def get_cell_value(cell):
    """Get cell value, handling merged cells."""
    if cell.value is not None:
        return cell.value
    return None


def analyze_sheet(ws):
    """Analyze a single worksheet."""
    info = {
        "dimensions": ws.dimensions,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "merged_cells": str(ws.merged_cells.ranges) if ws.merged_cells.ranges else "None",
    }

    # Read first 15 rows to understand headers/layout
    rows_data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(15, ws.max_row), values_only=False), start=1):
        row_vals = []
        for cell in row:
            val = get_cell_value(cell)
            if val is not None:
                row_vals.append(f"[{cell.column_letter}{cell.row}]={val}")
        if row_vals:
            rows_data.append(f"  Row {row_idx}: {' | '.join(row_vals)}")

    info["first_rows"] = rows_data

    # Count non-empty data rows (skip first 5 rows as potential headers)
    data_rows = 0
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
        if any(v is not None for v in row):
            data_rows += 1
    info["data_rows_after_row5"] = data_rows

    # Check last 5 rows for totals/summaries
    last_rows = []
    start = max(1, ws.max_row - 4)
    for row in ws.iter_rows(min_row=start, max_row=ws.max_row, values_only=False):
        row_vals = []
        for cell in row:
            val = get_cell_value(cell)
            if val is not None:
                row_vals.append(f"[{cell.column_letter}{cell.row}]={val}")
        if row_vals:
            last_rows.append(f"  {' | '.join(row_vals)}")
    info["last_rows"] = last_rows

    return info


def analyze_file(filepath):
    """Analyze a single Excel file."""
    basename = os.path.basename(filepath)
    print(f"\n{'='*100}")
    print(f"FILE: {basename}")
    print(f"PATH: {filepath}")
    print(f"{'='*100}")

    if not os.path.exists(filepath):
        print("  *** FILE NOT FOUND ***")
        return

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=False)
    except Exception as e:
        print(f"  *** ERROR LOADING: {e} ***")
        return

    print(f"Sheet count: {len(wb.sheetnames)}")
    print(f"Sheet names: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n  --- SHEET: '{sheet_name}' ---")

        info = analyze_sheet(ws)
        print(f"  Dimensions: {info['dimensions']}, Max row: {info['max_row']}, Max col: {info['max_col']}")
        print(f"  Merged cells: {info['merged_cells']}")
        print(f"  Data rows (after row 5): {info['data_rows_after_row5']}")

        print(f"\n  FIRST 15 ROWS (header/layout):")
        for r in info["first_rows"]:
            # Truncate long values for readability
            if len(r) > 300:
                r = r[:300] + "..."
            print(f"    {r}")

        print(f"\n  LAST 5 ROWS (totals/summary):")
        for r in info["last_rows"]:
            if len(r) > 300:
                r = r[:300] + "..."
            print(f"    {r}")

    wb.close()


def main():
    print("=" * 100)
    print("EXCEL FILE STRUCTURE ANALYSIS - BATCH 1 (12 files)")
    print("=" * 100)

    for i, filepath in enumerate(FILES, 1):
        print(f"\n\n>>> ANALYZING FILE {i}/{len(FILES)} <<<")
        analyze_file(filepath)

    print("\n\n" + "=" * 100)
    print("ANALYSIS COMPLETE")
    print("=" * 100)


if __name__ == "__main__":
    main()
