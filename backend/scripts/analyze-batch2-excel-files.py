#!/usr/bin/env python3
"""Analyze structure of Batch 2 Excel files (11 files) for cost reconciliation."""

import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed")
    sys.exit(1)

FILES = [
    "/Users/bear1108/Documents/Tháng 3/HƯNG PHÁT/Debit note HƯNG PHÁT-5P T3.2026 - L2.xlsx",
    "/Users/bear1108/Documents/Tháng 3/HƯNG PHÁT/Debit note HƯNG PHÁT-5P T3.2026.xlsx",
    "/Users/bear1108/Documents/Tháng 3/KCVN/BangKe_KCIL_T3_2026_v14.xlsx",
    "/Users/bear1108/Documents/Tháng 3/KK/Debit Note.KK.MAR.2026. org.xlsx",
    "/Users/bear1108/Documents/Tháng 3/KWE/5P in MAR.2026. KWE rev.xlsx",
    "/Users/bear1108/Documents/Tháng 3/LAS/DebitNote_LGZHPH260781_LAS_DRAFT (13).xlsx",
    "/Users/bear1108/Documents/Tháng 3/LKV BD/Debit note SX LỌC KHÍ VIỆT BD-5P T3.2026 REV1.xlsx",
    "/Users/bear1108/Documents/Tháng 3/LKV MB/Debit note SX LỌC KHÍ VIỆT MIỀN BẮC-5P T3.2026  rev14.xlsx",
    "/Users/bear1108/Documents/Tháng 3/LOGIMARK/Debit_LOGIMARK_T3_2026 (2).xlsx",
    "/Users/bear1108/Documents/Tháng 3/LOGIMARK/Debit_LOGIMARK_T3_2026_updated (3).xlsx",
    "/Users/bear1108/Documents/Tháng 3/MESSER/Bảng kê MESSER 5P T3.2026.xlsx",
]


def get_cell_value(cell):
    """Get cell value, handling merged cells."""
    if cell.value is not None:
        return cell.value
    return None


def analyze_sheet(ws, sheet_name):
    """Analyze a single worksheet."""
    print(f"\n  --- Sheet: '{sheet_name}' ---")
    print(f"  Dimensions: {ws.dimensions}")
    print(f"  Max row: {ws.max_row}, Max col: {ws.max_column}")

    # Print merged cells info
    if ws.merged_cells.ranges:
        print(f"  Merged cells: {len(ws.merged_cells.ranges)} ranges")
        for mc in list(ws.merged_cells.ranges)[:10]:
            print(f"    {mc}")
        if len(ws.merged_cells.ranges) > 10:
            print(f"    ... and {len(ws.merged_cells.ranges) - 10} more")

    # Print first 25 rows to understand layout
    print(f"\n  First rows (up to 25):")
    rows_printed = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(25, ws.max_row), values_only=False), 1):
        values = []
        for cell in row:
            v = get_cell_value(cell)
            if v is not None:
                # Truncate long strings
                s = str(v).strip()
                if len(s) > 60:
                    s = s[:57] + "..."
                values.append(f"{cell.column_letter}{cell.row}={s}")
        if values:
            print(f"    Row {row_idx}: {' | '.join(values)}")
            rows_printed += 1

    # Print last 5 rows to see totals
    if ws.max_row > 25:
        print(f"\n  Last 5 rows (rows {ws.max_row-4} to {ws.max_row}):")
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=max(26, ws.max_row - 4), max_row=ws.max_row, values_only=False),
            max(26, ws.max_row - 4)
        ):
            values = []
            for cell in row:
                v = get_cell_value(cell)
                if v is not None:
                    s = str(v).strip()
                    if len(s) > 60:
                        s = s[:57] + "..."
                    values.append(f"{cell.column_letter}{cell.row}={s}")
            if values:
                print(f"    Row {row_idx}: {' | '.join(values)}")

    # Look for key patterns in all cells
    print(f"\n  Key field scan (searching all cells for keywords):")
    keywords = {
        'vat': [], 'thuế': [], 'tax': [],
        'tổng': [], 'total': [], 'cộng': [],
        'tờ khai': [], 'cd': [], 'customs': [],
        'bl': [], 'awb': [], 'bill': [],
        'bks': [], 'biển số': [], 'xe': [],
        'invoice': [], 'hóa đơn': [],
        'tuyến': [], 'route': [], 'chặng': [],
        'thu hộ': [], 'chi hộ': [],
        'phí': [], 'cước': [],
    }
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), values_only=False):
        for cell in row:
            v = get_cell_value(cell)
            if v is not None:
                s = str(v).lower().strip()
                for kw, locations in keywords.items():
                    if kw in s and len(locations) < 3:
                        locations.append(f"{cell.column_letter}{cell.row}='{str(v).strip()[:80]}'")

    for kw, locations in keywords.items():
        if locations:
            print(f"    '{kw}': {locations}")


def analyze_file(filepath):
    """Analyze a single Excel file."""
    fname = os.path.basename(filepath)
    print(f"\n{'='*100}")
    print(f"FILE: {fname}")
    print(f"PATH: {filepath}")
    print(f"{'='*100}")

    if not os.path.exists(filepath):
        print("  *** FILE NOT FOUND ***")
        return

    fsize = os.path.getsize(filepath)
    print(f"  File size: {fsize:,} bytes")

    try:
        wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True)
        print(f"  Sheet names: {wb.sheetnames}")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            analyze_sheet(ws, sheet_name)

        wb.close()
    except Exception as e:
        print(f"  ERROR reading file: {e}")
        import traceback
        traceback.print_exc()


def main():
    print("=" * 100)
    print("BATCH 2 - EXCEL FILE STRUCTURE ANALYSIS")
    print("=" * 100)

    for i, fpath in enumerate(FILES, 1):
        print(f"\n\n{'#'*100}")
        print(f"# FILE {i}/{len(FILES)}")
        print(f"{'#'*100}")
        analyze_file(fpath)

    print("\n\n" + "=" * 100)
    print("ANALYSIS COMPLETE")
    print("=" * 100)


if __name__ == "__main__":
    main()
