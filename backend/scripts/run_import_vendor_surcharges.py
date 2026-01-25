#!/usr/bin/env python3
"""
Script import vendor_surcharges từ Excel lên Supabase
Giả định bảng vendor_surcharges đã được tạo với đúng cấu trúc
"""

import os
import sys
import json
import re
from datetime import datetime
from typing import List, Dict, Any, Optional
from dotenv import load_dotenv
from supabase import create_client, Client
from openpyxl import load_workbook
import google.generativeai as genai

# Load environment variables
load_dotenv()


class VendorSurchargesImporter:
    """Class quản lý import vendor surcharges lên Supabase"""
    
    def __init__(self):
        """Khởi tạo Supabase client và AI model"""
        self.supabase: Client = self._get_supabase_client()
        self._setup_ai()
        self.stats = {
            'success': 0,
            'failed': 0,
            'errors': [],
        }
    
    def _get_supabase_client(self) -> Client:
        """Tạo Supabase client"""
        supabase_url = os.getenv('SUPABASE_URL')
        supabase_key = os.getenv('SUPABASE_SERVICE_ROLE_KEY')
        
        if not supabase_url or not supabase_key:
            raise ValueError("SUPABASE_URL và SUPABASE_SERVICE_ROLE_KEY phải được cấu hình trong .env")
        
        return create_client(supabase_url, supabase_key)
    
    def _setup_ai(self):
        """Setup AI model (Gemini)"""
        api_key = os.getenv('GOOGLE_GEMINI_API_KEY')
        if api_key:
            genai.configure(api_key=api_key)
            self.ai_model = genai.GenerativeModel('gemini-2.0-flash')
            self.ai_available = True
        else:
            print("⚠️ GOOGLE_GEMINI_API_KEY không được cấu hình, AI sẽ không hoạt động")
            self.ai_available = False
    
    def read_excel_sheet(self, file_path: str, sheet_name: str, max_rows: int = 100) -> str:
        """
        Đọc sheet từ Excel và chuyển thành text
        
        Args:
            file_path: Đường dẫn file Excel
            sheet_name: Tên sheet
            max_rows: Số dòng tối đa để đọc
        
        Returns:
            String chứa nội dung sheet
        """
        try:
            wb = load_workbook(file_path, data_only=True)
            sheet = wb[sheet_name]
            
            content = []
            for i, row in enumerate(sheet.iter_rows(max_row=max_rows), 1):
                row_text = " | ".join([str(cell.value)[:50] if cell.value else "" for cell in row])
                if row_text.strip(" |"):
                    content.append(f"Row {i}: {row_text}")
            
            return "\n".join(content)
            
        except Exception as e:
            print(f"❌ Lỗi khi đọc Excel: {str(e)}")
            return ""
    
    def parse_surcharges_with_ai(self, content: str) -> List[Dict[str, Any]]:
        """
        Sử dụng AI để parse surcharges từ Excel content
        
        Args:
            content: Nội dung Excel dạng text
        
        Returns:
            List của surcharge dictionaries
        """
        if not self.ai_available:
            print("⚠️ AI không khả dụng, không thể parse surcharges")
            return []
        
        example_json = '''[
  {"surcharge_code": "PHI_CAU_DUONG", "description": "Phí cầu đường", "amount": 50000, "unit": "lượt", "conditions": "Áp dụng cho các tuyến qua cầu"}
]'''
        
        prompt = f"""Analyze this Excel data containing Vietnamese trucking surcharges (phụ phí).

Extract surcharge information into a JSON array. Each surcharge should have:
- surcharge_code: Mã phụ phí (e.g. "PHI_CAU_DUONG", "PHI_CHO", "PHI_DO")
- description: Mô tả chi tiết
- amount: Số tiền (number only, in VND)
- unit: Đơn vị tính (e.g. "lượt", "km", "giờ", "ngày")
- conditions: Điều kiện áp dụng (nếu có)

Data:
```
{content}
```

IMPORTANT:
- Return ONLY valid JSON array
- Amount = number (remove commas/dots in thousands)
- Include ALL surcharge types found
- Skip header rows and notes

Example output:
{example_json}
"""
        
        try:
            response = self.ai_model.generate_content(prompt)
            response_text = response.text.strip()
            
            # Clean up response
            if "```json" in response_text:
                response_text = response_text.split("```json")[1].split("```")[0]
            elif "```" in response_text:
                lines = response_text.split("\n")
                start = next((i for i, l in enumerate(lines) if l.startswith("```")), 0) + 1
                end = next((i for i, l in enumerate(lines[start:]) if l.startswith("```")), len(lines)) + start
                response_text = "\n".join(lines[start:end])
            
            # Fix common JSON issues
            response_text = re.sub(r',\s*]', ']', response_text)
            response_text = re.sub(r',\s*}', '}', response_text)
            
            return json.loads(response_text)
            
        except Exception as e:
            print(f"⚠️ Lỗi khi parse surcharges với AI: {str(e)}")
            return []
    
    def get_vendor_id(self, vendor_name: str) -> Optional[int]:
        """
        Lấy vendor_id từ tên vendor
        
        Args:
            vendor_name: Tên vendor
        
        Returns:
            vendor_id hoặc None nếu không tìm thấy
        """
        try:
            result = self.supabase.table('vendors').select('vendor_id').ilike('company_name', f'%{vendor_name}%').or_(f"short_name.ilike.%{vendor_name}%").execute()
            
            if result.data:
                return result.data[0]['vendor_id']
            
            return None
            
        except Exception as e:
            print(f"❌ Lỗi khi tìm vendor: {str(e)}")
            return None
    
    def import_surcharges(self, file_path: str, vendor_name: str):
        """
        Import vendor surcharges từ Excel lên Supabase
        
        Args:
            file_path: Đường dẫn file Excel
            vendor_name: Tên vendor
        """
        print("\n" + "="*60)
        print(f"IMPORT VENDOR SURCHARGES - {vendor_name}")
        print("="*60)
        
        try:
            # Lấy vendor_id
            vendor_id = self.get_vendor_id(vendor_name)
            if not vendor_id:
                print(f"❌ Không tìm thấy vendor '{vendor_name}'")
                return
            
            print(f"✅ Vendor ID: {vendor_id}")
            
            # Đọc Excel
            wb = load_workbook(file_path, data_only=True)
            print(f"📑 Found sheets: {wb.sheetnames}")
            
            total_imported = 0
            
            for sheet_name in wb.sheetnames:
                # Chỉ xử lý sheet phụ phí
                if 'Phụ Phí' not in sheet_name:
                    continue
                
                print(f"\n🔄 Processing: {sheet_name}")
                
                # Đọc nội dung sheet
                content = self.read_excel_sheet(file_path, sheet_name)
                if not content:
                    print(f"⚠️ Không thể đọc sheet: {sheet_name}")
                    continue
                
                # Parse với AI
                print("🤖 Parsing with AI...")
                surcharges = self.parse_surcharges_with_ai(content)
                print(f"✅ Extracted {len(surcharges)} surcharge entries")
                
                # Import từng surcharge
                for i, surcharge in enumerate(surcharges, 1):
                    try:
                        # Tạo record
                        surcharge_data = {
                            'vendor_id': vendor_id,
                            'surcharge_code': surcharge.get('surcharge_code'),
                            'description': surcharge.get('description'),
                            'amount': surcharge.get('amount'),
                            'unit': surcharge.get('unit'),
                            'conditions': surcharge.get('conditions'),
                            'effective_date': datetime.now().date().isoformat()
                        }
                        
                        # Kiểm tra xem surcharge đã tồn tại chưa
                        existing = self.supabase.table('vendor_surcharges').select('id').eq('vendor_id', vendor_id).eq('surcharge_code', surcharge.get('surcharge_code')).execute()
                        
                        if existing.data:
                            # Update
                            self.supabase.table('vendor_surcharges').update(surcharge_data).eq('id', existing.data[0]['id']).execute()
                            print(f"  [{i}] Updated: {surcharge.get('surcharge_code')} | {surcharge.get('amount'):,} VND/{surcharge.get('unit')}")
                        else:
                            # Insert
                            self.supabase.table('vendor_surcharges').insert(surcharge_data).execute()
                            print(f"  [{i}] Inserted: {surcharge.get('surcharge_code')} | {surcharge.get('amount'):,} VND/{surcharge.get('unit')}")
                        
                        total_imported += 1
                        self.stats['success'] += 1
                    
                    except Exception as e:
                        error_msg = f"Error importing surcharge: {str(e)}"
                        print(f"  ❌ {error_msg}")
                        self.stats['failed'] += 1
                        self.stats['errors'].append(error_msg)
            
            print(f"\n✓ Import surcharges hoàn tất: {total_imported} surcharges")
            
        except Exception as e:
            error_msg = f"Lỗi khi import surcharges: {str(e)}"
            print(f"❌ {error_msg}")
            self.stats['errors'].append(error_msg)
    
    def print_summary(self):
        """In tổng kết import"""
        print("\n" + "="*60)
        print("TỔNG KẾT IMPORT VENDOR SURCHARGES")
        print("="*60)
        
        print(f"\n✓ Thành công: {self.stats['success']}")
        print(f"✗ Thất bại: {self.stats['failed']}")
        
        if self.stats['errors']:
            print(f"\nLỗi:")
            for error in self.stats['errors'][:5]:
                print(f"    - {error}")
            if len(self.stats['errors']) > 5:
                print(f"    ... và {len(self.stats['errors']) - 5} lỗi khác")
        
        print("\n" + "="*60)


def main():
    """Hàm chính"""
    try:
        if len(sys.argv) < 2:
            print("Usage: python run_import_vendor_surcharges.py <file_path> [vendor_name]")
            print("Example: python run_import_vendor_surcharges.py vendor_rates/Tam_bảo_092025.xlsx 'Tam Bảo'")
            print("\n⚠️  Lưu ý: Bảng vendor_surcharges phải được tạo trước với cấu trúc đúng!")
            print("Nếu gặp lỗi về cột 'amount', hãy chạy SQL sau trên Supabase Dashboard:")
            print("https://app.supabase.com/project/vpmsytbbsxmtdicnkytv")
            print("\nSQL cần chạy:")
            print("""
-- Xóa bảng phụ thuộc trước (nếu có)
DROP TABLE IF EXISTS public.vendor_surcharge_prices CASCADE;

-- Xóa bảng cũ
DROP TABLE IF EXISTS public.vendor_surcharges CASCADE;

-- Tạo bảng mới
CREATE TABLE public.vendor_surcharges (
    id SERIAL PRIMARY KEY,
    vendor_id INTEGER NOT NULL REFERENCES vendors(vendor_id) ON DELETE CASCADE,
    surcharge_code VARCHAR(50) NOT NULL,
    description TEXT,
    amount DECIMAL(18,2),
    unit VARCHAR(20),
    conditions TEXT,
    effective_date DATE DEFAULT CURRENT_DATE,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);

-- Index
CREATE INDEX idx_vendor_surcharges_vendor_id ON public.vendor_surcharges(vendor_id);
CREATE INDEX idx_vendor_surcharges_surcharge_code ON public.vendor_surcharges(surcharge_code);
""")
            sys.exit(1)
        
        file_path = sys.argv[1]
        vendor_name = sys.argv[2] if len(sys.argv) > 2 else "Tam Bảo"
        
        importer = VendorSurchargesImporter()
        
        # Import surcharges
        importer.import_surcharges(file_path, vendor_name)
        
        # In tổng kết
        importer.print_summary()
        
        # Exit code dựa trên kết quả
        sys.exit(0 if importer.stats['failed'] == 0 else 1)
        
    except Exception as e:
        print(f"\nFATAL ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
