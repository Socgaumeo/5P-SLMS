"""
Import March 2026 jobs from Excel debit notes into Supabase PostgreSQL.
Main file: helpers, DB insert logic, orchestration, simple parsers.
Complex parsers: import-march-2026-complex-parsers.py
"""
import os
import json
import datetime
import traceback

import psycopg2
import openpyxl
import xlrd

# ──────── CONFIG ────────
DB_URL = "postgresql://postgres.ooixntyflwmjaryxwakx:%21%40kHanh0112@aws-1-ap-southeast-1.pooler.supabase.com:6543/postgres"
BASE_DIR = "/Users/bear1108/Documents/Tháng 3"

CUSTOMER_MAP = {
    "DAINESE": 46, "DONSUNG": 58, "GANG THÉP TN": 44, "GLOREX": 18,
    "HƯNG PHÁT": 63, "KCVN": 61, "KK": 65, "KWE": 28, "LAS": 6,
    "LKV BD": 60, "LKV MB": 53, "LOGIMARK": 31, "MESSER": 22,
    "NIPPON": 64, "TDI": 20, "THÁI HOÀ": 45, "TVC": 59,
    "UTRACORN": 56, "VINTECH": 57, "XÂY LẮP VN": 2,
}

# ──────── HELPERS ────────

def s(v):
    """Safe string, strip newlines, truncate to 200 chars."""
    if v is None:
        return ""
    return str(v).strip().replace("\n", " ").replace("\r", "")[:200]


def n(v):
    """Safe numeric conversion; returns float."""
    if v is None:
        return 0.0
    if isinstance(v, bool):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0


def d(v, datemode=0):
    """Parse date from datetime, date, Excel serial, or string."""
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, (int, float)) and 40000 < v < 60000:
        try:
            return xlrd.xldate_as_datetime(v, datemode).date()
        except Exception:
            pass
    if isinstance(v, str):
        v = v.strip().split(" ")[0]  # take date part only
        for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"]:
            try:
                return datetime.datetime.strptime(v, fmt).date()
            except ValueError:
                continue
    return None


def cell(ws, r, c):
    """Get openpyxl cell value."""
    return ws.cell(row=r, column=c).value




# ──────── DB INSERTER ────────

def insert_jobs(conn, jobs_list):
    """
    Insert all jobs from jobs_list into DB.
    Each job_data dict:
      customer_id, date, description, svc_type, cd_no, bl_awb, invoice,
      origin, dest, route, weight, customs_type, customs_port, service_details,
      costs: list of {name, amount, vat_rate, is_reimbursement, invoice}
    Returns (inserted, errors) counts.
    """
    cur = conn.cursor()
    # Build per-customer sequence counters for job_no
    seq = {}
    inserted = 0
    errors = 0

    for job in jobs_list:
        cid = job["customer_id"]
        seq[cid] = seq.get(cid, 0) + 1
        job_no = f"{cid}-2503-{seq[cid]:03d}"

        date_val = job.get("date") or datetime.date(2026, 3, 1)
        svc_type = job.get("svc_type", "TRUCKING_DOM")

        # Build service_details JSONB summary
        costs = job.get("costs", [])
        svc_fees = [c for c in costs if not c.get("is_reimbursement")]
        reimbursements = [c for c in costs if c.get("is_reimbursement")]
        selling_price = sum(n(c["amount"]) for c in svc_fees)
        reimb_total = sum(n(c["amount"]) for c in reimbursements)
        # VAT on service fees (use 8% default)
        vat_rate_pct = svc_fees[0].get("vat_rate", 8) if svc_fees else 8
        vat_amount = sum(n(c["amount"]) * (n(c.get("vat_rate", 8)) / 100) for c in svc_fees)

        sd_extra = job.get("service_details", {})
        svc_details = {
            "selling_price": selling_price,
            "vat_rate": vat_rate_pct,
            "vat_amount": round(vat_amount, 2),
            "total_revenue": round(selling_price + vat_amount, 2),
            "reimbursement_total": reimb_total,
            "grand_total": round(selling_price + vat_amount + reimb_total, 2),
        }
        svc_details.update(sd_extra)

        # Route string
        origin = s(job.get("origin", ""))
        dest = s(job.get("dest", ""))
        route = s(job.get("route", ""))
        if not route and origin and dest:
            route = f"{origin} → {dest}"
        elif not route and origin:
            route = origin

        # Invoice numbers array
        inv_str = s(job.get("invoice", ""))
        inv_arr = [x.strip() for x in inv_str.replace(",", ";").split(";") if x.strip()] if inv_str else []

        try:
            # 1. Insert job
            cur.execute("""
                INSERT INTO jobs (job_no, customer_id, description, etd, eta,
                                  status_code, invoice_number, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, 'COMPLETED', %s, NOW(), NOW())
                RETURNING job_id
            """, (
                job_no, cid,
                s(job.get("description", svc_type))[:500],
                date_val, date_val,
                inv_arr[0] if inv_arr else None,
            ))
            job_id = cur.fetchone()[0]

            # 2. Insert job_service
            cur.execute("""
                INSERT INTO job_services (
                    job_id, service_type_code, scheduled_date, status_code,
                    origin_address, dest_address, route,
                    cd_no, bl_awb_no, customs_type, customs_port,
                    invoice_numbers, weight_kg,
                    service_details
                ) VALUES (%s, %s, %s, 'COMPLETED',
                          %s, %s, %s,
                          %s, %s, %s, %s,
                          %s, %s, %s)
                RETURNING svc_id
            """, (
                job_id, svc_type, date_val,
                origin or None, dest or None, route or None,
                s(job.get("cd_no", "")) or None,
                s(job.get("bl_awb", "")) or None,
                s(job.get("customs_type", "")) or None,
                s(job.get("customs_port", "")) or None,
                inv_arr if inv_arr else None,
                n(job.get("weight", 0)) or None,
                json.dumps(svc_details),
            ))
            svc_id = cur.fetchone()[0]

            # 3. Insert job_costs (one row per cost line)
            for cost in costs:
                amt = n(cost["amount"])
                if amt == 0 and not cost.get("is_reimbursement"):
                    continue  # skip zero service fees
                vat_r = n(cost.get("vat_rate", 8))
                is_reimb = bool(cost.get("is_reimbursement", False))
                cur.execute("""
                    INSERT INTO job_costs (
                        job_id, svc_id, cost_name, quantity, unit,
                        buying_rate, selling_rate, vat_rate, is_reimbursement
                    ) VALUES (%s, %s, %s, 1, 'TRIP', 0, %s, %s, %s)
                """, (
                    job_id, svc_id,
                    s(cost["name"])[:200],
                    amt, vat_r, is_reimb,
                ))

            conn.commit()
            inserted += 1

        except Exception as e:
            conn.rollback()
            print(f"  ERROR inserting {job_no} ({job.get('description','')[:60]}): {e}")
            errors += 1

    cur.close()
    return inserted, errors


# ──────── SIMPLE PARSERS ────────

def parse_lkv_bd():
    """LKV BD trucking: STT|Ngày|BKS|Loại xe|Dịch vụ|Đơn giá|SL|ĐVT|Thành tiền|Tax|Tổng|HĐ"""
    fp = f"{BASE_DIR}/LKV BD/Debit note SX LỌC KHÍ VIỆT BD-5P T3.2026 REV1.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb["LKV BẢNG KÊ CHI TIẾT"]
    jobs = []
    for r in range(14, ws.max_row + 1):
        stt = cell(ws, r, 1)
        if isinstance(stt, str) and "tổng" in stt.lower():
            break
        if not isinstance(stt, (int, float)):
            continue
        date_val = d(cell(ws, r, 2))
        bks = s(cell(ws, r, 3))
        vehicle = s(cell(ws, r, 4))
        desc = s(cell(ws, r, 5))
        pre_vat = n(cell(ws, r, 9))   # C9 = Thành tiền
        inv = s(cell(ws, r, 12))
        jobs.append({
            "customer_id": CUSTOMER_MAP["LKV BD"],
            "date": date_val or datetime.date(2026, 3, 1),
            "description": f"{desc} - BKS: {bks}",
            "svc_type": "TRUCKING_DOM",
            "origin": "", "dest": "",
            "route": desc,
            "bl_awb": bks,
            "invoice": inv,
            "service_details": {"vehicle_type": vehicle, "vehicle_plate": bks},
            "costs": [{"name": desc or "Cước vận chuyển", "amount": pre_vat, "vat_rate": 8}],
        })
    return jobs


def parse_lkv_mb():
    """LKV MB trucking: same format as LKV BD."""
    import glob
    folder = f"{BASE_DIR}/LKV MB"
    files = [f for f in os.listdir(folder) if f.endswith(".xlsx")]
    if not files:
        return []
    fp = f"{folder}/{files[0]}"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb["LKV BẢNG KÊ CHI TIẾT"]
    jobs = []
    for r in range(14, ws.max_row + 1):
        stt = cell(ws, r, 1)
        if isinstance(stt, str) and "tổng" in stt.lower():
            break
        if not isinstance(stt, (int, float)):
            continue
        date_val = d(cell(ws, r, 2))
        bks = s(cell(ws, r, 3))
        vehicle = s(cell(ws, r, 4))
        desc = s(cell(ws, r, 5))
        pre_vat = n(cell(ws, r, 9))
        inv = s(cell(ws, r, 12))
        jobs.append({
            "customer_id": CUSTOMER_MAP["LKV MB"],
            "date": date_val or datetime.date(2026, 3, 1),
            "description": f"{desc} - BKS: {bks}",
            "svc_type": "TRUCKING_DOM",
            "origin": "", "dest": "",
            "route": desc,
            "bl_awb": bks,
            "invoice": inv,
            "service_details": {"vehicle_type": vehicle, "vehicle_plate": bks},
            "costs": [{"name": desc or "Cước vận chuyển", "amount": pre_vat, "vat_rate": 8}],
        })
    return jobs


def parse_hung_phat():
    """Hưng Phát trucking: 2 files, 1 job each. Format: same as LKV."""
    files = [
        f"{BASE_DIR}/HƯNG PHÁT/Debit note HƯNG PHÁT-5P T3.2026.xlsx",
        f"{BASE_DIR}/HƯNG PHÁT/Debit note HƯNG PHÁT-5P T3.2026 - L2.xlsx",
    ]
    jobs = []
    for fp in files:
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb["BẢNG KÊ CHI TIẾT"]
        for r in range(14, ws.max_row + 1):
            stt = cell(ws, r, 1)
            if isinstance(stt, str) and "tổng" in stt.lower():
                break
            if not isinstance(stt, (int, float)):
                continue
            date_val = d(cell(ws, r, 2))
            bks = s(cell(ws, r, 3))
            vehicle = s(cell(ws, r, 4))
            desc = s(cell(ws, r, 5))
            pre_vat = n(cell(ws, r, 9))
            inv = s(cell(ws, r, 12))
            jobs.append({
                "customer_id": CUSTOMER_MAP["HƯNG PHÁT"],
                "date": date_val or datetime.date(2026, 3, 1),
                "description": f"{desc} - BKS: {bks}",
                "svc_type": "TRUCKING_DOM",
                "bl_awb": bks,
                "invoice": inv,
                "service_details": {"vehicle_type": vehicle},
                "costs": [{"name": desc or "Cước vận chuyển", "amount": pre_vat, "vat_rate": 8}],
            })
    return jobs


def parse_utracorn():
    """Utracorn trucking: STT|Ngày|Type|Điểm lấy|Điểm trả|BKS|SL|ĐVT|Đơn giá|Thành tiền|Chi phí khác|Tổng."""
    fp = f"{BASE_DIR}/UTRACORN/DebitNote_UTRACON_TRK1403_DRAFT (3).xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb["TRUCKING"]
    jobs = []
    for r in range(15, ws.max_row + 1):
        stt = cell(ws, r, 1)
        if isinstance(stt, str) and "tổng" in stt.lower():
            break
        if not isinstance(stt, (int, float)):
            continue
        date_val = d(cell(ws, r, 2))
        vtype = s(cell(ws, r, 3))
        origin = s(cell(ws, r, 4))
        dest = s(cell(ws, r, 5))
        bks = s(cell(ws, r, 6))
        pre_vat = n(cell(ws, r, 10))  # C10=Thành tiền
        extra = n(cell(ws, r, 11))    # C11=Chi phí khác
        desc = f"Cước vận chuyển {origin} → {dest}"
        costs = [{"name": "Cước vận chuyển", "amount": pre_vat, "vat_rate": 8}]
        if extra > 0:
            costs.append({"name": "Chi phí khác", "amount": extra, "vat_rate": 8})
        jobs.append({
            "customer_id": CUSTOMER_MAP["UTRACORN"],
            "date": date_val or datetime.date(2026, 3, 14),
            "description": desc,
            "svc_type": "TRUCKING_DOM",
            "origin": origin, "dest": dest,
            "bl_awb": bks,
            "service_details": {"vehicle_type": vtype, "vehicle_plate": bks},
            "costs": costs,
        })
    return jobs


def parse_vintech_trucking():
    """Vintech trucking: LKV BẢNG KÊ CHI TIẾT format."""
    fp = f"{BASE_DIR}/VINTECH/Debit note VINTECH-5P T3.2026.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb["LKV BẢNG KÊ CHI TIẾT"]
    jobs = []
    for r in range(14, ws.max_row + 1):
        stt = cell(ws, r, 1)
        if isinstance(stt, str) and "tổng" in stt.lower():
            break
        if not isinstance(stt, (int, float)):
            continue
        date_val = d(cell(ws, r, 2))
        bks = s(cell(ws, r, 3))
        vehicle = s(cell(ws, r, 4))
        desc = s(cell(ws, r, 5))
        pre_vat = n(cell(ws, r, 9))
        jobs.append({
            "customer_id": CUSTOMER_MAP["VINTECH"],
            "date": date_val or datetime.date(2026, 3, 1),
            "description": f"{desc} - BKS: {bks}",
            "svc_type": "TRUCKING_DOM",
            "bl_awb": bks,
            "service_details": {"vehicle_type": vehicle},
            "costs": [{"name": desc or "Cước vận chuyển", "amount": pre_vat, "vat_rate": 8}],
        })
    return jobs


def parse_vintech_air():
    """Vintech air import: DebitNote_VINTECH_NGB637324_DRAFT.xlsx, sheet 'AI'.
    Sectioned format: international fees (no VAT), VN fees (8%), trả hộ (no VAT)."""
    fp = f"{BASE_DIR}/VINTECH/DebitNote_VINTECH_NGB637324_DRAFT.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb["AI"]
    cid = CUSTOMER_MAP["VINTECH"]

    # Shipment info from header rows
    bl_awb = s(cell(ws, 16, 2))     # NGB637324
    cd_no = s(cell(ws, 16, 10))     # 108057770660
    weight = n(cell(ws, 17, 6))     # 562 kg
    origin = "Shanghai, China"
    dest = "Tân Thới Hiệp, TP. HCM"

    costs = []
    # Parse data rows R21 onwards (skip section headers and totals)
    for r in range(21, ws.max_row + 1):
        stt = cell(ws, r, 1)
        name = s(cell(ws, r, 2))
        vnd_total = n(cell(ws, r, 10))     # C10 = Tổng tiền VND
        vat_col = cell(ws, r, 9)            # C9 = VAT rate or 0

        # Skip section headers, totals, and empty rows
        if not name or vnd_total == 0:
            continue
        if isinstance(stt, str) and ("tổng" in stt.lower() or "total" in stt.lower()):
            continue
        if not isinstance(stt, (int, float)):
            continue

        # Determine VAT and reimbursement status
        vat_rate_val = n(vat_col)
        is_reimb = "trả hộ" in name.lower() or "thu hộ" in name.lower()
        hd = s(cell(ws, r, 11))  # invoice number

        if is_reimb:
            costs.append({"name": f"Thu hộ: {name} - HĐ: {hd}", "amount": vnd_total,
                          "vat_rate": 0, "is_reimbursement": True, "invoice": hd})
        elif vat_rate_val > 0:
            # VN fees: C10 is post-VAT, compute pre-VAT
            pre_vat = round(vnd_total / (1 + vat_rate_val))
            costs.append({"name": name, "amount": pre_vat, "vat_rate": 8, "invoice": hd})
        else:
            # International fees: no VAT, C10 is the amount
            costs.append({"name": name, "amount": vnd_total, "vat_rate": 0, "invoice": hd})

    if costs:
        return [{
            "customer_id": cid,
            "date": datetime.date(2026, 3, 1),
            "description": f"Air import - Bill: {bl_awb} - TK: {cd_no}",
            "svc_type": "AIR_IMP",
            "cd_no": cd_no, "bl_awb": bl_awb,
            "origin": origin, "dest": dest,
            "weight": weight,
            "customs_type": "IMPORT",
            "service_details": {"exchange_rate_usd": 26321, "pieces": 3},
            "costs": costs,
        }]
    return []


def parse_xay_lap_vn():
    """Xây Lắp VN trucking: BẢNG KÊ CHI TIẾT format."""
    fp = f"{BASE_DIR}/XÂY LẮP VN/Debit note XÂY LẮP VN-5P T3.2026.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb["BẢNG KÊ CHI TIẾT"]
    jobs = []
    for r in range(14, ws.max_row + 1):
        stt = cell(ws, r, 1)
        if isinstance(stt, str) and "tổng" in stt.lower():
            break
        if not isinstance(stt, (int, float)):
            continue
        date_val = d(cell(ws, r, 2))
        bks = s(cell(ws, r, 3))
        vehicle = s(cell(ws, r, 4))
        desc = s(cell(ws, r, 5))
        pre_vat = n(cell(ws, r, 9))
        jobs.append({
            "customer_id": CUSTOMER_MAP["XÂY LẮP VN"],
            "date": date_val or datetime.date(2026, 3, 30),
            "description": f"{desc} - BKS: {bks}",
            "svc_type": "TRUCKING_DOM",
            "bl_awb": bks,
            "service_details": {"vehicle_type": vehicle},
            "costs": [{"name": desc or "Cước vận chuyển", "amount": pre_vat, "vat_rate": 8}],
        })
    return jobs


def parse_logimark():
    """Logimark: 2 files, 1 customs declaration each.
    Header R12: STT|Ngày TK|Số TK/Bill|Luồng|Dịch vụ|... SL|Đơn giá|ĐVT|Thành tiền|Tax|Tổng|HĐ
    """
    files = [
        f"{BASE_DIR}/LOGIMARK/Debit_LOGIMARK_T3_2026 (2).xlsx",
        f"{BASE_DIR}/LOGIMARK/Debit_LOGIMARK_T3_2026_updated (3).xlsx",
    ]
    jobs = []
    for fp in files:
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb["LOGIMARK"]
        for r in range(14, ws.max_row + 1):
            stt = cell(ws, r, 1)
            if isinstance(stt, str) and any(k in stt.lower() for k in ["tổng", "total"]):
                break
            if not isinstance(stt, (int, float)):
                continue
            date_val = d(cell(ws, r, 2))
            cd_no = s(cell(ws, r, 3))
            luong = s(cell(ws, r, 4))
            desc = s(cell(ws, r, 5))
            pre_vat = n(cell(ws, r, 9))  # C9=Thành tiền
            inv = s(cell(ws, r, 12))
            jobs.append({
                "customer_id": CUSTOMER_MAP["LOGIMARK"],
                "date": date_val or datetime.date(2026, 3, 12),
                "description": f"{desc} - TK: {cd_no}",
                "svc_type": "CUS_EXPORT",
                "cd_no": cd_no,
                "invoice": inv,
                "customs_type": "EXPORT",
                "service_details": {"customs_channel": luong},
                "costs": [{"name": desc or "Phí dịch vụ hải quan", "amount": pre_vat, "vat_rate": 8}],
            })
    return jobs


def parse_tvc():
    """TVC warehouse handling: STT|Ngày|_|Nội dung|_|BKS|SL|ĐVT|Đơn giá|Thành tiền|Tổng."""
    fp = f"{BASE_DIR}/TVC/Debit Note. 5P. TVC. T03.2026.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb["BẢNG KÊ DỊCH VỤ"]
    jobs = []
    for r in range(15, ws.max_row + 1):
        stt = cell(ws, r, 1)
        if isinstance(stt, str) and "tổng" in stt.lower():
            break
        if not isinstance(stt, (int, float)):
            continue
        date_val = d(cell(ws, r, 2))
        desc = s(cell(ws, r, 4))
        qty = n(cell(ws, r, 7))
        unit = s(cell(ws, r, 8))
        unit_price = n(cell(ws, r, 9))
        pre_vat = n(cell(ws, r, 10))  # C10=Thành tiền
        jobs.append({
            "customer_id": CUSTOMER_MAP["TVC"],
            "date": date_val or datetime.date(2026, 3, 17),
            "description": desc or "Phí nâng hạ hàng",
            "svc_type": "WHS_HANDLE",
            "service_details": {"quantity": qty, "unit": unit, "unit_price": unit_price},
            "costs": [{"name": desc or "Phí nâng hạ hàng", "amount": pre_vat, "vat_rate": 8}],
        })
    return jobs


def parse_thai_hoa():
    """Thái Hòa: 2 files.
    File 1 (XNK TC sheet): 2 customs declarations (pre-VAT 600,000 each).
    File 2 (NHAP KHAU sheet): thu hộ lệ phí hải quan (20,000 each, no VAT).
    Strategy: create 1 job per declaration, attach thu hộ from file 2 by TK matching.
    """
    # File 1: customs
    fp1 = f"{BASE_DIR}/THÁI HOÀ/Debit_5PVN_THAI_HOA_T3_2026 (9).xlsx"
    wb1 = openpyxl.load_workbook(fp1, data_only=True)
    ws1 = wb1["XNK TC"]
    # File 2: thu hộ (2 items keyed by TK)
    fp2 = f"{BASE_DIR}/THÁI HOÀ/Debit_TCH_5PVN_THÁI HÒA_T3_2026_full (4) - Copy.xlsx"
    wb2 = openpyxl.load_workbook(fp2, data_only=True)
    ws2 = wb2["NHAP KHAU"]

    # Build thu hộ map: TK → {amount, gnt}
    thu_ho_map = {}
    for r in range(15, ws2.max_row + 1):
        stt2 = cell(ws2, r, 1)
        if isinstance(stt2, str) and "tổng" in stt2.lower():
            break
        if not isinstance(stt2, (int, float)):
            continue
        tk2 = s(cell(ws2, r, 7))  # C7=Tờ khai (truncated in file)
        amt2 = n(cell(ws2, r, 12))  # C12=Số tiền
        gnt = s(cell(ws2, r, 15))
        if tk2:
            thu_ho_map[tk2] = {"amount": amt2, "gnt": gnt}

    jobs = []
    for r in range(13, ws1.max_row + 1):
        stt = cell(ws1, r, 1)
        if isinstance(stt, str) and "tổng" in stt.lower():
            break
        if not isinstance(stt, (int, float)):
            continue
        date_val = d(cell(ws1, r, 2))
        cd_no = s(cell(ws1, r, 3))
        luong = s(cell(ws1, r, 4))
        inv = s(cell(ws1, r, 11))
        pre_vat = n(cell(ws1, r, 7))  # C7=Phí mở tờ khai

        costs = [{"name": "Phí dịch vụ hải quan XNK tại chỗ", "amount": pre_vat, "vat_rate": 8}]

        # Match thu hộ by partial TK number (file 2 has truncated TK)
        for tk2, th in thu_ho_map.items():
            if cd_no and (cd_no.startswith(tk2) or tk2.startswith(cd_no[:8])):
                costs.append({
                    "name": f"Thu hộ: Lệ phí hải quan - GNT: {th['gnt']}",
                    "amount": th["amount"],
                    "vat_rate": 0,
                    "is_reimbursement": True,
                })

        jobs.append({
            "customer_id": CUSTOMER_MAP["THÁI HOÀ"],
            "date": date_val or datetime.date(2026, 3, 23),
            "description": f"Thủ tục hải quan XNK tại chỗ - TK: {cd_no}",
            "svc_type": "CUS_EXPORT",
            "cd_no": cd_no,
            "invoice": inv,
            "customs_type": "EXPORT",
            "service_details": {"customs_channel": luong},
            "costs": costs,
        })
    return jobs


def parse_gang_thep():
    """Gang Thép TN: 2 files.
    File 1 (BangKe v10): ~20 customs declarations, 800,000 or 400,000 pre-VAT each.
    File 2 (Debit CPN): 2 rows grouped as 1 shipment (customs + trucking).
    """
    jobs = []

    # File 1
    fp1 = f"{BASE_DIR}/GANG THÉP TN/BangKe_GangThep_T3_2026_v10.xlsx"
    wb1 = openpyxl.load_workbook(fp1, data_only=True)
    ws1 = wb1["BẢNG KÊ DỊCH VỤ"]
    for r in range(15, ws1.max_row + 1):
        stt = cell(ws1, r, 1)
        if isinstance(stt, str) and "tổng" in stt.lower():
            break
        if not isinstance(stt, (int, float)):
            continue
        date_val = d(cell(ws1, r, 2))
        desc = s(cell(ws1, r, 4))
        cd_no = s(cell(ws1, r, 5))
        pre_vat = n(cell(ws1, r, 9))  # C9=Thành tiền
        # Detect service type from description
        svc = "CUS_EXPORT"
        if "nhập" in desc.lower() or "import" in desc.lower():
            svc = "CUS_IMPORT"
        # Detect port from description
        port = ""
        if "đồng nai" in desc.lower():
            port = "Đồng Nai"
        elif "thái nguyên" in desc.lower():
            port = "Thái Nguyên"
        jobs.append({
            "customer_id": CUSTOMER_MAP["GANG THÉP TN"],
            "date": date_val or datetime.date(2026, 3, 1),
            "description": desc or f"Thủ tục hải quan - TK: {cd_no}",
            "svc_type": svc,
            "cd_no": cd_no,
            "customs_type": "IMPORT" if svc == "CUS_IMPORT" else "EXPORT",
            "customs_port": port,
            "service_details": {},
            "costs": [{"name": desc or "Phí dịch vụ hải quan", "amount": pre_vat, "vat_rate": 8}],
        })

    # File 2: CPN - grouped 2 cost rows as 1 shipment
    fp2 = f"{BASE_DIR}/GANG THÉP TN/Debit_GangThep_CPN_T3_2026.xlsx"
    wb2 = openpyxl.load_workbook(fp2, data_only=True)
    ws2 = wb2["GANG THÉP"]
    # R14: customs row, R15: trucking row, R16: sub-total (Total shipment)
    # R17: TOTAL
    r = 14
    while r <= ws2.max_row:
        stt = cell(ws2, r, 1)
        if isinstance(stt, str) and "total" in stt.lower():
            break
        if not isinstance(stt, (int, float)):
            r += 1
            continue
        date_val = d(cell(ws2, r, 2))
        tk_bill = s(cell(ws2, r, 3))
        # Parse all sub-rows until next STT or total
        shipment_costs = []
        sr = r
        while sr <= ws2.max_row:
            sub_stt = cell(ws2, sr, 1)
            if sub_stt is not None and sr != r:
                if isinstance(sub_stt, str) and any(k in sub_stt.lower() for k in ["total", "tổng"]):
                    break
                if isinstance(sub_stt, (int, float)):
                    break
            sub_desc = s(cell(ws2, sr, 4))
            sub_pre_vat = n(cell(ws2, sr, 8))  # C8=Thành tiền (service)
            chi_ho = n(cell(ws2, sr, 13))       # C13=Thành tiền chi hộ
            chi_ho_inv = s(cell(ws2, sr, 12))   # C12=Số HĐ chi hộ
            if sub_desc and sub_pre_vat > 0:
                shipment_costs.append({
                    "name": sub_desc, "amount": sub_pre_vat, "vat_rate": 8
                })
            if chi_ho > 0:
                shipment_costs.append({
                    "name": f"Chi hộ: {sub_desc}",
                    "amount": chi_ho, "vat_rate": 0,
                    "is_reimbursement": True,
                    "invoice": chi_ho_inv,
                })
            sr += 1
        if shipment_costs:
            jobs.append({
                "customer_id": CUSTOMER_MAP["GANG THÉP TN"],
                "date": date_val or datetime.date(2026, 3, 24),
                "description": f"Thủ tục HQ + VC - TK/Bill: {tk_bill}",
                "svc_type": "CUS_EXPORT",
                "cd_no": tk_bill,
                "customs_type": "EXPORT",
                "service_details": {},
                "costs": shipment_costs,
            })
        r = sr
    return jobs


def parse_donsung():
    """Donsung warehouse: HDDT sheet, 3 rows (2 storage + 1 handling).
    Header R11: STT|Tên hàng|ĐVT|SL|Đơn giá|Phí khác|Thành tiền chưa VAT|Thuế suất|Tiền thuế|Thành tiền có VAT
    """
    fp = f"{BASE_DIR}/DONSUNG/BẢNG KÊ DỊCH VỤ KHO T3.2026 DONGSUNGrev.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb["HDDT"]
    jobs = []
    for r in range(12, ws.max_row + 1):
        stt = cell(ws, r, 1)
        if isinstance(stt, str) and "tổng" in stt.lower():
            break
        if not isinstance(stt, (int, float)):
            continue
        desc = s(cell(ws, r, 2))
        qty = n(cell(ws, r, 4))
        unit = s(cell(ws, r, 3))
        unit_price = n(cell(ws, r, 5))
        pre_vat = n(cell(ws, r, 7))   # C7=Thành tiền chưa VAT
        vat_rate_val = n(cell(ws, r, 8)) * 100  # 0.1 → 10, 0.08 → 8
        # Determine service type
        svc = "WHS_STORAGE" if "lưu kho" in desc.lower() else "WHS_HANDLE"
        jobs.append({
            "customer_id": CUSTOMER_MAP["DONSUNG"],
            "date": datetime.date(2026, 3, 1),
            "description": desc,
            "svc_type": svc,
            "service_details": {"quantity": qty, "unit": unit, "unit_price": unit_price},
            "costs": [{"name": desc, "amount": pre_vat, "vat_rate": vat_rate_val}],
        })
    return jobs


# ──────── ORCHESTRATOR ────────

def _load_complex_parsers():
    """Load complex parsers from kebab-case filename using importlib."""
    import importlib.util
    script_dir = os.path.dirname(os.path.abspath(__file__))
    module_path = os.path.join(script_dir, "import-march-2026-complex-parsers.py")
    spec = importlib.util.spec_from_file_location("complex_parsers", module_path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def run_parsers():
    """Run all parsers and collect all jobs."""
    cp = _load_complex_parsers()
    parse_dainese = cp.parse_dainese
    parse_tdi = cp.parse_tdi
    parse_glorex = cp.parse_glorex
    parse_nippon = cp.parse_nippon
    parse_messer = cp.parse_messer
    parse_las = cp.parse_las
    parse_kk = cp.parse_kk
    parse_kcvn = cp.parse_kcvn
    parse_kwe = cp.parse_kwe

    all_jobs = []
    parsers = [
        ("DAINESE", parse_dainese),
        ("TDI", parse_tdi),
        ("GLOREX", parse_glorex),
        ("NIPPON", parse_nippon),
        ("MESSER", parse_messer),
        ("LAS", parse_las),
        ("KK", parse_kk),
        ("KCVN", parse_kcvn),
        ("KWE", parse_kwe),
        ("LKV BD", parse_lkv_bd),
        ("LKV MB", parse_lkv_mb),
        ("HƯNG PHÁT", parse_hung_phat),
        ("UTRACORN", parse_utracorn),
        ("VINTECH trucking", parse_vintech_trucking),
        ("VINTECH air", parse_vintech_air),
        ("XÂY LẮP VN", parse_xay_lap_vn),
        ("LOGIMARK", parse_logimark),
        ("TVC", parse_tvc),
        ("THÁI HOÀ", parse_thai_hoa),
        ("GANG THÉP TN", parse_gang_thep),
        ("DONSUNG", parse_donsung),
    ]

    for name, fn in parsers:
        try:
            jobs = fn()
            total_rev = sum(
                sum(n(c["amount"]) for c in j.get("costs", []) if not c.get("is_reimbursement"))
                for j in jobs
            )
            print(f"  {name}: {len(jobs)} jobs, pre-VAT service = {total_rev:,.0f} VND")
            all_jobs.extend(jobs)
        except Exception as e:
            print(f"  ERROR in {name}: {e}")
            traceback.print_exc()

    return all_jobs


def delete_march_2026_jobs(conn):
    """Delete all March 2026 jobs (cascade deletes services + costs)."""
    cur = conn.cursor()
    cur.execute("""
        DELETE FROM jobs
        WHERE etd >= '2026-02-24' AND etd <= '2026-03-31'
    """)
    deleted = cur.rowcount
    conn.commit()
    cur.close()
    print(f"  Deleted {deleted} existing March 2026 jobs")


def print_summary(conn):
    """Print summary by customer."""
    cur = conn.cursor()
    cur.execute("""
        SELECT c.short_name, COUNT(*), SUM(j.total_revenue)
        FROM jobs j
        JOIN customers c ON j.customer_id = c.customer_id
        WHERE j.etd >= '2026-02-24' AND j.etd <= '2026-03-31'
        GROUP BY c.short_name
        ORDER BY SUM(j.total_revenue) DESC NULLS LAST
    """)
    print("\nSummary by customer (March 2026):")
    for row in cur.fetchall():
        rev = row[2] or 0
        print(f"  {row[0]}: {row[1]} jobs, {rev:,.0f} VND")
    cur.execute("""
        SELECT COUNT(*), SUM(total_revenue)
        FROM jobs
        WHERE etd >= '2026-02-24' AND etd <= '2026-03-31'
    """)
    row = cur.fetchone()
    total = row[1] or 0
    print(f"\n  TOTAL: {row[0]} jobs, {total:,.0f} VND")
    cur.close()


def main():
    print("=" * 65)
    print("IMPORT MARCH 2026 JOBS - FULL REWRITE")
    print("=" * 65)

    # Step 1: Parse all Excel files
    print("\n[1] PARSING EXCEL FILES...")
    all_jobs = run_parsers()
    print(f"\nTotal parsed: {len(all_jobs)} jobs")

    if not all_jobs:
        print("No jobs parsed. Exiting.")
        return

    # Step 2: Connect to DB
    print("\n[2] CONNECTING TO DB...")
    conn = psycopg2.connect(DB_URL)
    print("  Connected.")

    # Step 3: Delete existing data
    print("\n[3] DELETING EXISTING MARCH 2026 JOBS...")
    delete_march_2026_jobs(conn)

    # Step 4: Insert all jobs
    print("\n[4] INSERTING JOBS...")
    inserted, errors = insert_jobs(conn, all_jobs)
    print(f"\n  Inserted: {inserted}/{len(all_jobs)} | Errors: {errors}")

    # Step 5: Summary
    print("\n[5] SUMMARY:")
    print_summary(conn)

    conn.close()
    print("\nDone.")


if __name__ == "__main__":
    main()
