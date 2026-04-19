"""
Import Meiko T3 (March 2026) jobs from Excel file.
Sheets: EX (BORDER_EXP), IM (BORDER_IMP), IM Machine (BORDER_IMP)
Job_no format: {PREFIX}-{customer_id}-{DDMM}-{SEQ:4}
"""
import openpyxl
from datetime import datetime, date
import os, sys

# Add project root to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from supabase import create_client
from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), '..', '..', '.env'))

SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_SERVICE_ROLE_KEY') or os.getenv('SUPABASE_KEY')
client = create_client(SUPABASE_URL, SUPABASE_KEY)

MEIKO_CUSTOMER_ID = 32
EXCEL_PATH = '/Users/bear1108/Documents/THÁNG 3/MEIKO/Copy of Bảng kê 5P -Meiko T3 total final.xlsx'
DRY_RUN = False  # Set False to actually insert

# Track sequences per prefix-customer-date combo
seq_tracker = {}

def get_next_job_no(prefix: str, customer_id: int, job_date: date) -> str:
    """Generate job_no: {PREFIX}-{customer_id}-{DDMM}-{SEQ:4}"""
    date_part = job_date.strftime('%d%m')
    key = f"{prefix}-{customer_id}-{date_part}"

    if key not in seq_tracker:
        # Query DB for existing max sequence
        pattern = f"{key}-%"
        result = client.table('jobs').select('job_no').ilike('job_no', pattern).execute()
        max_seq = 0
        for row in result.data:
            try:
                seq = int(row['job_no'].rsplit('-', 1)[-1])
                max_seq = max(max_seq, seq)
            except (ValueError, IndexError):
                pass
        seq_tracker[key] = max_seq

    seq_tracker[key] += 1
    return f"{key}-{seq_tracker[key]:04d}"


def parse_ex_sheet(wb):
    """Parse EX sheet → BORDER_EXP jobs"""
    ws = wb['EX']
    jobs = []

    for row in range(5, 20):
        date_val = ws.cell(row, 1).value
        if not isinstance(date_val, datetime):
            continue

        factory = ws.cell(row, 2).value or ''
        invoice = ws.cell(row, 4).value or ''
        cs_no = ws.cell(row, 5).value or ''
        consignee = ws.cell(row, 6).value or ''
        dest = ws.cell(row, 7).value or ''
        bl = ws.cell(row, 8).value or ''
        mode = ws.cell(row, 9).value or ''

        # Cost columns R(18) to AE(31)
        cost_labels = ['Phí khai HQ', 'Phí seal HQ', 'Phí kiểm hoá', 'Lệ phí',
                       'Phí nâng hạ', 'Phí chờ giờ', 'Cước VC', 'Phí C/O',
                       'Phí hun trùng', 'Phí giám sát HQ', 'Phí khác',
                       'Phí hải quan (CUS)', 'Phí vận chuyển', 'Phí phát sinh']
        costs = {}
        for i, c in enumerate(range(18, 32)):
            v = ws.cell(row, c).value
            if v and isinstance(v, (int, float)) and v > 0:
                label = cost_labels[i] if i < len(cost_labels) else f'Phí khác {i}'
                costs[label] = v

        total = sum(costs.values())
        if total <= 0:
            continue

        jobs.append({
            'service_type': 'BORDER_EXP',
            'date': date_val.date() if isinstance(date_val, datetime) else date_val,
            'factory': factory,
            'invoice': str(invoice).strip(),
            'cs_no': str(cs_no).strip(),
            'consignee': consignee,
            'destination': dest,
            'bl_no': bl,
            'mode': mode,
            'costs': costs,
            'total': total,
            'description': f"XK đường bộ - {consignee} - {dest} - Inv: {invoice}"
        })

    return jobs


def parse_im_sheet(wb):
    """Parse IM sheet → BORDER_IMP jobs (1 row = 1 job)"""
    ws = wb['IM']
    jobs = []

    # Read cost column headers from row 13 (detail row)
    cost_labels = []
    for c in range(19, 32):  # S(19) to AE(31), AF(32)=Total
        h2 = ws.cell(13, c).value or ''
        label = str(h2).strip() or f'Phí {c}'
        cost_labels.append(label)

    for row in range(14, 100):
        date_val = ws.cell(row, 1).value
        if not hasattr(date_val, 'year'):
            continue

        factory = ws.cell(row, 2).value or ''
        invoice = ws.cell(row, 4).value or ''
        cs_no = ws.cell(row, 5).value or ''
        shipper = ws.cell(row, 6).value or ''
        depart = ws.cell(row, 7).value or ''
        bl = ws.cell(row, 8).value or ''
        vehicle_plate = ws.cell(row, 34).value or ''  # AH

        # Cost columns S(19) to AE(31)
        costs = {}
        for i, c in enumerate(range(19, 32)):
            v = ws.cell(row, c).value
            if v and isinstance(v, (int, float)) and v > 0:
                costs[cost_labels[i]] = v

        # Use AF(32) Total as authoritative total
        total_af = ws.cell(row, 32).value
        total = float(total_af) if total_af and isinstance(total_af, (int, float)) and total_af > 0 else sum(costs.values())

        jobs.append({
            'service_type': 'BORDER_IMP',
            'date': date_val.date() if isinstance(date_val, datetime) else date_val,
            'factory': str(factory).strip(),
            'invoice': str(invoice).strip(),
            'cs_no': str(cs_no).strip(),
            'shipper': str(shipper).strip(),
            'depart': str(depart).strip(),
            'bl': str(bl).strip(),
            'vehicle_plate': str(vehicle_plate).strip(),
            'costs': costs,
            'total': total,
            'description': f"NK đường bộ - {str(shipper).strip()[:50]} - {str(depart).strip()} - Xe: {str(vehicle_plate).strip()} - Inv: {str(invoice).strip()}"
        })

    return jobs


def parse_im_machine_sheet(wb):
    """Parse IM Machine sheet → BORDER_IMP jobs (service costs only, col V-AD)"""
    ws = wb['IM Machine']
    jobs = []

    # Service cost columns: V(22)=CS, W(23)=Handling, X(24)=Trucking,
    # Y(25)=Inspection, Z(26)=OT Customs, AA(27)=License fee,
    # AB(28)=Fuel surcharge, AC(29)=Arising fee, AD(30)=At cost
    cost_col_map = {
        22: 'CS', 23: 'Handling', 24: 'Trucking', 25: 'Inspection',
        26: 'OT Customs', 27: 'License fee', 28: 'Fuel surcharge',
        29: 'Arising fee', 30: 'At cost'
    }

    for row in range(17, 25):
        factory = ws.cell(row, 2).value or ws.cell(row, 3).value or ''
        invoice = ws.cell(row, 4).value or ''
        cs_no = ws.cell(row, 5).value or ''
        shipper = ws.cell(row, 6).value or ''
        bl = ws.cell(row, 8).value or ''

        if not invoice or not cs_no:
            continue

        # Service costs only (V to AD)
        costs = {}
        for c, label in cost_col_map.items():
            v = ws.cell(row, c).value
            if v and isinstance(v, (int, float)) and v > 0:
                costs[label] = v

        # Use AF(32) Total as authoritative
        total_af = ws.cell(row, 32).value
        total = float(total_af) if total_af and isinstance(total_af, (int, float)) and total_af > 0 else sum(costs.values())

        if total <= 0:
            continue

        jobs.append({
            'service_type': 'BORDER_IMP',
            'date': date(2026, 3, 1),
            'factory': str(factory).strip(),
            'invoice': str(invoice).strip(),
            'cs_no': str(cs_no).strip(),
            'shipper': str(shipper).strip()[:200],
            'bl_no': str(bl).strip(),
            'costs': costs,
            'total': total,
            'description': f"NK máy móc đường bộ - {str(shipper)[:60]} - Inv: {str(invoice).strip()}"
        })

    return jobs


def create_job_in_db(job_data, service_type, user_id=1):
    """Insert job + job_service + job_costs into DB"""
    prefix = 'BI' if service_type == 'BORDER_IMP' else 'BE'
    job_no = get_next_job_no(prefix, MEIKO_CUSTOMER_ID, job_data['date'])

    # Insert job
    job_record = {
        'job_no': job_no,
        'customer_id': MEIKO_CUSTOMER_ID,
        'description': job_data['description'][:500],
        'etd': job_data['date'].isoformat(),
        'status_code': 'COMPLETED',
        'created_by': user_id,
    }

    if DRY_RUN:
        print(f"  [DRY] INSERT job: {job_no} | {job_data['date']} | {job_data['total']:,.0f} VND")
        return None

    # Check if job_no already exists (resume after partial run)
    existing = client.table('jobs').select('job_id').eq('job_no', job_no).execute()
    if existing.data:
        job_id = existing.data[0]['job_id']
        print(f"  [SKIP] {job_no} already exists (job_id={job_id}), adding costs...")
    else:
        result = client.table('jobs').insert(job_record).execute()
        job = result.data[0]
        job_id = job['job_id']

    # Check if service already exists for this job
    existing_svc = client.table('job_services').select('svc_id').eq('job_id', job_id).execute()
    if existing_svc.data:
        # Check if costs exist too
        existing_costs = client.table('job_costs').select('cost_id').eq('job_id', job_id).execute()
        if existing_costs.data:
            print(f"  [SKIP] {job_no} already has services + costs, skipping")
            return job_id

    # Insert job_service
    invoice_str = ''
    if 'invoices' in job_data:
        invoice_str = ', '.join(job_data['invoices'])
    elif 'invoice' in job_data:
        invoice_str = job_data['invoice']

    svc_record = {
        'job_id': job_id,
        'service_type_code': service_type,
        'scheduled_date': job_data['date'].isoformat(),
        'origin_address': job_data.get('depart') or job_data.get('destination', ''),
        'dest_address': 'Meiko Hà Nội',
        'status_code': 'COMPLETED',
        'invoice_numbers': invoice_str,
        'seller_name': job_data.get('shipper', ''),
        'bl_awb_no': job_data.get('bl_no') or job_data.get('bl', ''),
    }
    if job_data.get('vehicle_plate'):
        svc_record['route'] = f"{job_data.get('depart', '')} → Meiko HN (Xe: {job_data['vehicle_plate']})"

    client.table('job_services').insert(svc_record).execute()

    # Insert job_costs (buying_amount/selling_amount are generated columns)
    # Excel `amount` is the SELLING price (what we charge MEIKO). Buying cost
    # comes from vendor invoices (separate import). Setting buying_rate=0 here
    # so profit shows '—' until real vendor cost is entered, instead of falsely
    # implying 0% margin (which happens when buy_rate==sell_rate).
    for cost_name, amount in job_data.get('costs', {}).items():
        if amount <= 0:
            continue
        client.table('job_costs').insert({
            'job_id': job_id,
            'cost_name': cost_name,
            'quantity': 1,
            'buying_rate': 0,
            'selling_rate': amount,
        }).execute()

    print(f"  [OK] Created {job_no} | job_id={job_id} | {job_data['total']:,.0f} VND")
    return job_id


def main():
    print("=" * 70)
    print(f"MEIKO T3 JOB IMPORT {'(DRY RUN)' if DRY_RUN else '(LIVE)'}")
    print(f"Excel: {EXCEL_PATH}")
    print(f"Customer: MEIKO (ID={MEIKO_CUSTOMER_ID})")
    print("=" * 70)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # Parse all sheets
    ex_jobs = parse_ex_sheet(wb)
    im_jobs = parse_im_sheet(wb)
    machine_jobs = parse_im_machine_sheet(wb)

    total_jobs = len(ex_jobs) + len(im_jobs) + len(machine_jobs)
    total_amount = sum(j['total'] for j in ex_jobs + im_jobs + machine_jobs)

    print(f"\nParsed: {len(ex_jobs)} EX + {len(im_jobs)} IM + {len(machine_jobs)} Machine = {total_jobs} jobs")
    print(f"Total amount: {total_amount:,.0f} VND")
    print()

    # Create EX jobs (BORDER_EXP)
    print(f"--- EX Sheet: {len(ex_jobs)} jobs (BORDER_EXP → BE) ---")
    for job in ex_jobs:
        create_job_in_db(job, 'BORDER_EXP')

    # Create IM jobs (BORDER_IMP)
    print(f"\n--- IM Sheet: {len(im_jobs)} jobs (BORDER_IMP → BI) ---")
    for job in im_jobs:
        create_job_in_db(job, 'BORDER_IMP')

    # Create Machine jobs (BORDER_IMP)
    print(f"\n--- IM Machine Sheet: {len(machine_jobs)} jobs (BORDER_IMP → BI) ---")
    for job in machine_jobs:
        create_job_in_db(job, 'BORDER_IMP')

    print(f"\n{'=' * 70}")
    print(f"DONE: {total_jobs} jobs {'would be created' if DRY_RUN else 'created'}")
    print(f"Total: {total_amount:,.0f} VND")
    if DRY_RUN:
        print("\n⚠️  Set DRY_RUN = False to actually insert into DB")


if __name__ == '__main__':
    main()
