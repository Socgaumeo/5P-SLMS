#!/usr/bin/env python3
"""
Script tạo bảng vendor_surcharges và import dữ liệu từ Excel lên Supabase
Sử dụng psycopg2 để thực thi SQL trực tiếp
"""

import os
import sys
import json
import re
from datetime import datetime
from typing import List, Dict, Any, Optional
from dotenv import load_dotenv
import psycopg2
from psycopg2 import sql
from openpyxl import load_workbook
import google.generativeai as genai

# Load environment variables
load_dotenv()


class VendorSurchargesManager:
    """Class quản lý tạo bảng và import vendor surcharges lên Supabase"""
    
    def __init__(self):
        """Khởi tạo database connection và AI model"""
        self.conn = self._get_db_connection()
        self._setup_ai()
        self.stats = {
            'success': 0,
            'failed': 0,
            'errors': [],
        }
    
    def _get_db_connection(self):
        """Tạo database connection"""
        database_url = os.getenv('DATABASE_URL')
        
        if not database_url:
            raise ValueError("DATABASE_URL phải được cấu hình trong .env")
        
        return psycopg2.connect(database_url)
    
    def _setup_ai(self):
        """Setup AI model (Gemini)"""
        api_key = os.getenv('GOOGLE_GEMINI_API_KEY')
        if api_key:
            genai.configure(api_key=api_key)
            self.ai_model = genai.GenerativeModel('gemini-2.0-flash')
            self.ai_available = True
        else:
            print("⚠️ GOOGLE_GEMINI_API_KEY không được cấu hình, AI sẽ không hoạt động")
            self.ai_available = False
    
    def create_vendor_surcharges_table(self):
        """Tạo bảng vendor_surcharges với cấu trúc đúng"""
        print("\n" + "="*60)
        print("TẠO BẢNG VENDOR_SURCHARGES")
        print("="*60)
        
        try:
            cursor = self.conn.cursor()
            
            # Xóa bảng cũ nếu tồn tại
            print("🗑️  Xóa bảng vendor_surcharges cũ (nếu có)...")
            cursor.execute("DROP TABLE IF EXISTS public.vendor_surcharges CASCADE;")
            
            # Tạo bảng mới
            print("🔨 Tạo bảng vendor_surcharges mới...")
            create_table_sql = """
                CREATE TABLE public.vendor_surcharges (
                    id SERIAL PRIMARY KEY,
                    vendor_id INTEGER NOT NULL REFERENCES vendors(vendor_id) ON DELETE CASCADE,
                    surcharge_code VARCHAR(50) NOT NULL,
                    description TEXT,
                    amount DECIMAL(18,2),
                    unit VARCHAR(20),
                    conditions TEXT,
                    effective_date DATE DEFAULT CURRENT_DATE,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """
            cursor.execute(create_table_sql)
            
            # Tạo index
            print("📊 Tạo index...")
            cursor.execute("CREATE INDEX idx_vendor_surcharges_vendor_id ON public.vendor_surcharges(vendor_id);")
            cursor.execute("CREATE INDEX idx_vendor_surcharges_surcharge_code ON public.vendor_surcharges(surcharge_code);")
            
            # Comment
            cursor.execute("COMMENT ON TABLE public.vendor_surcharges IS 'Bảng phụ phí của nhà vận chuyển';")
            
            # Commit
            self.conn.commit()
            cursor.close()
            
            print("✅ Bảng vendor_surcharges đã được tạo thành công!")
            
        except Exception as e:
            self.conn.rollback()
            error_msg = f"Lỗi khi tạo bảng vendor_surcharges: {str(e)}"
            print(f"❌ {error_msg}")
            self.stats['errors'].append(error_msg)
            raise
    
    def read_excel_sheet(self, file_path: str, sheet_name: str, max_rows: int = 100) -> str:
        """
        Đọc sheet từ Excel và chuyển thành text
        
        Args:
            file_path: Đường dẫn file Excel
            sheet_name: Tên sheet
            max_rows: Số dòng tối đa để đọc
        
        Returns:
            String chứa nội dung sheet
        """
        try:
            wb = load_workbook(file_path, data_only=True)
            sheet = wb[sheet_name]
            
            content = []
            for i, row in enumerate(sheet.iter_rows(max_row=max_rows), 1):
                row_text = " | ".join([str(cell.value)[:50] if cell.value else "" for cell in row])
                if row_text.strip(" |"):
                    content.append(f"Row {i}: {row_text}")
            
            return "\n".join(content)
            
        except Exception as e:
            print(f"❌ Lỗi khi đọc Excel: {str(e)}")
            return ""
    
    def parse_surcharges_with_ai(self, content: str) -> List[Dict[str, Any]]:
        """
        Sử dụng AI để parse surcharges từ Excel content
        
        Args:
            content: Nội dung Excel dạng text
        
        Returns:
            List của surcharge dictionaries
        """
        if not self.ai_available:
            print("⚠️ AI không khả dụng, không thể parse surcharges")
            return []
        
        example_json = '''[
  {"surcharge_code": "PHI_CAU_DUONG", "description": "Phí cầu đường", "amount": 50000, "unit": "lượt", "conditions": "Áp dụng cho các tuyến qua cầu"}
]'''
        
        prompt = f"""Analyze this Excel data containing Vietnamese trucking surcharges (phụ phí).

Extract surcharge information into a JSON array. Each surcharge should have:
- surcharge_code: Mã phụ phí (e.g. "PHI_CAU_DUONG", "PHI_CHO", "PHI_DO")
- description: Mô tả chi tiết
- amount: Số tiền (number only, in VND)
- unit: Đơn vị tính (e.g. "lượt", "km", "giờ", "ngày")
- conditions: Điều kiện áp dụng (nếu có)

Data:
```
{content}
```

IMPORTANT:
- Return ONLY valid JSON array
- Amount = number (remove commas/dots in thousands)
- Include ALL surcharge types found
- Skip header rows and notes

Example output:
{example_json}
"""
        
        try:
            response = self.ai_model.generate_content(prompt)
            response_text = response.text.strip()
            
            # Clean up response
            if "```json" in response_text:
                response_text = response_text.split("```json")[1].split("```")[0]
            elif "```" in response_text:
                lines = response_text.split("\n")
                start = next((i for i, l in enumerate(lines) if l.startswith("```")), 0) + 1
                end = next((i for i, l in enumerate(lines[start:]) if l.startswith("```")), len(lines)) + start
                response_text = "\n".join(lines[start:end])
            
            # Fix common JSON issues
            response_text = re.sub(r',\s*]', ']', response_text)
            response_text = re.sub(r',\s*}', '}', response_text)
            
            return json.loads(response_text)
            
        except Exception as e:
            print(f"⚠️ Lỗi khi parse surcharges với AI: {str(e)}")
            return []
    
    def get_vendor_id(self, vendor_name: str) -> Optional[int]:
        """
        Lấy vendor_id từ tên vendor
        
        Args:
            vendor_name: Tên vendor
        
        Returns:
            vendor_id hoặc None nếu không tìm thấy
        """
        try:
            cursor = self.conn.cursor()
            query = """
                SELECT vendor_id 
                FROM vendors 
                WHERE company_name ILIKE %s OR short_name ILIKE %s
                LIMIT 1
            """
            cursor.execute(query, (f'%{vendor_name}%', f'%{vendor_name}%'))
            result = cursor.fetchone()
            cursor.close()
            
            if result:
                return result[0]
            
            return None
            
        except Exception as e:
            print(f"❌ Lỗi khi tìm vendor: {str(e)}")
            return None
    
    def import_surcharges(self, file_path: str, vendor_name: str):
        """
        Import vendor surcharges từ Excel lên Supabase
        
        Args:
            file_path: Đường dẫn file Excel
            vendor_name: Tên vendor
        """
        print("\n" + "="*60)
        print(f"IMPORT VENDOR SURCHARGES - {vendor_name}")
        print("="*60)
        
        try:
            # Lấy vendor_id
            vendor_id = self.get_vendor_id(vendor_name)
            if not vendor_id:
                print(f"❌ Không tìm thấy vendor '{vendor_name}'")
                return
            
            print(f"✅ Vendor ID: {vendor_id}")
            
            # Đọc Excel
            wb = load_workbook(file_path, data_only=True)
            print(f"📑 Found sheets: {wb.sheetnames}")
            
            total_imported = 0
            
            for sheet_name in wb.sheetnames:
                # Chỉ xử lý sheet phụ phí
                if 'Phụ Phí' not in sheet_name:
                    continue
                
                print(f"\n🔄 Processing: {sheet_name}")
                
                # Đọc nội dung sheet
                content = self.read_excel_sheet(file_path, sheet_name)
                if not content:
                    print(f"⚠️ Không thể đọc sheet: {sheet_name}")
                    continue
                
                # Parse với AI
                print("🤖 Parsing with AI...")
                surcharges = self.parse_surcharges_with_ai(content)
                print(f"✅ Extracted {len(surcharges)} surcharge entries")
                
                # Import từng surcharge
                cursor = self.conn.cursor()
                for i, surcharge in enumerate(surcharges, 1):
                    try:
                        # Tạo record
                        surcharge_data = {
                            'vendor_id': vendor_id,
                            'surcharge_code': surcharge.get('surcharge_code'),
                            'description': surcharge.get('description'),
                            'amount': surcharge.get('amount'),
                            'unit': surcharge.get('unit'),
                            'conditions': surcharge.get('conditions'),
                            'effective_date': datetime.now().date().isoformat()
                        }
                        
                        # Kiểm tra xem surcharge đã tồn tại chưa
                        check_query = """
                            SELECT id FROM vendor_surcharges 
                            WHERE vendor_id = %s AND surcharge_code = %s
                        """
                        cursor.execute(check_query, (vendor_id, surcharge.get('surcharge_code')))
                        existing = cursor.fetchone()
                        
                        if existing:
                            # Update
                            update_query = """
                                UPDATE vendor_surcharges 
                                SET description = %s, amount = %s, unit = %s, conditions = %s, updated_at = CURRENT_TIMESTAMP
                                WHERE id = %s
                            """
                            cursor.execute(update_query, (
                                surcharge_data['description'],
                                surcharge_data['amount'],
                                surcharge_data['unit'],
                                surcharge_data['conditions'],
                                existing[0]
                            ))
                            print(f"  [{i}] Updated: {surcharge.get('surcharge_code')} | {surcharge.get('amount'):,} VND/{surcharge.get('unit')}")
                        else:
                            # Insert
                            insert_query = """
                                INSERT INTO vendor_surcharges (vendor_id, surcharge_code, description, amount, unit, conditions, effective_date)
                                VALUES (%s, %s, %s, %s, %s, %s, %s)
                            """
                            cursor.execute(insert_query, (
                                surcharge_data['vendor_id'],
                                surcharge_data['surcharge_code'],
                                surcharge_data['description'],
                                surcharge_data['amount'],
                                surcharge_data['unit'],
                                surcharge_data['conditions'],
                                surcharge_data['effective_date']
                            ))
                            print(f"  [{i}] Inserted: {surcharge.get('surcharge_code')} | {surcharge.get('amount'):,} VND/{surcharge.get('unit')}")
                        
                        total_imported += 1
                        self.stats['success'] += 1
                    
                    except Exception as e:
                        error_msg = f"Error importing surcharge: {str(e)}"
                        print(f"  ❌ {error_msg}")
                        self.stats['failed'] += 1
                        self.stats['errors'].append(error_msg)
                
                # Commit sau khi xử lý xong sheet
                self.conn.commit()
                cursor.close()
            
            print(f"\n✓ Import surcharges hoàn tất: {total_imported} surcharges")
            
        except Exception as e:
            self.conn.rollback()
            error_msg = f"Lỗi khi import surcharges: {str(e)}"
            print(f"❌ {error_msg}")
            self.stats['errors'].append(error_msg)
    
    def print_summary(self):
        """In tổng kết import"""
        print("\n" + "="*60)
        print("TỔNG KẾT IMPORT VENDOR SURCHARGES")
        print("="*60)
        
        print(f"\n✓ Thành công: {self.stats['success']}")
        print(f"✗ Thất bại: {self.stats['failed']}")
        
        if self.stats['errors']:
            print(f"\nLỗi:")
            for error in self.stats['errors'][:5]:
                print(f"    - {error}")
            if len(self.stats['errors']) > 5:
                print(f"    ... và {len(self.stats['errors']) - 5} lỗi khác")
        
        print("\n" + "="*60)
    
    def close(self):
        """Đóng database connection"""
        if self.conn:
            self.conn.close()


def main():
    """Hàm chính"""
    manager = None
    try:
        if len(sys.argv) < 2:
            print("Usage: python create_and_import_vendor_surcharges.py <file_path> [vendor_name]")
            print("Example: python create_and_import_vendor_surcharges.py vendor_rates/Tam_bảo_092025.xlsx 'Tam Bảo'")
            sys.exit(1)
        
        file_path = sys.argv[1]
        vendor_name = sys.argv[2] if len(sys.argv) > 2 else "Tam Bảo"
        
        manager = VendorSurchargesManager()
        
        # Tạo bảng vendor_surcharges
        manager.create_vendor_surcharges_table()
        
        # Import surcharges
        manager.import_surcharges(file_path, vendor_name)
        
        # In tổng kết
        manager.print_summary()
        
        # Đóng connection
        manager.close()
        
        # Exit code dựa trên kết quả
        sys.exit(0 if manager.stats['failed'] == 0 else 1)
        
    except Exception as e:
        print(f"\nFATAL ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        if manager:
            manager.close()
        sys.exit(1)


if __name__ == '__main__':
    main()
