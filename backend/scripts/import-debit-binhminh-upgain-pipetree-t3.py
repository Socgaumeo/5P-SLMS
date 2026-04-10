"""
Import T3 (March 2026) debit notes for 3 customers:
- Bình Minh (ID=68): 2 Excel files, 4 TRUCKING_DOM jobs, VAT 8%
- Upgain (ID=69): 1 Excel file, 3 TRUCKING_DOM jobs, VAT 0%
- Pipetree (ID=36): 1 Excel file, 5 TRUCKING_DOM + 1 AIR_DOM, VAT 8%

Job_no format: {PREFIX}-{customer_id}-{DDMM}-{SEQ:4}
"""
import openpyxl
from datetime import datetime, date
import os, sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from supabase import create_client
from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), '..', '..', '.env'))

SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_SERVICE_ROLE_KEY') or os.getenv('SUPABASE_KEY')
client = create_client(SUPABASE_URL, SUPABASE_KEY)

DRY_RUN = False
BASE_DIR = '/Users/bear1108/Documents/THÁNG 3'
PREFIX_MAP = {'TRUCKING_DOM': 'TRK', 'AIR_DOM': 'AD'}
seq_tracker = {}


def get_next_job_no(prefix: str, customer_id: int, job_date: date) -> str:
    date_part = job_date.strftime('%d%m')
    key = f"{prefix}-{customer_id}-{date_part}"
    if key not in seq_tracker:
        pattern = f"{key}-%"
        result = client.table('jobs').select('job_no').ilike('job_no', pattern).execute()
        max_seq = 0
        for row in result.data:
            try:
                seq = int(row['job_no'].rsplit('-', 1)[-1])
                max_seq = max(max_seq, seq)
            except (ValueError, IndexError):
                pass
        seq_tracker[key] = max_seq
    seq_tracker[key] += 1
    return f"{key}-{seq_tracker[key]:04d}"


def fix_march_date(val):
    """Parse date, fixing Excel DD/MM swap for March 2026 data.
    If Excel misread DD/MM as MM/DD, month won't be 3 → swap back."""
    if isinstance(val, str):
        # String format: '27/2/2026' or '14/3/2026'
        for fmt in ('%d/%m/%Y', '%d/%m/%y'):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
        return None
    if isinstance(val, datetime):
        d = val.date()
    elif isinstance(val, date):
        d = val
    else:
        return None
    # Fix swapped dates: all data is March 2026
    if d.year == 2026 and d.month != 3 and 1 <= d.day <= 3:
        try:
            return date(2026, d.day, d.month)
        except ValueError:
            pass
    return d


def create_job_in_db(job_data, customer_id, user_id=1):
    """Insert job + job_service + job_costs into DB"""
    service_type = job_data['service_type']
    prefix = PREFIX_MAP.get(service_type, 'TRK')
    job_no = get_next_job_no(prefix, customer_id, job_data['date'])

    job_record = {
        'job_no': job_no,
        'customer_id': customer_id,
        'description': job_data['description'][:500],
        'etd': job_data['date'].isoformat(),
        'status_code': 'COMPLETED',
        'created_by': user_id,
    }

    if DRY_RUN:
        print(f"  [DRY] {job_no} | {job_data['date']} | {job_data['total']:,.0f} VND")
        for cn, amt in job_data.get('costs', {}).items():
            print(f"         - {cn}: {amt:,.0f}")
        return None

    existing = client.table('jobs').select('job_id').eq('job_no', job_no).execute()
    if existing.data:
        job_id = existing.data[0]['job_id']
        existing_costs = client.table('job_costs').select('cost_id').eq('job_id', job_id).execute()
        if existing_costs.data:
            print(f"  [SKIP] {job_no} already exists with costs")
            return job_id
    else:
        result = client.table('jobs').insert(job_record).execute()
        job_id = result.data[0]['job_id']

    svc_record = {
        'job_id': job_id,
        'service_type_code': service_type,
        'scheduled_date': job_data['date'].isoformat(),
        'origin_address': job_data.get('origin', ''),
        'dest_address': job_data.get('dest', ''),
        'status_code': 'COMPLETED',
        'route': job_data.get('route', ''),
    }
    if job_data.get('bl_awb'):
        svc_record['bl_awb_no'] = job_data['bl_awb']
    client.table('job_services').insert(svc_record).execute()

    for cost_name, amount in job_data.get('costs', {}).items():
        if amount <= 0:
            continue
        client.table('job_costs').insert({
            'job_id': job_id,
            'cost_name': cost_name,
            'quantity': 1,
            'buying_rate': amount,
            'selling_rate': amount,
        }).execute()

    print(f"  [OK] {job_no} | job_id={job_id} | {job_data['total']:,.0f} VND")
    return job_id


# ============================================================
# BÌNH MINH (ID=68) - 2 files, TRUCKING_DOM
# Header R13: STT(1) Date(2) Type(3) Origin(4) Dest(5) Plate(6)
#   Qty(7) Unit(8) UnitPrice(9) Amount(10) OtherCost(11) Total(12)
# ============================================================
def parse_binhminh_file(filepath, label_col11='Chi phí khác'):
    """Parse a single Bình Minh debit file"""
    jobs = []
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    for row in range(15, ws.max_row + 1):
        # Check total column (12) — skip summary/VAT rows
        total_val = ws.cell(row, 12).value
        if not total_val or not isinstance(total_val, (int, float)) or total_val <= 0:
            continue
        # Skip summary rows ("Tổng", "Thuế VAT")
        cell_a = ws.cell(row, 1).value
        if isinstance(cell_a, str) and ('tổng' in cell_a.lower() or 'thuế' in cell_a.lower()):
            break

        job_date = fix_march_date(ws.cell(row, 2).value)
        if not job_date:
            continue

        cargo_type = str(ws.cell(row, 3).value or '').strip()
        origin = str(ws.cell(row, 4).value or '').strip()
        dest = str(ws.cell(row, 5).value or '').strip()
        plate = str(ws.cell(row, 6).value or '').strip()
        amount = ws.cell(row, 10).value or 0
        other_cost = ws.cell(row, 11).value or 0
        total = float(total_val)

        costs = {}
        if amount and float(amount) > 0:
            costs['Cước vận chuyển'] = float(amount)
        if other_cost and float(other_cost) > 0:
            costs[label_col11] = float(other_cost)

        jobs.append({
            'service_type': 'TRUCKING_DOM',
            'date': job_date,
            'origin': origin,
            'dest': dest,
            'route': f"{origin} → {dest} (Xe: {plate})",
            'costs': costs,
            'total': total,
            'vat_rate': 0.08,
            'description': f"VC nội địa - {cargo_type} - {origin} → {dest} - Xe: {plate}",
        })
    return jobs


def parse_binhminh():
    f1 = os.path.join(BASE_DIR, 'BÌNH MINH/Copy of Debit Note.Bình Minh. Mar.2026.xlsx')
    f2 = os.path.join(BASE_DIR, 'BÌNH MINH/Copy of Debit Note.Bình Minh. Mar.2026 L2.xlsx')
    jobs = parse_binhminh_file(f1, 'Chi phí khác')
    if os.path.exists(f2):
        jobs += parse_binhminh_file(f2, 'Phụ phí xăng dầu')
    return jobs


# ============================================================
# UPGAIN (ID=69) - 1 file, TRUCKING_DOM
# Header R13: STT(1) Date(2) Type(3) Origin(4) Dest(5) Plate(6)
#   Qty(7) Unit(8) UnitPrice(9) Amount(10) OtherCost(11) Surcharge(12) Total(13)
# ============================================================
def parse_upgain():
    jobs = []
    f = os.path.join(BASE_DIR, 'UPGAIN/Copy of Debit Note.UPGAIN.MARCH.2026 (.xlsx')
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb.active

    for row in range(15, ws.max_row + 1):
        total_val = ws.cell(row, 13).value
        if not total_val or not isinstance(total_val, (int, float)) or total_val <= 0:
            continue
        cell_a = ws.cell(row, 1).value
        if isinstance(cell_a, str) and ('tổng' in cell_a.lower() or 'thuế' in cell_a.lower()):
            break

        job_date = fix_march_date(ws.cell(row, 2).value)
        if not job_date:
            continue

        cargo_type = str(ws.cell(row, 3).value or '').strip()
        origin = str(ws.cell(row, 4).value or '').strip()
        dest = str(ws.cell(row, 5).value or '').strip()
        plate = str(ws.cell(row, 6).value or '').strip()
        amount = ws.cell(row, 10).value or 0
        other_cost = ws.cell(row, 11).value or 0
        surcharge = ws.cell(row, 12).value or 0
        total = float(total_val)

        costs = {}
        if amount and float(amount) > 0:
            costs['Cước vận chuyển'] = float(amount)
        if other_cost and float(other_cost) > 0:
            costs['Chi phí khác'] = float(other_cost)
        if surcharge and float(surcharge) > 0:
            costs['Phụ phí xăng dầu'] = float(surcharge)

        jobs.append({
            'service_type': 'TRUCKING_DOM',
            'date': job_date,
            'origin': origin,
            'dest': dest,
            'route': f"{origin} → {dest} (Xe: {plate})",
            'costs': costs,
            'total': total,
            'vat_rate': 0.0,
            'description': f"VC nội địa - {cargo_type} - {origin} → {dest} - Xe: {plate}",
        })
    return jobs


# ============================================================
# PIPETREE (ID=36) - 1 file, 2 sheets
# TRUCKING BỘ R13-14: STT(1) NgàyLấy(2) NgàyTrả(3) Type(4)
#   Origin(5) Dest(6) Plate(7) Hàng(8) [9] [10]
#   SL(11) ĐVT(12) ĐơnGiá(13) ThànhTiền(14) CPKhác(15) Tổng(16)
# AIR 1 R13: STT(1) AWB(2) Ngày(3) Kiện(4) Kg(5) Bill(6)
#   LoạiHàng(7) NơiĐi(8) NơiĐến(9) ĐơnGiá(10) ThànhTiền(11) PhụPhí(12) Tổng(13)
# ============================================================
def parse_pipetree():
    jobs = []
    f = os.path.join(BASE_DIR, 'PIPETREE/DEBIT NOTE PIPETREE MARCH.2026.xlsx')
    wb = openpyxl.load_workbook(f, data_only=True)

    # TRUCKING BỘ sheet
    ws = wb['TRUCKING BỘ']
    for row in range(15, ws.max_row + 1):
        total_val = ws.cell(row, 16).value
        if not total_val or not isinstance(total_val, (int, float)) or total_val <= 0:
            continue
        cell_a = ws.cell(row, 1).value
        if isinstance(cell_a, str) and ('tổng' in cell_a.lower() or 'thuế' in cell_a.lower()):
            break

        pickup_date = fix_march_date(ws.cell(row, 2).value)
        delivery_date = fix_march_date(ws.cell(row, 3).value)
        job_date = pickup_date or delivery_date
        if not job_date:
            continue

        cargo_type = str(ws.cell(row, 4).value or '').strip()
        origin = str(ws.cell(row, 5).value or '').strip()
        dest = str(ws.cell(row, 6).value or '').strip()
        plate = str(ws.cell(row, 7).value or '').strip()
        cargo = str(ws.cell(row, 8).value or '').strip()
        amount = ws.cell(row, 14).value or 0
        other_cost = ws.cell(row, 15).value or 0
        total = float(total_val)

        costs = {}
        if amount and float(amount) > 0:
            costs['Cước vận chuyển'] = float(amount)
        if other_cost and float(other_cost) > 0:
            costs['Chi phí khác'] = float(other_cost)

        jobs.append({
            'service_type': 'TRUCKING_DOM',
            'date': job_date,
            'origin': origin,
            'dest': dest,
            'route': f"{origin} → {dest} (Xe: {plate})",
            'costs': costs,
            'total': total,
            'vat_rate': 0.08,
            'description': f"VC nội địa - {cargo_type} {cargo} - {origin} → {dest} - Xe: {plate}",
        })

    # AIR 1 sheet (data starts at row 14, only 1 row)
    ws2 = wb['AIR 1']
    for row in range(14, ws2.max_row + 1):
        total_val = ws2.cell(row, 13).value
        if not total_val or not isinstance(total_val, (int, float)) or total_val <= 0:
            continue
        cell_a = ws2.cell(row, 1).value
        if isinstance(cell_a, str) and ('tổng' in cell_a.lower() or 'thuế' in cell_a.lower()):
            break

        awb_raw = ws2.cell(row, 2).value
        awb = str(int(awb_raw)) if isinstance(awb_raw, (int, float)) else str(awb_raw or '').strip()
        job_date = fix_march_date(ws2.cell(row, 3).value)
        if not job_date:
            continue

        cargo_type = str(ws2.cell(row, 7).value or '').strip()
        origin = str(ws2.cell(row, 8).value or '').strip()
        dest = str(ws2.cell(row, 9).value or '').strip()
        amount = ws2.cell(row, 11).value or 0
        surcharge = ws2.cell(row, 12).value or 0
        total = float(total_val)

        costs = {}
        if amount and float(amount) > 0:
            costs['Cước vận chuyển'] = float(amount)
        if surcharge and float(surcharge) > 0:
            costs['Phụ phí'] = float(surcharge)

        jobs.append({
            'service_type': 'AIR_DOM',
            'date': job_date,
            'origin': origin,
            'dest': dest,
            'bl_awb': awb,
            'route': f"{origin} → {dest} (AWB: {awb})",
            'costs': costs,
            'total': total,
            'vat_rate': 0.08,
            'description': f"HK nội địa - {cargo_type} - {origin} → {dest} - AWB: {awb}",
        })

    return jobs


def main():
    print("=" * 70)
    print(f"DEBIT IMPORT T3 - 3 CUSTOMERS {'(DRY RUN)' if DRY_RUN else '(LIVE)'}")
    print("=" * 70)

    # BÌNH MINH
    bm_jobs = parse_binhminh()
    bm_total = sum(j['total'] for j in bm_jobs)
    bm_vat = sum(j['total'] * j['vat_rate'] for j in bm_jobs)
    print(f"\n--- BÌNH MINH (ID=68): {len(bm_jobs)} jobs ---")
    print(f"    Subtotal: {bm_total:,.0f} | VAT 8%: {bm_vat:,.0f} | Grand: {bm_total + bm_vat:,.0f}")
    for job in bm_jobs:
        create_job_in_db(job, customer_id=68)

    # UPGAIN
    up_jobs = parse_upgain()
    up_total = sum(j['total'] for j in up_jobs)
    print(f"\n--- UPGAIN (ID=69): {len(up_jobs)} jobs ---")
    print(f"    Subtotal: {up_total:,.0f} | VAT 0%")
    for job in up_jobs:
        create_job_in_db(job, customer_id=69)

    # PIPETREE
    pt_jobs = parse_pipetree()
    pt_trk = [j for j in pt_jobs if j['service_type'] == 'TRUCKING_DOM']
    pt_air = [j for j in pt_jobs if j['service_type'] == 'AIR_DOM']
    pt_total = sum(j['total'] for j in pt_jobs)
    pt_vat = sum(j['total'] * j['vat_rate'] for j in pt_jobs)
    print(f"\n--- PIPETREE (ID=36): {len(pt_jobs)} jobs ({len(pt_trk)} trucking + {len(pt_air)} air) ---")
    print(f"    Subtotal: {pt_total:,.0f} | VAT 8%: {pt_vat:,.0f} | Grand: {pt_total + pt_vat:,.0f}")
    for job in pt_jobs:
        create_job_in_db(job, customer_id=36)

    # Summary
    all_jobs = bm_jobs + up_jobs + pt_jobs
    grand_total = sum(j['total'] for j in all_jobs)
    grand_vat = sum(j['total'] * j['vat_rate'] for j in all_jobs)
    print(f"\n{'=' * 70}")
    print(f"TOTAL: {len(all_jobs)} jobs | {grand_total:,.0f} + VAT {grand_vat:,.0f} = {grand_total + grand_vat:,.0f}")
    if DRY_RUN:
        print("\nSet DRY_RUN = False to actually insert into DB")


if __name__ == '__main__':
    main()
