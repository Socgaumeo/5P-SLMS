"""
Cross-check March 2026 jobs: DB data vs Excel source files.
Compares job count and total revenue per customer.
"""
import os, datetime, json
import psycopg2
import openpyxl
import xlrd

DB_URL = "postgresql://postgres.ooixntyflwmjaryxwakx:%21%40kHanh0112@aws-1-ap-southeast-1.pooler.supabase.com:6543/postgres"
BASE_DIR = "/Users/bear1108/Documents/Tháng 3"

# Map folder names to customer_ids (same as import)
CUSTOMER_MAP = {
    "DAINESE": 46, "DONSUNG": 58, "GANG THÉP TN": 44, "GLOREX": 18,
    "HƯNG PHÁT": 63, "KCVN": 61, "KWE": 28, "LAS": 6,
    "LKV BD": 60, "LKV MB": 53, "LOGIMARK": 31, "MESSER": 22,
    "NIPPON": 64, "TDI": 20, "THÁI HOÀ": 45, "TVC": 59,
    "UTRACORN": 56, "VINTECH": 57, "XÂY LẮP VN": 2,
}


def s(v):
    if v is None: return ""
    return str(v).strip()


def n(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).replace(",", "").strip())
    except: return 0


def d(v, wb_datemode=0):
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    if isinstance(v, (int, float)) and v > 40000:
        try: return xlrd.xldate_as_datetime(v, wb_datemode).date()
        except: pass
    if isinstance(v, str):
        for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]:
            try: return datetime.datetime.strptime(v.strip(), fmt).date()
            except: continue
    return None


def cell(ws, r, c):
    return ws.cell(row=r, column=c).value


# ──────── Read Excel totals per customer ────────

def read_excel_totals():
    """Read each customer's Excel files, extract total revenue and job count."""
    results = {}

    # ── DAINESE: trucking HĐ sheet ──
    fp = f"{BASE_DIR}/DAINESE/Copy of (DAINESE-5PVN) BẢNG KÊ TT T3.2026 bs2.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb["HĐ"]
    count, total = 0, 0
    for r in range(15, ws.max_row + 1):
        stt = cell(ws, r, 1)
        if isinstance(stt, str) and "tổng" in stt.lower(): break
        if not isinstance(stt, (int, float)): continue
        dt = d(cell(ws, r, 2))
        if not dt or not (dt.year == 2026 and dt.month in (2, 3)): continue
        total += n(cell(ws, r, 14))
        count += 1

    # DAINESE customs
    fp2 = f"{BASE_DIR}/DAINESE/Copy of Bảng kê xuất tháng 3.2026 final.xlsx"
    wb2 = openpyxl.load_workbook(fp2, data_only=True)
    for sn in wb2.sheetnames:
        ws2 = wb2[sn]
        if ws2.max_row < 12: continue
        hdr = None
        for r in range(1, min(20, ws2.max_row + 1)):
            if s(cell(ws2, r, 1)).upper() == "STT": hdr = r; break
        if not hdr: continue
        for r in range(hdr + 2, ws2.max_row + 1):
            stt = cell(ws2, r, 1)
            if isinstance(stt, str) and any(k in stt.lower() for k in ["tổng", "total"]): break
            if not isinstance(stt, (int, float)): continue
            dt = d(cell(ws2, r, 2))
            if not dt or not (dt.year == 2026 and dt.month in (2, 3)): continue
            rev = n(cell(ws2, r, 10)) or n(cell(ws2, r, 11))
            total += rev
            count += 1
    results["DAINESE"] = {"count": count, "total": round(total, 2)}

    # ── DONSUNG ──
    fp = f"{BASE_DIR}/DONSUNG/BẢNG KÊ DỊCH VỤ KHO T3.2026 DONGSUNGrev.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = None
    for sn in wb.sheetnames:
        if wb[sn].max_row > 10 and "kangatang" not in sn.lower():
            ws = wb[sn]; break
    count, total = 0, 0
    if ws:
        for r in range(12, ws.max_row + 1):
            stt = cell(ws, r, 1)
            if isinstance(stt, str) and "tổng" in stt.lower(): break
            if not isinstance(stt, (int, float)): continue
            total += n(cell(ws, r, 10))
            count += 1
    results["DONSUNG"] = {"count": count, "total": round(total, 2)}

    # ── GANG THÉP TN ──
    fp = f"{BASE_DIR}/GANG THÉP TN/BangKe_GangThep_T3_2026_v10.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb["BẢNG KÊ DỊCH VỤ"]
    count, total = 0, 0
    for r in range(15, ws.max_row + 1):
        stt = cell(ws, r, 1)
        if isinstance(stt, str) and "tổng" in stt.lower(): break
        if not isinstance(stt, (int, float)): continue
        total += n(cell(ws, r, 10))
        count += 1
    results["GANG THÉP TN"] = {"count": count, "total": round(total, 2)}

    # ── GLOREX quốc tế ──
    fp = f"{BASE_DIR}/GLOREX/Debit 5PVN_GLOREX 3.2026.QUỐC TẾ.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    count, total = 0, 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row < 12 or "kangatang" in sn.lower(): continue
        is_glorex = any("GLOREX" in s(cell(ws, r, c)).upper()
                        for r in range(1, 12) for c in range(1, 20) if cell(ws, r, c))
        if not is_glorex: continue
        r = 15
        while r <= ws.max_row:
            stt = cell(ws, r, 1)
            if isinstance(stt, str) and any(k in stt.lower() for k in ["tổng", "total"]): break
            if not isinstance(stt, (int, float)):
                r += 1; continue
            dt = d(cell(ws, r, 2))
            if not dt or not (dt.year == 2026 and dt.month in (2, 3)):
                r += 1; continue
            group_total = n(cell(ws, r, 14))
            sr = r + 1
            while sr <= ws.max_row:
                sub_stt = cell(ws, sr, 1)
                if sub_stt is not None: break
                group_total += n(cell(ws, sr, 14))
                sr += 1
            total += group_total
            count += 1
            r = sr
    glorex_qt = {"count": count, "total": round(total, 2)}

    # ── GLOREX tại chỗ ──
    fp = f"{BASE_DIR}/GLOREX/Debit 5PVN_GLOREX T3.2026 TẠI CHỖ.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb["XNK TC"]
    gc, gt = 0, 0
    for r in range(14, ws.max_row + 1):
        stt = cell(ws, r, 1)
        if isinstance(stt, str) and "tổng" in stt.lower(): break
        if not isinstance(stt, (int, float)): continue
        dt = d(cell(ws, r, 2))
        if not dt or not (dt.year == 2026 and dt.month in (2, 3)): continue
        t = n(cell(ws, r, 10))
        gt += t * 1.08 if t > 0 else 0
        gc += 1
    results["GLOREX"] = {
        "count": glorex_qt["count"] + gc,
        "total": round(glorex_qt["total"] + gt, 2)
    }

    # ── Standard format customers (Hưng Phát, LKV BD, LKV MB, Vintech, Utracorn, Xây Lắp VN) ──
    standard_files = {
        "HƯNG PHÁT": f"{BASE_DIR}/HƯNG PHÁT/Debit note HƯNG PHÁT-5P T3.2026.xlsx",
        "LKV BD": f"{BASE_DIR}/LKV BD/Debit note SX LỌC KHÍ VIỆT BD-5P T3.2026 REV1.xlsx",
        "LKV MB": f"{BASE_DIR}/LKV MB/Debit note SX LỌC KHÍ VIỆT MIỀN BẮC-5P T3.2026  rev14.xlsx",
        "VINTECH": f"{BASE_DIR}/VINTECH/Debit note VINTECH-5P T3.2026.xlsx",
        "UTRACORN": f"{BASE_DIR}/UTRACORN/DebitNote_UTRACON_TRK1403_DRAFT (3).xlsx",
        "XÂY LẮP VN": f"{BASE_DIR}/XÂY LẮP VN/Debit note XÂY LẮP VN-5P T3.2026.xlsx",
    }
    for folder, fp in standard_files.items():
        wb = openpyxl.load_workbook(fp, data_only=True)
        count, total = 0, 0
        for sn in wb.sheetnames:
            ws = wb[sn]
            if ws.max_row < 12 or "kangatang" in sn.lower(): continue
            hdr = None
            for r in range(10, min(15, ws.max_row + 1)):
                if s(cell(ws, r, 1)).upper() == "STT": hdr = r; break
            if not hdr: continue
            for r in range(hdr + 2, ws.max_row + 1):
                stt = cell(ws, r, 1)
                if isinstance(stt, str) and any(k in stt.lower() for k in ["tổng", "total"]): break
                if not isinstance(stt, (int, float)): continue
                dt = d(cell(ws, r, 2))
                if not dt or not (dt.year == 2026 and dt.month in (2, 3)): continue
                amount = n(cell(ws, r, 9))
                tax = n(cell(ws, r, 10))
                t = n(cell(ws, r, 11))
                rev = t if t > 0 else (amount + tax if amount > 0 else 0)
                total += rev
                count += 1
            break  # Only first valid sheet
        results[folder] = {"count": count, "total": round(total, 2)}

    # ── KCVN ──
    fp = f"{BASE_DIR}/KCVN/BangKe_KCIL_T3_2026_v14.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    count, total = 0, 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row < 12: continue
        for r in range(12, ws.max_row + 1):
            stt = cell(ws, r, 1)
            if isinstance(stt, str) and "tổng" in stt.lower(): break
            if not isinstance(stt, (int, float)): continue
            dt = d(cell(ws, r, 4))
            if not dt or not (dt.month == 3 and dt.year == 2026): continue
            total += n(cell(ws, r, 17))
            count += 1
        break
    results["KCVN"] = {"count": count, "total": round(total, 2)}

    # ── KWE ──
    fp = f"{BASE_DIR}/KWE/5P in MAR.2026. KWE rev.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = None
    for sn in wb.sheetnames:
        if "Accountant Sheet" == sn: ws = wb[sn]; break
    count, total = 0, 0
    if ws:
        for r in range(18, ws.max_row + 1):
            stt = cell(ws, r, 1)
            if isinstance(stt, str) and any(k in stt.lower() for k in ["tổng", "total"]): break
            if not isinstance(stt, (int, float)): continue
            total += n(cell(ws, r, 6))
            count += 1
    results["KWE"] = {"count": count, "total": round(total, 2)}

    # ── LOGIMARK ──
    fp = f"{BASE_DIR}/LOGIMARK/Debit_LOGIMARK_T3_2026_updated (3).xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb[wb.sheetnames[0]]
    count, total = 0, 0
    for r in range(14, ws.max_row + 1):
        stt = cell(ws, r, 1)
        if isinstance(stt, str) and any(k in stt.lower() for k in ["tổng", "total"]): break
        if not isinstance(stt, (int, float)): continue
        total += n(cell(ws, r, 11))
        count += 1
    results["LOGIMARK"] = {"count": count, "total": round(total, 2)}

    # ── MESSER ──
    fp = f"{BASE_DIR}/MESSER/Bảng kê MESSER 5P T3.2026.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    count, total = 0, 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row < 12: continue
        r = 14
        while r <= ws.max_row:
            stt = cell(ws, r, 1)
            if isinstance(stt, str) and any(k in stt.lower() for k in ["tổng", "total"]): break
            if not isinstance(stt, (int, float)):
                r += 1; continue
            svc_total = n(cell(ws, r, 10))
            chi_ho_total = n(cell(ws, r, 15))
            sr = r + 1
            while sr <= ws.max_row:
                sub_stt = cell(ws, sr, 1)
                if sub_stt is not None: break
                svc_total += n(cell(ws, sr, 10))
                chi_ho_total += n(cell(ws, sr, 15))
                sr += 1
            total += svc_total + chi_ho_total
            count += 1
            r = sr
    results["MESSER"] = {"count": count, "total": round(total, 2)}

    # ── NIPPON (.xls) ──
    fp = f"{BASE_DIR}/NIPPON/(THAI NGUYEN) BẢNG KÊ CHI PHÍ NIPPON THÁNG 3.2026.xls"
    wb = xlrd.open_workbook(fp)
    count, total = 0, 0
    for sn in wb.sheet_names():
        ws = wb.sheet_by_name(sn)
        if ws.nrows < 15: continue
        for r in range(18, ws.nrows):
            stt = ws.cell_value(r, 0)
            if isinstance(stt, str) and "tổng" in stt.lower(): break
            if not isinstance(stt, (int, float)) or stt < 1: continue
            dt = d(ws.cell_value(r, 3), wb.datemode)
            if not dt or not (dt.year == 2026 and dt.month in (2, 3)): continue
            total += n(ws.cell_value(r, 11))
            count += 1
    results["NIPPON"] = {"count": count, "total": round(total, 2)}

    # ── TDI ──
    fp = f"{BASE_DIR}/TDI/Copy of BangKe_TDI_AirT3_2026_ final1.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    count, total = 0, 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row < 12 or "CHI HỘ" in sn.upper(): continue
        r = 14
        while r <= ws.max_row:
            stt = cell(ws, r, 1)
            if isinstance(stt, str) and "total" in stt.lower(): break
            if not isinstance(stt, (int, float)):
                r += 1; continue
            dt = d(cell(ws, r, 2))
            if not dt or not (dt.year == 2026 and dt.month in (2, 3)):
                r += 1; continue
            total_r = n(cell(ws, r, 18))
            sr = r + 1
            while sr <= ws.max_row:
                sub_stt = cell(ws, sr, 1)
                if sub_stt is not None: break
                total_r += n(cell(ws, sr, 18))
                sr += 1
            if total_r == 0:
                total_r = n(cell(ws, r, 14))
                for sr2 in range(r + 1, sr):
                    total_r += n(cell(ws, sr2, 14))
            total += total_r
            count += 1
            r = sr
    results["TDI"] = {"count": count, "total": round(total, 2)}

    # ── THÁI HOÀ ──
    fp = f"{BASE_DIR}/THÁI HOÀ/Debit_5PVN_THAI_HOA_T3_2026 (9).xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    count, total = 0, 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row < 10 or "kangatang" in sn.lower(): continue
        is_th = any("THÁI" in s(cell(ws, r, c)).upper() or "THAI HOA" in s(cell(ws, r, c)).upper()
                     for r in range(1, 12) for c in range(1, 10) if cell(ws, r, c))
        if not is_th: continue
        for r in range(12, ws.max_row + 1):
            stt = cell(ws, r, 1)
            if isinstance(stt, str) and any(k in stt.lower() for k in ["tổng", "total"]): break
            if not isinstance(stt, (int, float)): continue
            dt = d(cell(ws, r, 2))
            if not dt or dt.year < 2026: continue
            t = n(cell(ws, r, 10))
            total += t * 1.08 if t > 0 else 0
            count += 1
    results["THÁI HOÀ"] = {"count": count, "total": round(total, 2)}

    # ── TVC ──
    fp = f"{BASE_DIR}/TVC/Debit Note. 5P. TVC. T03.2026.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb[wb.sheetnames[0]]
    count, total = 0, 0
    for r in range(15, ws.max_row + 1):
        stt = cell(ws, r, 1)
        if isinstance(stt, str) and "tổng" in stt.lower(): break
        if not isinstance(stt, (int, float)): continue
        dt = d(cell(ws, r, 2))
        if not dt or not (dt.year == 2026 and dt.month in (2, 3)): continue
        total += n(cell(ws, r, 11))
        count += 1
    results["TVC"] = {"count": count, "total": round(total, 2)}

    # ── LAS ──
    fp = f"{BASE_DIR}/LAS/DebitNote_LGZHPH260781_LAS_DRAFT (13).xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = None
    for sn in wb.sheetnames:
        if "dịch vụ" in sn.lower(): ws = wb[sn]; break
    if not ws: ws = wb[wb.sheetnames[0]]
    total = 0
    for r in range(20, ws.max_row + 1):
        stt = cell(ws, r, 1)
        if isinstance(stt, str) and any(k in stt.lower() for k in ["tổng", "total"]):
            total = n(cell(ws, r, 10)) or total
            break
        total += n(cell(ws, r, 10))
    results["LAS"] = {"count": 1, "total": round(total, 2)}

    return results


def read_db_totals():
    """Read totals per customer from DB."""
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()
    cur.execute("""
        SELECT c.company_name, COUNT(*), ROUND(SUM(j.total_revenue)::numeric, 2)
        FROM jobs j JOIN customers c ON j.customer_id = c.customer_id
        GROUP BY c.company_name, c.customer_id
        ORDER BY SUM(j.total_revenue) DESC
    """)
    db_data = {}
    cid_to_folder = {}
    # Reverse mapping: customer_id → folder name
    for folder, cid in CUSTOMER_MAP.items():
        cid_to_folder[cid] = folder

    cur.execute("""
        SELECT c.customer_id, c.company_name, COUNT(*), ROUND(SUM(j.total_revenue)::numeric, 2)
        FROM jobs j JOIN customers c ON j.customer_id = c.customer_id
        GROUP BY c.customer_id, c.company_name
        ORDER BY SUM(j.total_revenue) DESC
    """)
    for row in cur.fetchall():
        cid, name, cnt, rev = row
        folder = cid_to_folder.get(cid, name)
        db_data[folder] = {"name": name, "count": cnt, "total": float(rev)}

    cur.close()
    conn.close()
    return db_data


def main():
    print("=" * 80)
    print("CROSS-CHECK: DB vs Excel Files (Tháng 3/2026)")
    print("=" * 80)

    excel_data = read_excel_totals()
    db_data = read_db_totals()

    total_excel_rev = 0
    total_db_rev = 0
    total_excel_jobs = 0
    total_db_jobs = 0
    issues = []

    print(f"\n{'Customer':<16} | {'Excel Jobs':>10} | {'DB Jobs':>8} | {'Excel Revenue':>16} | {'DB Revenue':>16} | {'Diff':>12} | Status")
    print("-" * 110)

    for folder in sorted(set(list(excel_data.keys()) + list(db_data.keys()))):
        ex = excel_data.get(folder, {"count": 0, "total": 0})
        db = db_data.get(folder, {"count": 0, "total": 0})

        total_excel_rev += ex["total"]
        total_db_rev += db["total"]
        total_excel_jobs += ex["count"]
        total_db_jobs += db["count"]

        diff = db["total"] - ex["total"]
        pct = abs(diff / ex["total"] * 100) if ex["total"] > 0 else 0

        if ex["count"] != db["count"]:
            status = f"JOBS MISMATCH"
            issues.append(f"{folder}: Excel {ex['count']} jobs vs DB {db['count']} jobs")
        elif abs(diff) > 1:
            status = f"REV DIFF {pct:.1f}%"
            issues.append(f"{folder}: Revenue diff {diff:,.0f} VND ({pct:.1f}%)")
        else:
            status = "OK"

        print(f"{folder:<16} | {ex['count']:>10} | {db['count']:>8} | {ex['total']:>16,.0f} | {db['total']:>16,.0f} | {diff:>12,.0f} | {status}")

    print("-" * 110)
    total_diff = total_db_rev - total_excel_rev
    print(f"{'TOTAL':<16} | {total_excel_jobs:>10} | {total_db_jobs:>8} | {total_excel_rev:>16,.0f} | {total_db_rev:>16,.0f} | {total_diff:>12,.0f} |")

    print(f"\n{'='*80}")
    if issues:
        print("ISSUES FOUND:")
        for i in issues:
            print(f"  - {i}")
    else:
        print("ALL CUSTOMERS MATCH!")
    print(f"{'='*80}")


if __name__ == "__main__":
    main()
