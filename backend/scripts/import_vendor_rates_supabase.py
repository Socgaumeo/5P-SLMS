#!/usr/bin/env python3
"""
Script import vendor rates và vendor surcharges từ Excel lên Supabase
Sử dụng AI (Gemini/DeepSeek) để parse dữ liệu từ Excel
"""

import os
import sys
import json
import re
from datetime import datetime
from typing import List, Dict, Any, Optional
from dotenv import load_dotenv

# Thêm đường dẫn parent để import từ app
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from supabase import create_client, Client
from openpyxl import load_workbook
import google.generativeai as genai

# Load environment variables
load_dotenv()


class VendorRatesImporter:
    """Class quản lý import vendor rates và surcharges lên Supabase"""
    
    def __init__(self):
        """Khởi tạo Supabase client và AI model"""
        self.supabase: Client = self._get_supabase_client()
        self._setup_ai()
        self.stats = {
            'rates': {'success': 0, 'failed': 0, 'errors': []},
            'surcharges': {'success': 0, 'failed': 0, 'errors': []},
        }
    
    def _get_supabase_client(self) -> Client:
        """Tạo Supabase client"""
        supabase_url = os.getenv('SUPABASE_URL')
        supabase_key = os.getenv('SUPABASE_SERVICE_ROLE_KEY')
        
        if not supabase_url or not supabase_key:
            raise ValueError("SUPABASE_URL và SUPABASE_SERVICE_ROLE_KEY phải được cấu hình trong .env")
        
        return create_client(supabase_url, supabase_key)
    
    def _setup_ai(self):
        """Setup AI model (Gemini)"""
        api_key = os.getenv('GOOGLE_GEMINI_API_KEY')
        if api_key:
            genai.configure(api_key=api_key)
            self.ai_model = genai.GenerativeModel('gemini-2.0-flash')
            self.ai_available = True
        else:
            print("⚠️ GOOGLE_GEMINI_API_KEY không được cấu hình, AI sẽ không hoạt động")
            self.ai_available = False
    
    def read_excel_sheet(self, file_path: str, sheet_name: str, max_rows: int = 100) -> str:
        """
        Đọc sheet từ Excel và chuyển thành text
        
        Args:
            file_path: Đường dẫn file Excel
            sheet_name: Tên sheet
            max_rows: Số dòng tối đa để đọc
        
        Returns:
            String chứa nội dung sheet
        """
        try:
            wb = load_workbook(file_path, data_only=True)
            sheet = wb[sheet_name]
            
            content = []
            for i, row in enumerate(sheet.iter_rows(max_row=max_rows), 1):
                row_text = " | ".join([str(cell.value)[:50] if cell.value else "" for cell in row])
                if row_text.strip(" |"):
                    content.append(f"Row {i}: {row_text}")
            
            return "\n".join(content)
            
        except Exception as e:
            print(f"❌ Lỗi khi đọc Excel: {str(e)}")
            return ""
    
    def parse_rates_with_ai(self, content: str, rate_type: str) -> List[Dict[str, Any]]:
        """
        Sử dụng AI để parse rates từ Excel content
        
        Args:
            content: Nội dung Excel dạng text
            rate_type: Loại rate (STANDARD, REFRIGERATED)
        
        Returns:
            List của rate dictionaries
        """
        if not self.ai_available:
            print("⚠️ AI không khả dụng, không thể parse rates")
            return []
        
        prompt = f"""Analyze this Excel data containing Vietnamese trucking rates.
Rate type: {rate_type} (STANDARD = xe thường, REFRIGERATED = xe lạnh)

Extract pricing into a JSON array. Each rate should have:
- origin: Điểm đầu (pickup location name only)
- origin_province: Tỉnh đi
- destination: Điểm đến (destination name only)
- destination_province: Tỉnh đến
- rate_code: Mã giá (e.g. "TB18", "TB20")
- temperature_range: Yêu cầu nhiệt độ (for refrigerated only, e.g. "Từ 0 đến 10 độ")
- prices: Object with vehicle_type -> price mapping
  Example: {{"1.25T": 200000, "1.5T": 310000, "2.5T": 420000}}

Data:
```
{content}
```

IMPORTANT:
- Return ONLY valid JSON array
- Price = number (remove commas/dots in thousands)
- Include ALL vehicle types found (1.25T, 1.5T, 2.5T, 3.5T, 5T, 8T, 15T)
- rate_code is usually like TB18, TB20, TB25, etc.
- Skip header rows and notes

Example output:
[
  {{"origin": "Nội Bài", "origin_province": "Hà Nội", "destination": "KCN Quang Minh", "destination_province": "Hà Nội", "rate_code": "TB20", "temperature_range": null, "prices": {{"1.25T": 220000, "1.5T": 330000, "2.5T": 440000}}}}
]
"""
        
        try:
            response = self.ai_model.generate_content(prompt)
            response_text = response.text.strip()
            
            # Clean up response
            if "```json" in response_text:
                response_text = response_text.split("```json")[1].split("```")[0]
            elif "```" in response_text:
                lines = response_text.split("\n")
                start = next((i for i, l in enumerate(lines) if l.startswith("```")), 0) + 1
                end = next((i for i, l in enumerate(lines[start:]) if l.startswith("```")), len(lines)) + start
                response_text = "\n".join(lines[start:end])
            
            # Fix common JSON issues
            response_text = re.sub(r',\s*]', ']', response_text)
            response_text = re.sub(r',\s*}', '}', response_text)
            
            return json.loads(response_text)
            
        except Exception as e:
            print(f"⚠️ Lỗi khi parse với AI: {str(e)}")
            return []
    
    def parse_surcharges_with_ai(self, content: str) -> List[Dict[str, Any]]:
        """
        Sử dụng AI để parse surcharges từ Excel content
        
        Args:
            content: Nội dung Excel dạng text
        
        Returns:
            List của surcharge dictionaries
        """
        if not self.ai_available:
            print("⚠️ AI không khả dụng, không thể parse surcharges")
            return []
        
        prompt = f"""Analyze this Excel data containing Vietnamese trucking surcharges (phụ phí).

Extract surcharge information into a JSON array. Each surcharge should have:
- surcharge_type: Loại phụ phí (e.g. "Phí cầu đường", "Phí bến bãi", "Phí chờ", "Phí đêm")
- description: Mô tả chi tiết
- amount: Số tiền (number only, in VND)
- unit: Đơn vị tính (e.g. "lượt", "km", "giờ", "ngày")
- conditions: Điều kiện áp dụng (nếu có)

Data:
```
{content}
```

IMPORTANT:
- Return ONLY valid JSON array
- Amount = number (remove commas/dots in thousands)
- Include ALL surcharge types found
- Skip header rows and notes

Example output:
[
  {{"surcharge_type": "Phí cầu đường", "description": "Phí cầu đường Hà Nội", "amount": 50000, "unit": "lượt", "conditions": "Áp dụng cho các tuyến qua cầu"}}
]
"""
        
        try:
            response = self.ai_model.generate_content(prompt)
            response_text = response.text.strip()
            
            # Clean up response
            if "```json" in response_text:
                response_text = response_text.split("```json")[1].split("```")[0]
            elif "```" in response_text:
                lines = response_text.split("\n")
                start = next((i for i, l in enumerate(lines) if l.startswith("```")), 0) + 1
                end = next((i for i, l in enumerate(lines[start:]) if l.startswith("```")), len(lines)) + start
                response_text = "\n".join(lines[start:end])
            
            # Fix common JSON issues
            response_text = re.sub(r',\s*]', ']', response_text)
            response_text = re.sub(r',\s*}', '}', response_text)
            
            return json.loads(response_text)
            
        except Exception as e:
            print(f"⚠️ Lỗi khi parse surcharges với AI: {str(e)}")
            return []
    
    def get_or_create_route(self, origin: str, destination: str) -> Optional[int]:
        """
        Lấy hoặc tạo route trong master_routes
        
        Args:
            origin: Điểm đi
            destination: Điểm đến
        
        Returns:
            route_id hoặc None nếu thất bại
        """
        try:
            # Tìm route đã tồn tại
            result = self.supabase.table('master_routes').select('route_id').ilike('origin', f'%{origin}%').ilike('destination', f'%{destination}%').execute()
            
            if result.data:
                return result.data[0]['route_id']
            
            # Tạo route mới
            new_route = {
                'origin': origin,
                'destination': destination,
                'is_active': True
            }
            result = self.supabase.table('master_routes').insert(new_route).execute()
            
            if result.data:
                return result.data[0]['route_id']
            
            return None
            
        except Exception as e:
            print(f"❌ Lỗi khi lấy/tạo route: {str(e)}")
            return None
    
    def get_vendor_id(self, vendor_name: str) -> Optional[int]:
        """
        Lấy vendor_id từ tên vendor
        
        Args:
            vendor_name: Tên vendor
        
        Returns:
            vendor_id hoặc None nếu không tìm thấy
        """
        try:
            result = self.supabase.table('vendors').select('vendor_id').ilike('company_name', f'%{vendor_name}%').or_(f"short_name.ilike.%{vendor_name}%").execute()
            
            if result.data:
                return result.data[0]['vendor_id']
            
            return None
            
        except Exception as e:
            print(f"❌ Lỗi khi tìm vendor: {str(e)}")
            return None
    
    def import_rates(self, file_path: str, vendor_name: str):
        """
        Import vendor rates từ Excel lên Supabase
        
        Args:
            file_path: Đường dẫn file Excel
            vendor_name: Tên vendor
        """
        print("\n" + "="*60)
        print(f"IMPORT VENDOR RATES - {vendor_name}")
        print("="*60)
        
        try:
            # Lấy vendor_id
            vendor_id = self.get_vendor_id(vendor_name)
            if not vendor_id:
                print(f"❌ Không tìm thấy vendor '{vendor_name}'")
                return
            
            print(f"✅ Vendor ID: {vendor_id}")
            
            # Đọc Excel
            wb = load_workbook(file_path, data_only=True)
            print(f"📑 Found sheets: {wb.sheetnames}")
            
            total_imported = 0
            
            for sheet_name in wb.sheetnames:
                # Bỏ qua sheet phụ phí
                if 'Phụ Phí' in sheet_name or 'Phụ phí' in sheet_name:
                    print(f"\n⏭️  Bỏ qua sheet: {sheet_name} (surcharges sẽ xử lý riêng)")
                    continue
                
                # Xác định loại rate
                if 'Lạnh' in sheet_name or 'lạnh' in sheet_name:
                    rate_type = 'REFRIGERATED'
                else:
                    rate_type = 'STANDARD'
                
                print(f"\n🔄 Processing: {sheet_name} ({rate_type})")
                
                # Đọc nội dung sheet
                content = self.read_excel_sheet(file_path, sheet_name)
                if not content:
                    print(f"⚠️ Không thể đọc sheet: {sheet_name}")
                    continue
                
                # Parse với AI
                print("🤖 Parsing with AI...")
                rates = self.parse_rates_with_ai(content, rate_type)
                print(f"✅ Extracted {len(rates)} rate entries")
                
                # Import từng rate
                for i, rate in enumerate(rates, 1):
                    try:
                        # Lấy hoặc tạo route
                        route_id = self.get_or_create_route(
                            rate.get('origin'),
                            rate.get('destination')
                        )
                        
                        if not route_id:
                            print(f"⚠️ Không thể tạo route cho {rate.get('origin')} → {rate.get('destination')}")
                            continue
                        
                        # Import từng vehicle type
                        prices = rate.get('prices', {})
                        for vehicle_type, price in prices.items():
                            # Kiểm tra xem rate đã tồn tại chưa
                            existing = self.supabase.table('vendor_rates').select('id').eq('vendor_id', vendor_id).eq('route_id', route_id).eq('vehicle_type', vehicle_type).eq('rate_type', rate_type).execute()
                            
                            if existing.data:
                                # Update
                                rate_data = {
                                    'price': price,
                                    'rate_code': rate.get('rate_code'),
                                    'rate_type': rate_type,
                                    'temperature_range': rate.get('temperature_range'),
                                    'origin_province': rate.get('origin_province'),
                                    'destination_province': rate.get('destination_province'),
                                    'effective_date': datetime.now().date().isoformat()
                                }
                                self.supabase.table('vendor_rates').update(rate_data).eq('id', existing.data[0]['id']).execute()
                                print(f"  [{i}] Updated: {rate.get('origin')} → {rate.get('destination')} | {vehicle_type} | {price:,} VND")
                            else:
                                # Insert
                                rate_data = {
                                    'vendor_id': vendor_id,
                                    'route_id': route_id,
                                    'vehicle_type': vehicle_type,
                                    'price': price,
                                    'rate_code': rate.get('rate_code'),
                                    'rate_type': rate_type,
                                    'temperature_range': rate.get('temperature_range'),
                                    'origin_province': rate.get('origin_province'),
                                    'destination_province': rate.get('destination_province'),
                                    'effective_date': datetime.now().date().isoformat()
                                }
                                self.supabase.table('vendor_rates').insert(rate_data).execute()
                                print(f"  [{i}] Inserted: {rate.get('origin')} → {rate.get('destination')} | {vehicle_type} | {price:,} VND")
                            
                            total_imported += 1
                            self.stats['rates']['success'] += 1
                    
                    except Exception as e:
                        error_msg = f"Error importing rate: {str(e)}"
                        print(f"  ❌ {error_msg}")
                        self.stats['rates']['failed'] += 1
                        self.stats['rates']['errors'].append(error_msg)
            
            print(f"\n✓ Import rates hoàn tất: {total_imported} rates")
            
        except Exception as e:
            error_msg = f"Lỗi khi import rates: {str(e)}"
            print(f"❌ {error_msg}")
            self.stats['rates']['errors'].append(error_msg)
    
    def import_surcharges(self, file_path: str, vendor_name: str):
        """
        Import vendor surcharges từ Excel lên Supabase
        
        Args:
            file_path: Đường dẫn file Excel
            vendor_name: Tên vendor
        """
        print("\n" + "="*60)
        print(f"IMPORT VENDOR SURCHARGES - {vendor_name}")
        print("="*60)
        
        try:
            # Lấy vendor_id
            vendor_id = self.get_vendor_id(vendor_name)
            if not vendor_id:
                print(f"❌ Không tìm thấy vendor '{vendor_name}'")
                return
            
            print(f"✅ Vendor ID: {vendor_id}")
            
            # Đọc Excel
            wb = load_workbook(file_path, data_only=True)
            print(f"📑 Found sheets: {wb.sheetnames}")
            
            total_imported = 0
            
            for sheet_name in wb.sheetnames:
                # Chỉ xử lý sheet phụ phí
                if 'Phụ Phí' not in sheet_name and 'Phụ phí' not in sheet_name:
                    continue
                
                print(f"\n🔄 Processing: {sheet_name}")
                
                # Đọc nội dung sheet
                content = self.read_excel_sheet(file_path, sheet_name)
                if not content:
                    print(f"⚠️ Không thể đọc sheet: {sheet_name}")
                    continue
                
                # Parse với AI
                print("🤖 Parsing with AI...")
                surcharges = self.parse_surcharges_with_ai(content)
                print(f"✅ Extracted {len(surcharges)} surcharge entries")
                
                # Import từng surcharge
                for i, surcharge in enumerate(surcharges, 1):
                    try:
                        # Kiểm tra xem surcharge đã tồn tại chưa
                        existing = self.supabase.table('vendor_surcharges').select('id').eq('vendor_id', vendor_id).eq('surcharge_code', surcharge.get('surcharge_type')).execute()
                        
                        if existing.data:
                            # Update
                            surcharge_data = {
                                'description': surcharge.get('description'),
                                'amount': surcharge.get('amount'),
                                'unit': surcharge.get('unit'),
                                'conditions': surcharge.get('conditions'),
                                'effective_date': datetime.now().date().isoformat()
                            }
                            self.supabase.table('vendor_surcharges').update(surcharge_data).eq('id', existing.data[0]['id']).execute()
                            print(f"  [{i}] Updated: {surcharge.get('surcharge_type')} | {surcharge.get('amount'):,} VND/{surcharge.get('unit')}")
                        else:
                            # Insert
                            surcharge_data = {
                                'vendor_id': vendor_id,
                                'surcharge_type': surcharge.get('surcharge_type'),
                                'description': surcharge.get('description'),
                                'amount': surcharge.get('amount'),
                                'unit': surcharge.get('unit'),
                                'conditions': surcharge.get('conditions'),
                                'effective_date': datetime.now().date().isoformat()
                            }
                            self.supabase.table('vendor_surcharges').insert(surcharge_data).execute()
                            print(f"  [{i}] Inserted: {surcharge.get('surcharge_type')} | {surcharge.get('amount'):,} VND/{surcharge.get('unit')}")
                        
                        total_imported += 1
                        self.stats['surcharges']['success'] += 1
                    
                    except Exception as e:
                        error_msg = f"Error importing surcharge: {str(e)}"
                        print(f"  ❌ {error_msg}")
                        self.stats['surcharges']['failed'] += 1
                        self.stats['surcharges']['errors'].append(error_msg)
            
            print(f"\n✓ Import surcharges hoàn tất: {total_imported} surcharges")
            
        except Exception as e:
            error_msg = f"Lỗi khi import surcharges: {str(e)}"
            print(f"❌ {error_msg}")
            self.stats['surcharges']['errors'].append(error_msg)
    
    def print_summary(self):
        """In tổng kết import"""
        print("\n" + "="*60)
        print("TỔNG KẾT IMPORT VENDOR RATES & SURCHARGES")
        print("="*60)
        
        for table, stats in self.stats.items():
            print(f"\n{table.upper()}:")
            print(f"  ✓ Thành công: {stats['success']}")
            print(f"  ✗ Thất bại: {stats['failed']}")
            if stats['errors']:
                print(f"  Lỗi:")
                for error in stats['errors'][:5]:  # Chỉ hiển thị 5 lỗi đầu tiên
                    print(f"    - {error}")
                if len(stats['errors']) > 5:
                    print(f"    ... và {len(stats['errors']) - 5} lỗi khác")
        
        print("\n" + "="*60)


def main():
    """Hàm chính"""
    try:
        if len(sys.argv) < 2:
            print("Usage: python import_vendor_rates_supabase.py <file_path> [vendor_name]")
            print("Example: python import_vendor_rates_supabase.py vendor_rates/Tam_bảo_092025.xlsx 'Tam Bảo'")
            sys.exit(1)
        
        file_path = sys.argv[1]
        vendor_name = sys.argv[2] if len(sys.argv) > 2 else "Tam Bảo"
        
        importer = VendorRatesImporter()
        
        # Import rates
        importer.import_rates(file_path, vendor_name)
        
        # Import surcharges
        importer.import_surcharges(file_path, vendor_name)
        
        # In tổng kết
        importer.print_summary()
        
        # Exit code dựa trên kết quả
        total_failed = sum(stats['failed'] for stats in importer.stats.values())
        sys.exit(0 if total_failed == 0 else 1)
        
    except Exception as e:
        print(f"\nFATAL ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
