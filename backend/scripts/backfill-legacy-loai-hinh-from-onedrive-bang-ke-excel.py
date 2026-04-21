#!/usr/bin/env python3
"""
Backfill `job_services.loai_hinh` for the 164 legacy customs rows inserted
before the validator landed in Phase 6.

Strategy
--------
1. Pull all CUS_IMPORT / CUS_EXPORT / CUS rows from DB with loai_hinh IS NULL.
2. Scan T3/2026 bảng kê Excel files under OneDrive, indexing:
     cd_no → [hint_text_tokens]
   where hint tokens come from "Loại hình", "Note", "Tờ khai" neighborhood cells.
3. Map hint tokens + service_type_code → one of the 18 whitelisted VN codes.
4. Produce a dry-run JSON report ({svc_id, current, proposed, confidence, source}).
5. Apply updates with --apply flag. Uncertain rows stay NULL and are listed
   so a human can fill them in the UI.

Heuristic rules (ordered — first match wins):
- Hint contains "tạm xuất" → G14
- Hint contains "tái xuất" → G24
- Hint contains "tại chỗ" / "DOM" / "XNK TC":
    CUS_IMPORT → A41, CUS_EXPORT → B13
- Hint contains "gia công":
    CUS_IMPORT → E11, CUS_EXPORT → E52
- Hint contains "SXXK" / "sản xuất xuất khẩu":
    CUS_IMPORT → E21, CUS_EXPORT → E42
- Hint contains "SEA" / "OCEAN" / "AIR" / "DHL" / "EXPRESS" (international):
    CUS_IMPORT → A11, CUS_EXPORT → B11
- cd_no starts with 308 + no other signal, CUS_EXPORT → B13
- cd_no starts with 108 + no other signal, CUS_IMPORT → A41
- Otherwise → flag for manual review (confidence=LOW)

Usage:
    python backend/scripts/backfill-legacy-loai-hinh-from-onedrive-bang-ke-excel.py          # dry-run
    python backend/scripts/backfill-legacy-loai-hinh-from-onedrive-bang-ke-excel.py --apply  # write to DB
"""
import argparse
import json
import os
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# Make backend/ importable so we can reuse the shared validator's whitelist
HERE = Path(__file__).resolve().parent
BACKEND_ROOT = HERE.parent
sys.path.insert(0, str(BACKEND_ROOT))

import importlib  # noqa: E402
customs_validator = importlib.import_module(
    "app.core.vietnamese-customs-declaration-codes-and-validator"
)

import openpyxl  # noqa: E402
from app.db.supabase_client import get_supabase  # noqa: E402


ONEDRIVE_T3 = Path(
    "/Users/bear1108/Library/CloudStorage/OneDrive-Personal/5P/DOANH THU/THÁNG 3"
)

# Header-column matching rules. Each key maps to a tuple (exact_matches,
# must_not_contain). A cell qualifies if its stripped/normalized text EQUALS
# any item in exact_matches and does NOT contain any word from must_not_contain.
# This avoids false positives like "Phí mở tờ khai" being matched as cd_no.
COL_RULES: Dict[str, Tuple[Tuple[str, ...], Tuple[str, ...]]] = {
    "cd_no": (
        ("tờ khai", "số tờ khai", "so to khai", "to khai", "số tkhq", "tkhq",
         "cd_no", "customs declaration", "số tk"),
        ("phí", "phi ", "fee"),
    ),
    "loai_hinh": (("loại hình", "loai hinh", "mã loại hình"), ()),
    "note": (("note", "ghi chú", "noi dung", "nội dung", "ghi chu"), ()),
    "invoice": (
        ("hóa đơn thương mại", "hoa don thuong mai", "invoice", "số invoice",
         "so hoa don", "số hóa đơn"),
        (),
    ),
    "bl": (
        ("vận đơn", "van don", "số vận đơn", "bl/awb", "bill", "b/l", "awb",
         "số bill"),
        (),
    ),
}


def matches_col_rule(cell_text: str, key: str) -> bool:
    """Return True if `cell_text` (normalized) matches the column rule for `key`."""
    exact, negatives = COL_RULES[key]
    norm = cell_text
    if any(neg in norm for neg in negatives):
        return False
    # Accept both exact match and "contains as a whole token" for flexibility.
    for ex in exact:
        if norm == ex or f" {ex} " in f" {norm} " or norm.endswith(" " + ex) or norm.startswith(ex + " "):
            return True
    return False


def normalize_text(s) -> str:
    """Lowercase + trim, for keyword matching."""
    if s is None:
        return ""
    return str(s).lower().strip()


def find_header_row(ws, max_scan_rows: int = 25) -> Tuple[Optional[int], Dict[str, int]]:
    """
    Scan the first `max_scan_rows` rows of `ws` looking for a header row
    that contains both a cd_no-like column and at least one other interesting column.
    Returns (row_index_1_based, {col_key: zero_based_col_index}).
    """
    best_row = None
    best_cols: Dict[str, int] = {}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1):
        cols_found: Dict[str, int] = {}
        for ci, cell in enumerate(row):
            if cell is None:
                continue
            norm = normalize_text(cell)
            for key in COL_RULES:
                if key in cols_found:
                    continue
                if matches_col_rule(norm, key):
                    cols_found[key] = ci
        if "cd_no" in cols_found and len(cols_found) > len(best_cols):
            best_row = i
            best_cols = cols_found
    return best_row, best_cols


def cell(row: Tuple, ci: Optional[int]):
    """Safe row index access."""
    if ci is None:
        return None
    if ci < 0 or ci >= len(row):
        return None
    return row[ci]


def collect_hints_from_xlsx(path: Path) -> Dict[str, List[str]]:
    """
    Read one xlsx, return {cd_no_str: [hint_text_1, hint_text_2, ...]}.
    Hints = concatenation of Loại hình + Note + Invoice + BL cells of the same row.
    """
    out: Dict[str, List[str]] = defaultdict(list)
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        print(f"  [WARN] cannot open {path.name}: {e}")
        return out

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row, cols = find_header_row(ws)
        if not header_row or "cd_no" not in cols:
            continue
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            cd_raw = cell(row, cols["cd_no"])
            if not cd_raw:
                continue
            cd_str = str(cd_raw).strip()
            # Accept only real VN customs declaration numbers (10-13 digits).
            # Filters out STT row numbers, column widths, etc.
            if not re.fullmatch(r"\d{10,13}", cd_str):
                continue
            parts = []
            for key in ("loai_hinh", "note", "invoice", "bl"):
                v = cell(row, cols.get(key))
                if v:
                    parts.append(str(v))
            # Also include EVERY non-empty string cell from the row as a fallback.
            # Different customers put the "loại hình" hint in ad-hoc columns like
            # "Nguồn" / "Ghi chú" / phí description cells — keyword matcher later
            # will pick the signal out. Skip numeric-only cells to keep noise low.
            for v in row:
                if isinstance(v, str) and v.strip() and v.strip() not in parts:
                    parts.append(v.strip())
            hint = " | ".join(p for p in parts if p)
            if hint:
                out[cd_str].append(hint)
    wb.close()
    return out


def build_excel_hint_index() -> Dict[str, List[str]]:
    """Walk every .xlsx/.xls file under ONEDRIVE_T3 and merge hint indexes."""
    index: Dict[str, List[str]] = defaultdict(list)
    for customer_dir in sorted(ONEDRIVE_T3.iterdir()):
        if not customer_dir.is_dir():
            continue
        for f in sorted(customer_dir.iterdir()):
            if f.suffix.lower() not in (".xlsx", ".xls"):
                continue
            # Skip "Copy of" duplicates heuristically — they're redundant
            if f.name.lower().startswith("copy of") or " - copy" in f.name.lower():
                continue
            per_file = collect_hints_from_xlsx(f)
            for cd, hints in per_file.items():
                index[cd].extend(hints)
    return index


def infer_loai_hinh(
    service_type_code: str,
    cd_no: Optional[str],
    hints_text: str,
    notes: str = "",
) -> Tuple[Optional[str], str, str]:
    """
    Rule-based inference. Returns (loai_hinh_code_or_None, confidence, rule_used).
    confidence: "HIGH" / "MED" / "LOW" / "NONE"
    """
    is_import = service_type_code.upper().startswith("CUS_IMPORT") or service_type_code.upper() == "CUS"
    is_export = service_type_code.upper().startswith("CUS_EXPORT")
    text = normalize_text(hints_text) + " " + normalize_text(notes)

    # Explicit mentions first (HIGH confidence)
    if "tạm xuất" in text or "tam xuat" in text:
        return "G14", "HIGH", "keyword:tạm xuất"
    if "tái xuất" in text or "tai xuat" in text:
        return "G24", "HIGH", "keyword:tái xuất"
    if re.search(r"\b(dom|tại chỗ|tai cho|xnk tc|xnk tại chỗ)\b", text):
        if is_export:
            return "B13", "HIGH", "keyword:tại chỗ+EXPORT"
        if is_import:
            return "A41", "HIGH", "keyword:tại chỗ+IMPORT"
    if "gia công" in text or "gia cong" in text:
        if is_export:
            return "E52", "MED", "keyword:gia công+EXPORT"
        if is_import:
            return "E11", "MED", "keyword:gia công+IMPORT"
    if "sxxk" in text or "sản xuất xuất khẩu" in text or "san xuat xuat khau" in text:
        if is_export:
            return "E42", "MED", "keyword:SXXK+EXPORT"
        if is_import:
            return "E21", "MED", "keyword:SXXK+IMPORT"
    # International transport keywords → kinh doanh
    if re.search(r"\b(sea|ocean|fcl|lcl|container|air|dhl|express|awb)\b", text):
        if is_export:
            return "B11", "MED", "keyword:international+EXPORT"
        if is_import:
            return "A11", "MED", "keyword:international+IMPORT"

    # cd_no prefix fallback (low confidence)
    if cd_no:
        prefix = cd_no[:3]
        # 308 in VN = xuất tại chỗ class; 108 = nhập tại chỗ class (observed empirically)
        if prefix == "308" and is_export:
            return "B13", "LOW", "cd_prefix:308+EXPORT"
        if prefix == "108" and is_import:
            return "A41", "LOW", "cd_prefix:108+IMPORT"

    return None, "NONE", "no-match"


def run(apply_changes: bool):
    client = get_supabase()

    # 1. Fetch legacy rows
    print("Fetching legacy CUS_* rows with loai_hinh IS NULL...")
    result = client.table("job_services").select(
        "svc_id,job_id,service_type_code,cd_no,bl_awb_no,invoice_numbers,special_requirements,"
        "jobs(job_no,customer_id,customers(customer_code,short_name))"
    ).or_(
        "service_type_code.like.CUS_IMPORT%,service_type_code.like.CUS_EXPORT%,service_type_code.eq.CUS"
    ).is_("loai_hinh", "null").execute()
    rows = result.data
    print(f"  Found {len(rows)} rows")

    # 2. Build Excel hint index
    print(f"Scanning Excel files under {ONEDRIVE_T3} ...")
    hint_index = build_excel_hint_index()
    print(f"  Indexed hints for {len(hint_index)} unique cd_no values")

    # 3. Infer per row
    report = {"HIGH": [], "MED": [], "LOW": [], "NONE": []}
    for r in rows:
        cd = (r.get("cd_no") or "").strip()
        hints = " ".join(hint_index.get(cd, [])) if cd else ""
        notes = r.get("special_requirements") or ""
        code, conf, rule = infer_loai_hinh(
            r["service_type_code"], cd, hints, notes
        )
        entry = {
            "svc_id": r["svc_id"],
            "job_no": (r.get("jobs") or {}).get("job_no"),
            "customer": ((r.get("jobs") or {}).get("customers") or {}).get("customer_code"),
            "service_type": r["service_type_code"],
            "cd_no": cd or None,
            "bl_awb_no": r.get("bl_awb_no"),
            "invoice": r.get("invoice_numbers"),
            "excel_hints_sample": (hints[:200] if hints else ""),
            "proposed_loai_hinh": code,
            "confidence": conf,
            "rule": rule,
        }
        report[conf].append(entry)

    # 4. Print summary
    print("\n=== Inference summary ===")
    for k in ("HIGH", "MED", "LOW", "NONE"):
        print(f"  {k}: {len(report[k])}")
    by_code = defaultdict(int)
    for conf in ("HIGH", "MED", "LOW"):
        for e in report[conf]:
            by_code[e["proposed_loai_hinh"]] += 1
    print("  Distribution of proposed codes:")
    for code, n in sorted(by_code.items(), key=lambda x: -x[1]):
        print(f"    {code}: {n}")

    # 5. Write full JSON for review
    out_path = BACKEND_ROOT.parent / "plans" / "reports" / "loai-hinh-backfill-dryrun.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(report, ensure_ascii=False, indent=2))
    print(f"\nDry-run report: {out_path}")

    # 6. Apply if requested — HIGH + MED only (per user decision 2026-04-21).
    # LOW uses cd_no prefix heuristic only and needs manual review.
    # NONE rows have no cd_no — fix via UI.
    if apply_changes:
        to_apply = report["HIGH"] + report["MED"]
        print(f"\nApplying {len(to_apply)} updates ({len(report['HIGH'])} HIGH + "
              f"{len(report['MED'])} MED) ...")
        applied = 0
        for e in to_apply:
            if not e["proposed_loai_hinh"]:
                continue
            client.table("job_services").update(
                {"loai_hinh": e["proposed_loai_hinh"]}
            ).eq("svc_id", e["svc_id"]).execute()
            applied += 1
        print(f"  Applied: {applied}")
        print(f"  LOW ({len(report['LOW'])}) left untouched — review manually")
        print(f"  NONE ({len(report['NONE'])}) left untouched — no cd_no / no hint")
    else:
        print("\n(dry-run) Re-run with --apply to write HIGH + MED updates to DB.")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true",
                    help="Write inferred loai_hinh (HIGH + MED confidence only) to DB.")
    args = ap.parse_args()

    # Load .env so supabase client picks up credentials
    env_file = BACKEND_ROOT.parent / ".env"
    if env_file.exists():
        for line in env_file.read_text().splitlines():
            if "=" in line and not line.strip().startswith("#"):
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())

    run(args.apply)
