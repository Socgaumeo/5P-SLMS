"""Analyze Batch 3 Excel files - structure, headers, data fields, service types."""
import sys
import os

# Files to analyze
files = [
    "/Users/bear1108/Documents/Tháng 3/NIPPON/(THAI NGUYEN) BẢNG KÊ CHI PHÍ NIPPON THÁNG 3.2026.xls",
    "/Users/bear1108/Documents/Tháng 3/NIPPON/BẢNG KÊ CHI PHÍ NIPPON THÁNG 3.2026 rv.xlsx",
    "/Users/bear1108/Documents/Tháng 3/TDI/Copy of BangKe_TDI_AirT3_2026_ final1.xlsx",
    "/Users/bear1108/Documents/Tháng 3/TDI/TDI of BẢNG THEO DÕI T03.2026 bs_ZKL.xlsx",
    "/Users/bear1108/Documents/Tháng 3/THÁI HOÀ/Debit_5PVN_THAI_HOA_T3_2026 (9).xlsx",
    "/Users/bear1108/Documents/Tháng 3/THÁI HOÀ/Debit_TCH_5PVN_THÁI HÒA_T3_2026_full (4) - Copy.xlsx",
    "/Users/bear1108/Documents/Tháng 3/TVC/Debit Note. 5P. TVC. T03.2026.xlsx",
    "/Users/bear1108/Documents/Tháng 3/UTRACORN/DebitNote_UTRACON_TRK1403_DRAFT (3).xlsx",
    "/Users/bear1108/Documents/Tháng 3/VINTECH/Debit note VINTECH-5P T3.2026.xlsx",
    "/Users/bear1108/Documents/Tháng 3/VINTECH/DebitNote_VINTECH_NGB637324_DRAFT.xlsx",
    "/Users/bear1108/Documents/Tháng 3/XÂY LẮP VN/Debit note XÂY LẮP VN-5P T3.2026.xlsx",
]


def analyze_xlsx(filepath):
    """Analyze .xlsx file using openpyxl."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    results = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        row_count = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 30:  # Capture first 30 rows for header analysis
                rows.append(row)
            row_count += 1
        results[sheet_name] = {
            "total_rows": row_count,
            "first_rows": rows,
        }
    wb.close()
    return results


def analyze_xls(filepath):
    """Analyze .xls file using xlrd."""
    import xlrd
    wb = xlrd.open_workbook(filepath)
    results = {}
    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        rows = []
        for i in range(min(30, ws.nrows)):
            row = []
            for j in range(ws.ncols):
                cell = ws.cell(i, j)
                row.append(cell.value)
            rows.append(tuple(row))
        results[sheet_name] = {
            "total_rows": ws.nrows,
            "first_rows": rows,
        }
    return results


def format_cell(val):
    """Format cell value for display."""
    if val is None:
        return ""
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return f"{val:,.2f}"
    return str(val).strip()


def print_analysis(file_idx, filepath, sheet_data):
    """Print structured analysis."""
    fname = os.path.basename(filepath)
    print(f"\n{'='*120}")
    print(f"FILE {file_idx}: {fname}")
    print(f"PATH: {filepath}")
    print(f"{'='*120}")

    for sheet_name, data in sheet_data.items():
        print(f"\n  --- Sheet: '{sheet_name}' | Total rows: {data['total_rows']} ---")
        rows = data["first_rows"]
        if not rows:
            print("    (empty sheet)")
            continue

        # Print first 30 rows with row index
        max_cols = max(len(r) for r in rows) if rows else 0
        for i, row in enumerate(rows):
            cells = [format_cell(c) for c in row]
            # Truncate long cells for readability
            cells_display = []
            for c in cells:
                if len(c) > 50:
                    cells_display.append(c[:47] + "...")
                else:
                    cells_display.append(c)
            line = " | ".join(cells_display)
            print(f"    Row {i:3d}: {line}")

        print(f"    ... (total {data['total_rows']} rows)")


def main():
    for idx, filepath in enumerate(files, 1):
        if not os.path.exists(filepath):
            print(f"\n{'='*120}")
            print(f"FILE {idx}: {os.path.basename(filepath)}")
            print(f"  *** FILE NOT FOUND ***")
            continue

        ext = os.path.splitext(filepath)[1].lower()
        try:
            if ext == ".xls":
                data = analyze_xls(filepath)
            else:
                data = analyze_xlsx(filepath)
            print_analysis(idx, filepath, data)
        except Exception as e:
            print(f"\n{'='*120}")
            print(f"FILE {idx}: {os.path.basename(filepath)}")
            print(f"  *** ERROR: {e} ***")
            import traceback
            traceback.print_exc()


if __name__ == "__main__":
    main()
