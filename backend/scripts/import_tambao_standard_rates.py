#!/usr/bin/env python3
"""
Import Tam Bảo standard vehicle rates to Supabase
"""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dotenv import load_dotenv
import openpyxl

load_dotenv(os.path.join(os.path.dirname(__file__), '..', '..', '.env'))

from supabase import create_client

url = os.getenv('SUPABASE_URL')
key = os.getenv('SUPABASE_SERVICE_ROLE_KEY')
client = create_client(url, key)

# Tam Bảo vendor - using vendor_id 62 (TAMBAO2) which matches tax code 0108988307
VENDOR_ID = 62
FILE_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'vendor_rates', 'Tam bảo_092025.xlsx')

# Vehicle types mapping from header
VEHICLE_TYPES = ['1.25T', '1.5T', '2.5T', '3.5T', '5T', '7T', '8T', '10T', '15T', '20T', 'MOOC']
VEHICLE_COL_START = 8  # Column H


def main():
    print(f"Loading file: {FILE_PATH}")
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    ws = wb['Báo Giá Xe Thường']

    print(f"Sheet: {ws.title} - {ws.max_row} rows")

    # Parse data rows (start from row 14)
    rates = []

    for row_idx in range(14, ws.max_row + 1):
        stt = ws.cell(row=row_idx, column=1).value
        origin = ws.cell(row=row_idx, column=2).value
        origin_province = ws.cell(row=row_idx, column=3).value
        destination = ws.cell(row=row_idx, column=4).value
        dest_province = ws.cell(row=row_idx, column=5).value
        rate_code = ws.cell(row=row_idx, column=7).value

        # Skip invalid rows
        if not stt or not origin or not str(stt).strip().isdigit():
            continue

        # Parse prices for each vehicle type
        for idx, vehicle_type in enumerate(VEHICLE_TYPES):
            col = VEHICLE_COL_START + idx
            if col > ws.max_column:
                break

            price = ws.cell(row=row_idx, column=col).value
            if price is None or price == '':
                continue

            # Convert price to float
            try:
                price_value = float(price)
            except (ValueError, TypeError):
                continue

            rate = {
                'vendor_id': VENDOR_ID,
                'vehicle_type': vehicle_type,
                'price': price_value,
                'currency': 'VND',
                'unit': 'TRIP',
                'effective_date': '2025-09-01',
                'is_active': True,
                'rate_code': rate_code,
                'rate_type': 'STANDARD',
                'origin_province': origin_province,
                'destination_province': dest_province,
                'notes': f"{origin} -> {destination}"
            }
            rates.append(rate)

        if row_idx % 500 == 0:
            print(f"  Parsed {row_idx} rows...")

    print(f"\n=== Total rates to import: {len(rates)} ===")

    # Insert in batches
    batch_size = 100
    success = 0
    errors = 0

    for i in range(0, len(rates), batch_size):
        batch = rates[i:i + batch_size]
        try:
            result = client.table('vendor_rates').insert(batch).execute()
            success += len(batch)
            print(f"  Inserted batch {i // batch_size + 1}: {len(batch)} records")
        except Exception as e:
            print(f"  Error batch {i // batch_size + 1}: {e}")
            # Try inserting one by one
            for rate in batch:
                try:
                    client.table('vendor_rates').insert(rate).execute()
                    success += 1
                except Exception as e2:
                    errors += 1
                    print(f"    Error: {rate['rate_code']} - {e2}")

    print(f"\n=== DONE ===")
    print(f"Success: {success}, Errors: {errors}")

    # Verify
    result = client.table('vendor_rates').select('*', count='exact').eq(
        'vendor_id', VENDOR_ID
    ).eq('rate_type', 'STANDARD').execute()
    print(f"\nTotal STANDARD rates for TAMBAO2: {result.count}")


if __name__ == "__main__":
    main()
