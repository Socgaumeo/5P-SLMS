"""
Complex parsers for March 2026 import: DAINESE, TDI, GLOREX, NIPPON, MESSER,
LAS, KK, KCVN, KWE.
Each function returns list of job_data dicts (see main script for schema).
"""
import os
import datetime
import traceback

import openpyxl
import xlrd

BASE_DIR = "/Users/bear1108/Documents/Tháng 3"

CUSTOMER_MAP = {
    "DAINESE": 46, "GLOREX": 18, "KCVN": 61, "KK": 65, "KWE": 28,
    "LAS": 6, "MESSER": 22, "NIPPON": 64, "TDI": 20,
}


def s(v):
    if v is None:
        return ""
    return str(v).strip().replace("\n", " ").replace("\r", "")[:200]


def n(v):
    if v is None:
        return 0.0
    if isinstance(v, bool):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0


def d(v, datemode=0):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, (int, float)) and 40000 < v < 60000:
        try:
            return xlrd.xldate_as_datetime(v, datemode).date()
        except Exception:
            pass
    if isinstance(v, str):
        v = v.strip().split(" ")[0]
        for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"]:
            try:
                return datetime.datetime.strptime(v, fmt).date()
            except ValueError:
                continue
    return None


def cell(ws, r, c):
    return ws.cell(row=r, column=c).value


# ──────── DAINESE ────────

def parse_dainese():
    """
    5 files for DAINESE (customer_id=46):
    1. CO file (DỊCH VỤ sheet): 2 CO jobs with thu hộ lệ phí
    2. Import file (NHẬP sheet): multi-fee columns per shipment (SEA/AIR)
    3. TC/CPN file (tc, cpn sheet): DOM customs multi-fee columns
    4. Trucking file (HĐ sheet): route-based trucking jobs
    5. Export file (XUẤT sheet): export customs multi-fee columns
    """
    cid = CUSTOMER_MAP["DAINESE"]
    jobs = []

    # ── File 1: CO ──
    try:
        fp = f"{BASE_DIR}/DAINESE/BẢNG KÊ PHÍ CO DAINESE T3.2026. 5P.xlsx"
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb["DỊCH VỤ"]
        # Header R13, sub-header R14, data R15+
        for r in range(15, ws.max_row + 1):
            stt = cell(ws, r, 1)
            if isinstance(stt, str) and "tổng" in stt.lower():
                break
            if not isinstance(stt, (int, float)):
                continue
            date_val = d(cell(ws, r, 2))
            invoice = s(cell(ws, r, 3))  # C3=INVOICE
            co_no = s(cell(ws, r, 4))    # C4=SỐ CO
            unit_price = n(cell(ws, r, 7))  # C7=ĐƠN GIÁ
            pre_vat = n(cell(ws, r, 8))     # C8=THÀNH TIỀN
            hd = s(cell(ws, r, 9))          # C9=SỐ HÓA ĐƠN
            thu_ho_nd = s(cell(ws, r, 11))  # C11=NỘI DUNG thu chi hộ
            thu_ho_amt = n(cell(ws, r, 12)) # C12=SỐ TIỀN thu chi hộ
            thu_ho_bl = s(cell(ws, r, 13))  # C13=SỐ BIÊN LAI
            costs = [{"name": f"Phí C/O {invoice}", "amount": pre_vat, "vat_rate": 0}]
            if thu_ho_amt > 0:
                costs.append({
                    "name": f"Thu hộ: {thu_ho_nd} - BL: {thu_ho_bl}",
                    "amount": thu_ho_amt,
                    "vat_rate": 0,
                    "is_reimbursement": True,
                })
            jobs.append({
                "customer_id": cid,
                "date": date_val or datetime.date(2026, 3, 1),
                "description": f"C/O - Invoice: {invoice} - CO: {co_no}",
                "svc_type": "CUS_CO",
                "invoice": hd,
                "service_details": {"co_no": co_no, "unit_price": unit_price},
                "costs": costs,
            })
    except Exception as e:
        print(f"  DAINESE CO error: {e}")

    # ── File 2: Import (NHẬP) ──
    try:
        fp = f"{BASE_DIR}/DAINESE/Bảng kê nhập tháng 3.2026.sea.air.xlsx"
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb["NHẬP"]
        # Header R12: No.|TK|HĐTM|Vận đơn|Ngày TK|Tuyến đường|Loại hình|Kgs|No.Cont|Note
        # Fee cols C11-C22: C11=Phí mở TK,C12=Kiểm hóa,C13=Vận chuyển,C14=Làm hàng,
        #                   C15=Phát sinh,C16=Nước ngoài,C17=Cước QT,C18=THC,C19=CFS,C20=DO,C21=Đại lý,C22=Tổng
        FEE_COLS = {
            11: "Phí mở tờ khai HQ",
            12: "Phí kiểm hóa",
            13: "Phí vận chuyển",
            14: "Phí làm hàng",
            15: "Phí phát sinh khác",
            16: "Phí đầu nước ngoài",
            17: "Cước vận tải quốc tế",
            18: "Phí xếp dỡ (THC)",
            19: "Phí gom hàng lẻ (CFS)",
            20: "Phí lấy lệnh (DO)",
            21: "Phí đại lý",
        }
        LOCAL_CHARGE_COL = 23  # Local charge (extra)
        for r in range(13, ws.max_row + 1):
            stt = cell(ws, r, 1)
            if isinstance(stt, str) and any(k in stt.lower() for k in ["tổng", "total"]):
                break
            if not isinstance(stt, (int, float)):
                continue
            cd_no = s(cell(ws, r, 2))
            hd_tm = s(cell(ws, r, 3))
            awb = s(cell(ws, r, 4))
            date_val = d(cell(ws, r, 5))
            route = s(cell(ws, r, 6))
            mode = s(cell(ws, r, 7)).upper()  # SEA or AIR
            kgs = n(cell(ws, r, 8))
            note = s(cell(ws, r, 10))
            svc_type = "AIR_IMP" if mode == "AIR" else "SEA_IMP"
            costs = []
            for col, name in FEE_COLS.items():
                amt = n(cell(ws, r, col))
                if amt > 0:
                    costs.append({"name": name, "amount": amt, "vat_rate": 8})
            # Local charge (col 23) — some have it
            lc = n(cell(ws, r, LOCAL_CHARGE_COL)) if ws.max_column >= LOCAL_CHARGE_COL else 0
            if lc > 0:
                costs.append({"name": "Local charge", "amount": lc, "vat_rate": 8})
            if not costs:
                continue
            jobs.append({
                "customer_id": cid,
                "date": date_val or datetime.date(2026, 3, 1),
                "description": f"{svc_type.replace('_', ' ')} - TK: {cd_no} - Bill: {awb}",
                "svc_type": svc_type,
                "cd_no": cd_no,
                "bl_awb": awb,
                "route": route,
                "weight": kgs,
                "customs_type": "IMPORT",
                "service_details": {"invoice_hd": hd_tm, "note": note, "container": s(cell(ws, r, 9))},
                "costs": costs,
            })
    except Exception as e:
        print(f"  DAINESE import error: {e}")
        traceback.print_exc()

    # ── File 3: TC/CPN (tc, cpn sheet) ──
    try:
        fp = f"{BASE_DIR}/DAINESE/Bảng kê tháng 3.2026. tc.nhap cpn.xlsx"
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb["tc, cpn"]
        # Same column layout as NHẬP but mode=DOM, svc=CUS_EXPORT
        # Fee cols same C11-C22, col 22=Tổng
        FEE_COLS_CPN = {
            11: "Phí mở tờ khai HQ",
            12: "Phí kiểm hóa",
            13: "Phí vận chuyển",
            14: "Phí làm hàng",
            15: "Phí phát sinh khác",
            16: "Phí phục vụ kiểm hóa tại cảng",
            17: "Phí đầu nước ngoài",
            18: "Cước vận tải quốc tế",
            19: "Phí xếp dỡ (THC)",
            20: "Phí gom hàng lẻ (CFS)",
            21: "Phí lấy lệnh (DO)",
        }
        LEP_PHI_COL = 25  # C25=Lệ phí (thu hộ in some rows)
        for r in range(13, ws.max_row + 1):
            stt = cell(ws, r, 1)
            if isinstance(stt, str) and any(k in stt.lower() for k in ["tổng", "total"]):
                break
            if not isinstance(stt, (int, float)):
                continue
            cd_no = s(cell(ws, r, 2))
            hd_tm = s(cell(ws, r, 3))
            note = s(cell(ws, r, 4))
            date_val = d(cell(ws, r, 5))
            route = s(cell(ws, r, 6))
            mode = s(cell(ws, r, 7)).upper()
            kgs = n(cell(ws, r, 8))
            luong = s(cell(ws, r, 10))
            costs = []
            for col, name in FEE_COLS_CPN.items():
                amt = n(cell(ws, r, col))
                if amt > 0:
                    costs.append({"name": name, "amount": amt, "vat_rate": 8})
            # Lệ phí col (thu hộ)
            lp = n(cell(ws, r, LEP_PHI_COL)) if ws.max_column >= LEP_PHI_COL else 0
            if lp > 0:
                costs.append({
                    "name": "Thu hộ: Lệ phí hải quan",
                    "amount": lp, "vat_rate": 0, "is_reimbursement": True,
                })
            if not costs:
                continue
            jobs.append({
                "customer_id": cid,
                "date": date_val or datetime.date(2026, 3, 1),
                "description": f"XNK tại chỗ/CPN - TK: {cd_no}",
                "svc_type": "CUS_EXPORT",
                "cd_no": cd_no,
                "bl_awb": hd_tm,
                "route": route,
                "weight": kgs,
                "customs_type": "EXPORT",
                "service_details": {"note": note, "customs_channel": luong},
                "costs": costs,
            })
    except Exception as e:
        print(f"  DAINESE TC/CPN error: {e}")

    # ── File 4: Trucking (HĐ sheet) ──
    try:
        fp = f"{BASE_DIR}/DAINESE/Copy of (DAINESE-5PVN) BẢNG KÊ TT T3.2026 bs2.xlsx"
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb["HĐ"]
        # Header R13: STT|Ngày|Điểm lấy(C3,C4)|Điểm trả(C5,C6)|BKS|ĐVT|SL|Phát sinh|Cước VC|Phụ phí|Thành tiền|Tổng|Note|Yêu cầu|Số HĐ
        for r in range(15, ws.max_row + 1):
            stt = cell(ws, r, 1)
            if isinstance(stt, str) and any(k in stt.lower() for k in ["tổng", "total"]):
                break
            if not isinstance(stt, (int, float)):
                continue
            date_val = d(cell(ws, r, 2))
            origin = f"{s(cell(ws,r,3))}, {s(cell(ws,r,4))}".strip(", ")
            dest = f"{s(cell(ws,r,5))}, {s(cell(ws,r,6))}".strip(", ")
            bks = s(cell(ws, r, 7))
            unit = s(cell(ws, r, 8))
            qty = n(cell(ws, r, 9))
            extra = n(cell(ws, r, 10))    # C10=Phát sinh
            cuoc_vc = n(cell(ws, r, 11))  # C11=Cước vận chuyển
            thanh_tien = n(cell(ws, r, 13))  # C13=Thành tiền (pre-VAT total)
            hd = s(cell(ws, r, 17))       # C17=Số hóa đơn
            note = s(cell(ws, r, 15))
            costs = [{"name": f"Cước vận chuyển {origin} → {dest}", "amount": cuoc_vc, "vat_rate": 8}]
            if extra > 0:
                costs.append({"name": "Phí phát sinh", "amount": extra, "vat_rate": 8})
            jobs.append({
                "customer_id": cid,
                "date": date_val or datetime.date(2026, 3, 1),
                "description": f"Vận chuyển {origin} → {dest} - BKS: {bks}",
                "svc_type": "TRUCKING_DOM",
                "origin": origin, "dest": dest,
                "bl_awb": bks,
                "invoice": hd,
                "service_details": {
                    "vehicle_plate": bks, "vehicle_type": unit,
                    "quantity": qty, "note": note,
                },
                "costs": costs,
            })
    except Exception as e:
        print(f"  DAINESE trucking error: {e}")

    # ── File 5: Export (XUẤT sheet) ──
    try:
        fp = f"{BASE_DIR}/DAINESE/Copy of Bảng kê xuất tháng 3.2026 final.xlsx"
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb["XUẤT"]
        # Header R12: No.|TK|HĐTM|Note|Ngày TK|Tuyến đường|Loại hình|Kgs|No.Cont|Luồng
        # Fee cols C11-C23+: C11=Mở TK,C12=Kiểm hóa,C13=Vận chuyển,C14=Làm hàng,
        #   C15=Phát sinh,C16=Phục vụ KH,C17=Nước ngoài,C18=Cước QT,C19=THC,
        #   C20=CFS,C21=DO,C22=Đại lý,C23=Phụ phí XD,C24=Tổng,C25=Local charge
        FEE_COLS_EXP = {
            11: "Phí mở tờ khai HQ",
            12: "Phí kiểm hóa",
            13: "Phí vận chuyển",
            14: "Phí làm hàng",
            15: "Phí phát sinh khác",
            16: "Phí phục vụ kiểm hóa tại cảng",
            17: "Phí đầu nước ngoài",
            18: "Cước vận tải quốc tế",
            19: "Phí xếp dỡ (THC)",
            20: "Phí gom hàng lẻ (CFS)",
            21: "Phí lấy lệnh (DO)",
            22: "Phí đại lý",
            23: "Phụ phí xăng dầu",
        }
        for r in range(13, ws.max_row + 1):
            stt = cell(ws, r, 1)
            if isinstance(stt, str) and any(k in stt.lower() for k in ["tổng", "total"]):
                break
            if not isinstance(stt, (int, float)):
                continue
            cd_no = s(cell(ws, r, 2))
            hd_tm = s(cell(ws, r, 3))
            note = s(cell(ws, r, 4))
            date_val = d(cell(ws, r, 5))
            route = s(cell(ws, r, 6))
            mode = s(cell(ws, r, 7)).upper()
            kgs_raw = cell(ws, r, 8)
            kgs = n(kgs_raw) if not isinstance(kgs_raw, str) else 0
            luong = s(cell(ws, r, 10))
            costs = []
            for col, fname in FEE_COLS_EXP.items():
                amt = n(cell(ws, r, col))
                if amt > 0:
                    costs.append({"name": fname, "amount": amt, "vat_rate": 8})
            lc = n(cell(ws, r, 25)) if ws.max_column >= 25 else 0
            if lc > 0:
                costs.append({"name": "Local charge", "amount": lc, "vat_rate": 8})
            if not costs:
                continue
            svc_type = "CUS_EXPORT"
            if mode in ("DHL", "KNQ"):
                svc_type = "CUS_CO"
            jobs.append({
                "customer_id": cid,
                "date": date_val or datetime.date(2026, 3, 1),
                "description": f"Xuất khẩu - TK: {cd_no} - {note}",
                "svc_type": svc_type,
                "cd_no": cd_no,
                "bl_awb": hd_tm,
                "route": route,
                "weight": kgs,
                "customs_type": "EXPORT",
                "service_details": {"note": note, "customs_channel": luong, "mode": mode},
                "costs": costs,
            })
    except Exception as e:
        print(f"  DAINESE export error: {e}")
        traceback.print_exc()

    return jobs


# ──────── TDI ────────

def parse_tdi():
    """
    TDI (customer_id=20): 2 files.
    File 1 has 3 sheets:
      - Sheet 1 (TDI LÊN HÓA ĐƠN 1): main air import jobs with services
      - Sheet 2 (TDI LÊN HÓA ĐƠN  2): dangerous goods surcharge per bill
      - Sheet 3 (TDI THU CHI HỘ1 ): thu chi hộ items per bill
    File 2: ZKL customs declarations (T3.2026 sheet only)
    Strategy: build jobs from sheet 1, then merge sheet 2 & 3 by bill number.
    """
    cid = CUSTOMER_MAP["TDI"]
    jobs = []

    # ── File 1: Air import ──
    try:
        fp = f"{BASE_DIR}/TDI/Copy of BangKe_TDI_AirT3_2026_ final1.xlsx"
        wb = openpyxl.load_workbook(fp, data_only=True)

        # Build dangerous-goods lookup: bill → cost item
        dg_by_bill = {}
        ws2 = wb["TDI LÊN HÓA ĐƠN  2"]
        # Header R12-13, data R14+
        # Cols: STT|Ngày|Số TK|Số Bill|...|Dịch vụ|SL|Đơn giá|ĐVT|Thành tiền|Tax|Tổng|HĐ
        for r in range(14, ws2.max_row + 1):
            stt = cell(ws2, r, 1)
            if isinstance(stt, str) and any(k in stt.lower() for k in ["tổng", "total"]):
                break
            if not isinstance(stt, (int, float)):
                continue
            bill2 = s(cell(ws2, r, 4))
            svc_name = s(cell(ws2, r, 7))
            pre_vat2 = n(cell(ws2, r, 11))  # C11=Thành tiền
            hd2 = s(cell(ws2, r, 14))
            if bill2:
                key = _normalize_bill(bill2)
                if key not in dg_by_bill:
                    dg_by_bill[key] = []
                dg_by_bill[key].append({"name": svc_name, "amount": pre_vat2, "vat_rate": 8, "invoice": hd2})

        # Build thu chi hộ lookup: bill → list of costs
        tch_by_bill = {}
        ws3 = wb["TDI THU CHI HỘ1 "]
        # Cols: STT|Ngày|Số Bill|Dịch vụ|Số HĐ|Thành tiền|Tổng|Ghi chú
        for r in range(14, ws3.max_row + 1):
            stt = cell(ws3, r, 1)
            if isinstance(stt, str) and any(k in stt.lower() for k in ["tổng", "total"]):
                break
            if not isinstance(stt, (int, float)):
                continue
            bill3 = s(cell(ws3, r, 3))
            svc3 = s(cell(ws3, r, 4))
            hd3 = s(cell(ws3, r, 5))
            amt3 = n(cell(ws3, r, 6))
            if bill3 and amt3 > 0:
                key = _normalize_bill(bill3)
                if key not in tch_by_bill:
                    tch_by_bill[key] = []
                tch_by_bill[key].append({
                    "name": f"Thu hộ: {svc3}",
                    "amount": amt3,
                    "vat_rate": 0,
                    "is_reimbursement": True,
                    "invoice": hd3,
                })

        # Sheet 1: main jobs
        # Cols R12-13: STT|Ngày|Số TK|Số Bill|Số REF|Trọng lượng|Dịch vụ|SL|Đơn giá|Phụ phí XD|ĐVT|Thành tiền|Tax|Tổng|HĐ|Chi hộ Số HĐ|Chi hộ Thành tiền|Tổng TT|Ghi chú
        ws1 = wb["TDI LÊN HÓA ĐƠN 1"]
        r = 14
        while r <= ws1.max_row:
            stt = cell(ws1, r, 1)
            if isinstance(stt, str) and any(k in stt.lower() for k in ["tổng", "total"]):
                break
            if not isinstance(stt, (int, float)):
                r += 1
                continue
            date_val = d(cell(ws1, r, 2))
            cd_no = s(cell(ws1, r, 3))
            bill = s(cell(ws1, r, 4))
            ref = s(cell(ws1, r, 5))
            weight_raw = cell(ws1, r, 6)
            weight = n(str(weight_raw).split("\n")[0]) if weight_raw else 0
            note = s(cell(ws1, r, 19)) if ws1.max_column >= 19 else ""

            # Collect sub-rows (same bill, no STT)
            costs = []
            sr = r
            while sr <= ws1.max_row:
                sub_stt = cell(ws1, sr, 1)
                if sr != r and sub_stt is not None:
                    break
                svc_name = s(cell(ws1, sr, 7))
                pre_vat_sr = n(cell(ws1, sr, 12))  # C12=Thành tiền
                phuphi_xd = n(cell(ws1, sr, 10))   # C10=Phụ phí xăng dầu
                hd_sr = s(cell(ws1, sr, 15))
                total_amt = pre_vat_sr + phuphi_xd
                if svc_name and total_amt > 0:
                    costs.append({
                        "name": svc_name, "amount": total_amt,
                        "vat_rate": 8, "invoice": hd_sr,
                    })
                # Thu chi hộ inline (C16=Số HĐ, C17=Thành tiền)
                tch_inline_hd = s(cell(ws1, sr, 16))
                tch_inline_amt = n(cell(ws1, sr, 17))
                if tch_inline_amt > 0:
                    tch_name = s(cell(ws1, sr, 7)) or "Thu hộ"
                    costs.append({
                        "name": f"Thu hộ: {tch_name}",
                        "amount": tch_inline_amt,
                        "vat_rate": 0,
                        "is_reimbursement": True,
                        "invoice": tch_inline_hd,
                    })
                sr += 1

            # Merge dangerous goods
            bill_key = _normalize_bill(bill)
            for dg in dg_by_bill.get(bill_key, []):
                costs.append(dg)

            # Merge thu chi hộ from sheet 3
            for tch in tch_by_bill.get(bill_key, []):
                # Avoid duplicates from inline thu hộ
                costs.append(tch)

            if costs:
                jobs.append({
                    "customer_id": cid,
                    "date": date_val or datetime.date(2026, 3, 1),
                    "description": f"Air import - Bill: {bill} - TK: {cd_no}",
                    "svc_type": "AIR_IMP",
                    "cd_no": cd_no,
                    "bl_awb": bill,
                    "weight": weight,
                    "customs_type": "IMPORT",
                    "service_details": {"ref": ref, "note": note},
                    "costs": costs,
                })
            r = sr

    except Exception as e:
        print(f"  TDI air error: {e}")
        traceback.print_exc()

    # ── File 2: ZKL customs (T3.2026 sheet only) ──
    try:
        fp2 = f"{BASE_DIR}/TDI/TDI of BẢNG THEO DÕI T03.2026 bs_ZKL.xlsx"
        wb2 = openpyxl.load_workbook(fp2, data_only=True)
        ws_zkl = wb2["T3.2026"]
        # Header R20: STT|Dịch vụ|Ngày|Số quyết toán|Số TK|Phân luồng|Phí thông quan|Kiểm hóa|Ngoài giờ|Chi phí khác|Số tiền|8% VAT|Tổng|Thu hộ|Ghi chú
        for r in range(21, ws_zkl.max_row + 1):
            stt = cell(ws_zkl, r, 1)
            if isinstance(stt, str) and any(k in stt.lower() for k in ["tổng", "total"]):
                break
            if not isinstance(stt, (int, float)):
                continue
            svc_name = s(cell(ws_zkl, r, 2))
            date_val = d(cell(ws_zkl, r, 3))
            cd_no = s(cell(ws_zkl, r, 5))
            luong = s(cell(ws_zkl, r, 6))
            pre_vat = n(cell(ws_zkl, r, 11))   # C11=Số tiền (pre-VAT)
            thu_ho = n(cell(ws_zkl, r, 14))
            costs = [{"name": svc_name or "Dịch vụ thủ tục HQ", "amount": pre_vat, "vat_rate": 8}]
            if thu_ho > 0:
                costs.append({
                    "name": "Thu hộ: Lệ phí hải quan",
                    "amount": thu_ho, "vat_rate": 0, "is_reimbursement": True,
                })
            jobs.append({
                "customer_id": cid,
                "date": date_val or datetime.date(2026, 3, 1),
                "description": f"Thủ tục HQ xuất khẩu - TK: {cd_no} ({luong})",
                "svc_type": "CUS_EXPORT",
                "cd_no": cd_no,
                "customs_type": "EXPORT",
                "service_details": {"customs_channel": luong},
                "costs": costs,
            })
    except Exception as e:
        print(f"  TDI ZKL error: {e}")

    return jobs


def _normalize_bill(bill_str):
    """Normalize bill number for matching (strip spaces, newlines, dashes)."""
    return s(bill_str).replace("-", "").replace(" ", "").upper()[:20]


# ──────── GLOREX ────────

def parse_glorex():
    """
    GLOREX (customer_id=18): 3 files (skip GLOBAL file).
    File 1 (GLOREX sheet): international logistics, grouped cost rows per shipment.
    File 2 (XNK TC sheet): customs tại chỗ, 10 declarations.
    File 3 (NHAP KHAU sheet): thu hộ lệ phí HQ, 1 row per TK (match by TK).
    """
    cid = CUSTOMER_MAP["GLOREX"]
    jobs = []

    # File 3 first: build thu hộ map keyed by TK
    thu_ho_map = {}
    try:
        fp3 = f"{BASE_DIR}/GLOREX/Debit_TCH_5PVN_GLOREX_T3_2026_full (4).xlsx"
        wb3 = openpyxl.load_workbook(fp3, data_only=True)
        ws3 = wb3["NHAP KHAU"]
        # Cols: STT|Ngày|Loại xe|Bill|Lấy|Trả|Tờ khai|Nội dung|ĐVT|SL|Đơn giá|Số tiền|VAT|Tổng|Số GNT
        for r in range(15, ws3.max_row + 1):
            stt = cell(ws3, r, 1)
            if isinstance(stt, str) and "tổng" in stt.lower():
                break
            if not isinstance(stt, (int, float)):
                continue
            tk3 = s(cell(ws3, r, 7))   # C7=Tờ khai
            amt3 = n(cell(ws3, r, 12)) # C12=Số tiền
            gnt = s(cell(ws3, r, 15))
            # Row 9 in file has qty=2 for 2 TKs on same row
            qty3 = n(cell(ws3, r, 10))
            if tk3 and amt3 > 0:
                # Normalize TK key (truncated in file)
                key = tk3.strip()[:11]
                thu_ho_map[key] = {"amount": amt3 / max(1, qty3), "gnt": gnt, "full_row_amt": amt3}
    except Exception as e:
        print(f"  GLOREX thu hộ error: {e}")

    # File 1: International logistics
    try:
        fp1 = f"{BASE_DIR}/GLOREX/Debit 5PVN_GLOREX 3.2026.QUỐC TẾ.xlsx"
        wb1 = openpyxl.load_workbook(fp1, data_only=True)
        ws1 = wb1["GLOREX"]
        # Header R12-13: STT|Ngày|Loại xe|Booking/Bill/BKS|Lấy|Trả|TK|Nội dung|ĐVT|SL|Đơn giá|Số tiền|VAT|Tổng|Số HĐ|Note
        r = 15
        while r <= ws1.max_row:
            stt = cell(ws1, r, 1)
            if isinstance(stt, str) and any(k in stt.lower() for k in ["tổng", "total"]):
                break
            if not isinstance(stt, (int, float)):
                r += 1
                continue
            date_val = d(cell(ws1, r, 2))
            vehicle = s(cell(ws1, r, 3))
            bill = s(cell(ws1, r, 4))
            origin = s(cell(ws1, r, 5))
            dest = s(cell(ws1, r, 6))
            cd_no = s(cell(ws1, r, 7))

            # Collect all sub-rows until next STT
            costs = []
            sr = r
            while sr <= ws1.max_row:
                sub_stt = cell(ws1, sr, 1)
                if sr != r and sub_stt is not None:
                    break
                svc_name = s(cell(ws1, sr, 8))   # C8=Nội dung
                pre_vat_sr = n(cell(ws1, sr, 12)) # C12=Số tiền (pre-VAT)
                is_tra_ho = "trả hộ" in svc_name.lower() or "thu hộ" in svc_name.lower()
                hd_sr = s(cell(ws1, sr, 15))
                if svc_name and pre_vat_sr > 0:
                    costs.append({
                        "name": svc_name,
                        "amount": pre_vat_sr,
                        "vat_rate": 0 if is_tra_ho else 8,
                        "is_reimbursement": is_tra_ho,
                        "invoice": hd_sr,
                    })
                sr += 1

            if costs:
                jobs.append({
                    "customer_id": cid,
                    "date": date_val or datetime.date(2026, 3, 1),
                    "description": f"Logistics quốc tế - Bill: {bill} - TK: {cd_no}",
                    "svc_type": "SEA_IMP",
                    "bl_awb": bill,
                    "cd_no": cd_no,
                    "origin": origin, "dest": dest,
                    "service_details": {"vehicle_type": vehicle},
                    "costs": costs,
                })
            r = sr
    except Exception as e:
        print(f"  GLOREX quốc tế error: {e}")

    # File 2: Tại chỗ (XNK TC sheet)
    try:
        fp2 = f"{BASE_DIR}/GLOREX/Debit 5PVN_GLOREX T3.2026 TẠI CHỖ.xlsx"
        wb2 = openpyxl.load_workbook(fp2, data_only=True)
        ws2 = wb2["XNK TC"]
        # Header R12-13: STT|Ngày|TK|Luồng|Note|Số HĐ/PXK|Phí mở TK|Phí KH|Phí phát sinh|Tổng|Thu chi hộ: Vé bãi|Số tiền|Số HĐ
        for r in range(14, ws2.max_row + 1):
            stt = cell(ws2, r, 1)
            if isinstance(stt, str) and "tổng" in stt.lower():
                break
            if not isinstance(stt, (int, float)):
                continue
            date_val = d(cell(ws2, r, 2))
            cd_no = s(cell(ws2, r, 3))
            luong = s(cell(ws2, r, 4))
            inv = s(cell(ws2, r, 6))
            pre_vat = n(cell(ws2, r, 7))   # C7=Phí mở tờ khai

            costs = [{"name": "Phí dịch vụ hải quan XNK tại chỗ", "amount": pre_vat, "vat_rate": 8}]

            # Match thu hộ from file 3
            cd_key = s(cd_no)[:11]
            if cd_key in thu_ho_map:
                th = thu_ho_map[cd_key]
                costs.append({
                    "name": f"Thu hộ: Lệ phí hải quan - GNT: {th['gnt']}",
                    "amount": th["amount"],
                    "vat_rate": 0,
                    "is_reimbursement": True,
                })

            jobs.append({
                "customer_id": cid,
                "date": date_val or datetime.date(2026, 3, 1),
                "description": f"XNK tại chỗ - TK: {cd_no}",
                "svc_type": "CUS_EXPORT",
                "cd_no": cd_no,
                "invoice": inv,
                "customs_type": "EXPORT",
                "service_details": {"customs_channel": luong},
                "costs": costs,
            })
    except Exception as e:
        print(f"  GLOREX tại chỗ error: {e}")

    return jobs


# ──────── NIPPON ────────

def parse_nippon():
    """
    NIPPON (customer_id=64): 2 files.
    File 1: .xls (Thai Nguyen), ~10 declarations, 300,000 each.
    File 2: .xlsx rv file — checked, has NO March 2026 data (all Jan-Feb 2026). SKIP.
    """
    cid = CUSTOMER_MAP["NIPPON"]
    jobs = []

    # File 1: .xls
    try:
        fp = f"{BASE_DIR}/NIPPON/(THAI NGUYEN) BẢNG KÊ CHI PHÍ NIPPON THÁNG 3.2026.xls"
        wb = xlrd.open_workbook(fp)
        ws = wb.sheet_by_name("NIPPON")
        # Header R17-18 (0-indexed: 16-17), data from R19 (idx 18)
        # Cols: STT|Invoice|Số TKHQ|Ngày TK|Phân luồng|...|Phí mở TK|Phí khác|...|Tổng phí DV|VAT|Tổng TT|...|Phí chi hộ Số HĐ|Nội dung|Số tiền
        for r in range(18, ws.nrows):
            stt = ws.cell_value(r, 0)
            if isinstance(stt, str) and "tổng" in stt.lower():
                break
            if not isinstance(stt, (int, float)) or stt < 1:
                continue
            cd_no_raw = ws.cell_value(r, 2)
            cd_no = str(int(cd_no_raw)) if isinstance(cd_no_raw, float) else s(cd_no_raw)
            date_val = d(ws.cell_value(r, 3), wb.datemode)
            luong = s(ws.cell_value(r, 4))
            phi_mo_tk = n(ws.cell_value(r, 5))      # col F (idx 5)
            pre_vat = n(ws.cell_value(r, 9))        # col J=Tổng phí DV (idx 9)
            chi_ho_hd = s(ws.cell_value(r, 13)) if ws.ncols > 13 else ""
            chi_ho_nd = s(ws.cell_value(r, 14)) if ws.ncols > 14 else ""
            chi_ho_amt = n(ws.cell_value(r, 15)) if ws.ncols > 15 else 0

            costs = [{"name": "Phí dịch vụ thủ tục HQ", "amount": pre_vat or phi_mo_tk, "vat_rate": 8}]
            if chi_ho_amt > 0:
                costs.append({
                    "name": f"Thu hộ: {chi_ho_nd} - HĐ: {chi_ho_hd}",
                    "amount": chi_ho_amt, "vat_rate": 0, "is_reimbursement": True,
                })
            jobs.append({
                "customer_id": cid,
                "date": date_val or datetime.date(2026, 3, 1),
                "description": f"Thủ tục HQ XNK tại chỗ Thái Nguyên - TK: {cd_no}",
                "svc_type": "CUS_EXPORT",
                "cd_no": cd_no,
                "customs_type": "EXPORT",
                "customs_port": "Thái Nguyên",
                "service_details": {"customs_channel": luong},
                "costs": costs,
            })
    except Exception as e:
        print(f"  NIPPON xls error: {e}")

    # File 2 (rv.xlsx): verified no March 2026 rows → skip
    return jobs


# ──────── MESSER ────────

def parse_messer():
    """
    MESSER (customer_id=22): 1 file, MESSER HẢI DƯƠNG sheet.
    Grouped rows: 1 shipment = R14-R17 (customs + trucking + 2 chi hộ rows).
    Cols: Date|Ngày TK|Số TK|Dịch vụ|Đơn giá|SL|ĐVT|Thành tiền|Tax|Tổng|HĐ|Chi hộ Số HĐ|Số tiền|VAT|Thành tiền|TỔNG|Ghi chú
    """
    cid = CUSTOMER_MAP["MESSER"]
    jobs = []
    try:
        fp = f"{BASE_DIR}/MESSER/Bảng kê MESSER 5P T3.2026.xlsx"
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb["MESSER HẢI DƯƠNG"]

        r = 14
        while r <= ws.max_row:
            stt = cell(ws, r, 1)
            if isinstance(stt, str) and any(k in stt.lower() for k in ["tổng", "total"]):
                break
            if not isinstance(stt, (int, float)):
                r += 1
                continue

            date_val = d(cell(ws, r, 2))
            cd_no = s(cell(ws, r, 3))
            note = s(cell(ws, r, 17)) if ws.max_column >= 17 else ""
            costs = []

            # Collect sub-rows until next STT row
            sr = r
            while sr <= ws.max_row:
                sub_stt = cell(ws, sr, 1)
                col2 = s(cell(ws, sr, 2))
                # Stop at TOTAL row (may appear in column 1 or 2)
                if "total" in col2.lower() or "tổng" in col2.lower():
                    break
                if sr != r and sub_stt is not None:
                    if isinstance(sub_stt, (int, float)) or \
                       (isinstance(sub_stt, str) and "total" in sub_stt.lower()):
                        break
                svc_name = s(cell(ws, sr, 4))
                pre_vat_sr = n(cell(ws, sr, 8))  # C8=Thành tiền (service)
                hd_sr = s(cell(ws, sr, 11))      # C11=Số HĐ (service invoice)
                chi_ho_hd = s(cell(ws, sr, 12))  # C12=Chi hộ Số HĐ
                chi_ho_amt = n(cell(ws, sr, 13)) # C13=Chi hộ Số tiền (pre-VAT)
                if svc_name and pre_vat_sr > 0:
                    costs.append({
                        "name": svc_name, "amount": pre_vat_sr,
                        "vat_rate": 8, "invoice": hd_sr,
                    })
                if chi_ho_amt > 0:
                    costs.append({
                        "name": f"Chi hộ: {svc_name or 'Phí XLHH/lưu kho'} - HĐ: {chi_ho_hd}",
                        "amount": chi_ho_amt, "vat_rate": 0,
                        "is_reimbursement": True, "invoice": chi_ho_hd,
                    })
                sr += 1

            if costs:
                jobs.append({
                    "customer_id": cid,
                    "date": date_val or datetime.date(2026, 3, 1),
                    "description": f"Nhập khẩu - TK: {cd_no} - {note}",
                    "svc_type": "CUS_IMPORT",
                    "cd_no": cd_no,
                    "customs_type": "IMPORT",
                    "service_details": {"note": note},
                    "costs": costs,
                })
            r = sr
    except Exception as e:
        print(f"  MESSER error: {e}")
    return jobs


# ──────── LAS ────────

def parse_las():
    """
    LAS (customer_id=6): 1 file, 1 shipment. Sectioned format:
    Section 1: Phí tại nước ngoài (no VAT)
    Section 2: Chi phí tại Việt Nam (8% VAT on VN fees)
    Section 3: Chi phí trả hộ (no VAT, is_reimbursement=True)
    Cols: STT|Nội dung|SL|ĐVT|Đơn giá|Đồng tiền|Exchange rate|Tổng(trước VAT)|VAT|Tổng VND|Số HĐ|...|Đơn giá VNĐ
    """
    cid = CUSTOMER_MAP["LAS"]
    jobs = []
    try:
        fp = f"{BASE_DIR}/LAS/DebitNote_LGZHPH260781_LAS_DRAFT (13).xlsx"
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb["dịch vụ"]

        # Shipment info
        bill = s(cell(ws, 16, 2))
        cd_no = s(cell(ws, 16, 10))
        weight = n(cell(ws, 17, 6))

        costs = []
        current_section = 1  # 1=nước ngoài, 2=VN, 3=trả hộ

        for r in range(20, ws.max_row + 1):
            stt_v = cell(ws, r, 1)
            if stt_v is None:
                continue
            stt_str = s(stt_v)
            # Detect section headers
            if "tổng cộng" in stt_str.lower():
                break
            if "tổng (1)" in stt_str.lower() or "tổng(1)" in stt_str.lower():
                continue
            if "tổng (2)" in stt_str.lower() or "tổng(2)" in stt_str.lower():
                continue
            if "tổng (3)" in stt_str.lower() or "tổng(3)" in stt_str.lower():
                continue
            # Section 2 header
            if isinstance(stt_v, (int, float)) and float(stt_v) == 2.0:
                current_section = 2
                continue
            # Section 3 header — "Các chi phí trả hộ"
            nd = s(cell(ws, r, 2))
            if "trả hộ" in nd.lower() and isinstance(stt_v, (int, float)):
                current_section = 3
                continue
            if "xuất theo hđ" in nd.lower() and isinstance(stt_v, (int, float)):
                current_section = 3
                continue
            if isinstance(stt_v, (int, float)) and float(stt_v) == 3.0:
                # Could be section 3 header (appears twice in file)
                if "trả hộ" in nd.lower() or not nd:
                    current_section = 3
                    continue

            # Section detection from STT like 1.1, 2.1 etc.
            if isinstance(stt_v, float):
                if 1.0 <= stt_v < 2.0:
                    current_section = 1
                elif 2.0 <= stt_v < 3.0:
                    current_section = 2
                elif 3.0 <= stt_v < 4.0:
                    current_section = 3

            desc = s(cell(ws, r, 2))
            if not desc:
                continue
            # Amount: use C13=Đơn giá VNĐ (actual VND pre-VAT) if available
            vnd_unit = n(cell(ws, r, 13))
            qty = n(cell(ws, r, 3))
            pre_vat_usd = n(cell(ws, r, 8))  # USD amount (pre-VAT)
            vat_rate_cell = n(cell(ws, r, 9))  # 0.08 or 0
            # Total VND (post-VAT)
            total_vnd = n(cell(ws, r, 10))
            hd = s(cell(ws, r, 11))

            # Compute pre-VAT VND amount
            if vnd_unit > 0 and qty > 0:
                pre_vat_vnd = vnd_unit * qty
            elif total_vnd > 0 and vat_rate_cell > 0:
                pre_vat_vnd = total_vnd / (1 + vat_rate_cell)
            elif total_vnd > 0:
                pre_vat_vnd = total_vnd
            else:
                continue

            if pre_vat_vnd == 0:
                continue

            vat_pct = vat_rate_cell * 100  # 0.08 → 8

            if current_section == 3:
                costs.append({
                    "name": f"Thu hộ: {desc}",
                    "amount": pre_vat_vnd, "vat_rate": 0,
                    "is_reimbursement": True, "invoice": hd,
                })
            else:
                costs.append({
                    "name": desc, "amount": pre_vat_vnd,
                    "vat_rate": vat_pct, "invoice": hd,
                })

        if costs:
            jobs.append({
                "customer_id": cid,
                "date": datetime.date(2026, 3, 12),
                "description": f"Sea import LCL - Bill: {bill} - TK: {cd_no}",
                "svc_type": "SEA_IMP",
                "bl_awb": bill,
                "cd_no": cd_no,
                "weight": weight,
                "customs_type": "IMPORT",
                "service_details": {},
                "costs": costs,
            })
    except Exception as e:
        print(f"  LAS error: {e}")
        traceback.print_exc()
    return jobs


# ──────── KK ────────

def parse_kk():
    """
    KK (customer_id=65): 1 file, 3 sheets.
    TRUCKING VẢI: 3 trucking jobs (consol)
    TRUCKING CHỐNG ẨM: 1 trucking job
    SEA DOM: 11 container shipments
    Cols: STT|Ngày|Type|Điểm lấy|Điểm trả|BKS/Container|SL|ĐVT|Đơn giá|Thành tiền|Chi phí khác|Tổng|JOB|Ghi chú
    """
    cid = CUSTOMER_MAP["KK"]
    jobs = []
    try:
        fp = f"{BASE_DIR}/KK/Debit Note.KK.MAR.2026. org.xlsx"
        wb = openpyxl.load_workbook(fp, data_only=True)

        for sheet_name in ["TRUCKING VẢI", "TRUCKING CHỐNG ẨM", "SEA DOM"]:
            ws = wb[sheet_name]
            svc_type = "SEA_DOM" if "SEA" in sheet_name else "TRUCKING_DOM"

            for r in range(15, ws.max_row + 1):
                stt = cell(ws, r, 1)
                if isinstance(stt, str) and any(k in stt.lower() for k in ["tổng", "thuế", "total"]):
                    break
                if not isinstance(stt, (int, float)):
                    continue
                date_val = d(cell(ws, r, 2))
                vtype = s(cell(ws, r, 3))
                origin = s(cell(ws, r, 4))
                dest = s(cell(ws, r, 5))
                container_bks = s(cell(ws, r, 6))
                qty = n(cell(ws, r, 7))
                unit = s(cell(ws, r, 8))
                unit_price = n(cell(ws, r, 9))
                thanh_tien = n(cell(ws, r, 10))  # C10=Thành tiền
                extra = n(cell(ws, r, 11))       # C11=Chi phí khác / Phí khác
                job_no = s(cell(ws, r, 12)) if ws.max_column >= 12 else ""
                note = s(cell(ws, r, 13)) if ws.max_column >= 13 else ""

                costs = [{"name": f"Cước vận chuyển {origin} → {dest}", "amount": thanh_tien, "vat_rate": 8}]
                if extra > 0:
                    costs.append({"name": "Phí khác", "amount": extra, "vat_rate": 8})

                desc = f"{svc_type.replace('_', ' ')} - {origin} → {dest}"
                if container_bks:
                    desc += f" - {container_bks}"

                jobs.append({
                    "customer_id": cid,
                    "date": date_val or datetime.date(2026, 3, 20),
                    "description": desc,
                    "svc_type": svc_type,
                    "origin": origin, "dest": dest,
                    "bl_awb": container_bks,
                    "service_details": {
                        "vehicle_type": vtype, "quantity": qty,
                        "unit": unit, "unit_price": unit_price,
                        "job_ref": job_no, "note": note,
                    },
                    "costs": costs,
                })
    except Exception as e:
        print(f"  KK error: {e}")
    return jobs


# ──────── KCVN ────────

def parse_kcvn():
    """
    KCVN (customer_id=61): 1 file, 'Tháng 2.2026' sheet (actually March data).
    4 customs declarations, amounts in USD with exchange rate.
    Cols: STT|Invoice|Số TKHQ|Ngày TK|Bill|Note|SL|Cont|Weight|Invoice|Luồng|Đơn giá USD|Tỷ giá|Thành tiền USD|Tổng phí DV|VAT|Tổng TT|Phát sinh|Chi hộ...
    selling_rate = pre-VAT VND (C15=Tổng phí DV)
    """
    cid = CUSTOMER_MAP["KCVN"]
    jobs = []
    try:
        fp = f"{BASE_DIR}/KCVN/BangKe_KCIL_T3_2026_v14.xlsx"
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb["Tháng 2.2026"]

        for r in range(19, ws.max_row + 1):
            stt = cell(ws, r, 1)
            if isinstance(stt, str) and "tổng" in stt.lower():
                break
            if not isinstance(stt, (int, float)):
                continue
            date_val = d(cell(ws, r, 4))
            if not date_val or date_val.year != 2026:
                continue
            cd_no = s(cell(ws, r, 3))
            inv_no = s(cell(ws, r, 10))
            weight = n(cell(ws, r, 9))
            luong = s(cell(ws, r, 11))
            usd_rate = n(cell(ws, r, 12))    # Đơn giá USD
            ty_gia = n(cell(ws, r, 13))       # Tỷ giá
            total_usd = n(cell(ws, r, 14))    # Thành tiền USD
            pre_vat_vnd = n(cell(ws, r, 15)) # Tổng phí DV (pre-VAT VND)
            chi_ho_hd = s(cell(ws, r, 19)) if ws.max_column >= 19 else ""
            chi_ho_nd = s(cell(ws, r, 20)) if ws.max_column >= 20 else ""
            chi_ho_amt = n(cell(ws, r, 21)) if ws.max_column >= 21 else 0

            costs = [{
                "name": "Phí dịch vụ thủ tục XNK",
                "amount": pre_vat_vnd, "vat_rate": 8,
                "invoice": inv_no,
            }]
            if chi_ho_amt > 0:
                costs.append({
                    "name": f"Chi hộ: {chi_ho_nd} - HĐ: {chi_ho_hd}",
                    "amount": chi_ho_amt, "vat_rate": 0, "is_reimbursement": True,
                })

            jobs.append({
                "customer_id": cid,
                "date": date_val,
                "description": f"Thủ tục HQ XNK - TK: {cd_no} - {total_usd:.2f} USD",
                "svc_type": "CUS_EXPORT",
                "cd_no": cd_no,
                "invoice": inv_no,
                "weight": weight,
                "customs_type": "EXPORT",
                "service_details": {
                    "customs_channel": luong,
                    "usd_rate": usd_rate,
                    "exchange_rate": ty_gia,
                    "total_usd": total_usd,
                },
                "costs": costs,
            })
    except Exception as e:
        print(f"  KCVN error: {e}")
    return jobs


# ──────── KWE ────────

def parse_kwe():
    """
    KWE (customer_id=28): 1 file, 'Accountant Sheet' (current version).
    4 service lines (Storage, Stevedore, Inventory, Trucking) as 1 job.
    Cols: No.|Description|Amount(VND)|VAT(%)|VAT Amount|Total Amount
    """
    cid = CUSTOMER_MAP["KWE"]
    jobs = []
    try:
        fp = f"{BASE_DIR}/KWE/5P in MAR.2026. KWE rev.xlsx"
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb["Accountant Sheet"]

        costs = []
        for r in range(18, ws.max_row + 1):
            stt = cell(ws, r, 1)
            if stt is None:
                continue
            if isinstance(stt, str) and any(k in stt.lower() for k in ["grand total", "tổng", "say"]):
                break
            if not isinstance(stt, (int, float)):
                continue
            desc = s(cell(ws, r, 2))
            pre_vat = n(cell(ws, r, 3))   # C3=Amount(VND)
            vat_pct = n(cell(ws, r, 4)) * 100  # 0.08 → 8
            if desc and pre_vat > 0:
                svc = "WHS_STORAGE"
                if "stevedore" in desc.lower() or "nâng hạ" in desc.lower():
                    svc = "WHS_HANDLE"
                elif "trucking" in desc.lower() or "vận chuyển" in desc.lower():
                    svc = "TRUCKING_DOM"
                costs.append({"name": desc, "amount": pre_vat, "vat_rate": vat_pct})

        if costs:
            # 1 single job for all 4 services this month
            jobs.append({
                "customer_id": cid,
                "date": datetime.date(2026, 3, 31),
                "description": "Dịch vụ kho bãi và vận chuyển tháng 03/2026",
                "svc_type": "WHS_STORAGE",
                "service_details": {},
                "costs": costs,
            })
    except Exception as e:
        print(f"  KWE error: {e}")
    return jobs
