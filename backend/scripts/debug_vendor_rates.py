#!/usr/bin/env python3
"""
Debug script để kiểm tra import vendor rates
"""

import os
import sys
from dotenv import load_dotenv
from supabase import create_client
from openpyxl import load_workbook
import google.generativeai as genai

load_dotenv()

# Setup Gemini
genai.configure(api_key=os.getenv('GOOGLE_GEMINI_API_KEY'))
model = genai.GenerativeModel('gemini-2.0-flash')

# Setup Supabase
supabase = create_client(os.getenv('SUPABASE_URL'), os.getenv('SUPABASE_SERVICE_ROLE_KEY'))

# Đọc file Excel
file_path = os.path.join('..', 'vendor_rates', 'Tam bảo_092025.xlsx')
wb = load_workbook(file_path, data_only=True)
print(f'Sheets: {wb.sheetnames}')

# Đọc sheet 'Báo Giá Xe Thường'
sheet = wb['Báo Giá Xe Thường']

# Đọc nội dung sheet
content = []
for i, row in enumerate(sheet.iter_rows(max_row=50), 1):
    row_text = ' | '.join([str(cell.value)[:50] if cell.value else '' for cell in row])
    if row_text.strip(' |'):
        content.append(f'Row {i}: {row_text}')

excel_text = '\n'.join(content)
print(f'\nĐọc được {len(content)} dòng từ Excel')
print(f'\nNội dung Excel (đầu 1000 ký tự):\n{excel_text[:1000]}')

# Parse với AI
print('\n🤖 Parsing with AI...')
try:
    response = model.generate_content(f'''Analyze this Excel data containing Vietnamese trucking rates.
Extract pricing into a JSON array. Each rate should have:
- origin: Điểm đầu (pickup location name only)
- origin_province: Tỉnh đi
- destination: Điểm đến (destination name only)
- destination_province: Tỉnh đến
- rate_code: Mã giá (e.g. "TB18", "TB20")
- temperature_range: Yêu cầu nhiệt độ (for refrigerated only, e.g. "Từ 0 đến 10 độ")
- prices: Object with vehicle_type -> price mapping
  Example: {{"1.25T": 200000, "1.5T": 310000, "2.5T": 420000}}

Data:
```
{excel_text[:5000]}
```

IMPORTANT:
- Return ONLY valid JSON array
- Price = number (remove commas/dots in thousands)
- Include ALL vehicle types found (1.25T, 1.5T, 2.5T, 3.5T, 5T, 8T, 15T)
- rate_code is usually like TB18, TB20, TB25, etc.
- Skip header rows and notes

Example output:
[
  {{"origin": "Nội Bài", "origin_province": "Hà Nội", "destination": "KCN Quang Minh", "destination_province": "Hà Nội", "rate_code": "TB20", "temperature_range": null, "prices": {{"1.25T": 220000, "1.5T": 330000, "2.5T": 440000}}}}
]
''')
    response_text = response.text.strip()
    print(f'\n📄 AI Response (đầu 500 ký tự):\n{response_text[:500]}')
except Exception as e:
    print(f'❌ Lỗi khi parse với AI: {e}')
    import traceback
    traceback.print_exc()
