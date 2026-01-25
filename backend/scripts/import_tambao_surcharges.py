#!/usr/bin/env python3
"""
Import Tam Bảo surcharges to Supabase
"""

import os
import sys
import json

# Add parent path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dotenv import load_dotenv
import openpyxl

# Load .env from project root
load_dotenv(os.path.join(os.path.dirname(__file__), '..', '..', '.env'))

from supabase import create_client

url = os.getenv('SUPABASE_URL')
key = os.getenv('SUPABASE_SERVICE_ROLE_KEY')
client = create_client(url, key)

VENDOR_ID = 62  # TAMBAO2
FILE_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'vendor_rates', 'Tam bảo_092025.xlsx')


def parse_surcharges(ws, vehicle_category):
    """Parse surcharges from worksheet"""
    # Determine vehicle types from header
    vehicle_types = []
    header_row = 1
    for col_idx in range(3, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val:
            vehicle_types.append(str(val).strip())

    print(f"  Vehicle types: {vehicle_types}")

    surcharges = []
    for row_idx in range(2, ws.max_row + 1):
        stt = ws.cell(row=row_idx, column=1).value
        name = ws.cell(row=row_idx, column=2).value

        if not stt or not name:
            continue

        stt_str = str(stt).strip()
        if not stt_str.isdigit():
            continue

        name = str(name).strip()

        # Parse prices
        pricing = {}
        for idx, vtype in enumerate(vehicle_types):
            col_idx = idx + 3
            if col_idx <= ws.max_column:
                price = ws.cell(row=row_idx, column=col_idx).value
                if price is not None:
                    pricing[vtype] = price

        # Skip if no prices
        if not pricing:
            continue

        # Determine rate_type and unit
        rate_type = "FIXED"
        unit = "VND/chuyến"
        free_allowance = None

        if "/ 1KM" in name or "/1KM" in name or "/ KM" in name:
            rate_type = "PER_KM"
            unit = "VND/km"
        elif "/ 1 giờ" in name or "/giờ" in name:
            rate_type = "PER_HOUR"
            unit = "VND/giờ"
        elif "ca xe" in name.lower():
            rate_type = "PER_SHIFT"
            unit = "VND/ca"
        elif "Miễn phí" in name:
            rate_type = "FREE_ALLOWANCE"
            unit = "giờ"
            for v in pricing.values():
                if v:
                    free_allowance = str(v)
                    break

        # Generate surcharge code
        code_parts = name.split()[:3]
        code = '_'.join([p[:4].upper() for p in code_parts if len(p) > 1])
        code = f"{vehicle_category}_{code}_{stt_str}"

        surcharge = {
            "vendor_id": VENDOR_ID,
            "surcharge_code": code,
            "surcharge_name": name[:100],
            "description": name,
            "rate_type": rate_type,
            "pricing": pricing,
            "free_allowance": free_allowance,
            "unit": unit,
            "effective_date": "2025-09-01",
            "is_active": True,
            "notes": f"Imported from Tam Bảo price file - {vehicle_category}"
        }
        surcharges.append(surcharge)
        print(f"    {stt_str}. {name[:40]}... -> {code}")

    return surcharges


def main():
    print(f"Loading file: {FILE_PATH}")
    wb = openpyxl.load_workbook(FILE_PATH)

    # Parse both sheets
    print("\n=== Parsing Xe Thuong surcharges ===")
    surcharges_thuong = parse_surcharges(wb["Phụ Phí Cửa Khẩu Xe Thường"], "XE_THUONG")

    print("\n=== Parsing Xe Lanh surcharges ===")
    surcharges_lanh = parse_surcharges(wb["Phụ Phí Cửa Kh Xe Lạnh"], "XE_LANH")

    all_surcharges = surcharges_thuong + surcharges_lanh

    print(f"\n=== Importing {len(all_surcharges)} surcharges to Supabase ===")

    success = 0
    errors = 0

    for surcharge in all_surcharges:
        # Convert pricing to JSON string
        surcharge["pricing"] = json.dumps(surcharge["pricing"])

        try:
            result = client.table('vendor_surcharges').insert(surcharge).execute()
            print(f"  ✓ Inserted: {surcharge['surcharge_code']}")
            success += 1
        except Exception as e:
            print(f"  ✗ Error: {surcharge['surcharge_code']} - {e}")
            errors += 1

    print(f"\n=== DONE ===")
    print(f"Success: {success}, Errors: {errors}")

    # Verify
    result = client.table('vendor_surcharges').select('id, surcharge_code, surcharge_name').eq('vendor_id', VENDOR_ID).execute()
    print(f"\nTotal surcharges for TAMBAO2 (vendor_id={VENDOR_ID}): {len(result.data)}")
    for r in result.data:
        print(f"  - {r['surcharge_code']}: {r['surcharge_name'][:50]}")


if __name__ == "__main__":
    main()
