"""
Shared styling, formatting, and section-builder helpers used by ALL
customer Bảng Kê (statement) renderers.

Eliminates duplication across:
  - DAINESE renderers (5 templates)
  - Generic Family A (trucking), B (handling), C (customs), D (international)

Anything customer-specific lives in per-family renderer modules; anything
constant across the whole 5P brand (logo, company info, bank, fonts, colors,
number formats) lives here.
"""

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.cell import range_boundaries
from openpyxl.drawing.image import Image as XLImage


# ---------- 5P Brand constants ----------

FONT_NAME = "Times New Roman"

COMPANY_INFO = {
    "name": "CÔNG TY TNHH THƯƠNG MẠI VÀ DỊCH VỤ 5P VIỆT NAM",
    "address": "Số nhà 02 Ngõ 1H Phố Trần Quang Diệu, Phường Đống Đa, Thành phố Hà Nội, Việt Nam",
    "mst": "MST: 0110523309",
}

BANK_INFO = [
    "Thông tin chuyển khoản:",
    "Tài khoản: Công Ty TNHH Thương mại và dịch vụ 5P Việt Nam",
    "Số tài khoản: 346886",
    "Tại Ngân hàng: Ngân hàng TMCP Kỹ thương Việt Nam - Techcombank Chi nhánh Đông Đô",
]

# Path to the 5P star logo extracted from customer reference files
DEFAULT_LOGO_PATH = (
    Path(__file__).resolve().parent.parent.parent.parent.parent
    / "plans" / "reports" / "dainese-templates" / "logos"
    / "Bảng_kê_nhập_tháng_3_2026_sea_air__image1.png"
)


# ---------- Style primitives ----------

THIN = Side(style="thin", color="000000")
MEDIUM = Side(style="medium", color="000000")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_CC = Alignment(horizontal="center", vertical="center")
ALIGN_CC_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

NF_INT = "#,##0"
NF_DEC = "#,##0.00"
NF_DATE = "[$-409]d-mmm-yy;@"
NF_DATE_MDY = "mm-dd-yy"
NF_ACCT = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'

FILL_HEADER = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
FILL_DATA = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_DATA_ZEBRA = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_TOTAL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")


def font(size: int = 11, bold: bool = False, italic: bool = False) -> Font:
    return Font(name=FONT_NAME, size=size, bold=bold, italic=italic)


# ---------- Value/data helpers ----------

def safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def parse_details(svc: Dict[str, Any]) -> Dict[str, Any]:
    """service_details may be a string (legacy) or dict — normalize to dict."""
    raw = svc.get("service_details") or {}
    if isinstance(raw, str):
        try:
            return json.loads(raw)
        except Exception:
            return {}
    return raw


def format_month_title(month: Optional[str], style: str = "vi_long") -> str:
    """
    Format month token for use in titles.
    Styles:
      vi_long  → '03 NĂM 2026'  (DAINESE nhập)
      vi_short → '03/2026'      (BÌNH MINH/UPGAIN trucking)
      vi_dot   → '03.2026'      (DAINESE phí CO)
    """
    if month and len(month) == 7 and month[4] == "-":
        year, mon = month.split("-")
    else:
        now = datetime.now()
        year, mon = str(now.year), f"{now.month:02d}"

    if style == "vi_short":
        return f"{mon}/{year}"
    if style == "vi_dot":
        return f"{mon}.{year}"
    return f"{mon} NĂM {year}"


def split_address(addr: str) -> tuple[str, str]:
    """'KCN Yên Bình, Thái Nguyên' → ('KCN Yên Bình', 'Thái Nguyên')."""
    if not addr:
        return ("", "")
    parts = [p.strip() for p in addr.split(",") if p.strip()]
    if len(parts) >= 2:
        return (parts[0], parts[-1])
    return (parts[0] if parts else addr, "")


def apply_border_to_range(ws, rng: str, fill: Optional[PatternFill] = None) -> None:
    c1, r1, c2, r2 = range_boundaries(rng)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_ALL
            if fill is not None:
                cell.fill = fill


# ---------- Vietnamese amount-in-words ----------

_NUM_VI = ["không", "một", "hai", "ba", "bốn", "năm", "sáu", "bảy", "tám", "chín"]


def vn_money_words(n: int) -> str:
    """Vietnamese amount-in-words for invoice 'Tổng tiền bằng chữ' line."""
    if n == 0:
        return "Không đồng"
    units = [("", 1), ("nghìn", 1000), ("triệu", 1_000_000), ("tỷ", 1_000_000_000)]
    parts = []
    n = int(n)
    while n > 0 and len(parts) < 4:
        chunk = n % 1000
        n //= 1000
        if chunk:
            txt = []
            h, t, u = chunk // 100, (chunk // 10) % 10, chunk % 10
            if h:
                txt.append(f"{_NUM_VI[h]} trăm")
            if t > 1:
                txt.append(f"{_NUM_VI[t]} mươi")
            elif t == 1:
                txt.append("mười")
            elif h and u:
                txt.append("lẻ")
            if u:
                if t > 1 and u == 5:
                    txt.append("lăm")
                elif t > 0 and u == 1:
                    txt.append("mốt")
                else:
                    txt.append(_NUM_VI[u])
            unit = units[len(parts)][0]
            parts.append(" ".join(txt) + (f" {unit}" if unit else ""))
        else:
            parts.append("")
    parts = [p for p in reversed(parts) if p.strip()]
    s = " ".join(parts).strip().replace("  ", " ")
    return s[:1].upper() + s[1:] + " đồng chẵn./."


# ---------- Section builders (used by every renderer) ----------

def build_company_header_block(
    ws,
    logo_path: Optional[str] = None,
    logo_anchor: str = "A1",
    company_text_col: str = "C",
    title_size: int = 11,
) -> None:
    """
    Build the standard 5P top-of-sheet header:
      A1 logo (if available)
      C1: company name
      C2: company address
      C3: MST
    """
    if logo_path:
        try:
            img = XLImage(logo_path)
            img.width = 90
            img.height = 90
            ws.add_image(img, logo_anchor)
        except Exception:
            pass

    for r in (1, 2, 3, 4):
        ws.row_dimensions[r].height = 23

    ws[f"{company_text_col}1"] = COMPANY_INFO["name"]
    ws[f"{company_text_col}2"] = COMPANY_INFO["address"]
    ws[f"{company_text_col}3"] = COMPANY_INFO["mst"]
    for row in (1, 2, 3):
        cell = ws[f"{company_text_col}{row}"]
        cell.font = font(title_size, bold=True)
        cell.alignment = ALIGN_LEFT


def build_title_row(
    ws,
    row: int,
    text: str,
    merge_range: str,
    size: int = 18,
    height: int = 50,
) -> None:
    """Centered bold title row."""
    ws.merge_cells(merge_range)
    first = merge_range.split(":")[0]
    ws[first] = text
    ws[first].font = font(size, bold=True)
    ws[first].alignment = ALIGN_CC_WRAP
    ws.row_dimensions[row].height = height


def build_customer_block(
    ws,
    customer: Dict[str, Any],
    start_row: int = 7,
    label_size: int = 12,
    style: str = "kinh_gui",
) -> None:
    """
    Customer info block. Two styles:
    - 'kinh_gui':  'KÍNH GỬI : {NAME}' / 'Địa chỉ: ...' / 'Attn: ...'
    - 'customer':  'Họ tên người mua hàng (Customer): {NAME}' / 'Địa chỉ (Address): ...' / 'Mã số thuế (Tax-code): ...'
    """
    cust_name = (
        customer.get("company_name")
        or customer.get("short_name")
        or customer.get("customer_code", "")
    )
    cust_addr = customer.get("address") or ""
    cust_attn = customer.get("contact_name") or ""
    cust_mst = customer.get("tax_code") or ""

    if style == "customer":
        lines = [
            f"Họ tên người mua hàng (Customer): {cust_name}",
            f"Địa chỉ (Address): {cust_addr}",
            f"Mã số thuế (Tax- code): {cust_mst}",
        ]
    else:  # kinh_gui
        lines = [
            f"KÍNH GỬI :  {cust_name}",
            f"Địa chỉ : {cust_addr}",
            f"Attn:   {cust_attn}",
        ]

    for i, text in enumerate(lines):
        cell = ws.cell(start_row + i, 1, text)
        cell.font = font(label_size, bold=(style == "kinh_gui"))
        cell.alignment = ALIGN_LEFT


def build_bank_footer(
    ws,
    start_row: int,
    merge_last_col: str = "F",
    size: int = 12,
) -> None:
    """Standard 5P bank info footer (4 lines)."""
    for i, line in enumerate(BANK_INFO):
        r = start_row + i
        cell = ws.cell(r, 1, line)
        if i == 0:
            cell.font = font(size, bold=True, italic=True)
        else:
            cell.font = font(size)
        if i == len(BANK_INFO) - 1:
            ws.merge_cells(f"A{r}:{merge_last_col}{r}")
            cell.alignment = ALIGN_LEFT_WRAP


def build_bang_chu_row(ws, row: int, grand_total: float) -> None:
    """'Tổng số tiền bằng chữ: ...' line."""
    cell = ws.cell(row, 1, f"Tổng số tiền bằng chữ: {vn_money_words(int(grand_total))}")
    cell.font = font(12, bold=True, italic=True)


def build_signature_blocks(
    ws,
    row: int,
    customer_label: str = "ĐẠI DIỆN KHÁCH HÀNG",
    company_col: int = 8,
) -> None:
    """Two signature label cells side by side."""
    cust = ws.cell(row, 1, customer_label)
    cust.font = font(12, bold=True)
    own = ws.cell(row, company_col, COMPANY_INFO["name"])
    own.font = font(12, bold=True)
    own.alignment = ALIGN_LEFT


# ---------- Foreign-party heuristic (shared with DAINESE) ----------

FOREIGN_NAME_MARKERS = (
    "dainese s.p.a", "dainese spa", "s.p.a", "s.r.l", "co.,ltd",
    "corporation", "gmbh", "limited", "ltd.", "inc.", "sas",
    "pte ltd", "co. ltd",
)


def is_foreign_party(note: Optional[str]) -> bool:
    if not note:
        return False
    n = note.lower().strip()
    return any(m in n for m in FOREIGN_NAME_MARKERS)
