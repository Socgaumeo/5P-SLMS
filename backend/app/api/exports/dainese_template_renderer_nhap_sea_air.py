"""
Renderer for DAINESE 'Bảng kê chi tiết hàng nhập quốc tế' (SEA + AIR).

Reproduces layout/styles/formulas of the customer reference file:
  plans/reports/dainese-templates/file-01-analysis-report-bang-ke-nhap-quoc-te-sea-air.md

Pure rendering: takes already-fetched data + customer info, returns a Workbook.
DB access is the caller's responsibility (in dainese_customer_export_template.py).
"""

import json
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.drawing.image import Image as XLImage


# ---- Style constants ----

FONT_NAME = "Times New Roman"

THIN = Side(style="thin", color="000000")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_CC_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

NF_INT = "#,##0"
NF_DEC = "#,##0.00"
NF_ACCT = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'

FILL_HEADER = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
FILL_DATA = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_TOTAL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

COMPANY_INFO = {
    "name": "CÔNG TY TNHH THƯƠNG MẠI VÀ DỊCH VỤ 5P VIỆT NAM",
    "address": "Số nhà 02 Ngõ 1H Phố Trần Quang Diệu, Phường Đống Đa, Thành phố Hà Nội, Việt Nam",
    "mst": "MST: 0110523309",
}

BANK_INFO = [
    "Thông tin chuyển khoản:",
    "Tài khoản: Công Ty TNHH Thương mại và dịch vụ 5P Việt Nam",
    "Số tài khoản: 346886",
    "Tại Ngân hàng: Ngân hàng TMCP Kỹ thương Việt Nam - Techcombank Chi nhánh Đông Đô",
]

# Column widths for sheet "NHẬP" (matches the reference file exactly)
COL_WIDTHS = {
    "A": 5.4,  "B": 13.6, "C": 19.6, "D": 19.9, "E": 16.6, "F": 26.3, "G": 11.4,
    "H": 13.5, "I": 9.5,  "J": 18.5, "K": 15.3, "L": 14.5, "M": 16.4, "N": 15.2,
    "O": 14.2, "P": 18.8, "Q": 17.9, "R": 13.7, "S": 12.6, "T": 13.8, "U": 11.0,
    "V": 15.7, "W": 12.5, "X": 11.5, "Y": 20.4, "Z": 15.7, "AA": 18.5, "AB": 21.8,
    "AC": 8.2, "AD": 8.0,
}


# ---- Helpers ----

def _font(size: int = 11, bold: bool = False, italic: bool = False) -> Font:
    return Font(name=FONT_NAME, size=size, bold=bold, italic=italic)


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _parse_details(svc: Dict[str, Any]) -> Dict[str, Any]:
    """service_details JSONB may be a string in some rows."""
    raw = svc.get("service_details") or {}
    if isinstance(raw, str):
        try:
            return json.loads(raw)
        except Exception:
            return {}
    return raw


def _format_month_title(month: Optional[str]) -> str:
    """'2026-03' -> '03 NĂM 2026'. Defaults to current month."""
    if month and len(month) == 7 and month[4] == "-":
        year, mon = month.split("-")
        return f"{mon} NĂM {year}"
    now = datetime.now()
    return f"{now.month:02d} NĂM {now.year}"


def _apply_border_to_range(ws, rng: str, fill: Optional[PatternFill] = None) -> None:
    c1, r1, c2, r2 = range_boundaries(rng)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_ALL
            if fill is not None:
                cell.fill = fill


# ---- Section builders ----

def _build_header(ws, customer: Dict[str, Any], logo_path: Optional[str], title_month: str) -> None:
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w
    for r in (1, 2, 3, 4):
        ws.row_dimensions[r].height = 23
    ws.row_dimensions[5].height = 67
    ws.row_dimensions[6].height = 38
    for r in (7, 8, 9):
        ws.row_dimensions[r].height = 18
    ws.row_dimensions[10].height = 73
    ws.row_dimensions[11].height = 26
    ws.row_dimensions[12].height = 42

    if logo_path:
        try:
            img = XLImage(logo_path)
            img.width = 90
            img.height = 90
            ws.add_image(img, "A1")
        except Exception:
            pass

    ws["C1"] = COMPANY_INFO["name"]
    ws["C2"] = COMPANY_INFO["address"]
    ws["C3"] = COMPANY_INFO["mst"]
    for coord in ("C1", "C2", "C3"):
        ws[coord].font = _font(12, bold=True)
        ws[coord].alignment = ALIGN_LEFT

    ws.merge_cells("A5:AD5")
    ws["A5"] = f"BẢNG KÊ CHI TIẾT HÀNG NHẬP QUỐC TẾ THÁNG {title_month}"
    ws["A5"].font = _font(26, bold=True)
    ws["A5"].alignment = ALIGN_CC_WRAP

    cust_name = (
        customer.get("company_name")
        or customer.get("short_name")
        or customer.get("customer_code", "")
    )
    cust_addr = customer.get("address") or ""
    cust_attn = customer.get("contact_person") or ""

    ws["B7"] = f"KÍNH GỬI :  {cust_name}"
    ws["B8"] = f"Địa chỉ : {cust_addr}"
    ws["B9"] = f"Attn:   {cust_attn}"
    for coord in ("B7", "B8", "B9"):
        ws[coord].font = _font(14, bold=True)
        ws[coord].alignment = ALIGN_LEFT


def _build_table_header(ws) -> None:
    """Two-row header at rows 11-12 with merged group labels."""
    group_headers = [
        ("A11:G11", "Thông tin chung"),
        ("H11:I11", "Khối lượng"),
        ("J11:J12", "Note"),
        ("K11:V11", "Phí dịch vụ làm hàng  ( VND) "),
        ("W11:Z11", "Phí trả hộ"),
        ("AA11:AA12", "Tổng tiền phải trả "),
        ("AB11:AB12", "Số hóa đơn trả hộ"),
        ("AC11:AD12", "Số hóa đơn của forwarder"),
    ]
    for rng, label in group_headers:
        ws.merge_cells(rng)
        first = rng.split(":")[0]
        ws[first] = label
        ws[first].font = _font(12, bold=True, italic=True)
        ws[first].alignment = ALIGN_CC_WRAP
        _apply_border_to_range(ws, rng, fill=FILL_HEADER)

    col_headers = {
        "A12": "No.",
        "B12": "Tờ khai",
        "C12": "Hóa đơn thương mại",
        "D12": "Vận đơn/Note",
        "E12": "Ngày tờ khai",
        "F12": "Tuyến đường",
        "G12": "Loại hình vận chuyển",
        "H12": "Kgs",
        "I12": "No. Cont",
        "K12": "Phí mở tờ khai hải quan",
        "L12": "Phí kiểm hóa",
        "M12": "Phí vận chuyển",
        "N12": "Phí làm hàng",
        "O12": "Phí phát sinh khác",
        "P12": "Phí đầu nước ngoài",
        "Q12": "Cước vận tải quốc tế",
        "R12": "Phí xếp dỡ (THC)",
        "S12": "Phí gom hàng lẻ (CFS)/\nCIC/ LSS",
        "T12": "Phí lấy lệnh  (DO)",
        "U12": "Phí  đại lý,",
        "V12": "Tổng",
        "W12": "Local charge",
        "X12": "CSHT, thuế, vé bãi",
        "Y12": "Lưu kho giao nhận bốc xếp, nâng hạ",
        "Z12": "Tổng phí trả hộ",
    }
    for coord, val in col_headers.items():
        ws[coord] = val
        ws[coord].font = _font(10, bold=True)
        ws[coord].alignment = ALIGN_CC_WRAP
        ws[coord].fill = FILL_HEADER
        ws[coord].border = BORDER_ALL


def _service_to_row(svc: Dict[str, Any], job: Dict[str, Any], idx: int) -> Dict[str, Any]:
    """Map one service+job to the column dict used by the renderer."""
    d = _parse_details(svc)
    origin = svc.get("origin_address") or d.get("origin") or ""
    dest = svc.get("dest_address") or d.get("destination") or ""
    route = d.get("route") or (f"{origin} - {dest}" if origin or dest else "")

    # SEA vs AIR
    code = (svc.get("service_type_code") or "").upper()
    loai = "AIR" if "AIR" in code else "SEA"

    return {
        "no": idx,
        "to_khai": d.get("cd_no") or d.get("declaration_number") or "",
        "hd_tm": d.get("commercial_invoice") or job.get("invoice_number") or "",
        "van_don": d.get("bill_of_lading") or d.get("bl_awb_no") or "",
        "ngay_tk": str(d.get("cd_date") or svc.get("scheduled_date") or ""),
        "tuyen": route,
        "loai": loai,
        "kgs": d.get("weight_kg") or "",
        "cont": d.get("container_size") or d.get("cont_type") or "",
        "note": d.get("notes") or d.get("note") or "",
        "phi_mtk": _safe_float(d.get("fee_declaration")),
        "phi_kh": _safe_float(d.get("fee_inspection")),
        "phi_vc": _safe_float(d.get("fee_transport")),
        "phi_lh": _safe_float(d.get("fee_handling")),
        "phi_psk": _safe_float(d.get("fee_extra")),
        "phi_dnn": _safe_float(d.get("fee_overseas")),
        "cuoc_qt": _safe_float(d.get("fee_international") or d.get("fee_freight")),
        "thc": _safe_float(d.get("fee_thc")),
        "cfs": _safe_float(d.get("fee_cfs")),
        "do": _safe_float(d.get("fee_do") or d.get("fee_delivery_order")),
        "dly": _safe_float(d.get("fee_agency")),
        "local": _safe_float(d.get("fee_local")),
        "csht": _safe_float(d.get("fee_csht")),
        "kho": _safe_float(d.get("fee_storage")),
        "hd_traho": d.get("invoice_reimburse") or "",
        "hd_fwd_1": d.get("invoice_fwd_1") or "",
        "hd_fwd_2": d.get("invoice_fwd_2") or "",
    }


def _build_data_rows(ws, rows: List[Dict[str, Any]], start_row: int = 13) -> int:
    """Write data rows starting at start_row. Returns row index AFTER last data row."""
    common_font = _font(10, bold=True)

    for idx, r_data in enumerate(rows, start=1):
        r = start_row + idx - 1
        ws.row_dimensions[r].height = 37

        ws.cell(r, 1, idx)
        ws.cell(r, 2, r_data["to_khai"])
        ws.cell(r, 3, r_data["hd_tm"])
        ws.cell(r, 4, r_data["van_don"])
        ws.cell(r, 5, r_data["ngay_tk"])
        ws.cell(r, 6, r_data["tuyen"])
        ws.cell(r, 7, r_data["loai"])
        ws.cell(r, 8, r_data["kgs"])
        ws.cell(r, 9, r_data["cont"])
        ws.cell(r, 10, r_data["note"])

        # Cols K..U (11..21) - per-row fees
        fee_keys = ["phi_mtk", "phi_kh", "phi_vc", "phi_lh", "phi_psk",
                    "phi_dnn", "cuoc_qt", "thc", "cfs", "do", "dly"]
        for i, k in enumerate(fee_keys):
            v = r_data.get(k) or 0
            cell = ws.cell(r, 11 + i, v if v else None)
            cell.number_format = NF_DEC

        # V (22): =SUM(K:U)
        ws.cell(r, 22, f"=SUM(K{r}:U{r})").number_format = NF_ACCT

        # W..Y (23..25): reimbursable fees
        for i, k in enumerate(["local", "csht", "kho"]):
            v = r_data.get(k) or 0
            ws.cell(r, 23 + i, v if v else None).number_format = NF_DEC

        # Z (26): =SUM(W:Y)
        ws.cell(r, 26, f"=SUM(W{r}:Y{r})").number_format = NF_ACCT
        # AA (27): =V+Z
        ws.cell(r, 27, f"=+V{r}+Z{r}").number_format = NF_ACCT
        # AB, AC, AD (28..30)
        ws.cell(r, 28, r_data.get("hd_traho", ""))
        ws.cell(r, 29, r_data.get("hd_fwd_1", ""))
        ws.cell(r, 30, r_data.get("hd_fwd_2", ""))

        for c in range(1, 31):
            cell = ws.cell(r, c)
            cell.font = common_font
            cell.alignment = ALIGN_CC_WRAP
            cell.border = BORDER_ALL
            cell.fill = FILL_DATA

    return start_row + len(rows)


def _build_totals(ws, data_start: int, data_end_exclusive: int) -> int:
    """Build the 3-row totals block (Tổng / VAT / Tổng cộng). Returns last row."""
    last_data = data_end_exclusive - 1
    tong_row = data_end_exclusive
    vat_row = tong_row + 1
    tc_row = vat_row + 1

    def _label_row(row: int, label: str) -> None:
        ws.merge_cells(f"A{row}:J{row}")
        ws.cell(row, 1, label)
        ws.cell(row, 1).font = _font(12, bold=True, italic=True)
        ws.cell(row, 1).alignment = ALIGN_CC_WRAP
        for c in range(1, 11):
            ws.cell(row, c).fill = FILL_TOTAL
            ws.cell(row, c).border = BORDER_ALL

    # Tổng row: SUM of data
    _label_row(tong_row, "Tổng")
    for c in range(11, 28):  # K..AA
        col = get_column_letter(c)
        cell = ws.cell(tong_row, c, f"=SUM({col}{data_start}:{col}{last_data})")
        cell.font = _font(12, bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = FILL_TOTAL
        cell.border = BORDER_ALL
        cell.number_format = NF_INT

    # VAT row (zero by default - DAINESE statement-level VAT handled separately)
    _label_row(vat_row, "VAT")
    for c in range(11, 28):
        cell = ws.cell(vat_row, c, 0)
        cell.font = _font(12, bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = FILL_TOTAL
        cell.border = BORDER_ALL
        cell.number_format = NF_INT

    # Tổng cộng = Tổng + VAT
    _label_row(tc_row, "Tổng cộng")
    for c in range(11, 28):
        col = get_column_letter(c)
        cell = ws.cell(tc_row, c, f"=SUM({col}{tong_row}:{col}{vat_row})")
        cell.font = _font(12, bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = FILL_TOTAL
        cell.border = BORDER_ALL
        cell.number_format = NF_INT

    return tc_row


def _build_footer(ws, start_row: int) -> None:
    for i, line in enumerate(BANK_INFO):
        r = start_row + i
        ws.cell(r, 1, line)
        if i == len(BANK_INFO) - 1:
            ws.merge_cells(f"A{r}:F{r}")
            ws.cell(r, 1).alignment = ALIGN_LEFT_WRAP
        ws.cell(r, 1).font = _font(12)


# ---- Public entry point ----

def render_nhap_sea_air_workbook(
    customer: Dict[str, Any],
    services: List[Dict[str, Any]],
    jobs_map: Dict[Any, Dict[str, Any]],
    month: Optional[str],
    logo_path: Optional[str],
) -> Workbook:
    """Build the full workbook for File 1 (Bảng kê nhập SEA/AIR)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "NHẬP"

    title_month = _format_month_title(month)
    _build_header(ws, customer, logo_path, title_month)
    _build_table_header(ws)

    # Map services -> renderable rows. Sort by date for consistency.
    services_sorted = sorted(
        services,
        key=lambda s: (s.get("scheduled_date") or "", s.get("service_id") or 0),
    )
    rows = [
        _service_to_row(svc, jobs_map.get(svc["job_id"], {}), idx)
        for idx, svc in enumerate(services_sorted, start=1)
    ]

    after_data = _build_data_rows(ws, rows, start_row=13)
    last_total_row = _build_totals(ws, data_start=13, data_end_exclusive=after_data)
    _build_footer(ws, last_total_row + 4)

    return wb
