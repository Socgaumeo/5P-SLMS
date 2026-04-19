"""
Renderer for DAINESE 'Bảng kê thanh toán dịch vụ vận chuyển' (TT - trucking).

Reproduces layout/styles/formulas of customer reference file:
  plans/reports/dainese-templates/file-04-bang-ke-tt-summary.md

Sheet 'HĐ' only — the second 'PurchPurchaseOrder' sheet (DAINESE PO format)
is generated separately by their procurement system, not by this exporter.

Service filter: TRUCKING_DOM / TRUCKING_SHORT / TRUCKING_LONG.

Layout (17 cols A-Q):
  A=STT  B=Ngày  C=ĐiểmLấy  D=Tỉnh  E=ĐiểmTrả  F=Tỉnh  G=BKS  H=Đơn vị
  I=SL   J=Phát sinh   K=Cước VC   L=Phụ phí xăng dầu
  M=Thành tiền (=K*I)  N=Tổng (=J+L+M)  O=Note  P=Yêu cầu  Q=Số HĐ
Totals at row N+1: Tổng / VAT / Tổng cộng (only N column).
"""

import json
import re
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.drawing.image import Image as XLImage

# Reuse Vietnamese amount-in-words helper from phi_co renderer
from app.api.exports.dainese_template_renderer_phi_co import _vn_money_words


FONT_NAME = "Times New Roman"

THIN = Side(style="thin", color="000000")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_CC_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_CC = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

NF_INT = "#,##0"
NF_DATE = "[$-409]d-mmm-yy;@"
NF_DATE_MDY = "mm-dd-yy"
NF_ACCT = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'

FILL_HEADER = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
FILL_DATA = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_TOTAL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

COMPANY_INFO = {
    "name": "CÔNG TY TNHH THƯƠNG MẠI VÀ DỊCH VỤ 5P VIỆT NAM",
    "address": "Số nhà 02, Ngõ 1H, Phố Trần Quang Diệu, Phường Đống Đa, Thành phố Hà Nội, Việt Nam",
    "mst": "MST: 0110523309",
}

BANK_INFO = [
    "Thông tin chuyển khoản:",
    "Tài khoản: Công Ty TNHH Thương mại và dịch vụ 5P Việt Nam",
    "Số tài khoản: 346886",
    "Tại Ngân hàng: Ngân hàng TMCP Kỹ thương Việt Nam - Techcombank Chi nhánh Đông Đô",
]

COL_WIDTHS = {
    "A": 5.2,  "B": 12.0, "C": 17.4, "D": 14.4, "E": 19.8, "F": 14.2, "G": 13.1,
    "H": 7.9,  "I": 8.6,  "J": 10.7, "K": 12.1, "L": 12.0, "M": 18.0, "N": 18.0,
    "O": 20.8, "P": 2.2,  "Q": 11.2,
}


def _font(size: int = 12, bold: bool = False, italic: bool = False) -> Font:
    return Font(name=FONT_NAME, size=size, bold=bold, italic=italic)


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _parse_details(svc: Dict[str, Any]) -> Dict[str, Any]:
    raw = svc.get("service_details") or {}
    if isinstance(raw, str):
        try:
            return json.loads(raw)
        except Exception:
            return {}
    return raw


def _format_month_title(month: Optional[str]) -> str:
    if month and len(month) == 7 and month[4] == "-":
        year, mon = month.split("-")
        return f"{mon}/{year}"
    now = datetime.now()
    return f"{now.month:02d}/{now.year}"


def _split_address(addr: str) -> tuple[str, str]:
    """'KCN Yên Bình, Thái Nguyên' → ('KCN Yên Bình', 'Thái Nguyên')."""
    if not addr:
        return ("", "")
    parts = [p.strip() for p in addr.split(",") if p.strip()]
    if len(parts) >= 2:
        return (parts[0], parts[-1])
    return (parts[0] if parts else addr, "")


def _apply_border_to_range(ws, rng: str, fill: Optional[PatternFill] = None) -> None:
    c1, r1, c2, r2 = range_boundaries(rng)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_ALL
            if fill is not None:
                cell.fill = fill


# ---- Section builders ----

def _build_header(ws, customer: Dict[str, Any], logo_path: Optional[str], title_month: str) -> None:
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[6].height = 18
    ws.row_dimensions[12].height = 16
    for r in (13, 14):
        ws.row_dimensions[r].height = 16

    if logo_path:
        try:
            img = XLImage(logo_path)
            img.width = 80
            img.height = 80
            ws.add_image(img, "A1")
        except Exception:
            pass

    ws["C1"] = COMPANY_INFO["name"]
    ws["Q1"] = f"Ngày: {datetime.now().strftime('%d/%m/%Y')}"
    ws["Q1"].font = _font(12)
    ws["Q1"].alignment = ALIGN_RIGHT
    ws["C2"] = COMPANY_INFO["address"]
    ws["C3"] = COMPANY_INFO["mst"]
    for coord in ("C1", "C2", "C3"):
        ws[coord].font = _font(12, bold=True)
        ws[coord].alignment = ALIGN_LEFT

    # Title row 6 (merged A6:Q6)
    ws.merge_cells("A6:Q6")
    ws["A6"] = f"BẢNG KÊ THANH TOÁN DỊCH VỤ VẬN CHUYỂN THÁNG {title_month}"
    ws["A6"].font = _font(14, bold=True)
    ws["A6"].alignment = ALIGN_CC

    # Customer block
    cust_name = (
        customer.get("company_name")
        or customer.get("short_name")
        or customer.get("customer_code", "")
    )
    ws["A9"] = f"Họ tên người mua hàng (Customer): {cust_name}"
    ws["A10"] = f"Địa chỉ (Address): {customer.get('address') or ''}"
    ws["A11"] = f"Mã số thuế (Tax- code): {customer.get('tax_code') or ''}"
    for coord in ("A9", "A10", "A11"):
        ws[coord].font = _font(12)
        ws[coord].alignment = ALIGN_LEFT

    # Unit indicator
    ws["O12"] = "Đơn vị tính: VNĐ"
    ws["O12"].font = _font(12)
    ws["O12"].alignment = ALIGN_RIGHT


def _build_table_header(ws) -> None:
    """Row 13-14 header (most cols merged 13:14, except C/E split)."""
    # Merged single-col headers
    single_merges = [
        ("A13:A14", "STT"),
        ("B13:B14", "Ngày dịch vụ"),
        ("C13:D13", "Điểm lấy hàng"),
        ("E13:F13", "Điểm trả hàng"),
        ("G13:G14", "BKS"),
        ("H13:H14", "Đơn vị"),
        ("I13:I14", "Số lượng"),
        ("J13:J14", "Phát sinh"),
        ("K13:K14", "Cước vận chuyển"),
        ("L13:L14", "Phụ phí xăng dầu"),
        ("M13:M14", "Thành tiền"),
        ("N13:N14", "Tổng"),
        ("O13:O14", "Note"),
        ("Q13:Q14", "Số hóa đơn"),
    ]
    for rng, label in single_merges:
        ws.merge_cells(rng)
        first = rng.split(":")[0]
        ws[first] = label
        ws[first].font = _font(12, bold=True)
        ws[first].alignment = ALIGN_CC_WRAP
        _apply_border_to_range(ws, rng, fill=FILL_HEADER)

    # P13: Yêu cầu (single cell, no merge)
    ws["P13"] = "Yêu cầu"
    ws["P13"].font = _font(12)
    ws["P13"].alignment = ALIGN_LEFT
    ws["P13"].fill = FILL_HEADER
    ws["P13"].border = BORDER_ALL


_RE_PHAT_SINH = re.compile(r"phát\s*sinh|phat\s*sinh|extra", re.IGNORECASE)
_RE_XANG_DAU = re.compile(r"phụ\s*phí\s*xăng|phu\s*phi\s*xang|fuel|xăng\s*dầu|xang\s*dau", re.IGNORECASE)
_RE_CUOC = re.compile(r"cước|cuoc|trucking|transport|vận\s*chuyển|van\s*chuyen", re.IGNORECASE)


def _split_costs_for_truck(cost_rows: List[Dict[str, Any]]) -> Dict[str, float]:
    """Bucket trucking job_costs into Cước VC / Phát sinh / Phụ phí xăng dầu."""
    out = {"cuoc_vc": 0.0, "phat_sinh": 0.0, "xang_dau": 0.0}
    for c in cost_rows:
        if c.get("is_reimbursement"):
            continue
        name = c.get("cost_name") or ""
        amt = _safe_float(c.get("selling_amount"))
        if _RE_XANG_DAU.search(name):
            out["xang_dau"] += amt
        elif _RE_PHAT_SINH.search(name):
            out["phat_sinh"] += amt
        elif _RE_CUOC.search(name):
            out["cuoc_vc"] += amt
        else:
            # Unknown bucket → cước vận chuyển fallback (most common)
            out["cuoc_vc"] += amt
    return out


def _build_data_rows(ws, services_with_costs: List[tuple], jobs_map: Dict, start_row: int = 15) -> int:
    common_font = _font(12)
    for idx, (svc, cost_rows) in enumerate(services_with_costs, start=1):
        r = start_row + idx - 1
        ws.row_dimensions[r].height = 18

        d = _parse_details(svc)
        job = jobs_map.get(svc["job_id"], {})

        c_loc, c_prov = _split_address(svc.get("origin_address") or "")
        e_loc, e_prov = _split_address(svc.get("dest_address") or "")
        bks = d.get("vehicle_plate") or ""
        unit = d.get("vehicle_type") or ""
        qty = _safe_float(d.get("quantity")) or 1
        note = d.get("note") or ""

        buckets = _split_costs_for_truck(cost_rows)

        ws.cell(r, 1, idx)
        ws.cell(r, 2, svc.get("scheduled_date") or job.get("etd"))
        ws.cell(r, 3, c_loc)
        ws.cell(r, 4, c_prov)
        ws.cell(r, 5, e_loc)
        ws.cell(r, 6, e_prov)
        ws.cell(r, 7, bks)
        ws.cell(r, 8, unit)
        ws.cell(r, 9, int(qty) if qty == int(qty) else qty)
        # J: Phát sinh
        ws.cell(r, 10, buckets["phat_sinh"] if buckets["phat_sinh"] else None)
        # K: Cước vận chuyển
        ws.cell(r, 11, buckets["cuoc_vc"] if buckets["cuoc_vc"] else 0)
        # L: Phụ phí xăng dầu
        ws.cell(r, 12, buckets["xang_dau"] if buckets["xang_dau"] else None)
        # M: =K*I
        ws.cell(r, 13, f"=K{r}*I{r}")
        # N: =J+L+M (use IFERROR-style sums)
        ws.cell(r, 14, f"=IFERROR(J{r},0)+IFERROR(L{r},0)+IFERROR(M{r},0)")
        # O: Note
        ws.cell(r, 15, note)
        # P: Yêu cầu (empty)
        ws.cell(r, 16, "")
        # Q: Số hóa đơn (job.invoice_number) - only first row populated typically
        if idx == 1:
            ws.cell(r, 17, job.get("invoice_number") or "")

        for c in range(1, 18):
            cell = ws.cell(r, c)
            cell.font = common_font
            cell.alignment = ALIGN_CC_WRAP
            cell.border = BORDER_ALL
            cell.fill = FILL_DATA

        # Number formats
        ws.cell(r, 2).number_format = NF_DATE_MDY
        ws.cell(r, 9).number_format = "0"
        ws.cell(r, 10).number_format = NF_INT
        ws.cell(r, 11).number_format = NF_ACCT
        ws.cell(r, 12).number_format = NF_ACCT
        ws.cell(r, 13).number_format = NF_ACCT
        ws.cell(r, 14).number_format = NF_INT

    return start_row + len(services_with_costs)


def _build_totals(ws, data_start: int, data_end_exclusive: int, grand_total: float) -> int:
    last_data = data_end_exclusive - 1
    tong_row = data_end_exclusive
    vat_row = tong_row + 1
    tc_row = vat_row + 1

    def _label(row: int, text: str) -> None:
        ws.merge_cells(f"A{row}:K{row}")
        ws.cell(row, 1, text)
        ws.cell(row, 1).font = _font(12, bold=True)
        ws.cell(row, 1).alignment = ALIGN_CC
        for c in range(1, 12):
            ws.cell(row, c).fill = FILL_TOTAL
            ws.cell(row, c).border = BORDER_ALL

    _label(tong_row, "Tổng")
    cell = ws.cell(tong_row, 14, f"=SUM(N{data_start}:N{last_data})")
    cell.font = _font(12, bold=True)
    cell.alignment = ALIGN_RIGHT
    cell.fill = FILL_TOTAL
    cell.border = BORDER_ALL
    cell.number_format = NF_INT

    _label(vat_row, "VAT (8%)")
    cell = ws.cell(vat_row, 14, f"=N{tong_row}*0.08")
    cell.font = _font(12, bold=True)
    cell.alignment = ALIGN_RIGHT
    cell.fill = FILL_TOTAL
    cell.border = BORDER_ALL
    cell.number_format = NF_INT

    _label(tc_row, "Tổng cộng")
    cell = ws.cell(tc_row, 14, f"=N{tong_row}+N{vat_row}")
    cell.font = _font(12, bold=True)
    cell.alignment = ALIGN_RIGHT
    cell.fill = FILL_TOTAL
    cell.border = BORDER_ALL
    cell.number_format = NF_INT

    # Tổng số tiền bằng chữ (one row below grand total, no merge)
    words_row = tc_row + 2
    ws.cell(words_row, 1, f"Tổng số tiền bằng chữ: {_vn_money_words(int(grand_total))}")
    ws.cell(words_row, 1).font = _font(12, bold=True, italic=True)

    return words_row


def _build_footer(ws, start_row: int) -> None:
    for i, line in enumerate(BANK_INFO):
        r = start_row + i
        ws.cell(r, 1, line)
        if i == 0:
            ws.cell(r, 1).font = _font(12, bold=True, italic=True)
        else:
            ws.cell(r, 1).font = _font(12)
        if i == len(BANK_INFO) - 1:
            ws.merge_cells(f"A{r}:E{r}")
            ws.cell(r, 1).alignment = ALIGN_LEFT_WRAP


# ---- Public entry ----

def render_thanh_toan_truck_workbook(
    customer: Dict[str, Any],
    services: List[Dict[str, Any]],
    jobs_map: Dict[Any, Dict[str, Any]],
    costs_by_svc: Dict[Any, List[Dict[str, Any]]],
    month: Optional[str],
    logo_path: Optional[str],
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "HĐ"

    title_month = _format_month_title(month)
    _build_header(ws, customer, logo_path, title_month)
    _build_table_header(ws)

    services_with_costs = sorted(
        [(svc, costs_by_svc.get(svc["svc_id"], [])) for svc in services],
        key=lambda pair: (pair[0].get("scheduled_date") or "", pair[0].get("svc_id") or 0),
    )

    after_data = _build_data_rows(ws, services_with_costs, jobs_map, start_row=15)

    # Compute grand total for "bằng chữ" line
    grand_total = 0.0
    for svc, cost_rows in services_with_costs:
        b = _split_costs_for_truck(cost_rows)
        d = _parse_details(svc)
        qty = _safe_float(d.get("quantity")) or 1
        # N = Phát sinh + Phụ phí + Cước*qty
        grand_total += b["phat_sinh"] + b["xang_dau"] + b["cuoc_vc"] * qty
    grand_total *= 1.08  # +VAT 8%

    last_total = _build_totals(ws, data_start=15, data_end_exclusive=after_data, grand_total=grand_total)
    _build_footer(ws, last_total + 3)

    return wb
