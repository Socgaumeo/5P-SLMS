# backend/app/api/exports/meiko-customer-export-template.py
"""
MEIKO Customer Export Template
==============================
Generates Excel export matching MEIKO's 3-sheet template format:
- Sheet 1: Bảng kê IM (Border Import services)
- Sheet 2: Truck (Domestic trucking services)
- Sheet 3: Debit (Summary totals)
"""

import json
import tempfile
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse

from app.db.session import get_db, DatabaseSession

logger = logging.getLogger(__name__)
router = APIRouter()

# Styling constants
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
TOTAL_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def safe_float(val) -> float:
    """Safely convert value to float."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def parse_service_details(svc: Dict[str, Any]) -> Dict[str, Any]:
    """Parse service_details JSONB field."""
    details = svc.get('service_details') or {}
    if isinstance(details, str):
        try:
            details = json.loads(details)
        except Exception:
            details = {}
    return details


def parse_vendor_text_input(svc: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Parse vendor_text_input field - can be string, list, or dict."""
    vti_raw = svc.get('vendor_text_input')
    if isinstance(vti_raw, str):
        try:
            parsed = json.loads(vti_raw)
            if isinstance(parsed, list):
                return parsed
            elif isinstance(parsed, dict):
                return [parsed]
        except Exception:
            pass
    elif isinstance(vti_raw, list):
        return vti_raw
    elif isinstance(vti_raw, dict):
        return [vti_raw]
    return []


def create_border_import_sheet(ws, services: List[Dict], jobs_map: Dict) -> Dict[str, float]:
    """
    Create Sheet 1: Bảng kê IM (Border Import)
    Returns totals for summary sheet.
    """
    headers = [
        'STT', 'Số JOB', 'Loại hình', 'Ngày bắt đầu', 'Ngày kết thúc',
        'Số tờ khai', 'Số hóa đơn', 'Biển số xe TQ', 'Biển số xe VN',
        'Tuyến đường', 'Loại xe', 'Số lượng pallet', 'Trọng lượng (kg)',
        'Phí sang tải', 'Phí CSHT', 'Phí ra vào bãi', 'Phí mở tờ khai',
        'Phí vận tải nội địa', 'Thủ tục giao nhận', 'Phí chờ giờ', 'Phát sinh',
        'Tổng trước thuế', 'VAT', 'Tổng thanh toán', 'Ghi Chú'
    ]

    # Write headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    totals = {
        'subtotal': 0, 'vat': 0, 'total': 0,
        'fee_transload': 0, 'fee_csht': 0, 'fee_yard': 0,
        'fee_declaration': 0, 'fee_domestic': 0, 'fee_handling': 0,
        'fee_waiting': 0, 'fee_extra': 0
    }

    row_idx = 2
    for idx, svc in enumerate(services, 1):
        job = jobs_map.get(svc['job_id'], {})
        details = parse_service_details(svc)
        vti_list = parse_vendor_text_input(svc)

        # Extract license plates
        china_plates = [v.get('china_license_plate', '') for v in vti_list if v.get('china_license_plate')]
        vn_plates = [v.get('license_plate', '') or v.get('vietnam_license_plate', '') for v in vti_list]
        vn_plates = [p for p in vn_plates if p]

        # Extract fees
        fee_transload = safe_float(details.get('fee_transload'))
        fee_csht = safe_float(details.get('fee_csht'))
        fee_yard = safe_float(details.get('fee_yard'))
        fee_declaration = safe_float(details.get('fee_declaration'))
        fee_domestic = safe_float(details.get('fee_domestic_transport'))
        fee_handling = safe_float(details.get('fee_handling'))
        fee_waiting = safe_float(details.get('fee_waiting'))
        fee_extra = safe_float(details.get('fee_extra'))

        subtotal = fee_transload + fee_csht + fee_yard + fee_declaration + fee_domestic + fee_handling + fee_waiting + fee_extra
        vat_rate = safe_float(details.get('vat_rate', 0.10))
        vat_amount = subtotal * vat_rate
        total = subtotal + vat_amount

        # Update totals
        totals['fee_transload'] += fee_transload
        totals['fee_csht'] += fee_csht
        totals['fee_yard'] += fee_yard
        totals['fee_declaration'] += fee_declaration
        totals['fee_domestic'] += fee_domestic
        totals['fee_handling'] += fee_handling
        totals['fee_waiting'] += fee_waiting
        totals['fee_extra'] += fee_extra
        totals['subtotal'] += subtotal
        totals['vat'] += vat_amount
        totals['total'] += total

        row_data = [
            idx,
            job.get('job_no', ''),
            'Nhập khẩu đường bộ',
            str(svc.get('scheduled_date') or details.get('start_date') or ''),
            str(details.get('end_date') or ''),
            details.get('cd_no', ''),
            job.get('invoice_number') or details.get('invoice_number', ''),
            ', '.join(china_plates),
            ', '.join(vn_plates),
            details.get('route') or f"{svc.get('origin_address', '')} → {svc.get('dest_address', '')}",
            details.get('vehicle_type', ''),
            details.get('pallet_qty', ''),
            details.get('weight_kg', ''),
            fee_transload,
            fee_csht,
            fee_yard,
            fee_declaration,
            fee_domestic,
            fee_handling,
            fee_waiting,
            fee_extra,
            subtotal,
            vat_amount,
            total,
            details.get('notes', '')
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = THIN_BORDER
            if col >= 14 and col <= 24:  # Fee columns
                cell.number_format = '#,##0'
            if col == 13:  # Weight
                cell.number_format = '#,##0.0'

        row_idx += 1

    # Total row
    ws.cell(row=row_idx, column=1, value="TỔNG CỘNG").font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws.cell(row=row_idx, column=col).fill = TOTAL_FILL
        ws.cell(row=row_idx, column=col).border = THIN_BORDER
    ws.cell(row=row_idx, column=14, value=totals['fee_transload']).number_format = '#,##0'
    ws.cell(row=row_idx, column=15, value=totals['fee_csht']).number_format = '#,##0'
    ws.cell(row=row_idx, column=16, value=totals['fee_yard']).number_format = '#,##0'
    ws.cell(row=row_idx, column=17, value=totals['fee_declaration']).number_format = '#,##0'
    ws.cell(row=row_idx, column=18, value=totals['fee_domestic']).number_format = '#,##0'
    ws.cell(row=row_idx, column=19, value=totals['fee_handling']).number_format = '#,##0'
    ws.cell(row=row_idx, column=20, value=totals['fee_waiting']).number_format = '#,##0'
    ws.cell(row=row_idx, column=21, value=totals['fee_extra']).number_format = '#,##0'
    ws.cell(row=row_idx, column=22, value=totals['subtotal']).number_format = '#,##0'
    ws.cell(row=row_idx, column=23, value=totals['vat']).number_format = '#,##0'
    ws.cell(row=row_idx, column=24, value=totals['total']).number_format = '#,##0'

    # Column widths
    col_widths = [5, 14, 16, 12, 12, 14, 12, 12, 12, 20, 10, 8, 10,
                  12, 12, 12, 12, 12, 12, 12, 12, 14, 12, 14, 20]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = 'A2'
    return totals


def create_trucking_sheet(ws, services: List[Dict], jobs_map: Dict, customer_name: str) -> Dict[str, float]:
    """
    Create Sheet 2: Truck (Domestic Trucking)
    Returns totals for summary sheet.
    """
    headers = [
        'STT', 'Số JOB', 'Loại hình', 'Ngày thực hiện', 'Khách hàng',
        'Số hóa đơn', 'Biển số xe', 'Tuyến đường', 'Tỉnh', 'Loại xe',
        'Số lượng kiện', 'Trọng lượng (kg)', 'Phí vận chuyển', 'Phí bốc xếp',
        'Phí chờ giờ', 'Phụ phí', 'Phí khác', 'Tổng trước thuế',
        'VAT', 'Tổng thanh toán', 'Ghi Chú'
    ]

    # Write headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    totals = {
        'subtotal': 0, 'vat': 0, 'total': 0,
        'fee_transport': 0, 'fee_loading': 0, 'fee_waiting': 0,
        'fee_surcharge': 0, 'fee_other': 0
    }

    row_idx = 2
    for idx, svc in enumerate(services, 1):
        job = jobs_map.get(svc['job_id'], {})
        details = parse_service_details(svc)
        vti_list = parse_vendor_text_input(svc)

        # Extract license plates
        plates = [v.get('license_plate', '') for v in vti_list if v.get('license_plate')]

        # Extract fees
        fee_transport = safe_float(details.get('fee_transport') or details.get('selling_price') or details.get('total_revenue'))
        fee_loading = safe_float(details.get('fee_loading'))
        fee_waiting = safe_float(details.get('fee_waiting'))
        fee_surcharge = safe_float(details.get('fee_surcharge'))
        fee_other = safe_float(details.get('fee_other'))

        subtotal = fee_transport + fee_loading + fee_waiting + fee_surcharge + fee_other
        vat_rate = safe_float(details.get('vat_rate', 0.10))
        vat_amount = subtotal * vat_rate
        total = subtotal + vat_amount

        # Update totals
        totals['fee_transport'] += fee_transport
        totals['fee_loading'] += fee_loading
        totals['fee_waiting'] += fee_waiting
        totals['fee_surcharge'] += fee_surcharge
        totals['fee_other'] += fee_other
        totals['subtotal'] += subtotal
        totals['vat'] += vat_amount
        totals['total'] += total

        row_data = [
            idx,
            job.get('job_no', ''),
            'Vận tải nội địa',
            str(svc.get('scheduled_date') or ''),
            customer_name,
            job.get('invoice_number') or details.get('invoice_number', ''),
            ', '.join(plates),
            details.get('route') or f"{svc.get('origin_address', '')} → {svc.get('dest_address', '')}",
            details.get('province', ''),
            details.get('vehicle_type', ''),
            details.get('package_qty', ''),
            details.get('weight_kg', ''),
            fee_transport,
            fee_loading,
            fee_waiting,
            fee_surcharge,
            fee_other,
            subtotal,
            vat_amount,
            total,
            details.get('notes', '')
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = THIN_BORDER
            if col >= 13 and col <= 20:  # Fee columns
                cell.number_format = '#,##0'
            if col == 12:  # Weight
                cell.number_format = '#,##0.0'

        row_idx += 1

    # Total row
    ws.cell(row=row_idx, column=1, value="TỔNG CỘNG").font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws.cell(row=row_idx, column=col).fill = TOTAL_FILL
        ws.cell(row=row_idx, column=col).border = THIN_BORDER
    ws.cell(row=row_idx, column=13, value=totals['fee_transport']).number_format = '#,##0'
    ws.cell(row=row_idx, column=14, value=totals['fee_loading']).number_format = '#,##0'
    ws.cell(row=row_idx, column=15, value=totals['fee_waiting']).number_format = '#,##0'
    ws.cell(row=row_idx, column=16, value=totals['fee_surcharge']).number_format = '#,##0'
    ws.cell(row=row_idx, column=17, value=totals['fee_other']).number_format = '#,##0'
    ws.cell(row=row_idx, column=18, value=totals['subtotal']).number_format = '#,##0'
    ws.cell(row=row_idx, column=19, value=totals['vat']).number_format = '#,##0'
    ws.cell(row=row_idx, column=20, value=totals['total']).number_format = '#,##0'

    # Column widths
    col_widths = [5, 14, 14, 12, 16, 12, 12, 20, 10, 10, 8, 10,
                  12, 12, 12, 12, 12, 14, 12, 14, 20]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = 'A2'
    return totals


def create_debit_sheet(ws, border_totals: Dict, truck_totals: Dict, customer_name: str, month: str):
    """
    Create Sheet 3: Debit (Summary)
    """
    ws.merge_cells('A1:D1')
    ws.cell(row=1, column=1, value=f"BẢNG TỔNG HỢP DỊCH VỤ - {customer_name}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:D2')
    ws.cell(row=2, column=1, value=f"Tháng: {month}")
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')

    # Headers
    headers = ['Loại dịch vụ', 'Tổng trước thuế', 'VAT (10%)', 'Tổng thanh toán']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    # Border Import row
    ws.cell(row=5, column=1, value="Nhập khẩu đường bộ").border = THIN_BORDER
    ws.cell(row=5, column=2, value=border_totals.get('subtotal', 0)).border = THIN_BORDER
    ws.cell(row=5, column=2).number_format = '#,##0'
    ws.cell(row=5, column=3, value=border_totals.get('vat', 0)).border = THIN_BORDER
    ws.cell(row=5, column=3).number_format = '#,##0'
    ws.cell(row=5, column=4, value=border_totals.get('total', 0)).border = THIN_BORDER
    ws.cell(row=5, column=4).number_format = '#,##0'

    # Trucking row
    ws.cell(row=6, column=1, value="Vận tải nội địa").border = THIN_BORDER
    ws.cell(row=6, column=2, value=truck_totals.get('subtotal', 0)).border = THIN_BORDER
    ws.cell(row=6, column=2).number_format = '#,##0'
    ws.cell(row=6, column=3, value=truck_totals.get('vat', 0)).border = THIN_BORDER
    ws.cell(row=6, column=3).number_format = '#,##0'
    ws.cell(row=6, column=4, value=truck_totals.get('total', 0)).border = THIN_BORDER
    ws.cell(row=6, column=4).number_format = '#,##0'

    # Grand total row
    grand_subtotal = border_totals.get('subtotal', 0) + truck_totals.get('subtotal', 0)
    grand_vat = border_totals.get('vat', 0) + truck_totals.get('vat', 0)
    grand_total = border_totals.get('total', 0) + truck_totals.get('total', 0)

    for col in range(1, 5):
        ws.cell(row=7, column=col).fill = TOTAL_FILL
        ws.cell(row=7, column=col).border = THIN_BORDER
        ws.cell(row=7, column=col).font = Font(bold=True)

    ws.cell(row=7, column=1, value="TỔNG CỘNG")
    ws.cell(row=7, column=2, value=grand_subtotal).number_format = '#,##0'
    ws.cell(row=7, column=3, value=grand_vat).number_format = '#,##0'
    ws.cell(row=7, column=4, value=grand_total).number_format = '#,##0'

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18


@router.get("/exports/meiko")
def export_meiko_template(
    customer_id: int = Query(..., description="Customer ID"),
    month: Optional[str] = Query(None, description="Month filter YYYY-MM"),
    from_date: Optional[str] = Query(None, description="Start date YYYY-MM-DD"),
    to_date: Optional[str] = Query(None, description="End date YYYY-MM-DD"),
    db: DatabaseSession = Depends(get_db)
):
    """
    Export jobs for MEIKO customer using their 3-sheet template format.

    Sheets:
    1. Bảng kê IM - Border Import (BORDER_IMP services)
    2. Truck - Domestic Trucking (TRUCKING_DOM, TRUCKING_SHORT, TRUCKING_LONG)
    3. Debit - Summary totals
    """
    try:
        # Get customer info
        db.execute("""
            SELECT customer_id, customer_code, short_name, company_name
            FROM customers WHERE customer_id = %s
        """, (customer_id,))
        customer = db.fetchone()
        if not customer:
            raise HTTPException(404, "Customer not found")
        customer = dict(customer)
        customer_name = customer.get('short_name') or customer.get('company_name')

        # Build date filter
        date_filter = ""
        params = [customer_id]
        if month:
            import calendar
            year, mon = month.split('-')
            last_day = calendar.monthrange(int(year), int(mon))[1]
            date_filter = " AND j.etd >= %s AND j.etd <= %s"
            params.extend([f"{year}-{mon}-01", f"{year}-{mon}-{last_day}"])
        else:
            if from_date:
                date_filter += " AND j.etd >= %s"
                params.append(from_date)
            if to_date:
                date_filter += " AND j.etd <= %s"
                params.append(to_date)

        # Get all jobs and services
        db.execute(f"""
            SELECT j.job_id, j.job_no, j.status_code, j.etd, j.created_at, j.invoice_number
            FROM jobs j
            WHERE j.customer_id = %s {date_filter}
            ORDER BY j.created_at DESC
        """, tuple(params))
        jobs = [dict(r) for r in db.fetchall()]
        jobs_map = {j['job_id']: j for j in jobs}

        if not jobs:
            raise HTTPException(404, "No jobs found for this customer in the specified period")

        job_ids = list(jobs_map.keys())
        placeholders = ','.join(['%s'] * len(job_ids))

        # Get services
        db.execute(f"""
            SELECT js.*, v.vendor_code, v.short_name as vendor_name
            FROM job_services js
            LEFT JOIN vendors v ON js.vendor_id = v.vendor_id
            WHERE js.job_id IN ({placeholders})
            ORDER BY js.scheduled_date
        """, tuple(job_ids))
        all_services = [dict(r) for r in db.fetchall()]

        # Split by service type
        border_services = [s for s in all_services if s.get('service_type_code') == 'BORDER_IMP']
        truck_services = [s for s in all_services if s.get('service_type_code') in
                         ['TRUCKING_DOM', 'TRUCKING_SHORT', 'TRUCKING_LONG']]

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Sheet 1: Bảng kê IM
        ws1 = wb.create_sheet("Bảng kê IM")
        border_totals = create_border_import_sheet(ws1, border_services, jobs_map)

        # Sheet 2: Truck
        ws2 = wb.create_sheet("Truck")
        truck_totals = create_trucking_sheet(ws2, truck_services, jobs_map, customer_name)

        # Sheet 3: Debit
        ws3 = wb.create_sheet("Debit")
        month_display = month or datetime.now().strftime("%Y-%m")
        create_debit_sheet(ws3, border_totals, truck_totals, customer_name, month_display)

        # Save
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()

        filename = f"MEIKO_{customer.get('customer_code')}_{month_display}_export.xlsx"
        return FileResponse(
            path=temp_file.name,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"MEIKO export error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(500, f"Export error: {str(e)}")
