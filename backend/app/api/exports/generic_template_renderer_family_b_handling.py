"""
FAMILY B — Dịch vụ làm hàng / handling Bảng Kê renderer (generic).

Covers customers with the cost-per-line handling-debit pattern:
  HƯNG PHÁT, VINTECH, XÂY LẮP VN, LOGIMARK, MESSER, GANG THÉP CPN, LAS, …

Layout (variant of clusters C3/C5/C6/C14):
  STT | Ngày | BKS/Loại xe | Nội dung dịch vụ | SL | ĐVT | Đơn giá |
  Thành tiền | VAT% | Tổng tiền | Số HĐ | (optional Thu hộ cols) | Ghi chú

Each service is expanded into ONE ROW PER COST (not per service), since
the customer wants every fee line itemized. Reimbursement rows can show
in a separate THU HỘ section if config.include_thu_ho = True.

Priority: full fee data + ghi chú correctness.
"""

from typing import Any, Dict, List, Optional

from openpyxl import Workbook

from app.api.exports.common_bang_ke_styling_and_formatting import (
    ALIGN_CC, ALIGN_CC_WRAP, ALIGN_LEFT, ALIGN_LEFT_WRAP, ALIGN_RIGHT,
    BORDER_ALL, FILL_DATA_ZEBRA, FILL_HEADER, FILL_TOTAL,
    NF_DATE_MDY, NF_INT,
    apply_border_to_range,
    build_bang_chu_row,
    build_bank_footer,
    build_company_header_block,
    build_customer_block,
    build_title_row,
    font,
    format_month_title,
    parse_details,
    safe_float,
)


COL_WIDTHS = {
    "A": 5.2,   # STT
    "B": 12.0,  # Ngày
    "C": 14.0,  # BKS / Loại xe
    "D": 38.0,  # Nội dung dịch vụ
    "E": 9.0,   # SL
    "F": 10.0,  # ĐVT
    "G": 13.0,  # Đơn giá
    "H": 14.0,  # Thành tiền
    "I": 8.0,   # VAT%
    "J": 14.0,  # Tổng tiền (sau VAT)
    "K": 14.0,  # Số HĐ
    "L": 22.0,  # Ghi chú
}


def _build_table_header(ws, header_row: int) -> None:
    headers = {
        "A": "STT",
        "B": "Ngày",
        "C": "BKS / Loại xe",
        "D": "Nội dung dịch vụ",
        "E": "Số lượng",
        "F": "ĐVT",
        "G": "Đơn giá",
        "H": "Thành tiền",
        "I": "VAT %",
        "J": "Tổng tiền (sau VAT)",
        "K": "Số HĐ",
        "L": "Ghi chú",
    }
    for col, label in headers.items():
        cell = ws[f"{col}{header_row}"]
        cell.value = label
        cell.font = font(12, bold=True)
        cell.alignment = ALIGN_CC_WRAP
        cell.fill = FILL_HEADER
        cell.border = BORDER_ALL
    ws.row_dimensions[header_row].height = 32


def _expand_service_to_cost_rows(
    svc: Dict[str, Any],
    job: Dict[str, Any],
    cost_rows: List[Dict[str, Any]],
    include_reim: bool,
) -> List[Dict[str, Any]]:
    """One job_costs row → one Excel row. Skips empty/zero rows."""
    d = parse_details(svc)
    bks_or_vehicle = d.get("vehicle_plate") or d.get("vehicle_type") or ""
    invoice = job.get("invoice_number") or svc.get("invoice_numbers") or ""
    note = d.get("note") or d.get("notes") or ""

    out = []
    for c in cost_rows:
        if not include_reim and c.get("is_reimbursement"):
            continue
        sell_amt = safe_float(c.get("selling_amount"))
        if sell_amt == 0:
            continue
        qty = safe_float(c.get("quantity")) or 1
        unit = c.get("unit") or "lô"
        unit_price = safe_float(c.get("selling_rate")) or (sell_amt / qty if qty else sell_amt)
        vat_rate = safe_float(c.get("vat_rate"))  # 0..100 in DB
        out.append({
            "ngay": svc.get("scheduled_date") or job.get("etd"),
            "bks": bks_or_vehicle,
            "noi_dung": c.get("cost_name") or "Dịch vụ",
            "qty": qty,
            "unit": unit,
            "don_gia": unit_price,
            "thanh_tien": sell_amt,
            "vat_pct": vat_rate,
            "invoice": invoice,
            "note": note,
            "is_reim": bool(c.get("is_reimbursement")),
        })

    # Fallback: if no cost rows had non-zero selling, synthesize ONE row from
    # service_details (selling_price/unit_price/grand_total). Keeps the bảng kê
    # non-empty when job_costs wasn't populated during import.
    if not out:
        unit_price_fb = safe_float(d.get("unit_price")) or safe_float(d.get("selling_price"))
        qty_fb = safe_float(d.get("quantity")) or 1
        vat_rate_fb = safe_float(d.get("vat_rate"))
        sell_amt_fb = unit_price_fb * qty_fb
        if sell_amt_fb == 0:
            gt = safe_float(d.get("total_revenue")) or safe_float(d.get("grand_total"))
            if gt > 0:
                sell_amt_fb = gt / (1 + vat_rate_fb / 100.0) if vat_rate_fb else gt
                unit_price_fb = sell_amt_fb
                qty_fb = 1
        if sell_amt_fb > 0:
            out.append({
                "ngay": svc.get("scheduled_date") or job.get("etd"),
                "bks": bks_or_vehicle,
                "noi_dung": d.get("note") or f"Dịch vụ {svc.get('service_type_code', '')}",
                "qty": qty_fb,
                "unit": d.get("unit") or "lô",
                "don_gia": unit_price_fb,
                "thanh_tien": sell_amt_fb,
                "vat_pct": vat_rate_fb,
                "invoice": invoice,
                "note": note,
                "is_reim": False,
            })
    return out


def _build_data_rows(ws, rows: List[Dict[str, Any]], start_row: int) -> int:
    common_font = font(11)
    for idx, r_data in enumerate(rows, start=1):
        r = start_row + idx - 1
        ws.row_dimensions[r].height = 28

        ws.cell(r, 1, idx)
        ws.cell(r, 2, r_data["ngay"])
        ws.cell(r, 3, r_data["bks"])
        ws.cell(r, 4, r_data["noi_dung"])
        ws.cell(r, 5, r_data["qty"])
        ws.cell(r, 6, r_data["unit"])
        ws.cell(r, 7, r_data["don_gia"])
        ws.cell(r, 8, f"=E{r}*G{r}")
        ws.cell(r, 9, r_data["vat_pct"] or 0)
        ws.cell(r, 10, f"=H{r}*(1+I{r}/100)")
        ws.cell(r, 11, r_data["invoice"])
        ws.cell(r, 12, r_data["note"])

        for c in range(1, 13):
            cell = ws.cell(r, c)
            cell.font = common_font
            cell.alignment = ALIGN_CC_WRAP
            cell.border = BORDER_ALL
            cell.fill = FILL_DATA_ZEBRA
        # Left-align nội dung + ghi chú
        ws.cell(r, 4).alignment = ALIGN_LEFT_WRAP
        ws.cell(r, 12).alignment = ALIGN_LEFT_WRAP

        ws.cell(r, 2).number_format = NF_DATE_MDY
        ws.cell(r, 5).number_format = "#,##0.##"
        ws.cell(r, 7).number_format = NF_INT
        ws.cell(r, 8).number_format = NF_INT
        ws.cell(r, 9).number_format = "0.##"
        ws.cell(r, 10).number_format = NF_INT

    return start_row + len(rows)


def _build_totals(ws, data_start: int, data_end_exclusive: int) -> int:
    """Tổng trước VAT (col H) + Tổng tiền sau VAT (col J)."""
    last_data = data_end_exclusive - 1
    tt_r = data_end_exclusive
    tc_r = tt_r + 1

    def _label(r: int, text: str) -> None:
        ws.merge_cells(f"A{r}:G{r}")
        ws.cell(r, 1, text)
        ws.cell(r, 1).font = font(12, bold=True, italic=True)
        ws.cell(r, 1).alignment = ALIGN_RIGHT
        for c in range(1, 13):
            ws.cell(r, c).fill = FILL_TOTAL
            ws.cell(r, c).border = BORDER_ALL

    _label(tt_r, "Tổng (trước VAT)")
    cell = ws.cell(tt_r, 8, f"=SUM(H{data_start}:H{last_data})")
    cell.font = font(12, bold=True)
    cell.alignment = ALIGN_RIGHT
    cell.fill = FILL_TOTAL
    cell.border = BORDER_ALL
    cell.number_format = NF_INT

    _label(tc_r, "Tổng thanh toán (đã VAT)")
    cell = ws.cell(tc_r, 10, f"=SUM(J{data_start}:J{last_data})")
    cell.font = font(12, bold=True)
    cell.alignment = ALIGN_RIGHT
    cell.fill = FILL_TOTAL
    cell.border = BORDER_ALL
    cell.number_format = NF_INT

    return tc_r


# ---------- Public entry point ----------

def render_handling_workbook(
    customer: Dict[str, Any],
    services: List[Dict[str, Any]],
    jobs_map: Dict[Any, Dict[str, Any]],
    costs_by_svc: Dict[Any, List[Dict[str, Any]]],
    month: Optional[str],
    logo_path: Optional[str] = None,
    config: Optional[Dict[str, Any]] = None,
) -> Workbook:
    """
    Build handling-debit workbook.

    config (optional):
      sheet_name        → str
      title_template    → str with {month}, default 'BẢNG KÊ DỊCH VỤ LÀM HÀNG THÁNG {month}'
      include_reim      → bool, default True (include is_reimbursement rows)
      include_bang_chu  → bool, default True
      include_bank      → bool, default True
    """
    cfg = config or {}
    wb = Workbook()
    ws = wb.active
    ws.title = (cfg.get("sheet_name") or customer.get("short_name") or "Bảng kê")[:31]

    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    build_company_header_block(ws, logo_path=logo_path)

    title_month = format_month_title(month, style="vi_short")
    title_template = cfg.get("title_template") or "BẢNG KÊ DỊCH VỤ LÀM HÀNG THÁNG {month}"
    build_title_row(ws, row=6, text=title_template.format(month=title_month),
                    merge_range="A6:L6", size=14, height=28)

    build_customer_block(ws, customer, start_row=8, style="customer")

    header_row = 12
    _build_table_header(ws, header_row=header_row)

    services_sorted = sorted(
        services,
        key=lambda s: (str(s.get("scheduled_date") or ""), s.get("svc_id") or 0),
    )

    include_reim = cfg.get("include_reim", True)
    all_rows: List[Dict[str, Any]] = []
    for svc in services_sorted:
        cost_rows = costs_by_svc.get(svc["svc_id"], [])
        all_rows.extend(_expand_service_to_cost_rows(
            svc, jobs_map.get(svc["job_id"], {}), cost_rows, include_reim
        ))

    after_data = _build_data_rows(ws, all_rows, start_row=header_row + 1)
    last_total = _build_totals(ws, data_start=header_row + 1, data_end_exclusive=after_data)

    if cfg.get("include_bang_chu", True):
        grand_total = sum(r["thanh_tien"] * (1 + (r["vat_pct"] / 100.0)) for r in all_rows)
        build_bang_chu_row(ws, last_total + 2, grand_total)
        last_total += 2

    if cfg.get("include_bank", True):
        build_bank_footer(ws, last_total + 3, merge_last_col="F")

    return wb
