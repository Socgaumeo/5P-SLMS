"""
Renderer for DAINESE 'Bảng kê chi tiết dịch vụ hàng xuất' (XUẤT - exports).

Reproduces customer reference layout:
  plans/reports/dainese-templates/file-05-bang-ke-xuat-summary.md

Sheet 'XUẤT'. 32 cols A-AF. Same general structure as File 3 but with:
- Extra W col 'Phụ phí xăng dầu' (formula =M*0.145)
- AD/AE 'Thu chi hộ' (HĐ TCH / Total) instead of single AC col
- AF 'Số hóa đơn của forwarder'

Service filter: CUS_EXPORT for DAINESE (could be extended to
SEA_EXP/AIR_EXP/BORDER_EXP if those become populated).
"""

import json
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.drawing.image import Image as XLImage

from app.api.exports.dainese_cost_name_to_column_mapper import (
    aggregate_costs_into_columns,
)


FONT_NAME = "Times New Roman"

THIN = Side(style="thin", color="000000")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_CC_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_CC = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

NF_INT = "#,##0"
NF_DATE = "d/m/yyyy"
NF_ACCT = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'

FILL_HEADER = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
FILL_DATA = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_TOTAL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

COMPANY_INFO = {
    "name": "CÔNG TY TNHH THƯƠNG MẠI VÀ DỊCH VỤ 5P VIỆT NAM",
    "address": "Số nhà 02 Ngõ 1H Phố Trần Quang Diệu, Phường Đống Đa, Thành phố Hà Nội, Việt Nam",
    "mst": "MST: 0110523309",
}

BANK_INFO = [
    "Thông tin chuyển khoản:",
    "Tài khoản: Công Ty TNHH Thương mại và dịch vụ 5P Việt Nam",
    "Số tài khoản: 346886",
    "Tại Ngân hàng: Ngân hàng TMCP Kỹ thương Việt Nam - Techcombank Chi nhánh Đông Đô",
]

COL_WIDTHS = {
    "A": 5.4,  "B": 16.0, "C": 21.2, "D": 32.5, "E": 13.7, "F": 27.7, "G": 11.4,
    "H": 13.5, "I": 13.4, "J": 12.8, "K": 16.2, "L": 14.0, "M": 15.4, "N": 11.0,
    "O": 11.0, "P": 11.0, "Q": 14.5, "R": 14.5, "S": 13.7, "T": 12.6, "U": 13.8,
    "V": 11.0, "W": 11.0, "X": 12.1, "Y": 11.5, "Z": 11.5, "AA": 14.2, "AB": 14.2,
    "AC": 18.5, "AD": 16.7, "AE": 17.6, "AF": 16.9,
}


def _font(size: int = 11, bold: bool = False, italic: bool = False) -> Font:
    return Font(name=FONT_NAME, size=size, bold=bold, italic=italic)


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _parse_details(svc: Dict[str, Any]) -> Dict[str, Any]:
    raw = svc.get("service_details") or {}
    if isinstance(raw, str):
        try:
            return json.loads(raw)
        except Exception:
            return {}
    return raw


def _format_month_title(month: Optional[str]) -> str:
    if month and len(month) == 7 and month[4] == "-":
        year, mon = month.split("-")
        return f"{mon} NĂM {year}"
    now = datetime.now()
    return f"{now.month:02d} NĂM {now.year}"


def _apply_border_to_range(ws, rng: str, fill: Optional[PatternFill] = None) -> None:
    c1, r1, c2, r2 = range_boundaries(rng)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_ALL
            if fill is not None:
                cell.fill = fill


def _format_luong(channel: Optional[str]) -> str:
    if not channel:
        return ""
    c = channel.strip()
    if c.lower().startswith("luồng"):
        return c
    return c  # File 5 uses just 'Xanh'/'Vàng'/'Đỏ' (no 'Luồng' prefix)


# ---- Reim splitter for File 5 (3 buckets) ----

def _split_reim(cost_rows: List[Dict[str, Any]]) -> Dict[str, float]:
    """Bucket reim costs into Local / CSHT / Lưu kho for cols Y/Z/AA."""
    out = {"local": 0.0, "csht": 0.0, "luu_kho": 0.0}
    for c in cost_rows:
        if not c.get("is_reimbursement"):
            continue
        name = (c.get("cost_name") or "").lower()
        amt = _safe_float(c.get("selling_amount"))
        if "local charge" in name or "lcc" in name:
            out["local"] += amt
        elif "csht" in name or "thuế" in name or "thue" in name or "vé bãi" in name:
            out["csht"] += amt
        elif "lưu kho" in name or "luu kho" in name or "bốc xếp" in name or "boc xep" in name or "nâng" in name or "nang" in name or "hạ hàng" in name or "ha hang" in name or "thu hộ" in name or "thu ho" in name:
            out["luu_kho"] += amt
        else:
            out["luu_kho"] += amt
    return out


# ---- Section builders ----

def _build_header(ws, customer: Dict[str, Any], logo_path: Optional[str], title_month: str) -> None:
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w
    for r in (1, 2, 3, 4):
        ws.row_dimensions[r].height = 23
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[10].height = 16
    ws.row_dimensions[11].height = 30
    ws.row_dimensions[12].height = 36

    if logo_path:
        try:
            img = XLImage(logo_path)
            img.width = 80
            img.height = 80
            ws.add_image(img, "A1")
        except Exception:
            pass

    ws["C1"] = COMPANY_INFO["name"]
    ws["C2"] = COMPANY_INFO["address"]
    ws["C3"] = COMPANY_INFO["mst"]
    for coord in ("C1", "C2", "C3"):
        ws[coord].font = _font(12, bold=True)
        ws[coord].alignment = ALIGN_LEFT

    ws.merge_cells("A5:AF5")
    ws["A5"] = f"BẢNG KÊ CHI TIẾT DỊCH VỤ HÀNG XUẤT THÁNG {title_month}"
    ws["A5"].font = _font(18, bold=True)
    ws["A5"].alignment = ALIGN_CC

    cust_name = (
        customer.get("company_name")
        or customer.get("short_name")
        or customer.get("customer_code", "")
    )
    cust_addr = customer.get("address") or ""
    cust_attn = customer.get("contact_name") or ""

    ws["B7"] = f"KÍNH GỬI :  {cust_name}"
    ws["B8"] = f"Địa chỉ : {cust_addr}"
    ws["B9"] = f"Attn:   {cust_attn}"
    for coord in ("B7", "B8", "B9"):
        ws[coord].font = _font(12, bold=True)
        ws[coord].alignment = ALIGN_LEFT


def _build_table_header(ws) -> None:
    """Row 11 group headers + row 12 column headers."""
    group_headers = [
        ("A11:G11", "Thông tin chung"),
        ("H11:I11", "Khối lượng"),
        ("J11:J12", "Luồng"),
        ("K11:X11", "Phí dịch vụ làm hàng  ( VND) "),
        ("Y11:AB11", "Phí trả hộ"),
        ("AC11:AC12", "Tổng tiền phải trả "),
        ("AD11:AE11", "Thu chi hộ"),
        ("AF11:AF12", "Số hóa đơn của forwarder"),
    ]
    for rng, label in group_headers:
        ws.merge_cells(rng)
        first = rng.split(":")[0]
        ws[first] = label
        ws[first].font = _font(12, bold=True, italic=True)
        ws[first].alignment = ALIGN_CC_WRAP
        _apply_border_to_range(ws, rng, fill=FILL_HEADER)

    col_headers = {
        "A12": "No.",
        "B12": "Tờ khai",
        "C12": "Hóa đơn thương mại",
        "D12": "Note",
        "E12": "Ngày tờ khai",
        "F12": "Tuyến đường",
        "G12": "Loại hình",
        "H12": "Kgs",
        "I12": "No. Cont",
        "K12": "Phí mở tờ khai hải quan",
        "L12": "Phí kiểm hóa",
        "M12": "Phí vận chuyển",
        "N12": "Phí làm hàng",
        "O12": "Phí phát sinh khác",
        "P12": "Phí phục vụ kiểm hóa tại cảng",
        "Q12": "Phí đầu nước ngoài",
        "R12": "Cước vận tải quốc tế",
        "S12": "Phí xếp dỡ (THC)",
        "T12": "Phí gom hàng lẻ (CFS)/\nCIC/ LSS",
        "U12": "Phí lấy lệnh  (DO)",
        "V12": "Phí  đại lý,",
        "W12": "Phụ phí xăng dầu",
        "X12": "Tổng",
        "Y12": "Local charge",
        "Z12": "CSHT, thuế",
        "AA12": "Lưu kho giao nhận bốc xếp, nâng hạ",
        "AB12": "Tổng phí trả hộ",
        "AD12": "HĐ TCH",
        "AE12": "Total",
    }
    for coord, val in col_headers.items():
        ws[coord] = val
        ws[coord].font = _font(12, bold=True)
        ws[coord].alignment = ALIGN_CC_WRAP
        ws[coord].fill = FILL_HEADER
        ws[coord].border = BORDER_ALL


def _service_to_row(svc, job, cost_rows, idx) -> Dict[str, Any]:
    d = _parse_details(svc)
    fees = aggregate_costs_into_columns(cost_rows)
    reim = _split_reim(cost_rows)
    mode = (d.get("mode") or "").upper()
    loai = mode if mode in ("DHL", "SEA", "AIR", "KNQ") else (mode or "")
    return {
        "no": idx,
        "to_khai": svc.get("cd_no") or "",
        "hd_tm": svc.get("invoice_numbers") or d.get("invoice_numbers") or "",
        "note_d": svc.get("buyer_name") or svc.get("seller_name") or d.get("note") or "",
        "ngay_tk": svc.get("scheduled_date") or job.get("etd") or "",
        "tuyen": svc.get("route") or d.get("route") or "",
        "loai": loai,
        "kgs": svc.get("weight_kg") or d.get("weight_kg") or "",
        "cont": d.get("container") or "",
        "luong": _format_luong(d.get("customs_channel")),
        # K-V (12 cols own fees, mapped from buckets)
        "phi_mtk": fees["phi_mtk"],
        "phi_kh": fees["phi_kh"],
        "phi_vc": fees["phi_vc"],
        "phi_lh": fees["phi_lh"],
        "phi_psk": fees["phi_psk"],
        "phi_pv_kh_tc": 0,
        "phi_dnn": fees["phi_dnn"],
        "cuoc_qt": fees["cuoc_qt"],
        "thc": fees["thc"],
        "cfs": fees["cfs"],
        "do": fees["do"],
        "dly": fees["dly"],
        # Y-AA reim (3 cols)
        "y_local": reim["local"],
        "z_csht": reim["csht"],
        "aa_luu_kho": reim["luu_kho"],
        # AD/AE/AF
        "hd_tch": d.get("invoice_hd") or "",
        "fwd_total": "",
        "hd_fwd": "",
    }


def _build_data_rows(ws, services_with_costs: List[tuple], jobs_map: Dict, start_row: int = 13) -> int:
    common_font = _font(12, bold=True)
    for idx, (svc, cost_rows) in enumerate(services_with_costs, start=1):
        r = start_row + idx - 1
        ws.row_dimensions[r].height = 33

        rd = _service_to_row(svc, jobs_map.get(svc["job_id"], {}), cost_rows, idx)

        ws.cell(r, 1, idx)
        ws.cell(r, 2, rd["to_khai"])
        ws.cell(r, 3, rd["hd_tm"])
        ws.cell(r, 4, rd["note_d"])
        ws.cell(r, 5, rd["ngay_tk"])
        ws.cell(r, 6, rd["tuyen"])
        ws.cell(r, 7, rd["loai"])
        ws.cell(r, 8, rd["kgs"])
        ws.cell(r, 9, rd["cont"])
        ws.cell(r, 10, rd["luong"])

        # K..V (cols 11..22) own fees
        fee_keys = [
            "phi_mtk", "phi_kh", "phi_vc", "phi_lh", "phi_psk",
            "phi_pv_kh_tc", "phi_dnn", "cuoc_qt", "thc", "cfs", "do", "dly",
        ]
        for i, k in enumerate(fee_keys):
            v = rd.get(k) or 0
            cell = ws.cell(r, 11 + i, v if v else None)
            cell.number_format = NF_INT

        # W (23): Phụ phí xăng dầu = M*0.145 (per reference formula)
        ws.cell(r, 23, f"=M{r}*0.145").number_format = NF_ACCT
        # X (24): =SUM(K:W)
        ws.cell(r, 24, f"=SUM(K{r}:W{r})").number_format = NF_INT

        # Y..AA (25..27) reim
        for i, k in enumerate(["y_local", "z_csht", "aa_luu_kho"]):
            v = rd.get(k) or 0
            ws.cell(r, 25 + i, v if v else None).number_format = NF_INT
        # AB (28): =SUM(Y:AA)
        ws.cell(r, 28, f"=SUM(Y{r}:AA{r})").number_format = NF_INT
        # AC (29): =X+AB
        ws.cell(r, 29, f"=+X{r}+AB{r}").number_format = NF_INT
        # AD (30) HĐ TCH, AE (31) Total, AF (32) Số HĐ FWD
        ws.cell(r, 30, rd.get("hd_tch", ""))
        ws.cell(r, 31, rd.get("fwd_total", ""))
        ws.cell(r, 32, rd.get("hd_fwd", ""))

        for c in range(1, 33):
            cell = ws.cell(r, c)
            cell.font = common_font
            cell.alignment = ALIGN_CC_WRAP
            cell.border = BORDER_ALL

        ws.cell(r, 5).number_format = NF_DATE

    return start_row + len(services_with_costs)


def _build_totals(ws, data_start: int, data_end_exclusive: int) -> int:
    last_data = data_end_exclusive - 1
    tong_row = data_end_exclusive
    vat_row = tong_row + 1
    tc_row = vat_row + 1

    def _label(row: int, text: str) -> None:
        ws.merge_cells(f"A{row}:I{row}")
        ws.cell(row, 1, text)
        ws.cell(row, 1).font = _font(12, bold=True, italic=True)
        ws.cell(row, 1).alignment = ALIGN_CC
        for c in range(1, 10):
            ws.cell(row, c).fill = FILL_TOTAL
            ws.cell(row, c).border = BORDER_ALL

    _label(tong_row, "Tổng")
    for c in range(11, 30):  # K..AC
        col = get_column_letter(c)
        cell = ws.cell(tong_row, c, f"=SUM({col}{data_start}:{col}{last_data})")
        cell.font = _font(12, bold=True)
        cell.alignment = ALIGN_CC
        cell.fill = FILL_TOTAL
        cell.border = BORDER_ALL
        cell.number_format = NF_INT

    _label(vat_row, "VAT")
    for c in range(11, 30):
        cell = ws.cell(vat_row, c, 0)
        cell.font = _font(12, bold=True)
        cell.alignment = ALIGN_CC
        cell.fill = FILL_TOTAL
        cell.border = BORDER_ALL
        cell.number_format = NF_INT

    _label(tc_row, "Tổng cộng")
    for c in range(11, 30):
        col = get_column_letter(c)
        cell = ws.cell(tc_row, c, f"=SUM({col}{tong_row}:{col}{vat_row})")
        cell.font = _font(12, bold=True)
        cell.alignment = ALIGN_CC
        cell.fill = FILL_TOTAL
        cell.border = BORDER_ALL
        cell.number_format = NF_INT

    return tc_row


def _build_footer(ws, start_row: int) -> None:
    for i, line in enumerate(BANK_INFO):
        r = start_row + i
        ws.cell(r, 1, line)
        if i == 0:
            ws.cell(r, 1).font = _font(12, bold=True, italic=True)
        else:
            ws.cell(r, 1).font = _font(12)
        if i == len(BANK_INFO) - 1:
            ws.merge_cells(f"A{r}:F{r}")
            ws.cell(r, 1).alignment = ALIGN_LEFT_WRAP


# ---- Public entry point ----

def render_xuat_workbook(
    customer: Dict[str, Any],
    services: List[Dict[str, Any]],
    jobs_map: Dict[Any, Dict[str, Any]],
    costs_by_svc: Dict[Any, List[Dict[str, Any]]],
    month: Optional[str],
    logo_path: Optional[str],
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "XUẤT"

    title_month = _format_month_title(month)
    _build_header(ws, customer, logo_path, title_month)
    _build_table_header(ws)

    # Filter to real exports: services where buyer/seller (service_details.note)
    # is a foreign company (DAINESE S.P.A., MAVETS S.R.L., TOPKEY CORP, etc.).
    # Services with VN buyer/seller go to File 3 (xuất nhập tại chỗ).
    _FOREIGN_NAME_MARKERS = (
        "dainese s.p.a", "dainese spa", "s.p.a", "s.r.l", "co.,ltd",
        "corporation", "gmbh", "limited", "ltd.", "inc.", "sas",
        "pte ltd", "co. ltd",
    )

    def _is_real_export(svc):
        code = (svc.get("service_type_code") or "").upper()
        d = _parse_details(svc)
        # Real CO services go to File 2
        if d.get("co_no"):
            return False
        note = (d.get("note") or "").lower().strip()
        is_foreign = any(m in note for m in _FOREIGN_NAME_MARKERS)
        # Match if foreign party + has a customs declaration
        if str(svc.get("cd_no") or "").strip() and is_foreign:
            return True
        # Future-proof: service codes for real exports
        if code in ("SEA_EXP", "AIR_EXP", "BORDER_EXP"):
            return True
        return False

    services_with_costs = sorted(
        [(svc, costs_by_svc.get(svc["svc_id"], []))
         for svc in services
         if _is_real_export(svc)],
        key=lambda pair: (pair[0].get("scheduled_date") or "", pair[0].get("svc_id") or 0),
    )

    after_data = _build_data_rows(ws, services_with_costs, jobs_map, start_row=13)
    last_total = _build_totals(ws, data_start=13, data_end_exclusive=after_data)
    _build_footer(ws, last_total + 9)

    return wb
