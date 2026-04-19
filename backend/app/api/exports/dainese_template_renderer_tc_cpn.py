"""
Renderer for DAINESE 'Bảng kê hàng xuất nhập tại chỗ và hàng nhập CPN'.

Reproduces layout/styles/formulas of customer reference file:
  plans/reports/dainese-templates/file-03-bang-ke-tc-nhap-cpn-summary.md

Service filter: services with service_type_code='CUS_CO' AND
service_details.mode IN ('KNQ', 'DHL') AND no co_no
(those are TC for KNQ = nhập khẩu tại chỗ, CPN for DHL/express).

Layout vs File 1 (NHẬP):
- Title row 5 (slightly different text)
- Extra column J = Luồng (customs channel)
- Cols K..V = 12 own-fee columns (extra P = Phí phục vụ kiểm hóa tại cảng)
- W = SUM(K:U), AA = SUM(X:Z), AB = W + AA
- Headers font 14pt (vs 12pt)
- Totals at row N+1 (row 50 in reference for 37 data rows)
"""

import json
import re
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.drawing.image import Image as XLImage

from app.api.exports.dainese_cost_name_to_column_mapper import (
    aggregate_costs_into_columns,
)


# ---- Style constants ----

FONT_NAME = "Times New Roman"

THIN = Side(style="thin", color="000000")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_CC_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_CC = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

NF_INT = "#,##0"
NF_DEC = "#,##0.00"
NF_DATE = "mm-dd-yy"
NF_ACCT = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'

FILL_HEADER = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
FILL_DATA = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_TOTAL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

COMPANY_INFO = {
    "name": "CÔNG TY TNHH THƯƠNG MẠI VÀ DỊCH VỤ 5P VIỆT NAM",
    "address": "Số nhà 02 Ngõ 1H Phố Trần Quang Diệu, Phường Đống Đa, Thành phố Hà Nội, Việt Nam",
    "mst": "MST: 0110523309",
}

BANK_INFO = [
    "Thông tin chuyển khoản:",
    "Tài khoản: Công Ty TNHH Thương mại và dịch vụ 5P Việt Nam",
    "Số tài khoản: 346886",
    "Tại Ngân hàng: Ngân hàng TMCP Kỹ thương Việt Nam - Techcombank Chi nhánh Đông Đô",
]

# Column widths for sheet "tc, cpn" (matches reference)
COL_WIDTHS = {
    "A": 5.4,  "B": 18.7, "C": 28.0, "D": 63.9, "E": 19.9, "F": 36.5, "G": 11.4,
    "H": 13.5, "I": 13.4, "J": 17.1, "K": 16.2, "L": 14.0, "M": 17.4, "N": 11.0,
    "O": 14.0, "P": 11.0, "Q": 14.5, "R": 14.5, "S": 13.7, "T": 12.6, "U": 13.8,
    "V": 11.0, "W": 18.1, "X": 11.5, "Y": 11.5, "Z": 14.2, "AA": 14.2, "AB": 18.5,
    "AC": 15.9, "AD": 14.5,
}


# ---- Helpers ----

def _font(size: int = 11, bold: bool = False, italic: bool = False) -> Font:
    return Font(name=FONT_NAME, size=size, bold=bold, italic=italic)


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _parse_details(svc: Dict[str, Any]) -> Dict[str, Any]:
    raw = svc.get("service_details") or {}
    if isinstance(raw, str):
        try:
            return json.loads(raw)
        except Exception:
            return {}
    return raw


def _format_month_title(month: Optional[str]) -> str:
    if month and len(month) == 7 and month[4] == "-":
        year, mon = month.split("-")
        return f"{mon} NĂM {year}"
    now = datetime.now()
    return f"{now.month:02d} NĂM {now.year}"


def _apply_border_to_range(ws, rng: str, fill: Optional[PatternFill] = None) -> None:
    c1, r1, c2, r2 = range_boundaries(rng)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_ALL
            if fill is not None:
                cell.fill = fill


# ---- Filter for TC + CPN services ----

def is_tc_or_cpn_service(svc: Dict[str, Any], cost_rows: List[Dict[str, Any]]) -> bool:
    """
    A service belongs in Bảng kê TC/CPN (xuất nhập tại chỗ + CPN) if:
    - It has a customs declaration (cd_no) AND
    - It's NOT a real C/O service (covered by File 2), AND EITHER
      a) cd_no starts with '108' (nhập tại chỗ from Vietnamese sellers), OR
      b) service_details.mode is KNQ/DHL (CUS_CO TC + CPN imports), OR
      c) service is CUS_IMPORT (express courier).

    Real international exports (cd_no starting with '308' on CUS_EXPORT
    without KNQ/DHL mode) go to File 5.
    """
    d = _parse_details(svc)
    # Exclude real CO services (covered by File 2)
    if d.get("co_no"):
        return False
    for c in cost_rows:
        name_lc = (c.get("cost_name") or "").lower()
        if "phí c/o" in name_lc or "phi co" in name_lc:
            return False

    cd_no = str(svc.get("cd_no") or "").strip()
    if not cd_no:
        return False

    code = (svc.get("service_type_code") or "").upper()
    mode = (d.get("mode") or "").upper()

    # CUS_CO with KNQ/DHL = TC/CPN imports
    if code == "CUS_CO" and mode in ("KNQ", "DHL", "TC", "CPN"):
        return True

    # CUS_IMPORT = always express courier import → File 3
    if code == "CUS_IMPORT":
        return True

    # CUS_EXPORT with cd_no starting '108' = xuất nhập tại chỗ from VN seller
    if code == "CUS_EXPORT" and cd_no.startswith("108"):
        return True

    return False


def _format_luong(channel: Optional[str]) -> str:
    """'Xanh' → 'Luồng xanh', etc."""
    if not channel:
        return ""
    c = channel.strip()
    if c.lower().startswith("luồng"):
        return c
    return f"Luồng {c.lower()}"


def _format_loai_hinh(mode: Optional[str]) -> str:
    """KNQ → 'DOM' (tại chỗ), DHL → 'CPN'."""
    if not mode:
        return ""
    m = mode.upper()
    if m == "KNQ" or m == "TC":
        return "DOM"
    if m == "DHL" or m == "CPN":
        return "CPN"
    return m


# ---- Section builders ----

def _build_header(ws, customer: Dict[str, Any], logo_path: Optional[str], title_month: str) -> None:
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w
    for r in (1, 2, 3, 4):
        ws.row_dimensions[r].height = 23
    ws.row_dimensions[5].height = 52
    for r in (7, 8, 9, 10):
        ws.row_dimensions[r].height = 18
    ws.row_dimensions[11].height = 38
    ws.row_dimensions[12].height = 41

    if logo_path:
        try:
            img = XLImage(logo_path)
            img.width = 90
            img.height = 90
            ws.add_image(img, "A1")
        except Exception:
            pass

    ws["C1"] = COMPANY_INFO["name"]
    ws["C2"] = COMPANY_INFO["address"]
    ws["C3"] = COMPANY_INFO["mst"]
    for coord in ("C1", "C2", "C3"):
        ws[coord].font = _font(12, bold=True)
        ws[coord].alignment = ALIGN_LEFT

    # Title row 5 (merged A5:AD5)
    ws.merge_cells("A5:AD5")
    ws["A5"] = (
        f"BẢNG KÊ CHI TIẾT HÀNG XUẤT NHẬP TẠI CHỖ VÀ HÀNG NHẬP CPN THÁNG {title_month}"
    )
    ws["A5"].font = _font(22, bold=True)
    ws["A5"].alignment = ALIGN_CC_WRAP

    cust_name = (
        customer.get("company_name")
        or customer.get("short_name")
        or customer.get("customer_code", "")
    )
    cust_addr = customer.get("address") or ""
    cust_attn = customer.get("contact_name") or ""

    ws["B7"] = f"KÍNH GỬI :  {cust_name}"
    ws["B8"] = f"Địa chỉ : {cust_addr}"
    ws["B9"] = f"Attn:   {cust_attn}"
    for coord in ("B7", "B8", "B9"):
        ws[coord].font = _font(14, bold=True)
        ws[coord].alignment = ALIGN_LEFT


def _build_table_header(ws) -> None:
    """Two-row header at rows 11-12 with merged group labels."""
    group_headers = [
        ("A11:G11", "Thông tin chung"),
        ("H11:I11", "Khối lượng"),
        ("J11:J12", "Luồng"),
        ("K11:W11", "Phí dịch vụ làm hàng  ( VND) "),
        ("X11:AA11", "Phí trả hộ"),
        ("AB11:AB12", "Tổng tiền phải trả "),
        ("AC11:AC12", "Số hóa đơn trả hộ"),
        ("AD11:AD12", "Số hóa đơn của forwarder"),
    ]
    for rng, label in group_headers:
        ws.merge_cells(rng)
        first = rng.split(":")[0]
        ws[first] = label
        ws[first].font = _font(14, bold=True, italic=True)
        ws[first].alignment = ALIGN_CC_WRAP
        _apply_border_to_range(ws, rng, fill=FILL_HEADER)

    col_headers = {
        "A12": "No.",
        "B12": "Tờ khai",
        "C12": "Hóa đơn thương mại",
        "D12": "Note",
        "E12": "Ngày tờ khai",
        "F12": "Tuyến đường",
        "G12": "Loại hình",
        "H12": "Kgs",
        "I12": "No. Cont",
        "K12": "Phí mở tờ khai hải quan",
        "L12": "Phí kiểm hóa",
        "M12": "Phí vận chuyển",
        "N12": "Phí làm hàng",
        "O12": "Phí phát sinh khác",
        "P12": "Phí phục vụ kiểm hóa tại cảng",
        "Q12": "Phí đầu nước ngoài",
        "R12": "Cước vận tải quốc tế",
        "S12": "Phí xếp dỡ (THC)",
        "T12": "Phí gom hàng lẻ (CFS)/\nCIC/ LSS",
        "U12": "Phí lấy lệnh  (DO)",
        "V12": "Phí  đại lý,",
        "W12": "Tổng",
        "X12": "Local charge",
        "Y12": "Lệ phí",
        "Z12": "Vé bãi kiểm hóa",
        "AA12": "Tổng phí trả hộ",
    }
    for coord, val in col_headers.items():
        ws[coord] = val
        ws[coord].font = _font(14, bold=True)
        ws[coord].alignment = ALIGN_CC_WRAP
        ws[coord].fill = FILL_HEADER
        ws[coord].border = BORDER_ALL


# Bucket → reim column key for File 3:
# - "Local charge"        → X (local)
# - "Lệ phí" (Thu hộ)     → Y (le_phi)
# - "Vé bãi", "CSHT"      → Z (ve_bai)
def _split_reim_for_file3(cost_rows: List[Dict[str, Any]]) -> Dict[str, float]:
    """Walk reimbursement rows and bucket into local / le_phi / ve_bai."""
    out = {"local": 0.0, "le_phi": 0.0, "ve_bai": 0.0}
    for c in cost_rows:
        if not c.get("is_reimbursement"):
            continue
        name = (c.get("cost_name") or "").lower()
        amt = _safe_float(c.get("selling_amount"))
        if "local charge" in name or "lcc" in name:
            out["local"] += amt
        elif "lệ phí" in name or "le phi" in name or "thu hộ" in name or "thu ho" in name:
            out["le_phi"] += amt
        elif "vé bãi" in name or "ve bai" in name or "csht" in name:
            out["ve_bai"] += amt
        else:
            # Default unknown reim → vé bãi bucket
            out["ve_bai"] += amt
    return out


def _service_to_row(svc, job, cost_rows, idx) -> Dict[str, Any]:
    d = _parse_details(svc)
    fees = aggregate_costs_into_columns(cost_rows)
    reim = _split_reim_for_file3(cost_rows)
    return {
        "no": idx,
        "to_khai": svc.get("cd_no") or d.get("cd_no") or "",
        "hd_tm": svc.get("invoice_numbers") or d.get("invoice_numbers") or "",
        "note_d": svc.get("seller_name") or d.get("note") or "",
        "ngay_tk": svc.get("scheduled_date") or job.get("etd") or "",
        "tuyen": svc.get("route") or d.get("route") or "",
        "loai": _format_loai_hinh(d.get("mode")),
        "kgs": svc.get("weight_kg") or d.get("weight_kg") or "",
        "cont": d.get("container") or "",
        "luong": _format_luong(d.get("customs_channel")),
        # K-V (12 fees)
        "phi_mtk": fees["phi_mtk"],
        "phi_kh": fees["phi_kh"],
        "phi_vc": fees["phi_vc"],
        "phi_lh": fees["phi_lh"],
        "phi_psk": fees["phi_psk"],
        "phi_pv_kh_tc": 0,  # bucket not in current mapper - reserved
        "phi_dnn": fees["phi_dnn"],
        "cuoc_qt": fees["cuoc_qt"],
        "thc": fees["thc"],
        "cfs": fees["cfs"],
        "do": fees["do"],
        "dly": fees["dly"],
        # X-Z (3 reim)
        "x_local": reim["local"],
        "y_le_phi": reim["le_phi"],
        "z_ve_bai": reim["ve_bai"],
        # AC, AD: invoice numbers (data not in DB normally)
        "hd_traho": d.get("invoice_hd") or "",
        "hd_fwd": "",
    }


def _build_data_rows(ws, rows: List[Dict[str, Any]], start_row: int = 13) -> int:
    """Write data rows. Returns row index AFTER last data row."""
    common_font = _font(11)
    for idx, r_data in enumerate(rows, start=1):
        r = start_row + idx - 1
        ws.row_dimensions[r].height = 52

        ws.cell(r, 1, idx)
        ws.cell(r, 2, r_data["to_khai"])
        ws.cell(r, 3, r_data["hd_tm"])
        ws.cell(r, 4, r_data["note_d"])
        ws.cell(r, 5, r_data["ngay_tk"])
        ws.cell(r, 6, r_data["tuyen"])
        ws.cell(r, 7, r_data["loai"])
        ws.cell(r, 8, r_data["kgs"])
        ws.cell(r, 9, r_data["cont"])
        ws.cell(r, 10, r_data["luong"])

        # K..V (cols 11..22): 12 own fees
        fee_keys = [
            "phi_mtk", "phi_kh", "phi_vc", "phi_lh", "phi_psk",
            "phi_pv_kh_tc", "phi_dnn", "cuoc_qt", "thc", "cfs", "do", "dly",
        ]
        for i, k in enumerate(fee_keys):
            v = r_data.get(k) or 0
            cell = ws.cell(r, 11 + i, v if v else None)
            cell.number_format = NF_INT

        # W (23): =SUM(K:U) - only K..U (11..21), excludes V (đại lý). Match File 1 formula.
        # Per reference: W13 = =SUM(K13:U13)
        ws.cell(r, 23, f"=SUM(K{r}:U{r})").number_format = NF_INT

        # X..Z (24..26): reim
        for i, k in enumerate(["x_local", "y_le_phi", "z_ve_bai"]):
            v = r_data.get(k) or 0
            ws.cell(r, 24 + i, v if v else None).number_format = NF_INT

        # AA (27): =SUM(X:Z)
        ws.cell(r, 27, f"=SUM(X{r}:Z{r})").number_format = NF_INT
        # AB (28): =W+AA
        ws.cell(r, 28, f"=+W{r}+AA{r}").number_format = NF_INT
        # AC (29), AD (30)
        ws.cell(r, 29, r_data.get("hd_traho", ""))
        ws.cell(r, 30, r_data.get("hd_fwd", ""))

        for c in range(1, 31):
            cell = ws.cell(r, c)
            cell.font = common_font
            cell.alignment = ALIGN_CC_WRAP
            cell.border = BORDER_ALL
            cell.fill = FILL_DATA

        # Date format for col E
        ws.cell(r, 5).number_format = NF_DATE

    return start_row + len(rows)


def _build_totals(ws, data_start: int, data_end_exclusive: int) -> int:
    last_data = data_end_exclusive - 1
    tong_row = data_end_exclusive
    vat_row = tong_row + 1
    tc_row = vat_row + 1

    def _label_row(row: int, label: str) -> None:
        ws.merge_cells(f"A{row}:I{row}")
        ws.cell(row, 1, label)
        ws.cell(row, 1).font = _font(14, bold=True, italic=True)
        ws.cell(row, 1).alignment = ALIGN_CC_WRAP
        for c in range(1, 10):
            ws.cell(row, c).fill = FILL_TOTAL
            ws.cell(row, c).border = BORDER_ALL

    _label_row(tong_row, "Tổng")
    for c in range(11, 28):  # K..AA
        col = get_column_letter(c)
        cell = ws.cell(tong_row, c, f"=SUM({col}{data_start}:{col}{last_data})")
        cell.font = _font(14, bold=True)
        cell.alignment = ALIGN_CC
        cell.fill = FILL_TOTAL
        cell.border = BORDER_ALL
        cell.number_format = NF_INT

    _label_row(vat_row, "VAT")
    for c in range(11, 28):
        cell = ws.cell(vat_row, c, 0)
        cell.font = _font(14, bold=True)
        cell.alignment = ALIGN_CC
        cell.fill = FILL_TOTAL
        cell.border = BORDER_ALL
        cell.number_format = NF_INT

    _label_row(tc_row, "Tổng cộng")
    for c in range(11, 28):
        col = get_column_letter(c)
        cell = ws.cell(tc_row, c, f"=SUM({col}{tong_row}:{col}{vat_row})")
        cell.font = _font(14, bold=True)
        cell.alignment = ALIGN_CC
        cell.fill = FILL_TOTAL
        cell.border = BORDER_ALL
        cell.number_format = NF_INT

    return tc_row


def _build_footer(ws, start_row: int) -> None:
    for i, line in enumerate(BANK_INFO):
        r = start_row + i
        ws.cell(r, 1, line)
        if i == 0:
            ws.cell(r, 1).font = _font(14, bold=True, italic=True)
        else:
            ws.cell(r, 1).font = _font(14)
        if i == len(BANK_INFO) - 1:
            ws.merge_cells(f"A{r}:D{r}")
            ws.cell(r, 1).alignment = ALIGN_LEFT_WRAP


# ---- Public entry point ----

def render_tc_cpn_workbook(
    customer: Dict[str, Any],
    services: List[Dict[str, Any]],
    jobs_map: Dict[Any, Dict[str, Any]],
    costs_by_svc: Dict[Any, List[Dict[str, Any]]],
    month: Optional[str],
    logo_path: Optional[str],
) -> Workbook:
    """Build the workbook for File 3 (Bảng kê TC + CPN)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "tc, cpn"

    title_month = _format_month_title(month)
    _build_header(ws, customer, logo_path, title_month)
    _build_table_header(ws)

    # Filter to ONLY TC/CPN services
    tc_cpn_services = []
    for svc in services:
        cost_rows = costs_by_svc.get(svc["svc_id"], [])
        if is_tc_or_cpn_service(svc, cost_rows):
            tc_cpn_services.append((svc, cost_rows))

    tc_cpn_services.sort(
        key=lambda pair: (pair[0].get("scheduled_date") or "", pair[0].get("svc_id") or 0)
    )

    rows = [
        _service_to_row(svc, jobs_map.get(svc["job_id"], {}), cost_rows, idx)
        for idx, (svc, cost_rows) in enumerate(tc_cpn_services, start=1)
    ]

    after_data = _build_data_rows(ws, rows, start_row=13)
    last_total_row = _build_totals(ws, data_start=13, data_end_exclusive=after_data)
    _build_footer(ws, last_total_row + 9)

    return wb
