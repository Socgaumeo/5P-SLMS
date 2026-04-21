"""
FAMILY C — Customs declaration Bảng Kê renderer (generic, multi-customer).

Covers customers whose bảng kê is a customs-declaration fee statement:
  GLOREX, THÁI HOÀ, GANG THÉP TN (customs variant), and future similar customers.

Layout (12 cols — aligns with the "XNK TC" reference sheet used by
GLOREX + THAIHOA):
  A  STT
  B  Ngày
  C  Tờ khai          — cd_no
  D  Luồng tờ khai    — customs channel (xanh/vàng/đỏ)
  E  Note             — DOM / tại chỗ / SEA / AIR
  F  Số hóa đơn/PXK   — invoice / delivery note
  G  Phí mở tờ khai   — phi_mtk bucket
  H  Phí kiểm hóa     — phi_kh bucket
  I  Phí phát sinh    — phi_psk bucket + other non-bucketed own fees
  J  Tổng             — G + H + I (formula)
  K  Thu chi hộ       — reimbursement total (sum selling of is_reimbursement)
  L  Số GNT / hóa đơn — cost_name hint containing "GNT" / invoice for reim

Multi-sheet split — following the convention documented in the DAINESE plan:
  cd_no prefix 108xxx → sheet "NHAP KHAU"   (nhập kinh doanh tại chỗ, A41 class)
  cd_no prefix 308xxx → sheet "XNK TAI CHO" (xuất kinh doanh tại chỗ, B13 class)
  other prefixes      → sheet "CUSTOMS"     (fallback bucket)

If only one bucket has data, a single sheet is emitted.

Priority per user: data correctness + ghi chú completeness > pixel-perfect layout.
"""

import importlib
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.api.exports.common_bang_ke_styling_and_formatting import (
    ALIGN_CC, ALIGN_CC_WRAP, ALIGN_LEFT, ALIGN_RIGHT,
    BORDER_ALL, FILL_DATA_ZEBRA, FILL_HEADER, FILL_TOTAL,
    NF_DATE_MDY, NF_INT,
    build_bang_chu_row,
    build_bank_footer,
    build_company_header_block,
    build_customer_block,
    build_title_row,
    font,
    format_month_title,
    parse_details,
    safe_float,
)
from app.api.exports.dainese_cost_name_to_column_mapper import (
    aggregate_costs_into_columns,
)

# Shared customs code whitelist (kebab-case filename → importlib)
_customs_validator = importlib.import_module(
    "app.core.vietnamese-customs-declaration-codes-and-validator"
)


# ---------- Layout / column widths ----------

COL_WIDTHS = {
    "A": 5.4,   # STT
    "B": 12.0,  # Ngày
    "C": 15.5,  # Tờ khai
    "D": 12.5,  # Luồng tờ khai
    "E": 10.0,  # Note
    "F": 14.0,  # Số hóa đơn/PXK
    "G": 14.0,  # Phí mở tờ khai
    "H": 13.0,  # Phí kiểm hóa
    "I": 14.0,  # Phí phát sinh khác
    "J": 14.0,  # Tổng
    "K": 14.0,  # Thu chi hộ
    "L": 16.0,  # Số GNT / hóa đơn
}

_HEADER_ROW = 13
_DATA_START_ROW = 14


# ---------- Sheet split by cd_no prefix ----------

def _split_services_by_cd_prefix(
    services: List[Dict[str, Any]],
) -> Dict[str, List[Dict[str, Any]]]:
    """
    Group services into sheet buckets based on cd_no prefix.
    Services without cd_no fall into "CUSTOMS" so nothing gets dropped.
    """
    buckets: Dict[str, List[Dict[str, Any]]] = {
        "NHAP KHAU": [],
        "XNK TAI CHO": [],
        "CUSTOMS": [],
    }
    for svc in services:
        cd = (svc.get("cd_no") or "").strip()
        prefix = cd[:3] if cd else ""
        if prefix == "108":
            buckets["NHAP KHAU"].append(svc)
        elif prefix == "308":
            buckets["XNK TAI CHO"].append(svc)
        else:
            buckets["CUSTOMS"].append(svc)
    return buckets


# ---------- Per-row data extraction ----------

def _extract_reimbursement_invoice(cost_rows: List[Dict[str, Any]]) -> str:
    """
    Pull a "Số GNT" / hóa đơn hint from reimbursement cost lines.
    Real data has cost_name like 'Thu hộ: Lệ phí hải quan - GNT: 0360356';
    we extract the tail (after '- GNT:' or ': ') so the customer sees
    the receipt number.
    """
    hints: List[str] = []
    for c in cost_rows:
        if not c.get("is_reimbursement"):
            continue
        name = c.get("cost_name") or ""
        # Look for common Vietnamese receipt-number markers
        for marker in ("GNT:", "GNT :", "HĐ:", "HD:", "Hóa đơn:", "BL:"):
            if marker in name:
                tail = name.split(marker, 1)[1].strip().rstrip(",;")
                if tail:
                    hints.append(tail)
                    break
    return ", ".join(hints[:3])  # cap to 3 to avoid overflow


_LUMP_SUM_NAMES = (
    "doanh thu cơ bản",
    "doanh thu co ban",
    "phí dịch vụ hải quan",
    "phi dich vu hai quan",
    "phí dịch vụ làm thủ tục hải quan",
    "phi dich vu lam thu tuc hai quan",
    "phí mở tờ khai",  # already direct match but include for safety
)


def _is_customs_lump_sum(name: Optional[str]) -> bool:
    """
    Detect a generic customs-service selling line (e.g. "Doanh thu cơ bản" synced
    from a quotation). Used to route the amount to "Phí mở tờ khai" instead of
    the fallback "Phí phát sinh khác" bucket, since for customs declarations
    the flat fee IS the mở tờ khai fee.
    """
    if not name:
        return False
    norm = name.lower().strip()
    return any(pat in norm for pat in _LUMP_SUM_NAMES)


def _service_to_row(
    svc: Dict[str, Any],
    job: Dict[str, Any],
    cost_rows: List[Dict[str, Any]],
    idx: int,
) -> Dict[str, Any]:
    """
    Map service + job + costs → column dict for the customs layout.

    Revenue resolution order (first non-zero wins):
      1. job_costs aggregated via aggregate_costs_into_columns (itemized — best)
         with a customs-specific override: generic lump-sum cost names
         ("Doanh thu cơ bản" etc.) are re-routed to Phí mở tờ khai.
      2. service_details.selling_price (fallback when costs not yet imported)
      3. service_details.total_revenue minus VAT
    """
    d = parse_details(svc)

    # Pre-filter: pull lump-sum customs-service rows OUT of cost_rows so the
    # generic mapper doesn't dump them into phi_psk. Their selling_amount goes
    # straight to phi_mtk below.
    lump_sum_own = 0.0
    non_lump_costs: List[Dict[str, Any]] = []
    for c in cost_rows:
        if not c.get("is_reimbursement") and _is_customs_lump_sum(c.get("cost_name")):
            lump_sum_own += safe_float(c.get("selling_amount"))
        else:
            non_lump_costs.append(c)

    buckets = aggregate_costs_into_columns(non_lump_costs)

    phi_mtk = buckets.get("phi_mtk", 0.0) + lump_sum_own
    phi_kh = buckets.get("phi_kh", 0.0)
    phi_psk = buckets.get("phi_psk", 0.0)
    # Other own-side fees that don't map to mtk/kh/psk get folded into "phát sinh"
    # so money doesn't vanish from the total.
    other_own = (
        buckets.get("phi_vc", 0.0)
        + buckets.get("phi_lh", 0.0)
        + buckets.get("phi_dnn", 0.0)
        + buckets.get("cuoc_qt", 0.0)
        + buckets.get("thc", 0.0)
        + buckets.get("cfs", 0.0)
        + buckets.get("do", 0.0)
        + buckets.get("dly", 0.0)
    )
    phi_psk_total = phi_psk + other_own
    reim_total = buckets.get("subtotal_reimburse", 0.0)

    # Fallback: no job_costs at all → try service_details
    own_sum = phi_mtk + phi_kh + phi_psk_total
    if own_sum == 0:
        fb = safe_float(d.get("selling_price"))
        if fb > 0:
            phi_mtk = fb
        else:
            gt = safe_float(d.get("total_revenue")) or safe_float(d.get("grand_total"))
            if gt > 0:
                vat_rate = safe_float(d.get("vat_rate")) or 0
                phi_mtk = gt / (1 + vat_rate / 100.0) if vat_rate else gt

    cd_no = svc.get("cd_no") or ""
    luong = d.get("customs_channel") or d.get("luong") or d.get("luong_to_khai") or ""
    # "Note" column — transport type hint (DOM / SEA / AIR / DHL / tại chỗ)
    # Prefer explicit note, fall back to service_type-derived label.
    note_text = (
        d.get("note")
        or d.get("notes")
        or d.get("transport_mode")
        or _note_from_service_type(svc.get("service_type_code"), cd_no)
    )
    invoice = svc.get("invoice_numbers") or d.get("invoice_number") or ""
    # Strip PostgreSQL array literal braces/quotes when invoice_numbers is a
    # TEXT[] column leaking through as e.g. `{00000022}` or `{"INV1","INV2"}`.
    if isinstance(invoice, str) and invoice.startswith("{") and invoice.endswith("}"):
        invoice = invoice[1:-1].replace('"', '').strip()
    reim_inv = _extract_reimbursement_invoice(cost_rows)

    return {
        "stt": idx,
        "ngay": svc.get("scheduled_date") or job.get("etd"),
        "cd_no": cd_no,
        "luong": luong,
        "note": note_text,
        "invoice": invoice,
        "phi_mtk": phi_mtk,
        "phi_kh": phi_kh,
        "phi_psk": phi_psk_total,
        "reim_total": reim_total,
        "reim_invoice": reim_inv,
    }


def _note_from_service_type(svc_code: Optional[str], cd_no: str) -> str:
    """Derive the Note column hint when user didn't fill one."""
    prefix = cd_no[:3] if cd_no else ""
    if prefix in ("108", "308"):
        return "DOM"
    code = (svc_code or "").upper()
    if code.startswith("SEA_"):
        return "SEA"
    if code.startswith("AIR_"):
        return "AIR"
    if code.startswith("BORDER"):
        return "BORDER"
    return ""


# ---------- Sheet builders ----------

def _build_table_header(ws, header_row: int = _HEADER_ROW) -> None:
    """Single-row table header: bold + header fill + border + wrap."""
    headers = {
        "A": "STT",
        "B": "Ngày",
        "C": "Tờ khai",
        "D": "Luồng tờ khai",
        "E": "Note",
        "F": "Số hóa đơn/PXK",
        "G": "Phí mở tờ khai",
        "H": "Phí kiểm hóa",
        "I": "Phí phát sinh khác",
        "J": "Tổng tiền",
        "K": "Thu chi hộ",
        "L": "Số hóa đơn/GNT",
    }
    for col, label in headers.items():
        cell = ws[f"{col}{header_row}"]
        cell.value = label
        cell.font = font(12, bold=True)
        cell.alignment = ALIGN_CC_WRAP
        cell.fill = FILL_HEADER
        cell.border = BORDER_ALL
    ws.row_dimensions[header_row].height = 32


def _build_data_rows(ws, rows: List[Dict[str, Any]], start_row: int = _DATA_START_ROW) -> int:
    """Write data rows. Returns index AFTER last data row."""
    common_font = font(11)
    for idx, r_data in enumerate(rows, start=1):
        r = start_row + idx - 1
        ws.row_dimensions[r].height = 22

        ws.cell(r, 1, idx)
        ws.cell(r, 2, r_data["ngay"])
        ws.cell(r, 3, r_data["cd_no"])
        ws.cell(r, 4, r_data["luong"])
        ws.cell(r, 5, r_data["note"])
        ws.cell(r, 6, r_data["invoice"])
        ws.cell(r, 7, r_data["phi_mtk"] or None)
        ws.cell(r, 8, r_data["phi_kh"] or None)
        ws.cell(r, 9, r_data["phi_psk"] or None)
        # J: Tổng = G + H + I
        ws.cell(r, 10, f"=IFERROR(G{r},0)+IFERROR(H{r},0)+IFERROR(I{r},0)")
        ws.cell(r, 11, r_data["reim_total"] or None)
        ws.cell(r, 12, r_data["reim_invoice"])

        for c in range(1, 13):
            cell = ws.cell(r, c)
            cell.font = common_font
            cell.alignment = ALIGN_CC_WRAP
            cell.border = BORDER_ALL
            cell.fill = FILL_DATA_ZEBRA

        ws.cell(r, 2).number_format = NF_DATE_MDY
        for col_idx in (7, 8, 9, 10, 11):
            ws.cell(r, col_idx).number_format = NF_INT
        # Left-align text columns for readability
        for col_idx in (3, 5, 6, 12):
            ws.cell(r, col_idx).alignment = ALIGN_LEFT

    return start_row + len(rows)


def _build_totals(
    ws,
    data_start: int,
    data_end_exclusive: int,
    vat_rate: float = 0.08,
) -> int:
    """Tổng / VAT / Tổng cộng over columns G-K (Phí mở TK through Thu chi hộ)."""
    last_data = data_end_exclusive - 1
    if last_data < data_start:
        return data_end_exclusive
    tong_r = data_end_exclusive
    vat_r = tong_r + 1
    tc_r = vat_r + 1

    def _label(r: int, text: str) -> None:
        ws.merge_cells(f"A{r}:F{r}")
        ws.cell(r, 1, text)
        ws.cell(r, 1).font = font(12, bold=True, italic=True)
        ws.cell(r, 1).alignment = ALIGN_CC
        for c in range(1, 13):
            ws.cell(r, c).fill = FILL_TOTAL
            ws.cell(r, c).border = BORDER_ALL

    def _sum_cell(r: int, col: int, formula: str) -> None:
        cell = ws.cell(r, col, formula)
        cell.font = font(12, bold=True)
        cell.alignment = ALIGN_RIGHT
        cell.fill = FILL_TOTAL
        cell.border = BORDER_ALL
        cell.number_format = NF_INT

    _label(tong_r, "Tổng")
    for col in (7, 8, 9, 10, 11):
        letter = get_column_letter(col)
        _sum_cell(tong_r, col, f"=SUM({letter}{data_start}:{letter}{last_data})")

    vat_pct = int(vat_rate * 100)
    _label(vat_r, f"VAT ({vat_pct}%)")
    # VAT on own-side total (column J = Tổng). Reimbursement (K) stays VAT-free.
    _sum_cell(vat_r, 10, f"=J{tong_r}*{vat_rate}")

    _label(tc_r, "Tổng cộng (sau VAT)")
    _sum_cell(tc_r, 10, f"=J{tong_r}+J{vat_r}")
    _sum_cell(tc_r, 11, f"=K{tong_r}")  # reim pass-through, no VAT

    return tc_r


def _build_one_sheet(
    ws,
    sheet_title: str,
    customer: Dict[str, Any],
    services: List[Dict[str, Any]],
    jobs_map: Dict[Any, Dict[str, Any]],
    costs_by_svc: Dict[Any, List[Dict[str, Any]]],
    month: Optional[str],
    logo_path: Optional[str],
    cfg: Dict[str, Any],
) -> None:
    """Populate one worksheet with headers + data + totals + footer."""
    ws.title = sheet_title[:31]
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    # 1. Company header
    build_company_header_block(ws, logo_path=logo_path)

    # 2. Title row
    title_month = format_month_title(month, style="vi_short")
    default_title = "BẢNG KÊ THU CHI HỘ LỆ PHÍ HẢI QUAN THÁNG {month}"
    title_template = cfg.get("title_template") or default_title
    title_text = title_template.format(month=title_month)
    build_title_row(ws, row=6, text=title_text, merge_range="A6:L6", size=14, height=28)

    # 3. Customer block
    build_customer_block(ws, customer, start_row=8, style="customer")

    # 4. Table header
    _build_table_header(ws, header_row=_HEADER_ROW)

    # 5. Data rows
    services_sorted = sorted(
        [(svc, costs_by_svc.get(svc["svc_id"], [])) for svc in services],
        key=lambda pair: (str(pair[0].get("scheduled_date") or ""), pair[0].get("svc_id") or 0),
    )
    rows = [
        _service_to_row(svc, jobs_map.get(svc["job_id"], {}), cost_rows, idx)
        for idx, (svc, cost_rows) in enumerate(services_sorted, start=1)
    ]
    after_data = _build_data_rows(ws, rows, start_row=_DATA_START_ROW)

    # 6. Totals
    vat_rate = float(cfg.get("vat_rate", 0.08))
    last_total = _build_totals(
        ws, data_start=_DATA_START_ROW, data_end_exclusive=after_data, vat_rate=vat_rate
    )

    # 7. Bằng chữ
    if cfg.get("include_bang_chu", True) and rows:
        grand_total_estimate = sum(
            (r["phi_mtk"] + r["phi_kh"] + r["phi_psk"]) * (1 + vat_rate)
            + r["reim_total"]
            for r in rows
        )
        build_bang_chu_row(ws, last_total + 2, grand_total_estimate)
        last_total += 2

    # 8. Bank footer
    if cfg.get("include_bank", True):
        build_bank_footer(ws, last_total + 3, merge_last_col="F")


# ---------- Public entry point ----------

def render_customs_workbook(
    customer: Dict[str, Any],
    services: List[Dict[str, Any]],
    jobs_map: Dict[Any, Dict[str, Any]],
    costs_by_svc: Dict[Any, List[Dict[str, Any]]],
    month: Optional[str],
    logo_path: Optional[str] = None,
    config: Optional[Dict[str, Any]] = None,
) -> Workbook:
    """
    Build customs-declaration workbook for any customer.

    Produces 1..3 sheets, one per cd_no-prefix bucket that actually has data:
      NHAP KHAU (108xxx), XNK TAI CHO (308xxx), CUSTOMS (fallback).

    config (all optional):
      title_template    → override title
      vat_rate          → default 0.08
      include_bang_chu  → default True
      include_bank      → default True
      force_single_sheet → default False. If True, merge all into one "CUSTOMS" sheet.
    """
    cfg = config or {}
    wb = Workbook()
    # Remove the default sheet that openpyxl creates — we add our own below.
    default = wb.active
    wb.remove(default)

    if cfg.get("force_single_sheet"):
        buckets = {"CUSTOMS": services}
    else:
        buckets = _split_services_by_cd_prefix(services)

    # Keep deterministic sheet order matching the reference OneDrive files.
    sheet_order = ("NHAP KHAU", "XNK TAI CHO", "CUSTOMS")
    any_written = False
    for sheet_name in sheet_order:
        svc_list = buckets.get(sheet_name, [])
        if not svc_list:
            continue
        ws = wb.create_sheet(title=sheet_name)
        _build_one_sheet(
            ws, sheet_name, customer, svc_list, jobs_map, costs_by_svc,
            month, logo_path, cfg,
        )
        any_written = True

    # If nothing matched (empty dataset), still emit a blank sheet so the
    # download works and the user sees headers.
    if not any_written:
        ws = wb.create_sheet(title="CUSTOMS")
        _build_one_sheet(
            ws, "CUSTOMS", customer, [], jobs_map, costs_by_svc,
            month, logo_path, cfg,
        )

    return wb
