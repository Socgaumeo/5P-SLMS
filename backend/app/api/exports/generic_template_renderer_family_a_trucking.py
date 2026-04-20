"""
FAMILY A — Trucking nội địa Bảng Kê renderer (generic, multi-customer).

Covers customers with the trucking-debit pattern:
  BÌNH MINH, K+K (DONSUNG/KK), PIPETREE, UPGAIN, UTRACORN, TVC, …

Layout based on cluster C1 of customer-bang-ke scan:
  STT | Ngày | Type | Điểm lấy | Điểm trả | BKS | Số lượng | Đơn vị |
  Đơn giá | Thành tiền | Phụ phí xăng dầu | Tổng | JOB | Ghi chú

Per-customer differences are passed via `config` dict so this single module
serves every trucking customer. Add new customer → 1 line in the registry.

Priority per user: full fee data + ghi chú correctness > pixel-perfect layout.
"""

import re
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.api.exports.common_bang_ke_styling_and_formatting import (
    ALIGN_CC, ALIGN_CC_WRAP, ALIGN_LEFT, ALIGN_RIGHT,
    BORDER_ALL, FILL_DATA_ZEBRA, FILL_HEADER, FILL_TOTAL,
    NF_DATE_MDY, NF_INT,
    apply_border_to_range,
    build_bang_chu_row,
    build_bank_footer,
    build_company_header_block,
    build_customer_block,
    build_title_row,
    font,
    format_month_title,
    parse_details,
    safe_float,
    split_address,
)


# ---------- Cost-name to bucket mapping (trucking) ----------

_RE_PHAT_SINH = re.compile(r"phát\s*sinh|phat\s*sinh|extra", re.IGNORECASE)
_RE_XANG_DAU = re.compile(r"phụ\s*phí\s*xăng|phu\s*phi\s*xang|fuel|xăng\s*dầu|xang\s*dau", re.IGNORECASE)
_RE_BOC_XEP = re.compile(r"bốc\s*xếp|boc\s*xep|loading|hạ\s*hàng|ha\s*hang|nâng|nang", re.IGNORECASE)
_RE_CUOC = re.compile(r"cước|cuoc|trucking|transport|vận\s*chuyển|van\s*chuyen", re.IGNORECASE)


def _bucket_trucking_costs(cost_rows: List[Dict[str, Any]]) -> Dict[str, float]:
    """
    Sum job_costs into 4 buckets for the trucking layout:
      cuoc_vc       → cột Đơn giá / Thành tiền
      xang_dau      → cột Phụ phí xăng dầu
      phat_sinh     → cột Chi phí khác / Phát sinh
      boc_xep_other → cộng vào "Phụ phí" or "Chi phí khác"
    """
    out = {"cuoc_vc": 0.0, "xang_dau": 0.0, "phat_sinh": 0.0, "boc_xep": 0.0}
    for c in cost_rows:
        if c.get("is_reimbursement"):
            continue
        name = c.get("cost_name") or ""
        amt = safe_float(c.get("selling_amount"))
        if _RE_XANG_DAU.search(name):
            out["xang_dau"] += amt
        elif _RE_PHAT_SINH.search(name):
            out["phat_sinh"] += amt
        elif _RE_BOC_XEP.search(name):
            out["boc_xep"] += amt
        elif _RE_CUOC.search(name):
            out["cuoc_vc"] += amt
        else:
            # Fallback: any unmatched non-reim cost → cước vận chuyển
            out["cuoc_vc"] += amt
    return out


# ---------- Layout / column widths ----------

# Reference column widths (from BÌNH MINH/UPGAIN/etc.)
COL_WIDTHS = {
    "A": 5.2,   # STT
    "B": 12.0,  # Ngày
    "C": 14.0,  # Type
    "D": 18.0,  # Điểm lấy
    "E": 18.0,  # Điểm trả
    "F": 13.0,  # BKS
    "G": 8.5,   # SL
    "H": 8.5,   # Đơn vị
    "I": 12.0,  # Đơn giá
    "J": 14.0,  # Thành tiền
    "K": 14.0,  # Phụ phí xăng dầu
    "L": 14.0,  # Tổng
    "M": 15.0,  # JOB
    "N": 22.0,  # Ghi chú
}


# ---------- Builders ----------

def _build_table_header(ws, header_row: int = 13) -> None:
    """Single-row table header. Bold + light-blue fill + border + wrap."""
    headers = {
        "A": "STT",
        "B": "Ngày dịch vụ",
        "C": "Type",
        "D": "Điểm lấy hàng",
        "E": "Điểm trả hàng",
        "F": "Biển số",
        "G": "Số lượng",
        "H": "Đơn vị",
        "I": "Đơn giá",
        "J": "Thành tiền",
        "K": "Phụ phí xăng dầu",
        "L": "Tổng",
        "M": "JOB",
        "N": "Ghi chú",
    }
    for col, label in headers.items():
        cell = ws[f"{col}{header_row}"]
        cell.value = label
        cell.font = font(12, bold=True)
        cell.alignment = ALIGN_CC_WRAP
        cell.fill = FILL_HEADER
        cell.border = BORDER_ALL
    ws.row_dimensions[header_row].height = 32


def _service_to_row(svc: Dict[str, Any], job: Dict[str, Any], cost_rows: List[Dict], idx: int) -> Dict[str, Any]:
    """
    Map one service+job+cost_rows to renderer column dict.

    Revenue resolution order (first non-zero wins):
      1. job_costs aggregation (itemized — best quality)
      2. service_details.unit_price / selling_price (single-line fallback)
      3. service_details.grand_total minus vat (computed fallback)

    This lets us export ANY customer even when costs haven't been imported
    into job_costs — as long as at least service_details has pricing.
    """
    d = parse_details(svc)

    buckets = _bucket_trucking_costs(cost_rows)

    # Fallback #1: service_details.unit_price / selling_price when cost rows empty
    if buckets["cuoc_vc"] == 0 and buckets["xang_dau"] == 0 and buckets["phat_sinh"] == 0:
        unit_price_fb = safe_float(d.get("unit_price")) or safe_float(d.get("selling_price"))
        if unit_price_fb > 0:
            buckets["cuoc_vc"] = unit_price_fb
        else:
            # Fallback #2: grand_total (before VAT) — back-compute from vat
            gt = safe_float(d.get("total_revenue")) or safe_float(d.get("grand_total"))
            if gt > 0:
                vat_rate = safe_float(d.get("vat_rate")) or 0
                # If grand_total includes VAT, strip it out
                buckets["cuoc_vc"] = gt / (1 + vat_rate / 100.0) if vat_rate else gt

    qty = safe_float(d.get("quantity")) or 1
    unit = d.get("vehicle_type") or "Chuyến"
    bks = d.get("vehicle_plate") or ""
    note = d.get("note") or d.get("notes") or ""

    pickup_full = svc.get("origin_address") or ""
    drop_full = svc.get("dest_address") or ""

    type_str = (svc.get("service_type_code") or "").replace("TRUCKING_", "")

    return {
        "stt": idx,
        "ngay": svc.get("scheduled_date") or job.get("etd"),
        "type": type_str or "DOM",
        "pickup": pickup_full,
        "drop": drop_full,
        "bks": bks,
        "qty": int(qty) if qty == int(qty) else qty,
        "unit": unit,
        "don_gia": buckets["cuoc_vc"],
        "phat_sinh": buckets["phat_sinh"] + buckets["boc_xep"],
        "xang_dau": buckets["xang_dau"],
        "job_no": job.get("job_no", ""),
        "note": note,
    }


def _build_data_rows(ws, rows: List[Dict[str, Any]], start_row: int = 14) -> int:
    """Write data rows. Returns row index AFTER last data row."""
    common_font = font(11)
    for idx, r_data in enumerate(rows, start=1):
        r = start_row + idx - 1
        ws.row_dimensions[r].height = 22

        ws.cell(r, 1, idx)
        ws.cell(r, 2, r_data["ngay"])
        ws.cell(r, 3, r_data["type"])
        ws.cell(r, 4, r_data["pickup"])
        ws.cell(r, 5, r_data["drop"])
        ws.cell(r, 6, r_data["bks"])
        ws.cell(r, 7, r_data["qty"])
        ws.cell(r, 8, r_data["unit"])
        ws.cell(r, 9, r_data["don_gia"] or 0)
        # J (10): Thành tiền = Đơn giá × Số lượng
        ws.cell(r, 10, f"=I{r}*G{r}")
        ws.cell(r, 11, r_data["xang_dau"] or None)
        # L (12): Tổng = Thành tiền + Phụ phí xăng dầu + Phát sinh
        ws.cell(r, 12, f"=IFERROR(J{r},0)+IFERROR(K{r},0)+{r_data['phat_sinh'] or 0}")
        ws.cell(r, 13, r_data["job_no"])
        ws.cell(r, 14, r_data["note"])

        for c in range(1, 15):
            cell = ws.cell(r, c)
            cell.font = common_font
            cell.alignment = ALIGN_CC_WRAP
            cell.border = BORDER_ALL
            cell.fill = FILL_DATA_ZEBRA

        ws.cell(r, 2).number_format = NF_DATE_MDY
        ws.cell(r, 7).number_format = "0"
        for col_idx in (9, 10, 11, 12):
            ws.cell(r, col_idx).number_format = NF_INT

    return start_row + len(rows)


def _build_totals(ws, data_start: int, data_end_exclusive: int, vat_rate: float = 0.08) -> int:
    """Tổng / VAT / Tổng cộng (sum L column only)."""
    last_data = data_end_exclusive - 1
    tong_r = data_end_exclusive
    vat_r = tong_r + 1
    tc_r = vat_r + 1

    def _label(r: int, text: str) -> None:
        ws.merge_cells(f"A{r}:K{r}")
        ws.cell(r, 1, text)
        ws.cell(r, 1).font = font(12, bold=True, italic=True)
        ws.cell(r, 1).alignment = ALIGN_CC
        for c in range(1, 12):
            ws.cell(r, c).fill = FILL_TOTAL
            ws.cell(r, c).border = BORDER_ALL

    _label(tong_r, "Tổng")
    cell = ws.cell(tong_r, 12, f"=SUM(L{data_start}:L{last_data})")
    cell.font = font(12, bold=True)
    cell.alignment = ALIGN_RIGHT
    cell.fill = FILL_TOTAL
    cell.border = BORDER_ALL
    cell.number_format = NF_INT

    vat_pct = int(vat_rate * 100)
    _label(vat_r, f"VAT ({vat_pct}%)")
    cell = ws.cell(vat_r, 12, f"=L{tong_r}*{vat_rate}")
    cell.font = font(12, bold=True)
    cell.alignment = ALIGN_RIGHT
    cell.fill = FILL_TOTAL
    cell.border = BORDER_ALL
    cell.number_format = NF_INT

    _label(tc_r, "Tổng cộng (sau VAT)")
    cell = ws.cell(tc_r, 12, f"=L{tong_r}+L{vat_r}")
    cell.font = font(12, bold=True)
    cell.alignment = ALIGN_RIGHT
    cell.fill = FILL_TOTAL
    cell.border = BORDER_ALL
    cell.number_format = NF_INT

    return tc_r


# ---------- Public entry point ----------

def render_trucking_workbook(
    customer: Dict[str, Any],
    services: List[Dict[str, Any]],
    jobs_map: Dict[Any, Dict[str, Any]],
    costs_by_svc: Dict[Any, List[Dict[str, Any]]],
    month: Optional[str],
    logo_path: Optional[str] = None,
    config: Optional[Dict[str, Any]] = None,
) -> Workbook:
    """
    Build trucking-debit workbook for any customer.

    config (all optional, with sensible defaults):
      sheet_name        → str, default = customer short_name or 'TRUCKING'
      title_template    → str with '{month}' placeholder
      vat_rate          → float, default 0.08
      include_bang_chu  → bool, default True
      include_bank      → bool, default True
      include_signature → bool, default False
    """
    cfg = config or {}
    wb = Workbook()
    ws = wb.active
    ws.title = (cfg.get("sheet_name") or customer.get("short_name") or "TRUCKING")[:31]

    # Column widths
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    # 1. Company header
    build_company_header_block(ws, logo_path=logo_path)

    # 2. Title row 6 (merged A6:N6)
    title_month = format_month_title(month, style="vi_short")
    title_template = cfg.get("title_template") or "BẢNG KÊ DỊCH VỤ VẬN CHUYỂN ĐƯỜNG BỘ THÁNG {month}"
    title_text = title_template.format(month=title_month)
    build_title_row(ws, row=6, text=title_text, merge_range="A6:N6", size=14, height=28)

    # 3. Customer block rows 8-10
    build_customer_block(ws, customer, start_row=8, style="customer")

    # 4. Table header row 13
    _build_table_header(ws, header_row=13)

    # 5. Data rows
    services_with_costs = sorted(
        [(svc, costs_by_svc.get(svc["svc_id"], [])) for svc in services],
        # str() coerces mixed datetime.date / str values to comparable strings
        key=lambda pair: (str(pair[0].get("scheduled_date") or ""), pair[0].get("svc_id") or 0),
    )
    rows = [
        _service_to_row(svc, jobs_map.get(svc["job_id"], {}), cost_rows, idx)
        for idx, (svc, cost_rows) in enumerate(services_with_costs, start=1)
    ]
    after_data = _build_data_rows(ws, rows, start_row=14)

    # 6. Totals
    vat_rate = float(cfg.get("vat_rate", 0.08))
    last_total = _build_totals(ws, data_start=14, data_end_exclusive=after_data, vat_rate=vat_rate)

    # 7. Bằng chữ (compute grand total approximate)
    if cfg.get("include_bang_chu", True):
        grand_total_estimate = sum(
            (r["don_gia"] * (r["qty"] if isinstance(r["qty"], (int, float)) else 1))
            + r["xang_dau"] + r["phat_sinh"]
            for r in rows
        )
        grand_total_estimate *= (1 + vat_rate)
        build_bang_chu_row(ws, last_total + 2, grand_total_estimate)
        last_total += 2

    # 8. Bank footer
    if cfg.get("include_bank", True):
        build_bank_footer(ws, last_total + 3, merge_last_col="F")

    return wb
