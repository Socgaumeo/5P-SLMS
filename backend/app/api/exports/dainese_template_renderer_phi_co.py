"""
Renderer for DAINESE 'Bảng kê dịch vụ xin CO' (Certificate of Origin).

Reproduces layout/styles/formulas of the customer reference file:
  plans/reports/dainese-templates/file-02-bang-ke-phi-co-summary.md

Filter logic: only services with `service_details.co_no` set OR job_costs
having a "Phí C/O ..." row are included (NOT all CUS_CO services, since
that code is also used for KNQ/DHL customs declarations).

Pure rendering: takes already-fetched data + customer info, returns a Workbook.
"""

import json
import re
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.drawing.image import Image as XLImage


# ---- Style constants ----

FONT_NAME = "Times New Roman"

THIN = Side(style="thin", color="000000")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_CC_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_CC = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

NF_INT = "#,##0"
NF_DATE = "[$-409]d-mmm-yy;@"
NF_ACCT = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'

FILL_HEADER = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
FILL_TOTAL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

COMPANY_INFO = {
    "name": "CÔNG TY TNHH THƯƠNG MẠI VÀ DỊCH VỤ 5P VIỆT NAM",
    "address": "Số nhà 02 Ngõ 1H Phố Trần Quang Diệu, Phường Đống Đa, Thành phố Hà Nội, Việt Nam",
    "mst": "MST: 0110523309",
}

BANK_INFO = [
    "Thông tin chuyển khoản:",
    "Tài khoản: Công Ty TNHH Thương mại và dịch vụ 5P Việt Nam",
    "Số tài khoản: 346886",
    "Tại Ngân hàng: Ngân hàng TMCP Kỹ thương Việt Nam - Techcombank Chi nhánh Đông Đô",
]

# Column widths for sheet "DỊCH VỤ" (matches reference file exactly)
COL_WIDTHS = {
    "A": 5.4,  "B": 13.8, "C": 23.7, "D": 18.6, "E": 9.3,  "F": 12.5, "G": 12.4,
    "H": 13.9, "I": 13.5, "J": 11.7, "K": 11.5, "L": 9.1,  "M": 13.9,
}


# ---- Number-to-Vietnamese-words (for "Tổng số tiền bằng chữ") ----

_NUM_VI = ["không", "một", "hai", "ba", "bốn", "năm", "sáu", "bảy", "tám", "chín"]


def _vn_money_words(n: int) -> str:
    """Cheap-and-cheerful Vietnamese amount-in-words. Adequate for invoices up to ~999B."""
    if n == 0:
        return "Không đồng"
    units = [("", 1), ("nghìn", 1000), ("triệu", 1_000_000), ("tỷ", 1_000_000_000)]
    parts = []
    n = int(n)
    while n > 0 and len(parts) < 4:
        chunk = n % 1000
        n //= 1000
        if chunk:
            txt = []
            h, t, u = chunk // 100, (chunk // 10) % 10, chunk % 10
            if h: txt.append(f"{_NUM_VI[h]} trăm")
            if t > 1: txt.append(f"{_NUM_VI[t]} mươi")
            elif t == 1: txt.append("mười")
            elif h and u: txt.append("lẻ")
            if u:
                if t > 1 and u == 5: txt.append("lăm")
                elif t > 0 and u == 1: txt.append("mốt")
                else: txt.append(_NUM_VI[u])
            unit = units[len(parts)][0]
            parts.append(" ".join(txt) + (f" {unit}" if unit else ""))
        else:
            parts.append("")
    parts = [p for p in reversed(parts) if p.strip()]
    s = " ".join(parts).strip().replace("  ", " ")
    return s[:1].upper() + s[1:] + " đồng chẵn./."


# ---- Helpers ----

def _font(size: int = 11, bold: bool = False, italic: bool = False) -> Font:
    return Font(name=FONT_NAME, size=size, bold=bold, italic=italic)


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _parse_details(svc: Dict[str, Any]) -> Dict[str, Any]:
    raw = svc.get("service_details") or {}
    if isinstance(raw, str):
        try:
            return json.loads(raw)
        except Exception:
            return {}
    return raw


def _format_month_title(month: Optional[str]) -> str:
    if month and len(month) == 7 and month[4] == "-":
        year, mon = month.split("-")
        return f"{mon}.{year}"
    now = datetime.now()
    return f"{now.month:02d}.{now.year}"


def _apply_border_to_range(ws, rng: str, fill: Optional[PatternFill] = None) -> None:
    c1, r1, c2, r2 = range_boundaries(rng)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_ALL
            if fill is not None:
                cell.fill = fill


# ---- CO-specific data extraction ----

_BL_RE = re.compile(r"BL[:\s]*([A-Z0-9\-]+)", re.IGNORECASE)


def is_real_co_service(svc: Dict[str, Any], cost_rows: List[Dict[str, Any]]) -> bool:
    """A service belongs in Bảng kê CO if it has a co_no OR a 'Phí C/O' cost row."""
    d = _parse_details(svc)
    if d.get("co_no"):
        return True
    for c in cost_rows:
        name = (c.get("cost_name") or "").lower()
        if "phí c/o" in name or "phi co" in name or "phí co " in name:
            return True
    return False


def extract_reimbursement(cost_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Pull the first reimbursement entry for the K/L/M columns.
    Cost name format: 'Thu hộ: <NỘI DUNG> - BL: <SỐ BIÊN LAI>'
    """
    for c in cost_rows:
        if not c.get("is_reimbursement"):
            continue
        name = c.get("cost_name") or ""
        # Strip 'Thu hộ:' prefix and BL suffix
        content = re.sub(r"^thu\s*hộ\s*:\s*", "", name, flags=re.IGNORECASE)
        bl_match = _BL_RE.search(content)
        bl_no = bl_match.group(1) if bl_match else ""
        content = _BL_RE.sub("", content).strip(" -")
        return {
            "noi_dung": content,
            "so_tien": _safe_float(c.get("selling_amount")),
            "bien_lai": bl_no,
        }
    return {"noi_dung": "", "so_tien": 0.0, "bien_lai": ""}


def extract_co_fee(cost_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Pull the 'Phí C/O' fee row (qty + unit price)."""
    for c in cost_rows:
        if c.get("is_reimbursement"):
            continue
        name = (c.get("cost_name") or "").lower()
        if "phí c/o" in name or "phi co" in name:
            qty = float(c.get("quantity") or 1)
            sell = _safe_float(c.get("selling_amount"))
            unit_price = sell / qty if qty else sell
            return {"qty": qty, "unit_price": unit_price, "total": sell}
    # Fallback: sum of all non-reim
    total = sum(_safe_float(c.get("selling_amount")) for c in cost_rows if not c.get("is_reimbursement"))
    return {"qty": 1, "unit_price": total, "total": total}


# ---- Section builders ----

def _build_header(ws, customer: Dict[str, Any], logo_path: Optional[str], title_month: str) -> None:
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w
    for r in (1, 2, 3):
        ws.row_dimensions[r].height = 18
    ws.row_dimensions[6].height = 22
    ws.row_dimensions[12].height = 16
    ws.row_dimensions[13].height = 28
    ws.row_dimensions[14].height = 22

    if logo_path:
        try:
            img = XLImage(logo_path)
            img.width = 80
            img.height = 80
            ws.add_image(img, "A1")
        except Exception:
            pass

    ws["C1"] = COMPANY_INFO["name"]
    ws["C2"] = COMPANY_INFO["address"]
    ws["C3"] = COMPANY_INFO["mst"]
    for coord in ("C1", "C2", "C3"):
        ws[coord].font = _font(11, bold=True)
        ws[coord].alignment = ALIGN_LEFT

    # Title row 6
    ws.merge_cells("A6:M6")
    ws["A6"] = f"BẢNG KÊ DỊCH VỤ XIN CO THÁNG {title_month}"
    ws["A6"].font = _font(18, bold=True)
    ws["A6"].alignment = ALIGN_CC

    # Customer block rows 9-11
    cust_name = customer.get("company_name") or customer.get("short_name") or ""
    cust_addr = customer.get("address") or ""
    cust_mst = customer.get("tax_code") or ""

    ws["A9"] = f"Họ tên người mua hàng (Customer):  {cust_name}"
    ws["A10"] = f"Địa chỉ (Address): {cust_addr}"
    ws["A11"] = f"Mã số thuế (Tax- code): {cust_mst}"
    for coord in ("A9", "A10", "A11"):
        ws[coord].font = _font(11)
        ws[coord].alignment = Alignment(horizontal="left", vertical="center")


def _build_table_header(ws) -> None:
    """Row 13 main headers + row 14 sub-headers for K/L/M (THU CHI HỘ)."""
    # K13:M13 merged for "THU CHI HỘ"
    ws.merge_cells("K13:M13")
    ws["K13"] = "THU CHI HỘ"

    headers_main = {
        "A13": ("STT", True),
        "B13": ("NGÀY C/O", True),
        "C13": ("INVOICE", False),
        "D13": ("SỐ CO", False),
        "E13": ("FORM", False),
        "F13": ("SỐ LƯỢNG", True),
        "G13": ("ĐƠN GIÁ (VND)", True),
        "H13": ("THÀNH TIỀN", False),
        "I13": ("SỐ HÓA ĐƠN", True),
        "J13": ("GHI CHÚ", False),
        "K13": ("THU CHI HỘ", True),
    }
    for coord, (val, _wrap) in headers_main.items():
        ws[coord] = val
        ws[coord].font = _font(11, bold=True)
        ws[coord].alignment = ALIGN_CC_WRAP
        ws[coord].fill = FILL_HEADER
        ws[coord].border = BORDER_ALL

    # Apply header fill+border to L13/M13 (covered by merge)
    for coord in ("L13", "M13"):
        ws[coord].fill = FILL_HEADER
        ws[coord].border = BORDER_ALL

    # Row 14 sub-headers (only K-M)
    for coord, val in (("K14", "NỘI DUNG"), ("L14", "SỐ TIỀN"), ("M14", "SỐ BIÊN LAI")):
        ws[coord] = val
        ws[coord].font = _font(11, bold=True)
        ws[coord].alignment = ALIGN_CC_WRAP
        ws[coord].fill = FILL_HEADER
        ws[coord].border = BORDER_ALL

    # A14:J14 also need borders/fill (for visual continuity of merged headers row 13)
    for col in range(1, 11):
        cell = ws.cell(14, col)
        cell.fill = FILL_HEADER
        cell.border = BORDER_ALL


def _build_data_rows(ws, rows: List[Dict[str, Any]], start_row: int = 15) -> int:
    common_font = _font(11)
    for idx, r_data in enumerate(rows, start=1):
        r = start_row + idx - 1
        ws.row_dimensions[r].height = 21

        ws.cell(r, 1, idx)
        ws.cell(r, 2, r_data.get("ngay_co"))
        ws.cell(r, 3, r_data.get("invoice"))
        ws.cell(r, 4, r_data.get("so_co"))
        ws.cell(r, 5, r_data.get("form") or "EUR1")
        ws.cell(r, 6, r_data.get("so_luong") or 1)
        ws.cell(r, 7, r_data.get("don_gia") or 0)
        ws.cell(r, 8, f"=F{r}*G{r}")
        ws.cell(r, 9, r_data.get("so_hoa_don") or "")
        ws.cell(r, 10, r_data.get("ghi_chu") or "")
        ws.cell(r, 11, r_data.get("noi_dung") or "")
        ws.cell(r, 12, r_data.get("so_tien_th") or 0 or None)
        ws.cell(r, 13, r_data.get("bien_lai") or "")

        for c in range(1, 14):
            cell = ws.cell(r, c)
            cell.font = common_font
            cell.border = BORDER_ALL
            cell.alignment = ALIGN_CC

        # Date format
        ws.cell(r, 2).number_format = NF_DATE
        # Number formats
        ws.cell(r, 6).number_format = NF_INT
        ws.cell(r, 7).number_format = NF_ACCT
        ws.cell(r, 8).number_format = NF_ACCT
        ws.cell(r, 12).number_format = NF_ACCT

    return start_row + len(rows)


def _build_totals(ws, data_start: int, data_end_exclusive: int, grand_total: float) -> int:
    """Build TỔNG TRƯỚC VAT / VAT (0%) / TỔNG SAU VAT / TỔNG THANH TOÁN. Returns last row."""
    last_data = data_end_exclusive - 1
    r17 = data_end_exclusive
    r18 = r17 + 1
    r19 = r18 + 1
    r20 = r19 + 1
    r21 = r20 + 1

    def _label_row(r: int, text: str, merge_to: str = "G") -> None:
        rng = f"A{r}:{merge_to}{r}"
        ws.merge_cells(rng)
        ws.cell(r, 1, text)
        ws.cell(r, 1).font = _font(11, bold=True)
        ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")
        for c in range(1, ord(merge_to) - ord("A") + 2):
            ws.cell(r, c).border = BORDER_ALL

    _label_row(r17, "TỔNG TRƯỚC VAT", "G")
    ws.cell(r17, 8, f"=SUM(H{data_start}:H{last_data})")
    ws.cell(r17, 8).font = _font(11, bold=True)
    ws.cell(r17, 8).alignment = ALIGN_CC
    ws.cell(r17, 8).number_format = NF_ACCT
    ws.cell(r17, 8).border = BORDER_ALL

    _label_row(r18, "VAT (0%)", "G")
    ws.cell(r18, 8, 0)
    ws.cell(r18, 8).font = _font(11, bold=True)
    ws.cell(r18, 8).alignment = ALIGN_CC
    ws.cell(r18, 8).number_format = NF_ACCT
    ws.cell(r18, 8).border = BORDER_ALL

    _label_row(r19, "TỔNG SAU VAT", "G")
    ws.cell(r19, 8, f"=H{r17}+H{r18}")
    ws.cell(r19, 8).font = _font(11, bold=True)
    ws.cell(r19, 8).alignment = ALIGN_CC
    ws.cell(r19, 8).number_format = NF_ACCT
    ws.cell(r19, 8).border = BORDER_ALL
    # L19 = SUM of reimbursements
    ws.cell(r19, 12, f"=SUM(L{data_start}:L{last_data})")
    ws.cell(r19, 12).font = _font(11, bold=True)
    ws.cell(r19, 12).alignment = ALIGN_CC
    ws.cell(r19, 12).number_format = NF_ACCT
    ws.cell(r19, 12).border = BORDER_ALL

    # TỔNG THANH TOÁN row: merge A:H label, I:M formula
    ws.merge_cells(f"A{r20}:H{r20}")
    ws.cell(r20, 1, "TỔNG THANH TOÁN")
    ws.cell(r20, 1).font = _font(11, bold=True)
    ws.cell(r20, 1).alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 9):
        ws.cell(r20, c).border = BORDER_ALL
    ws.merge_cells(f"I{r20}:M{r20}")
    ws.cell(r20, 9, f"=H{r19}+L{r19}")
    ws.cell(r20, 9).font = _font(11, bold=True)
    ws.cell(r20, 9).alignment = ALIGN_CC
    ws.cell(r20, 9).number_format = NF_ACCT
    for c in range(9, 14):
        ws.cell(r20, c).border = BORDER_ALL

    # Tổng tiền bằng chữ
    ws.cell(r21, 1, f"Tổng số tiền bằng chữ: {_vn_money_words(int(grand_total))}")
    ws.cell(r21, 1).font = _font(12, bold=True, italic=True)

    return r21


def _build_signature_and_footer(ws, start_row: int) -> None:
    # Signature blocks at start_row+2
    sig_row = start_row + 2
    cust_label = ws.cell(sig_row, 1, "CÔNG TY TNHH DAINESE VIỆT NAM")
    cust_label.font = _font(12, bold=True)
    own_label = ws.cell(sig_row, 8, COMPANY_INFO["name"])
    own_label.font = _font(12, bold=True)
    own_label.alignment = ALIGN_LEFT

    # Bank info: 4 lines starting at sig_row + 9
    bank_start = sig_row + 9
    ws.cell(bank_start, 1, BANK_INFO[0]).font = _font(12, bold=True, italic=True)
    for i, line in enumerate(BANK_INFO[1:], start=1):
        cell = ws.cell(bank_start + i, 1, line)
        cell.font = _font(12)
        if i == len(BANK_INFO) - 1:
            ws.merge_cells(f"A{bank_start + i}:D{bank_start + i}")
            cell.alignment = ALIGN_LEFT_WRAP


# ---- Public entry point ----

def render_phi_co_workbook(
    customer: Dict[str, Any],
    services: List[Dict[str, Any]],
    jobs_map: Dict[Any, Dict[str, Any]],
    costs_by_svc: Dict[Any, List[Dict[str, Any]]],
    month: Optional[str],
    logo_path: Optional[str],
) -> Workbook:
    """Build the workbook for File 2 (Bảng kê phí CO)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "DỊCH VỤ"

    title_month = _format_month_title(month)
    _build_header(ws, customer, logo_path, title_month)
    _build_table_header(ws)

    # Filter to ONLY real CO services
    real_co_services = []
    for svc in services:
        cost_rows = costs_by_svc.get(svc["svc_id"], [])
        if is_real_co_service(svc, cost_rows):
            real_co_services.append((svc, cost_rows))

    real_co_services.sort(
        key=lambda pair: (pair[0].get("scheduled_date") or "", pair[0].get("svc_id") or 0)
    )

    rows: List[Dict[str, Any]] = []
    grand_total_for_words = 0.0
    for svc, cost_rows in real_co_services:
        d = _parse_details(svc)
        co_fee = extract_co_fee(cost_rows)
        reim = extract_reimbursement(cost_rows)
        job = jobs_map.get(svc["job_id"], {})

        rows.append({
            "ngay_co": svc.get("scheduled_date") or job.get("etd"),
            "invoice": svc.get("invoice_numbers") or d.get("invoice_numbers") or "",
            "so_co": d.get("co_no") or svc.get("co_no") or "",
            "form": d.get("form") or "EUR1",
            "so_luong": co_fee["qty"],
            "don_gia": co_fee["unit_price"],
            "so_hoa_don": job.get("invoice_number") or "",
            "ghi_chu": d.get("note") or "",
            "noi_dung": reim["noi_dung"],
            "so_tien_th": reim["so_tien"] if reim["so_tien"] else None,
            "bien_lai": reim["bien_lai"],
        })
        grand_total_for_words += co_fee["total"] + reim["so_tien"]

    if not rows:
        # Empty bảng kê — still render headers and zero totals
        rows = []

    after_data = _build_data_rows(ws, rows, start_row=15)
    last_total_row = _build_totals(ws, data_start=15, data_end_exclusive=after_data, grand_total=grand_total_for_words)
    _build_signature_and_footer(ws, last_total_row)

    return wb
