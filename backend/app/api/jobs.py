"""
Jobs API - Job management endpoints
"""

from fastapi import APIRouter, HTTPException, Request, Query
from pydantic import BaseModel
from typing import Optional, Dict, Any
from datetime import date, time, datetime
import logging
import traceback
import json

from app.services.data_service import get_data_service
from app.ai.utils.smart_parser import parse_date, parse_time
from app.ai.utils.service_type_detector import normalize_service_code
from app.services.message_service import get_message_service
from app.db.supabase_client import get_supabase
from app.api.dependencies import get_current_user_optional
import importlib
import re

# Shared Vietnamese customs codes + validator (kebab-case filename → importlib)
customs_validator = importlib.import_module(
    "app.core.vietnamese-customs-declaration-codes-and-validator"
)

logger = logging.getLogger(__name__)


SEA_DOM_FIELDS = [
    'container_seal', 'qty', 'gross_weight_kg', 'vessel', 'route',
    'pol', 'pod', 'place_of_delivery', 'shipper', 'consignee'
]


def validate_sea_dom_fields(details: dict) -> list:
    """Return list of missing SEA_DOM fields (warn, don't block)."""
    return [f for f in SEA_DOM_FIELDS if not details.get(f)]


def is_sea_dom_service(service_type_code: str) -> bool:
    """Check if service type is SEA_DOM or SD prefix."""
    if not service_type_code:
        return False
    code = str(service_type_code).upper()
    return code == 'SEA_DOM' or code.startswith('SD')


def parse_package_quantity(value) -> tuple:
    """
    Parse package_quantity from various formats.

    Examples:
    - 5 -> (5, None)
    - "9 cartons, 1 pallet" -> (10, "9 cartons, 1 pallet")
    - "2 kiện" -> (2, "kiện")
    - "15" -> (15, None)

    Returns: (quantity: int, description: str or None)
    """
    if value is None:
        return (None, None)

    if isinstance(value, int):
        return (value, None)

    if isinstance(value, str):
        # Try to extract all numbers and sum them
        numbers = re.findall(r'\d+', value)
        if numbers:
            total = sum(int(n) for n in numbers)
            # Return original string as description if it has text
            if re.search(r'[a-zA-Zàáảãạăằắẳẵặâầấẩẫậèéẻẽẹêềếểễệìíỉĩịòóỏõọôồốổỗộơờớởỡợùúủũụưừứửữựỳýỷỹỵđ]', value, re.IGNORECASE):
                return (total, value)
            return (total, None)
        return (None, value)

    return (None, None)
router = APIRouter()


# Models

class JobCreateFromChatRequest(BaseModel):
    """Request from Chat UI with AI-extracted entities"""
    entities: Dict[str, Any]
    enriched_data: Optional[Dict[str, Any]] = None


class JobCreateRequest(BaseModel):
    customer_id: int
    booking_date: date
    pickup_time: Optional[time] = None
    route_id: Optional[int] = None
    vehicle_type: str
    vendor_id: Optional[int] = None
    invoice_numbers: Optional[str] = None
    cargo_type: Optional[str] = None
    package_info: Optional[str] = None
    cost_amount: Optional[float] = None
    revenue_amount: Optional[float] = None
    pickup_address: Optional[str] = None
    delivery_address: Optional[str] = None


class VehicleAssignRequest(BaseModel):
    license_plate: str
    driver_name: str
    driver_phone: Optional[str] = None
    driver_id_card: Optional[str] = None
    vendor_name: Optional[str] = None


class JobResponse(BaseModel):
    success: bool = True
    job_id: Optional[int] = None
    job_number: Optional[str] = None
    status: Optional[str] = None
    message: Optional[str] = None
    enriched_data: Optional[Dict[str, Any]] = None
    # Validation error details (e.g. missing_loai_hinh) — includes error code
    # + suggestion list so frontend can render chip UI without hardcoding codes.
    error: Optional[str] = None
    suggestions: Optional[list] = None

# Endpoints

@router.get("/by-number/{job_number}")
async def get_job_by_number(job_number: str):
    """
    Get job by job_number (e.g., TRK-2201-0002)
    """
    try:
        client = get_supabase()
        result = client.table('jobs').select(
            'job_id, job_no, customer_id, status_code'
        ).eq('job_no', job_number).limit(1).execute()

        if result.data:
            row = result.data[0]
            return {
                "success": True,
                "job_id": row["job_id"],
                "id": row["job_id"],
                "job_number": row["job_no"],
                "customer_id": row["customer_id"],
                "status": row["status_code"]
            }
        else:
            return {"success": False, "message": f"Job '{job_number}' not found"}
    except Exception as e:
        logger.error(f"Error looking up job: {e}")
        return {"success": False, "message": str(e)}

@router.post("/create", response_model=JobResponse)
async def create_job(request: JobCreateFromChatRequest, req: Request):
    """
    Create new job from Chat UI with AI-extracted entities
    """
    try:
        data_service = get_data_service()

        # Auth — accept either a valid Bearer token (direct REST caller) or
        # enriched_data.created_by (internal chat loopback from unified_processor,
        # which itself requires auth at /api/message). Reject if neither.
        current_user = await get_current_user_optional(req)
        entities = request.entities
        enriched = request.enriched_data or {}
        if current_user:
            user_id = current_user['user_id']
        else:
            # Internal loopback fallback — trust created_by when set.
            created_by_claim = enriched.get('created_by')
            if created_by_claim is None:
                raise HTTPException(
                    status_code=401,
                    detail="Yêu cầu đăng nhập để tạo job.",
                )
            try:
                user_id = int(created_by_claim)
            except (TypeError, ValueError):
                raise HTTPException(status_code=401, detail="created_by không hợp lệ.")
        
        # If user_id or user_code provided in enriched_data (e.g., from chat bot),
        # use it instead of default admin (1)
        if not current_user:
            if enriched.get('user_id'):
                user_id = enriched['user_id']
            elif enriched.get('user_code'):
                # Lookup user_id from users table (NOT employees)
                try:
                    user_result = data_service.supabase.table('users').select(
                        'user_id'
                    ).eq('user_code', enriched['user_code']).execute()
                    if user_result.data:
                        user_id = user_result.data[0]['user_id']
                        logger.info(f"Resolved user_code '{enriched['user_code']}' to user_id {user_id}")
                    else:
                        logger.warning(f"user_code '{enriched['user_code']}' not found in users table")
                except Exception as e:
                    logger.warning(f"Failed to resolve user_code: {e}")
        
        logger.info(f"Creating job from entities: {entities}")
        
        # Parse booking date (smart parser handles multiple formats: dd.mm.yyyy, yyyy-mm-dd, etc.)
        # Accept both booking_date and scheduled_date as aliases
        booking_date_str = (entities.get('booking_date') or entities.get('scheduled_date') or 
                           enriched.get('booking_date') or enriched.get('scheduled_date'))
        if booking_date_str:
            parsed = parse_date(booking_date_str)
            booking_date = parsed if parsed else date.today()
        else:
            booking_date = date.today()

        # Parse pickup time (smart parser handles: 11:00, 11h00, 11 giờ 30, etc.)
        # Accept both pickup_time and scheduled_time as aliases
        pickup_time_str = (entities.get('pickup_time') or entities.get('scheduled_time') or
                          enriched.get('pickup_time') or enriched.get('scheduled_time'))
        pickup_time = parse_time(pickup_time_str) if pickup_time_str else None
        
        # Handle invoice numbers - AI may return 'invoices' or 'invoice_numbers'
        inv_nums = (entities.get('invoice_numbers') or entities.get('invoices') or 
                   enriched.get('invoice_numbers') or enriched.get('invoices'))
        if isinstance(inv_nums, list):
            invoice_numbers = ', '.join(str(i) for i in inv_nums)
        else:
            invoice_numbers = inv_nums
        
        # Resolve customer_id from customer_code if not provided
        customer_id = enriched.get('customer_id')
        customer_code = entities.get('customer_code') or enriched.get('customer_code')
        
        if not customer_id and customer_code:
            # Look up customer by code
            client = get_supabase()
            result = client.table('customers').select('customer_id').eq(
                'customer_code', customer_code
            ).eq('is_active', True).limit(1).execute()

            if result.data:
                customer_id = result.data[0]['customer_id']
                logger.info(f"Resolved customer '{customer_code}' -> ID {customer_id}")
        
        if not customer_id:
            return JobResponse(
                success=False,
                message=f"Không tìm thấy khách hàng '{customer_code}' trong DB. Vui lòng chọn khách hàng đúng."
            )
        
        # Build job data with structured cargo fields
        job_data = {
            'customer_id': customer_id,
            'booking_date': booking_date,
            'pickup_time': pickup_time,
            'route_id': enriched.get('route_id'),
            'vehicle_type': entities.get('vehicle_type') or enriched.get('vehicle_type') or '1.25T',
            'vendor_id': enriched.get('vendor_id'),
            'invoice_numbers': invoice_numbers,
            'customer_code': entities.get('customer_code') or enriched.get('customer_code'),
            
            # Service type - check services array first, then service_type
            'services': entities.get('services') or [],  # Array of services from AI
            'service_type_code': normalize_service_code(
                (entities.get('services') or [None])[0] or entities.get('service_type') or enriched.get('service_type')
            ),
            
            # Structured cargo data (single item) — check both entities and enriched_data
            'cargo_type': entities.get('cargo_type') or enriched.get('cargo_type'),
            # Parse package_quantity - handle both int and string formats like "9 cartons, 1 pallet"
            'package_quantity': parse_package_quantity(entities.get('package_quantity') or enriched.get('package_quantity'))[0],
            'package_description': parse_package_quantity(entities.get('package_quantity') or enriched.get('package_quantity'))[1],
            'package_unit': entities.get('package_unit') or enriched.get('package_unit'),
            'weight_kg': entities.get('weight_kg') or enriched.get('weight_kg'),
            'dimension_length_cm': entities.get('dimension_length_cm'),
            'dimension_width_cm': entities.get('dimension_width_cm'),
            'dimension_height_cm': entities.get('dimension_height_cm'),
            
            # Multi-item cargo (from Excel/multiple lines)
            'cargo_items': entities.get('cargo_items', []),
            
            # Totals (for multi-item)
            'total_packages': entities.get('total_packages'),
            'total_weight_kg': entities.get('total_weight_kg'),
            'total_cbm': entities.get('total_cbm'),
            
            # Addresses
            'pickup_address': entities.get('pickup_address') or enriched.get('pickup_address'),
            'delivery_address': entities.get('delivery_address') or entities.get('delivery_details') or enriched.get('delivery_address'),
            
            # Special requirements
            'special_requirements': entities.get('special_requirements') or enriched.get('special_requirements'),
            
            # Warehouse-specific
            'storage_start_date': entities.get('storage_start_date'),
            'storage_end_date': entities.get('storage_end_date'),
            
            # Customs-specific
            'cd_no': entities.get('cd_no'),
            'declaration_datetime': entities.get('declaration_datetime'),
            'loai_hinh': entities.get('loai_hinh'),
            'customs_type': entities.get('customs_type'),
            'customs_port': entities.get('customs_port'),
            'buyer_name': entities.get('buyer_name'),
            'seller_name': entities.get('seller_name'),
            'hs_code': entities.get('hs_code'),
            'bl_awb_no': entities.get('bl_awb_no'),
            'co_no': entities.get('co_no'),
            
            # Packing-specific
            'packing_type': entities.get('packing_type'),
            'items_count': entities.get('items_count'),
            'packages_output': entities.get('packages_output'),
            'before_length_cm': entities.get('before_length_cm'),
            'before_width_cm': entities.get('before_width_cm'),
            'before_height_cm': entities.get('before_height_cm'),
            'after_length_cm': entities.get('after_length_cm'),
            'after_width_cm': entities.get('after_width_cm'),
            'after_height_cm': entities.get('after_height_cm'),
            'shrink_wrap': entities.get('shrink_wrap'),
            'vacuum_pack': entities.get('vacuum_pack'),
            'lashing': entities.get('lashing'),
            'fumigation': entities.get('fumigation'),
            
            # Multi-item packing
            'packing_items': entities.get('packing_items', []),
            'total_weight_kg': entities.get('total_weight_kg'),
            
            # Vendor
            'vendor_code': entities.get('vendor_code'),
            
            # Pricing
            'selling_price': enriched.get('selling_price') or entities.get('selling_price'),
            'buying_price': enriched.get('buying_price') or entities.get('buying_price'),
        }
        
        logger.info(f"Job data to create: {job_data}")

        # --- VALIDATOR: customs jobs MUST have a valid loai_hinh (mã loại hình HQ).
        # Defense-in-depth: duplicates service-layer check so malformed requests
        # don't even reach the DB layer. Shared whitelist in
        # app.core.vietnamese-customs-declaration-codes-and-validator.
        loai_hinh_error = customs_validator.validate_loai_hinh_for_service(
            job_data.get('service_type_code'),
            job_data.get('loai_hinh'),
        )
        if loai_hinh_error is not None:
            return JobResponse(
                success=False,
                message=loai_hinh_error["message"],
                error=loai_hinh_error["error"],
                suggestions=loai_hinh_error["suggestions"],
            )

        # --- DUPLICATE CHECK (strict: match cargo details, not just route) ---
        warnings = []
        if customer_id and booking_date:
            try:
                client = get_supabase()
                dup_check = client.table('jobs').select('job_id, job_no').eq(
                    'customer_id', customer_id
                ).eq('etd', str(booking_date)).neq(
                    'status_code', 'CANCELLED'
                ).execute()
                
                if dup_check.data:
                    new_inv = str(invoice_numbers or '').strip().lower()
                    new_qty = job_data.get('package_quantity')
                    new_weight = job_data.get('weight_kg')
                    new_cargo = str(job_data.get('cargo_type') or '').strip().lower()
                    
                    for dup in dup_check.data:
                        dup_detail = client.table('job_services').select('service_details').eq(
                            'job_id', dup['job_id']
                        ).limit(1).execute()
                        if not dup_detail.data:
                            continue
                        dd = dup_detail.data[0].get('service_details') or {}
                        if isinstance(dd, str):
                            dd = json.loads(dd)
                        
                        # Match score: need at least 2 matching fields
                        match_score = 0
                        
                        # Invoice/bill number match (strongest signal)
                        dup_inv = str(dd.get('invoice_numbers') or '').strip().lower()
                        if new_inv and dup_inv and new_inv == dup_inv:
                            match_score += 3  # Strong match
                        
                        # Package quantity match
                        dup_qty = dd.get('package_quantity')
                        if new_qty and dup_qty and int(new_qty) == int(dup_qty):
                            match_score += 1
                        
                        # Weight match
                        dup_weight = dd.get('weight_kg')
                        if new_weight and dup_weight and abs(float(new_weight) - float(dup_weight)) < 1:
                            match_score += 1
                        
                        # Cargo description match
                        dup_cargo = str(dd.get('cargo_type') or '').strip().lower()
                        if new_cargo and dup_cargo and (new_cargo == dup_cargo or new_cargo in dup_cargo or dup_cargo in new_cargo):
                            match_score += 1
                        
                        if match_score >= 2:
                            warnings.append(f"⚠️ Có thể trùng với {dup['job_no']} (cùng khách, cùng ngày, trùng thông tin hàng)")
            except Exception as e:
                logger.warning(f"Duplicate check failed: {e}")

        # --- DOCUMENT NUMBER DUPLICATE CHECK ---
        doc_fields = {
            'cd_no': job_data.get('cd_no'),
            'bl_awb_no': job_data.get('bl_awb_no'),
            'co_no': job_data.get('co_no'),
            'invoice_numbers': invoice_numbers,
        }
        doc_values = {k: v.strip() for k, v in doc_fields.items() if v and str(v).strip()}
        if doc_values:
            try:
                from app.db.session import get_db_context
                with get_db_context() as db:
                    conditions = []
                    params = []
                    field_labels = {'cd_no': 'Tờ khai', 'bl_awb_no': 'BL/AWB', 'co_no': 'CO', 'invoice_numbers': 'Invoice'}
                    for field, val in doc_values.items():
                        if field == 'invoice_numbers':
                            conditions.append(f"js.{field} ILIKE %s")
                            params.append(f"%{val}%")
                        else:
                            conditions.append(f"js.{field} = %s")
                            params.append(val)
                    where = " OR ".join(conditions)
                    db.execute(f"""
                        SELECT j.job_no, js.cd_no, js.bl_awb_no, js.co_no, js.invoice_numbers
                        FROM job_services js JOIN jobs j ON js.job_id = j.job_id
                        WHERE ({where}) LIMIT 5
                    """, params)
                    for row in db.fetchall():
                        matched = []
                        for field, val in doc_values.items():
                            db_val = str(row.get(field) or '').strip()
                            if field == 'invoice_numbers':
                                if val.lower() in db_val.lower():
                                    matched.append(field_labels[field])
                            elif db_val == val:
                                matched.append(field_labels[field])
                        if matched:
                            warnings.append(f"⚠️ Chứng từ trùng: {', '.join(matched)} đã tồn tại trong {row['job_no']}")
            except Exception as e:
                logger.warning(f"Document duplicate check failed: {e}")

        # --- PRICE ANOMALY CHECK ---
        selling_price = job_data.get('selling_price')
        buying_price = job_data.get('buying_price')
        if customer_id and (selling_price or buying_price):
            try:
                client = get_supabase()
                # Get recent jobs for same customer (last 30 days)
                from datetime import timedelta
                cutoff = (booking_date - timedelta(days=60)).isoformat()
                recent = client.table('jobs').select('job_id').eq(
                    'customer_id', customer_id
                ).gte('etd', cutoff).neq('status_code', 'CANCELLED').limit(20).execute()
                
                if recent.data:
                    recent_ids = [r['job_id'] for r in recent.data]
                    prices = []
                    for rid in recent_ids[:10]:
                        svc_r = client.table('job_services').select('service_details').eq('job_id', rid).limit(1).execute()
                        if svc_r.data:
                            sd = svc_r.data[0].get('service_details') or {}
                            if isinstance(sd, str):
                                sd = json.loads(sd)
                            sp = sd.get('selling_price') or sd.get('total_revenue')
                            bp = sd.get('buying_price') or sd.get('total_cost')
                            if sp: prices.append({'sell': float(sp), 'buy': float(bp) if bp else 0})
                    
                    if prices:
                        avg_sell = sum(p['sell'] for p in prices) / len(prices)
                        avg_buy = sum(p['buy'] for p in prices) / len(prices)
                        
                        if selling_price and avg_sell > 0:
                            diff_pct = abs(selling_price - avg_sell) / avg_sell * 100
                            if diff_pct > 30:
                                warnings.append(f"⚠️ Giá bán {selling_price:,.0f} chênh {diff_pct:.0f}% so với TB gần đây ({avg_sell:,.0f})")
                        
                        if buying_price and avg_buy > 0:
                            diff_pct = abs(buying_price - avg_buy) / avg_buy * 100
                            if diff_pct > 30:
                                warnings.append(f"⚠️ Giá mua {buying_price:,.0f} chênh {diff_pct:.0f}% so với TB gần đây ({avg_buy:,.0f})")
            except Exception as e:
                logger.warning(f"Price check failed: {e}")
        
        # Create job in database (pass user_id for created_by tracking)
        job = await data_service.create_job(job_data, user_id=user_id)
        
        # Log job creation source (chat bot vs web UI)
        source = "chat" if enriched.get('user_code') or enriched.get('created_by') else "api"
        logger.info(f"Job created: {job} | source={source} | user_id={user_id} | user_code={enriched.get('user_code', 'N/A')}")
        
        # --- AUTO-SAVE PRICING + SEA_DOM FIELDS to service_details ---
        service_type_code_for_check = normalize_service_code(
            (entities.get('services') or [None])[0] or entities.get('service_type') or enriched.get('service_type')
        ) or ''
        sea_dom_data = {}
        if is_sea_dom_service(service_type_code_for_check):
            for f in SEA_DOM_FIELDS:
                val = entities.get(f) or enriched.get(f)
                if val is not None:
                    sea_dom_data[f] = val
            missing = [f for f in SEA_DOM_FIELDS if not sea_dom_data.get(f)]
            if missing:
                warnings.append(f"⚠️ SEA_DOM: Thiếu các trường: {', '.join(missing)}")

        if job.get("id") and (selling_price or buying_price or sea_dom_data):
            try:
                client = get_supabase()
                svc_result = client.table('job_services').select('svc_id, service_details').eq(
                    'job_id', job['id']
                ).limit(1).execute()
                
                if svc_result.data:
                    svc = svc_result.data[0]
                    details = svc.get('service_details') or {}
                    if isinstance(details, str):
                        details = json.loads(details)
                    
                    if selling_price:
                        details['selling_price'] = selling_price
                    if buying_price:
                        details['buying_price'] = buying_price
                    
                    # Merge SEA_DOM fields
                    if sea_dom_data:
                        details.update(sea_dom_data)
                    
                    total_cost = (buying_price or 0) + sum(
                        float(x.get('amount', 0)) for x in (details.get('extra_costs') or [])
                    )
                    total_revenue = (selling_price or 0) + sum(
                        float(x.get('amount', 0)) for x in (details.get('extra_revenues') or [])
                    )
                    details['total_cost'] = total_cost
                    details['total_revenue'] = total_revenue
                    
                    client.table('job_services').update({
                        'service_details': json.dumps(details, ensure_ascii=False)
                    }).eq('svc_id', svc['svc_id']).execute()
                    
                    # Update job totals
                    profit = total_revenue - total_cost
                    client.table('jobs').update({
                        'total_revenue': total_revenue,
                        'total_cost': total_cost,
                        'profit': profit
                    }).eq('job_id', job['id']).execute()
                    
                    logger.info(f"Auto-saved pricing for job {job['id']}: sell={selling_price}, buy={buying_price}")
            except Exception as e:
                logger.warning(f"Auto-save pricing failed (job still created): {e}")
        
        msg = "Job đã được tạo thành công!"
        if warnings:
            msg += "\n" + "\n".join(warnings)
        
        return JobResponse(
            success=True,
            job_id=job.get("id"),
            job_number=job.get("job_number"),
            status="PENDING",
            message=msg
        )
        
    except HTTPException:
        # Let auth 401s / other explicit HTTP errors propagate with proper status
        raise
    except Exception as e:
        logger.error(f"Error creating job: {e}")
        logger.error(traceback.format_exc())
        return JobResponse(
            success=False,
            message=f"Lỗi tạo job: {str(e)}"
        )


@router.post("/{job_id}/assign-vehicle", response_model=JobResponse)
async def assign_vehicle(job_id: int, request: VehicleAssignRequest, req: Request):
    """
    Assign vehicle to job and generate customer confirmation message
    """
    try:
        data_service = get_data_service()

        # Check job exists
        job_data = await data_service.get_job(job_id)
        if not job_data:
            raise HTTPException(404, f"Job {job_id} not found")

        # Resolve user from JWT (no hardcode user_id=1)
        current_user = await get_current_user_optional(req)
        user_id = current_user['user_id'] if current_user else 1

        # Assign vehicle
        await data_service.assign_vehicle(job_id, request.dict(), user_id=user_id)
        
        # Get updated job with full details for frontend template
        # Re-using get_job_details logic to fetch everything needed for the message
        details = await get_job_details(job_id)
        job_info = details["job"]
        services = details["services"]
        
        # Get first service for time/address details (or aggregate if multiple)
        first_service = services[0] if services else {}
        
        # Map fields to match what frontend expects in enriched_data
        enriched = {
            "customer_code": job_info.get("customer_code"),
            "customer_name": job_info.get("customer_name"),
            "booking_date": str(first_service.get("scheduled_date") or job_info.get("etd") or ""),
            "scheduled_date": str(first_service.get("scheduled_date") or job_info.get("etd") or ""),
            "pickup_date": str(first_service.get("scheduled_date") or job_info.get("etd") or ""),
            "pickup_time": str(first_service.get("scheduled_time") or ""),
            "pickup_address": first_service.get("origin_address"),
            "delivery_address": first_service.get("dest_address"),
            "vehicle_type": request.model_dump().get("vehicle_type"),
            "invoice_numbers": [], # Will be populated below
            "cargo_type": "", # Will be populated below
            # Include vehicle info we just assigned
            "license_plate": request.license_plate,
            "driver_name": request.driver_name,
            "driver_phone": request.driver_phone,
            "vendor_name": request.vendor_name 
        }

        # Extract cargo items and invoices from services
        all_cargo_items = []
        invoices = []
        package_quantity = 0
        package_unit = "kiện"
        cargo_descriptions = set()
        dimensions = None
        
        if services:
            for svc in services:
                service_details = svc.get("service_details")
                
                # Parse service_details if it's a string
                if isinstance(service_details, str):
                    try:
                        import json
                        service_details = json.loads(service_details)
                    except:
                        service_details = {}
                
                # Extract invoices from service_details JSONB first
                if isinstance(service_details, dict):
                    if service_details.get("invoice_numbers"):
                        inv_list = service_details["invoice_numbers"]
                        if isinstance(inv_list, str):
                            invoices.extend([i.strip() for i in inv_list.split(",")])
                        elif isinstance(inv_list, list):
                            invoices.extend(inv_list)
                    
                    # Extract cargo items
                    if service_details.get("cargo_items"):
                        all_cargo_items.extend(service_details["cargo_items"])
                    
                    # Aggregate package quantities
                    if service_details.get("package_quantity"):
                        package_quantity += service_details["package_quantity"]
                    
                    if service_details.get("package_unit"):
                        package_unit = service_details["package_unit"]
                    
                    # Collect cargo descriptions
                    if service_details.get("cargo_type"):
                        cargo_descriptions.add(service_details["cargo_type"])
                    
                    # Dimensions
                    if service_details.get("dimension_length_cm"):
                        dimensions = {
                            "length": service_details.get("dimension_length_cm"),
                            "width": service_details.get("dimension_width_cm"),
                            "height": service_details.get("dimension_height_cm")
                        }
                
                # FALLBACK: If service_details is empty, read from columns (for old jobs)
                if not invoices and svc.get("invoice_numbers"):
                    inv_col = svc["invoice_numbers"]
                    if isinstance(inv_col, str):
                        invoices.extend([i.strip() for i in inv_col.split(",") if i.strip()])
                    elif isinstance(inv_col, list):
                        invoices.extend(inv_col)
                
                # FALLBACK: Get addresses from columns if not in enriched yet
                if not enriched.get("pickup_address") and svc.get("origin_address"):
                    enriched["pickup_address"] = svc["origin_address"]
                if not enriched.get("delivery_address") and svc.get("dest_address"):
                    enriched["delivery_address"] = svc["dest_address"]
        
        enriched["invoice_numbers"] = invoices
        enriched["cargo_items"] = all_cargo_items
        enriched["package_quantity"] = package_quantity
        enriched["package_unit"] = package_unit
        
        # If we have cargo items, use their details
        if all_cargo_items:
            # Extract invoices from cargo_items  
            for item in all_cargo_items:
                if item.get("invoice_no") and item["invoice_no"] not in invoices:
                    invoices.append(item["invoice_no"])
            enriched["invoice_numbers"] = invoices  # Update with cargo item invoices
            
            cargo_desc = ", ".join(set(item.get("description", "") for item in all_cargo_items if item.get("description")))
            enriched["cargo_type"] = cargo_desc or list(cargo_descriptions)[0] if cargo_descriptions else ""
        elif cargo_descriptions:
            enriched["cargo_type"] = ", ".join(cargo_descriptions)
        
        if dimensions:
            enriched["dimensions"] = dimensions
        
        # Important: The frontend checks `msg.entities.cargo_items` OR `msg.enriched_data`.
        # We should put cargo info in enriched_data.
        
        return JobResponse(
            success=True,
            job_id=job_id,
            job_number=job_data.get("job_number"),
            status="DISPATCHED",
            message="Đã gán xe thành công!",
            enriched_data=enriched
        )
    except Exception as e:
        logger.error(f"Error assigning vehicle: {e}")
        logger.error(traceback.format_exc())
        return JobResponse(
            success=False,
            message=f"Lỗi gán xe: {str(e)}"
        )


@router.post("/{job_id}/complete", response_model=JobResponse)
async def complete_job(job_id: int, delivery_time: Optional[datetime] = None):
    """
    Mark job as completed
    """
    data_service = get_data_service()
    
    # Check job exists
    job_data = await data_service.get_job(job_id)
    if not job_data:
        raise HTTPException(404, f"Job {job_id} not found")
    
    # Update job status to COMPLETED in DB
    try:
        client = get_supabase()
        update_data = {"status_code": "COMPLETED"}
        if delivery_time:
            update_data["actual_delivery_time"] = delivery_time.isoformat()
        client.table('jobs').update(update_data).eq('job_id', job_id).execute()
    except Exception as e:
        logger.error(f"Failed to complete job {job_id}: {e}")
        raise HTTPException(500, f"Không cập nhật được trạng thái: {e}")

    return JobResponse(
        success=True,
        job_id=job_id,
        job_number=job_data.get("job_no") or job_data.get("job_number"),
        status="COMPLETED",
        message="Job đã hoàn thành!"
    )


# ========================================
# EXCEL EXPORT BY CUSTOMER/VENDOR
# NOTE: This route MUST be defined BEFORE /{job_id} to avoid path matching conflict
# ========================================
@router.get("/exports/entity")
async def export_jobs_by_entity_v2(
    entity_type: str,  # customer or vendor
    entity_id: int,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    month: Optional[str] = None,
    job_ids: Optional[str] = None,  # comma-separated job_ids for multi-select
    service_type: Optional[str] = None,  # filter by service_type_code
):
    """Export jobs pivot format: cost names as columns, jobs as rows.
    Separate sheets for Revenue (Doanh thu) and Cost (Chi phí)."""
    from fastapi.responses import FileResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import tempfile
    import calendar

    if entity_type not in ["customer", "vendor"]:
        raise HTTPException(400, "entity_type phải là customer hoặc vendor")

    try:
        client = get_supabase()

        # Build date filter
        date_start = from_date
        date_end = to_date
        if month:
            year, mon = month.split('-')
            date_start = f"{year}-{mon}-01"
            last_day = calendar.monthrange(int(year), int(mon))[1]
            date_end = f"{year}-{mon}-{last_day}"

        # Get entity info
        if entity_type == "customer":
            entity_result = client.table('customers').select(
                'customer_id, customer_code, short_name, company_name'
            ).eq('customer_id', entity_id).limit(1).execute()
        else:
            entity_result = client.table('vendors').select(
                'vendor_id, vendor_code, short_name, company_name'
            ).eq('vendor_id', entity_id).limit(1).execute()

        if not entity_result.data:
            raise HTTPException(404, f"Không tìm thấy {entity_type}")

        entity = entity_result.data[0]
        entity_code = entity.get('customer_code') or entity.get('vendor_code')

        # Get jobs - support multi-select via job_ids param
        selected_job_ids = None
        if job_ids:
            selected_job_ids = [int(x.strip()) for x in job_ids.split(',') if x.strip()]

        if entity_type == "customer":
            query = client.table('jobs').select(
                '*, customers(customer_code, short_name)'
            ).eq('customer_id', entity_id)
            if selected_job_ids:
                query = query.in_('job_id', selected_job_ids)
            elif date_start:
                query = query.gte('etd', date_start)
            if not selected_job_ids and date_end:
                query = query.lte('etd', date_end)
        else:
            svc_query = client.table('job_services').select('job_id').eq('vendor_id', entity_id)
            svc_result = svc_query.execute()
            vendor_job_ids = list(set([r['job_id'] for r in svc_result.data]))
            if not vendor_job_ids:
                raise HTTPException(404, "Không có dữ liệu để xuất")
            if selected_job_ids:
                vendor_job_ids = [jid for jid in vendor_job_ids if jid in selected_job_ids]
            query = client.table('jobs').select(
                '*, customers(customer_code, short_name)'
            ).in_('job_id', vendor_job_ids)

        jobs = query.order('etd').execute().data
        if not jobs:
            raise HTTPException(404, "Không có dữ liệu để xuất")

        all_job_ids = [j['job_id'] for j in jobs]

        # Get services and costs
        services = client.table('job_services').select(
            '*, vendors(vendor_code, short_name)'
        ).in_('job_id', all_job_ids).order('scheduled_date').execute().data

        costs = client.table('job_costs').select('*').in_('job_id', all_job_ids).execute().data

        # Filter by service_type if specified
        if service_type:
            svc_job_ids = set(
                s['job_id'] for s in services
                if s.get('service_type_code') == service_type
            )
            jobs = [j for j in jobs if j['job_id'] in svc_job_ids]
            all_job_ids = [j['job_id'] for j in jobs]
            services = [s for s in services if s['job_id'] in set(all_job_ids)]
            costs = [c for c in costs if c['job_id'] in set(all_job_ids)]
            if not jobs:
                raise HTTPException(404, "Không có dữ liệu để xuất")

        # Group data by job
        svc_by_job = {}
        for s in services:
            svc_by_job.setdefault(s['job_id'], []).append(s)
        cost_by_job = {}
        for c in costs:
            cost_by_job.setdefault(c['job_id'], []).append(c)

        # ------------------------------------------------------------------
        # Normalize cost_name so near-duplicate variants don't spawn one
        # column per cost. Two kinds of noise get stripped:
        #
        # (1) Receipt numbers — "Thu hộ: Lệ phí hải quan - GNT: 0360356"
        #     both collapse to "Thu hộ: Lệ phí hải quan"; the numbers go into
        #     a separate "Chứng từ" column.
        #
        # (2) Route info after a transport-fee prefix — "Cước vận chuyển KCN
        #     Yên Bình → KCN Đồng Lạng" collapses to just "Cước vận chuyển"
        #     because the route is already visible in the Điểm đi / Điểm đến
        #     columns. Rule: if the name starts with a known transport prefix
        #     AND has more content after it, keep only the prefix.
        # ------------------------------------------------------------------
        _RECEIPT_MARKERS = (
            r"\s*-\s*GNT\s*:",
            r"\s*-\s*HĐ\s*:",
            r"\s*-\s*HD\s*:",
            r"\s*-\s*Hóa\s+đơn\s*:",
            r"\s*-\s*Hoa\s+don\s*:",
            r"\s*-\s*BL\s*:",
            r"\s*-\s*Bill\s*:",
            r"\s*-\s*Invoice\s*:",
        )
        _RECEIPT_SPLIT_RE = re.compile("|".join(_RECEIPT_MARKERS), re.IGNORECASE)

        # Two classes of prefix rules:
        # - DISCARD: tail is route info that's already visible in Điểm đi /
        #   Điểm đến columns — throw it away.
        # - CAPTURE: tail is a document reference (CO number, tờ khai number,
        #   invoice ref) — canonicalize and push the tail into the Chứng từ
        #   column so the receipt list stays per-job.
        _DISCARD_TAIL_PREFIXES = (
            "Cước vận chuyển",
            "Cước vận tải",
            "Phí vận chuyển",
            "Phí vận tải",
            "Chi phí vận chuyển",
            "Chi phí vận tải",
            "Cước xe",
        )
        _CAPTURE_TAIL_PREFIXES = (
            "Phí C/O",
            "Phí CO",
            "Phí mở tờ khai",
            "Phí mở TK",
        )

        def _build_prefix_re(prefixes):
            # Matches when `prefix` is followed by at least one non-prefix
            # character — signaling there's a tail to strip.
            return re.compile(
                r"^(" + "|".join(re.escape(p) for p in prefixes) + r")\b[\s\-–—:,.]*(\S.*)?$",
                re.IGNORECASE,
            )

        _DISCARD_TAIL_RE = _build_prefix_re(_DISCARD_TAIL_PREFIXES)
        _CAPTURE_TAIL_RE = _build_prefix_re(_CAPTURE_TAIL_PREFIXES)

        def _canonical_cost_name(raw: str) -> tuple[str, str]:
            """
            Split `raw` cost_name into (canonical_name, receipt_number).
            Handles 3 kinds of noise:
              1. explicit receipt markers  — "- GNT: 0360356"
              2. discard-tail prefixes     — route suffix after "Cước vận chuyển ..."
              3. capture-tail prefixes     — doc ref after "Phí C/O VEI2600046"
            """
            if not raw:
                return '', ''
            parts = _RECEIPT_SPLIT_RE.split(raw, maxsplit=1)
            canonical = (parts[0] or '').strip()
            receipt = (parts[1].strip() if len(parts) > 1 else '')

            m = _DISCARD_TAIL_RE.match(canonical)
            if m and m.group(2):
                canonical = m.group(1)

            m = _CAPTURE_TAIL_RE.match(canonical)
            if m and m.group(2):
                tail = m.group(2).strip().rstrip(",;")
                # Only treat the tail as a document reference if it has 3+
                # chars AND at least one digit. Keeps abbreviations like "HQ"
                # (hải quan) or "VN" from polluting the Chứng từ column.
                is_ref_like = len(tail) >= 3 and any(ch.isdigit() for ch in tail)
                canonical = m.group(1)
                if is_ref_like and tail and not receipt:
                    receipt = tail

            # Normalize leading letter casing so "cước vận chuyển" and
            # "Cước vận chuyển" land in the same pivot column.
            if canonical:
                canonical = canonical[:1].upper() + canonical[1:]

            return canonical, receipt

        # Build pivot data: collect unique CANONICAL cost names for revenue + cost columns
        revenue_names = []  # ordered unique canonical names with selling_amount > 0
        cost_names = []     # ordered unique canonical names with buying_amount > 0
        rev_set = set()
        cost_set = set()
        # receipts_by_job[job_id] = ordered list of receipt numbers (no dup)
        receipts_by_job: Dict[int, list] = {}
        has_any_receipt = False
        for c in costs:
            raw_name = (c.get('cost_name') or '').strip()
            if not raw_name:
                continue
            canonical, receipt = _canonical_cost_name(raw_name)
            if not canonical:
                continue
            if float(c.get('selling_amount') or 0) > 0 or float(c.get('selling_rate') or 0) > 0:
                if canonical not in rev_set:
                    rev_set.add(canonical)
                    revenue_names.append(canonical)
            if float(c.get('buying_amount') or 0) > 0 or float(c.get('buying_rate') or 0) > 0:
                if canonical not in cost_set:
                    cost_set.add(canonical)
                    cost_names.append(canonical)
            if receipt:
                has_any_receipt = True
                bucket = receipts_by_job.setdefault(c.get('job_id'), [])
                if receipt not in bucket:
                    bucket.append(receipt)

        # Helper: extract vehicle plate from service
        def _get_plate(svc):
            sd = svc.get('service_details') or {}
            if isinstance(sd, str):
                try:
                    sd = json.loads(sd)
                except Exception:
                    sd = {}
            plate = sd.get('vehicle_plate', '')
            if not plate and svc.get('route'):
                route_str = svc['route']
                if '(Xe:' in route_str:
                    plate = route_str.split('(Xe:')[-1].rstrip(')')
            if not plate:
                vti = svc.get('vendor_text_input')
                vti_list = []
                if isinstance(vti, str):
                    try:
                        parsed = json.loads(vti)
                        vti_list = parsed if isinstance(parsed, list) else [parsed] if isinstance(parsed, dict) else []
                    except Exception:
                        pass
                elif isinstance(vti, list):
                    vti_list = vti
                elif isinstance(vti, dict):
                    vti_list = [vti]
                if vti_list:
                    plates = [v.get('license_plate', '') for v in vti_list if v.get('license_plate')]
                    plate = ', '.join(plates)
            return (plate or '').strip()

        # Helper: get vendor name from service
        def _get_vendor(svc):
            vendor = svc.get('vendors') or {}
            return vendor.get('short_name') or vendor.get('vendor_code') or ''

        # Excel styles
        wb = openpyxl.Workbook()
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        tot_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        money_fmt = '#,##0'

        # Fixed columns for both sheets
        fixed_headers = ['STT', 'Mã Job', 'Ngày', 'Loại DV', 'Điểm đi', 'Điểm đến',
                         'Biển số xe', 'NCC/Vendor', 'BL/AWB']
        fixed_count = len(fixed_headers)

        # ── SHEET 1: DOANH THU (Revenue) ──
        ws_rev = wb.active
        ws_rev.title = "Doanh thu"
        # "Chứng từ" column goes after the pivot columns (only when at least one
        # cost has a receipt marker — keeps headers tidy for trucking customers).
        tail_headers = ['TỔNG DOANH THU', 'VAT', 'TỔNG SAU VAT']
        if has_any_receipt:
            tail_headers = ['Chứng từ'] + tail_headers
        rev_headers = fixed_headers + revenue_names + tail_headers
        for col, h in enumerate(rev_headers, 1):
            cell = ws_rev.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        row = 2
        for idx, job in enumerate(jobs, 1):
            job_id = job['job_id']
            job_svcs = svc_by_job.get(job_id, [])
            job_costs_list = cost_by_job.get(job_id, [])
            first_svc = job_svcs[0] if job_svcs else {}

            # Fixed columns
            ws_rev.cell(row=row, column=1, value=idx).border = border
            ws_rev.cell(row=row, column=2, value=job.get('job_no')).border = border
            ws_rev.cell(row=row, column=3, value=str(job.get('etd') or '')).border = border
            ws_rev.cell(row=row, column=4, value=first_svc.get('service_type_code', '')).border = border
            ws_rev.cell(row=row, column=5, value=first_svc.get('origin_address', '')).border = border
            ws_rev.cell(row=row, column=6, value=first_svc.get('dest_address', '')).border = border
            ws_rev.cell(row=row, column=7, value=_get_plate(first_svc)).border = border
            ws_rev.cell(row=row, column=8, value=_get_vendor(first_svc)).border = border
            ws_rev.cell(row=row, column=9, value=first_svc.get('bl_awb_no', '')).border = border

            # Revenue columns: pivot by CANONICAL cost_name so GNT-numbered
            # variants of the same fee collapse to a single column.
            rev_by_name = {}
            for c in job_costs_list:
                canonical, _ = _canonical_cost_name(c.get('cost_name') or '')
                if not canonical:
                    continue
                amt = float(c.get('selling_amount') or 0)
                if amt > 0:
                    rev_by_name[canonical] = rev_by_name.get(canonical, 0) + amt

            for ci, rn in enumerate(revenue_names):
                col_idx = fixed_count + ci + 1
                val = rev_by_name.get(rn, 0)
                cell = ws_rev.cell(row=row, column=col_idx, value=val if val else None)
                cell.border = border
                cell.number_format = money_fmt

            # "Chứng từ" column (if any cost carries a GNT/HĐ/BL receipt number)
            cursor = fixed_count + len(revenue_names)
            if has_any_receipt:
                cursor += 1
                cell = ws_rev.cell(
                    row=row, column=cursor,
                    value=", ".join(receipts_by_job.get(job_id, [])) or None,
                )
                cell.border = border
                cell.alignment = Alignment(wrap_text=True)

            # TỔNG DOANH THU (SUM of pivot columns — excludes Chứng từ column)
            total_col = cursor + 1
            if revenue_names:
                first_cl = get_column_letter(fixed_count + 1)
                last_cl = get_column_letter(fixed_count + len(revenue_names))
                cell = ws_rev.cell(row=row, column=total_col)
                cell.value = f'=SUM({first_cl}{row}:{last_cl}{row})'
            else:
                cell = ws_rev.cell(row=row, column=total_col, value=0)
            cell.border = border
            cell.number_format = money_fmt
            cell.font = Font(bold=True)

            # VAT — calculate from each cost line's vat_rate
            vat_col = total_col + 1
            vat_amt = 0
            for c in job_costs_list:
                sell_amt = float(c.get('selling_amount') or 0)
                vat_rate = float(c.get('vat_rate') or 0) / 100
                vat_amt += sell_amt * vat_rate
            cell = ws_rev.cell(row=row, column=vat_col, value=round(vat_amt))
            cell.border = border
            cell.number_format = money_fmt

            # TỔNG SAU VAT
            after_vat_col = vat_col + 1
            t_cl = get_column_letter(total_col)
            v_cl = get_column_letter(vat_col)
            cell = ws_rev.cell(row=row, column=after_vat_col)
            cell.value = f'={t_cl}{row}+{v_cl}{row}'
            cell.border = border
            cell.number_format = money_fmt
            cell.font = Font(bold=True)

            row += 1

        # Totals row — sum over numeric columns (pivot fees + TỔNG + VAT +
        # TỔNG SAU VAT). Skip the "Chứng từ" column since it's text.
        if row > 2:
            tr = row
            ws_rev.cell(row=tr, column=1, value="TỔNG CỘNG").font = Font(bold=True)
            chungtu_col = (fixed_count + len(revenue_names) + 1) if has_any_receipt else None
            for c in range(fixed_count + 1, len(rev_headers) + 1):
                if c == chungtu_col:
                    continue  # Chứng từ is text, don't SUM
                cl = get_column_letter(c)
                cell = ws_rev.cell(row=tr, column=c)
                cell.value = f'=SUM({cl}2:{cl}{tr-1})'
                cell.number_format = money_fmt
                cell.font = Font(bold=True)
            for c in range(1, len(rev_headers) + 1):
                ws_rev.cell(row=tr, column=c).fill = tot_fill
                ws_rev.cell(row=tr, column=c).border = border

        # Column widths
        for i in range(1, fixed_count + 1):
            ws_rev.column_dimensions[get_column_letter(i)].width = [5, 18, 12, 14, 20, 20, 14, 14, 16][i-1]
        # Pivot cost columns
        for i in range(fixed_count + 1, fixed_count + len(revenue_names) + 1):
            ws_rev.column_dimensions[get_column_letter(i)].width = 16
        # Chứng từ column — wider because it may hold multiple receipt numbers
        if has_any_receipt:
            chungtu_col_letter = get_column_letter(fixed_count + len(revenue_names) + 1)
            ws_rev.column_dimensions[chungtu_col_letter].width = 26
        # Tail total/vat columns
        for i in range(fixed_count + len(revenue_names) + (2 if has_any_receipt else 1),
                       len(rev_headers) + 1):
            ws_rev.column_dimensions[get_column_letter(i)].width = 16
        ws_rev.freeze_panes = 'A2'

        # ── SHEET 2: CHI PHÍ (Cost) ──
        ws_cost = wb.create_sheet("Chi phí")
        cost_headers = fixed_headers + cost_names + ['TỔNG CHI PHÍ']
        for col, h in enumerate(cost_headers, 1):
            cell = ws_cost.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        row = 2
        for idx, job in enumerate(jobs, 1):
            job_id = job['job_id']
            job_svcs = svc_by_job.get(job_id, [])
            job_costs_list = cost_by_job.get(job_id, [])
            first_svc = job_svcs[0] if job_svcs else {}

            # Fixed columns
            ws_cost.cell(row=row, column=1, value=idx).border = border
            ws_cost.cell(row=row, column=2, value=job.get('job_no')).border = border
            ws_cost.cell(row=row, column=3, value=str(job.get('etd') or '')).border = border
            ws_cost.cell(row=row, column=4, value=first_svc.get('service_type_code', '')).border = border
            ws_cost.cell(row=row, column=5, value=first_svc.get('origin_address', '')).border = border
            ws_cost.cell(row=row, column=6, value=first_svc.get('dest_address', '')).border = border
            ws_cost.cell(row=row, column=7, value=_get_plate(first_svc)).border = border
            ws_cost.cell(row=row, column=8, value=_get_vendor(first_svc)).border = border
            ws_cost.cell(row=row, column=9, value=first_svc.get('bl_awb_no', '')).border = border

            # Cost columns: pivot by CANONICAL cost_name
            cost_by_name = {}
            for c in job_costs_list:
                canonical, _ = _canonical_cost_name(c.get('cost_name') or '')
                if not canonical:
                    continue
                amt = float(c.get('buying_amount') or 0)
                if amt > 0:
                    cost_by_name[canonical] = cost_by_name.get(canonical, 0) + amt

            for ci, cn in enumerate(cost_names):
                col_idx = fixed_count + ci + 1
                val = cost_by_name.get(cn, 0)
                cell = ws_cost.cell(row=row, column=col_idx, value=val if val else None)
                cell.border = border
                cell.number_format = money_fmt

            # TỔNG CHI PHÍ
            total_col = fixed_count + len(cost_names) + 1
            if cost_names:
                first_cl = get_column_letter(fixed_count + 1)
                last_cl = get_column_letter(fixed_count + len(cost_names))
                cell = ws_cost.cell(row=row, column=total_col)
                cell.value = f'=SUM({first_cl}{row}:{last_cl}{row})'
            else:
                cell = ws_cost.cell(row=row, column=total_col, value=0)
            cell.border = border
            cell.number_format = money_fmt
            cell.font = Font(bold=True)

            row += 1

        # Totals row
        if row > 2:
            tr = row
            ws_cost.cell(row=tr, column=1, value="TỔNG CỘNG").font = Font(bold=True)
            for c in range(fixed_count + 1, fixed_count + len(cost_names) + 2):
                cl = get_column_letter(c)
                cell = ws_cost.cell(row=tr, column=c)
                cell.value = f'=SUM({cl}2:{cl}{tr-1})'
                cell.number_format = money_fmt
                cell.font = Font(bold=True)
            for c in range(1, len(cost_headers) + 1):
                ws_cost.cell(row=tr, column=c).fill = tot_fill
                ws_cost.cell(row=tr, column=c).border = border

        # Column widths
        for i in range(1, fixed_count + 1):
            ws_cost.column_dimensions[get_column_letter(i)].width = [5, 18, 12, 14, 20, 20, 14, 14, 16][i-1]
        for i in range(fixed_count + 1, len(cost_headers) + 1):
            ws_cost.column_dimensions[get_column_letter(i)].width = 16
        ws_cost.freeze_panes = 'A2'

        # ── SHEET 3: Tổng hợp (Summary: revenue, cost, profit per job) ──
        ws3 = wb.create_sheet("Tổng hợp")
        h3 = ['STT', 'Mã Job', 'Ngày', 'Loại DV', 'Tuyến', 'Biển số',
               'NCC/Vendor', 'BL/AWB', 'Doanh thu', 'Chi phí', 'Lợi nhuận', 'Margin %', 'Trạng thái']
        for col, h in enumerate(h3, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.border = border

        r3 = 2
        for idx, job in enumerate(jobs, 1):
            job_id = job['job_id']
            job_svcs = svc_by_job.get(job_id, [])
            job_costs_list = cost_by_job.get(job_id, [])
            first_svc = job_svcs[0] if job_svcs else {}

            revenue = sum(float(c.get('selling_amount') or 0) for c in job_costs_list)
            cost_total = sum(float(c.get('buying_amount') or 0) for c in job_costs_list)
            profit = revenue - cost_total
            margin = (profit / revenue * 100) if revenue > 0 else 0

            route_display = first_svc.get('route') or ''
            if '(Xe:' in route_display:
                route_display = route_display.split('(Xe:')[0].strip()

            row_data = [
                idx, job.get('job_no'), str(job.get('etd') or ''),
                first_svc.get('service_type_code', ''),
                route_display, _get_plate(first_svc),
                _get_vendor(first_svc), first_svc.get('bl_awb_no', ''),
                revenue, cost_total, profit, round(margin, 1),
                job.get('status_code', '')
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws3.cell(row=r3, column=col, value=val)
                cell.border = border
                if col in [9, 10, 11]:
                    cell.number_format = money_fmt
                if col == 12:
                    cell.number_format = '0.0"%"'
            r3 += 1

        # Summary totals
        if r3 > 2:
            ws3.cell(row=r3, column=1, value="TỔNG CỘNG").font = Font(bold=True)
            for c in [9, 10, 11]:
                cl = get_column_letter(c)
                cell = ws3.cell(row=r3, column=c)
                cell.value = f'=SUM({cl}2:{cl}{r3-1})'
                cell.number_format = money_fmt
                cell.font = Font(bold=True)
            for c in range(1, 14):
                ws3.cell(row=r3, column=c).fill = tot_fill
                ws3.cell(row=r3, column=c).border = border

        w3 = [5, 18, 12, 14, 30, 14, 14, 16, 16, 16, 16, 10, 12]
        for i, w in enumerate(w3, 1):
            ws3.column_dimensions[get_column_letter(i)].width = w
        ws3.freeze_panes = 'A2'

        # Save
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()

        period = month or f"{from_date or ''}_{to_date or ''}"
        filename = f"{entity_code}_{period}_export.xlsx"
        return FileResponse(
            path=temp_file.name,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Export error: {e}")
        logger.error(traceback.format_exc())
        raise HTTPException(500, f"Lỗi xuất Excel: {str(e)}")


# NOTE: This route MUST be defined BEFORE /{job_id} to avoid path matching conflict
@router.get("/recent")
async def get_recent_jobs(limit: int = 10, status: Optional[str] = None, offset: int = 0):
    """
    Get recent jobs for dashboard, optionally filtered by status.
    limit: max 500 (default 10 for dashboard widget)
    offset: for pagination
    """
    try:
        # Cap limit for safety
        limit = max(1, min(int(limit or 10), 500))
        offset = max(0, int(offset or 0))
        client = get_supabase()
        # Fetch jobs with customer info (created_by join added when column exists)
        query = client.table('jobs').select(
            'job_id, job_no, status_code, etd, created_at, updated_at, customer_id, created_by, updated_by, '
            'total_revenue, total_cost, '
            'customers(customer_code, short_name)'
        )
        if status:
            query = query.eq('status_code', status)
        result = query.order('created_at', desc=True).range(offset, offset + limit - 1).execute()

        # Get user info for created_by and updated_by if any jobs have them
        creator_ids = [r.get('created_by') for r in result.data if r.get('created_by')]
        updater_ids = [r.get('updated_by') for r in result.data if r.get('updated_by')]
        all_user_ids = list(set(creator_ids + updater_ids))
        user_map = {}
        if all_user_ids:
            users_result = client.table('users').select(
                'user_id, user_code, full_name'
            ).in_('user_id', all_user_ids).execute()
            for u in users_result.data:
                user_map[u['user_id']] = {'user_code': u['user_code'], 'full_name': u['full_name']}

        jobs = []
        for row in result.data:
            customer = row.get('customers') or {}
            creator = user_map.get(row.get('created_by'), {})
            updater = user_map.get(row.get('updated_by'), {})
            jobs.append({
                'job_id': row['job_id'],
                'job_no': row['job_no'],
                'status_code': row['status_code'],
                'etd': row['etd'],
                'created_at': row['created_at'],
                'updated_at': row.get('updated_at'),
                'customer_id': row.get('customer_id'),
                'customer_code': customer.get('customer_code'),
                'customer_name': customer.get('short_name'),
                'total_revenue': float(row.get('total_revenue') or 0),
                'total_cost': float(row.get('total_cost') or 0),
                'created_by': row.get('created_by'),
                'creator_code': creator.get('user_code'),
                'creator_name': creator.get('full_name'),
                'updated_by': row.get('updated_by'),
                'updater_code': updater.get('user_code'),
                'updater_name': updater.get('full_name'),
            })

        # Get service types and reimbursement totals for each job
        if jobs:
            job_ids = [j['job_id'] for j in jobs]
            svc_result = client.table('job_services').select(
                'job_id, service_type_code'
            ).in_('job_id', job_ids).execute()

            svc_map = {}
            for svc in svc_result.data:
                if svc['job_id'] not in svc_map:
                    svc_map[svc['job_id']] = svc['service_type_code']

            # Get reimbursement (at-cost / chi hộ) totals from job_costs — split selling vs buying
            reimb_rev_map = {}   # selling side → "CHI HỘ" billed to customer
            reimb_cost_map = {}  # buying side → pass-through cost paid to vendor
            try:
                costs_result = client.table('job_costs').select(
                    'job_id, selling_amount, buying_amount, is_reimbursement'
                ).in_('job_id', job_ids).eq('is_reimbursement', True).execute()
                for c in costs_result.data:
                    jid = c['job_id']
                    reimb_rev_map[jid] = reimb_rev_map.get(jid, 0) + float(c.get('selling_amount') or 0)
                    reimb_cost_map[jid] = reimb_cost_map.get(jid, 0) + float(c.get('buying_amount') or 0)
            except Exception:
                pass

            for job in jobs:
                jid = job['job_id']
                reimb_rev = reimb_rev_map.get(jid, 0)
                reimb_cost = reimb_cost_map.get(jid, 0)
                # jobs.total_revenue / total_cost are kept by a DB trigger that
                # already excludes is_reimbursement rows — do NOT subtract again.
                total_rev = float(job.get('total_revenue') or 0)
                total_cost = float(job.get('total_cost') or 0)
                job['service_type'] = svc_map.get(jid)
                job['reimbursement_total'] = reimb_rev          # legacy name for BC
                job['reimbursement_cost_total'] = reimb_cost
                job['net_revenue'] = total_rev                  # already net of chi hộ
                job['net_cost'] = total_cost                    # already net of chi hộ
                job['profit'] = total_rev - total_cost

        return {"jobs": jobs}

    except Exception as e:
        logger.error(f"Error fetching recent jobs: {e}")
        return {"jobs": []}


@router.get("/{job_id}")
async def get_job(job_id: int):
    """
    Get job details
    """
    data_service = get_data_service()
    job = await data_service.get_job(job_id)
    
    if not job:
        raise HTTPException(404, f"Job {job_id} not found")
    
    return job


@router.get("/{job_id}/details")
async def get_job_details(job_id: int):
    """
    Get job with all services for detail modal
    Returns complete job info with vendor/employee names
    """
    try:
        client = get_supabase()

        # Get job info with customer details
        job_result = client.table('jobs').select(
            '*, customers(short_name, customer_code, company_name)'
        ).eq('job_id', job_id).limit(1).execute()

        if not job_result.data:
            raise HTTPException(404, f"Job {job_id} not found")

        job_row = job_result.data[0]
        customer = job_row.pop('customers', {}) or {}
        job = {
            **job_row,
            'customer_name': customer.get('short_name'),
            'customer_code': customer.get('customer_code'),
            'customer_full_name': customer.get('company_name'),
            'status_display': job_row.get('status_code')
        }

        # Get all services with vendor/employee/driver details
        svc_result = client.table('job_services').select(
            '*, vendors(short_name, company_name), employees(full_name), drivers(full_name, phone, id_card, license_plate, vehicle_type, date_of_birth)'
        ).eq('job_id', job_id).order('svc_id').execute()

        services = []
        for row in svc_result.data:
            vendor = row.pop('vendors', {}) or {}
            employee = row.pop('employees', {}) or {}
            driver = row.pop('drivers', {}) or {}
            svc = {
                **row,
                'vendor_name': vendor.get('short_name'),
                'vendor_full_name': vendor.get('company_name'),
                'employee_name': employee.get('full_name'),
                'service_type_name': row.get('service_type_code'),
                # Driver info from drivers table (linked via driver_id)
                'db_driver_name': driver.get('full_name'),
                'db_driver_phone': driver.get('phone'),
                'db_driver_id_card': driver.get('id_card'),
                'db_driver_license_plate': driver.get('license_plate'),
                'db_driver_vehicle_type': driver.get('vehicle_type'),
            }

            # Parse vendor_text_input for vehicle info
            if svc.get('vendor_text_input'):
                try:
                    vehicle_data = svc['vendor_text_input']
                    if isinstance(vehicle_data, str):
                        vehicle_data = json.loads(vehicle_data)

                    if isinstance(vehicle_data, dict):
                        svc['license_plate'] = vehicle_data.get('license_plate')
                        svc['driver_name'] = vehicle_data.get('driver_name')
                        svc['driver_phone'] = vehicle_data.get('driver_phone')
                        svc['driver_id_card'] = vehicle_data.get('driver_id_card')
                        # Keep vendor_name and vehicle_info separate for proper display
                except Exception as e:
                    logger.error(f"Error parsing vendor_text_input for svc {svc.get('svc_id')}: {e}")

            # Parse service_details JSONB for invoice/quantity/cargo info
            if svc.get('service_details'):
                try:
                    details = svc['service_details']
                    if isinstance(details, str):
                        details = json.loads(details)

                    if isinstance(details, dict):
                        # Extract invoice_numbers
                        inv = details.get('invoice_numbers')
                        if inv and not svc.get('invoice_numbers'):
                            if isinstance(inv, list):
                                svc['invoice_numbers'] = ', '.join(str(i) for i in inv)
                            else:
                                svc['invoice_numbers'] = inv

                        # Extract package quantity and unit
                        if details.get('package_quantity') and not svc.get('package_quantity'):
                            svc['package_quantity'] = details['package_quantity']
                        if details.get('package_unit') and not svc.get('package_unit'):
                            svc['package_unit'] = details['package_unit']

                        # Extract cargo_type
                        if details.get('cargo_type') and not svc.get('cargo_type'):
                            svc['cargo_type'] = details['cargo_type']

                        # Extract quotation/pricing fields
                        if details.get('buying_price') is not None:
                            svc['buying_price'] = details['buying_price']
                        if details.get('buying_rate_id') is not None:
                            svc['buying_rate_id'] = details['buying_rate_id']
                        if details.get('selling_price') is not None:
                            svc['selling_price'] = details['selling_price']
                        if details.get('selling_rate_id') is not None:
                            svc['selling_rate_id'] = details['selling_rate_id']

                        # Extract extra costs, revenues, and custom info
                        if details.get('extra_costs'):
                            svc['extra_costs'] = details['extra_costs']
                        if details.get('extra_revenues'):
                            svc['extra_revenues'] = details['extra_revenues']
                        if details.get('extra_info'):
                            svc['extra_info'] = details['extra_info']

                        # Extract calculated profit
                        if details.get('profit') is not None:
                            svc['profit'] = details['profit']
                        if details.get('total_cost') is not None:
                            svc['total_cost'] = details['total_cost']
                        if details.get('total_revenue') is not None:
                            svc['total_revenue'] = details['total_revenue']

                        # Extract SEA_DOM specific fields
                        for sea_field in SEA_DOM_FIELDS:
                            if details.get(sea_field) is not None:
                                svc[sea_field] = details[sea_field]
                except Exception as e:
                    logger.error(f"Error parsing service_details for svc {svc.get('svc_id')}: {e}")

            services.append(svc)

        # Get job costs if available
        costs = []
        try:
            costs_result = client.table('job_costs').select(
                '*, vendors(short_name)'
            ).eq('job_id', job_id).order('cost_id').execute()

            for row in costs_result.data:
                vendor = row.pop('vendors', {}) or {}
                costs.append({
                    **row,
                    'vendor_name': vendor.get('short_name')
                })
        except Exception:
            pass  # job_costs table might not exist yet

        return {
            "job": job,
            "services": services,
            "costs": costs
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching job details: {e}")
        raise HTTPException(500, str(e))



@router.get("/pending/list")
async def list_pending_jobs():
    """
    List pending jobs (waiting for vehicle assignment)
    """
    data_service = get_data_service()
    # TODO: Implement
    return {"jobs": []}


@router.get("/find")
async def find_job(
    job_no: Optional[str] = None,
    invoice: Optional[str] = None,
    bl_awb: Optional[str] = None,
    customer_code: Optional[str] = None,
    date: Optional[str] = None
):
    """
    Find job by job_no, invoice, B/L, or customer+date
    """
    try:
        client = get_supabase()

        # Priority 1: Job number
        if job_no:
            result = client.table('jobs').select(
                '*, customers(short_name)'
            ).ilike('job_no', f'%{job_no}%').limit(1).execute()

            if result.data:
                job = result.data[0]
                customer = job.pop('customers', {}) or {}
                job['customer_name'] = customer.get('short_name')
                return {"found": True, "job": job}

        # Priority 2: Invoice number - need to search in job_services
        if invoice:
            svc_result = client.table('job_services').select(
                'job_id'
            ).ilike('invoice_numbers', f'%{invoice}%').limit(1).execute()

            if svc_result.data:
                job_id = svc_result.data[0]['job_id']
                result = client.table('jobs').select(
                    '*, customers(short_name)'
                ).eq('job_id', job_id).limit(1).execute()

                if result.data:
                    job = result.data[0]
                    customer = job.pop('customers', {}) or {}
                    job['customer_name'] = customer.get('short_name')
                    return {"found": True, "job": job}

        # Priority 3: B/L or AWB
        if bl_awb:
            svc_result = client.table('job_services').select(
                'job_id'
            ).ilike('bl_awb_no', f'%{bl_awb}%').limit(1).execute()

            if svc_result.data:
                job_id = svc_result.data[0]['job_id']
                result = client.table('jobs').select(
                    '*, customers(short_name)'
                ).eq('job_id', job_id).limit(1).execute()

                if result.data:
                    job = result.data[0]
                    customer = job.pop('customers', {}) or {}
                    job['customer_name'] = customer.get('short_name')
                    return {"found": True, "job": job}

        # Priority 4: Customer + date
        if customer_code and date:
            # First find customer
            cust_result = client.table('customers').select(
                'customer_id'
            ).ilike('customer_code', f'%{customer_code}%').limit(1).execute()

            if cust_result.data:
                cust_id = cust_result.data[0]['customer_id']
                result = client.table('jobs').select(
                    '*, customers(short_name)'
                ).eq('customer_id', cust_id).eq('etd', date).order(
                    'created_at', desc=True
                ).limit(1).execute()

                if result.data:
                    job = result.data[0]
                    customer = job.pop('customers', {}) or {}
                    job['customer_name'] = customer.get('short_name')
                    return {"found": True, "job": job}

        return {"found": False, "message": "Job not found"}
    except Exception as e:
        logger.error(f"Error finding job: {e}")
        return {"found": False, "message": str(e)}


@router.post("/update")
async def update_job_info(request: Request):
    """
    Update job info and/or status
    """
    data = await request.json()
    entities = data.get("entities", {})

    try:
        client = get_supabase()

        # Find the job first
        job_id = None
        job_no = entities.get("job_number")
        invoice_ref = entities.get("invoice_ref")
        bl_awb_ref = entities.get("bl_awb_ref")

        if job_no:
            result = client.table('jobs').select(
                'job_id, job_no'
            ).ilike('job_no', f'%{job_no}%').limit(1).execute()

            if result.data:
                job_id = result.data[0]["job_id"]
                job_no = result.data[0]["job_no"]

        if not job_id and invoice_ref:
            svc_result = client.table('job_services').select(
                'job_id'
            ).ilike('invoice_numbers', f'%{invoice_ref}%').limit(1).execute()

            if svc_result.data:
                job_id = svc_result.data[0]["job_id"]
                job_result = client.table('jobs').select('job_no').eq('job_id', job_id).limit(1).execute()
                if job_result.data:
                    job_no = job_result.data[0]["job_no"]

        if not job_id and bl_awb_ref:
            svc_result = client.table('job_services').select(
                'job_id'
            ).ilike('bl_awb_no', f'%{bl_awb_ref}%').limit(1).execute()

            if svc_result.data:
                job_id = svc_result.data[0]["job_id"]
                job_result = client.table('jobs').select('job_no').eq('job_id', job_id).limit(1).execute()
                if job_result.data:
                    job_no = job_result.data[0]["job_no"]

        if not job_id:
            return {"success": False, "message": "Không tìm thấy job. Vui lòng cung cấp job_number, invoice hoặc B/L."}

        # Update status if provided - sync both job and services
        new_status = entities.get("new_status")
        if new_status:
            valid_statuses = ["PENDING", "CONFIRMED", "DISPATCHED", "IN_TRANSIT", "ASSIGNED", "COMPLETED", "CANCELLED"]
            if new_status.upper() in valid_statuses:
                # Update job status
                job_update = {
                    'status_code': new_status.upper(),
                    'updated_at': datetime.now().isoformat()
                }
                update_user_id = entities.get("user_id")
                if update_user_id:
                    job_update['updated_by'] = update_user_id
                client.table('jobs').update(job_update).eq('job_id', job_id).execute()

                # Also update all job_services status to keep in sync
                client.table('job_services').update({
                    'status_code': new_status.upper(),
                    'updated_at': datetime.now().isoformat()
                }).eq('job_id', job_id).execute()

                logger.info(f"Updated job {job_no} and services status to {new_status}")

        # Update job_services fields
        svc_updates = {}

        if entities.get("update_pickup_time"):
            svc_updates['scheduled_time'] = entities["update_pickup_time"]

        if entities.get("update_delivery_address"):
            svc_updates['dest_address'] = entities["update_delivery_address"]

        if entities.get("update_special_requirements"):
            svc_updates['special_requirements'] = entities["update_special_requirements"]

        # Vehicle info update
        if entities.get("license_plate") or entities.get("driver_name"):
            vendor_text = json.dumps({
                'license_plate': entities.get('license_plate', ''),
                'driver_name': entities.get('driver_name', ''),
                'driver_phone': entities.get('driver_phone', '')
            })
            svc_updates['vendor_text_input'] = vendor_text

        if svc_updates:
            svc_updates['updated_at'] = datetime.now().isoformat()
            client.table('job_services').update(svc_updates).eq('job_id', job_id).execute()

        # Handle update_notes separately (append to special_requirements)
        if entities.get("update_notes"):
            # Get current special_requirements
            current = client.table('job_services').select(
                'svc_id, special_requirements'
            ).eq('job_id', job_id).execute()

            for svc in current.data:
                current_req = svc.get('special_requirements') or ''
                new_req = f"{current_req} {entities['update_notes']}".strip()
                client.table('job_services').update({
                    'special_requirements': new_req,
                    'updated_at': datetime.now().isoformat()
                }).eq('svc_id', svc['svc_id']).execute()

        return {
            "success": True,
            "job_id": job_id,
            "job_number": job_no,
            "message": f"Đã cập nhật job {job_no} thành công!"
        }

    except Exception as e:
        logger.error(f"Error updating job: {e}")
        return {"success": False, "message": str(e)}


# ========================================
# DASHBOARD ENDPOINTS
# ========================================

# Dashboard stats endpoint - mounted at /api/dashboard/stats in main.py
async def get_dashboard_stats():
    """
    Get dashboard statistics via PostgreSQL RPC (fast, single query)
    """
    try:
        client = get_supabase()
        result = client.rpc('get_dashboard_stats').execute()
        stats = result.data if result.data else {}

        # Format revenue
        revenue_total = float(stats.get('revenue_total') or 0)
        if revenue_total >= 1_000_000_000:
            revenue_mtd = f"{revenue_total / 1_000_000_000:.1f}B"
        elif revenue_total >= 1_000_000:
            revenue_mtd = f"{revenue_total / 1_000_000:.0f}M"
        elif revenue_total >= 1_000:
            revenue_mtd = f"{revenue_total / 1_000:.0f}K"
        else:
            revenue_mtd = f"{int(revenue_total)}"

        return {
            "jobs_today": stats.get('jobs_today', 0),
            "trucking": stats.get('trucking', 0),
            "sea": stats.get('sea', 0),
            "air": stats.get('air', 0),
            "revenue": revenue_mtd,
            "status_counts": stats.get('status_counts', {})
        }

    except Exception as e:
        logger.error(f"Error fetching dashboard stats: {e}")
        return {
            "jobs_today": 0,
            "trucking": 0,
            "sea": 0,
            "air": 0,
            "revenue": "0",
            "status_counts": {}
        }


# Service-specific endpoints - mounted at /api/services/{type} in main.py
async def get_service_data(service_type: str):
    """
    Get service-specific data from views
    """
    service_type_codes = {
        # Logistics - Road Transport
        "trucking": ['TRUCKING_DOM', 'BORDER_IMP'],
        "container": ['LIFT_ON', 'LIFT_OFF'],
        # Air Freight
        "air": ['AIR_IMP', 'AIR_EXP'],
        # Sea Freight
        "sea": ['SEA_IMP', 'SEA_EXP'],
        # Warehouse
        "warehouse": ['WHS_STORAGE', 'WHS_VAS'],
        "handling": ['WHS_HANDLE'],
        # Customs
        "customs": ['CUS_IMPORT', 'CUS_EXPORT', 'CUS_TRANSIT'],
        "co": ['CUS_CO'],
        # Value-Added
        "packing": ['SVC_PACK'],
        "special": ['SVC_FUMI', 'SVC_VACUUM', 'SVC_SHRINK', 'SVC_LASHING', 'SVC_LASHING_TRUCK', 'SVC_LASHING_CONT', 'SVC_LASHING_FR'],
    }

    codes = service_type_codes.get(service_type)
    if not codes:
        return {"services": [], "error": f"Unknown service type: {service_type}"}

    try:
        client = get_supabase()

        # Query job_services with related data
        result = client.table('job_services').select(
            '*, jobs(job_no, customer_id, customers(customer_code, short_name, company_name)), '
            'vendors(short_name), employees(full_name)'
        ).in_('service_type_code', codes).order(
            'scheduled_date', desc=True, nullsfirst=False
        ).limit(50).execute()

        services = []
        for row in result.data:
            job = row.pop('jobs', {}) or {}
            vendor = row.pop('vendors', {}) or {}
            employee = row.pop('employees', {}) or {}
            customer = job.pop('customers', {}) or {} if job else {}

            # Determine assignment type
            if row.get('vendor_id'):
                assignment_type = 'VENDOR'
                assigned_to = vendor.get('short_name')
            elif row.get('employee_id'):
                assignment_type = 'EMPLOYEE'
                assigned_to = employee.get('full_name')
            else:
                assignment_type = 'UNASSIGNED'
                assigned_to = None

            svc = {
                **row,
                'job_no': job.get('job_no'),
                'customer_code': customer.get('customer_code'),
                'customer': customer.get('short_name') or customer.get('company_name'),
                'vendor_name': vendor.get('short_name'),
                'employee_name': employee.get('full_name'),
                'assignment_type': assignment_type,
                'assigned_to': assigned_to
            }

            # Parse vehicle info for trucking services
            if service_type == "trucking" and svc.get('vendor_text_input'):
                try:
                    vehicle_data = svc['vendor_text_input']
                    if isinstance(vehicle_data, str):
                        vehicle_data = json.loads(vehicle_data)

                    if isinstance(vehicle_data, dict):
                        svc['license_plate'] = vehicle_data.get('license_plate')
                        svc['driver_name'] = vehicle_data.get('driver_name')

                        # Update assigned_to if no vendor assigned
                        if not svc.get('assigned_to') and svc.get('license_plate'):
                            info = f"{svc['license_plate']}"
                            if svc.get('driver_name'):
                                info += f" - {svc['driver_name']}"

                            svc['assigned_to'] = info
                            svc['assignment_type'] = 'VENDOR'

                except Exception as e:
                    logger.error(f"Error parsing vendor_text_input: {e}")

            services.append(svc)

        return {"services": services}

    except Exception as e:
        logger.error(f"Error fetching {service_type} services: {e}")
        return {"services": [], "error": str(e)}


# ========================================
# EXCEL EXPORT ENDPOINT
# ========================================

@router.get("/export/{service_type}")
async def export_services_excel(
    service_type: str,
    customer_id: Optional[int] = None,
    vendor_id: Optional[int] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    month: Optional[str] = None  # Format: 2026-01
):
    """
    Export services to Excel with filters
    Returns Excel file with structured fields (extracted from JSON) for payment reconciliation
    """
    from fastapi.responses import FileResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import tempfile

    valid_types = ["trucking", "container", "warehouse", "handling", "customs", "co", "packing", "special", "air", "sea"]
    if service_type not in valid_types:
        raise HTTPException(400, f"Unknown service type: {service_type}")

    service_type_codes = {
        # Logistics - Road Transport
        "trucking": ['TRUCKING_DOM', 'BORDER_IMP'],
        "container": ['LIFT_ON', 'LIFT_OFF'],
        # Air Freight
        "air": ['AIR_IMP', 'AIR_EXP'],
        # Sea Freight
        "sea": ['SEA_IMP', 'SEA_EXP'],
        # Warehouse
        "warehouse": ['WHS_STORAGE', 'WHS_VAS'],
        "handling": ['WHS_HANDLE'],
        # Customs
        "customs": ['CUS_IMPORT', 'CUS_EXPORT', 'CUS_TRANSIT'],
        "co": ['CUS_CO'],
        # Value-Added
        "packing": ['SVC_PACK'],
        "special": ['SVC_FUMI', 'SVC_VACUUM', 'SVC_SHRINK', 'SVC_LASHING', 'SVC_LASHING_TRUCK', 'SVC_LASHING_CONT', 'SVC_LASHING_FR'],
    }

    try:
        client = get_supabase()

        # Build query with Supabase SDK
        query = client.table('job_services').select(
            '*, jobs(job_no, customer_id, customers(customer_code, short_name)), vendors(short_name, company_name)'
        ).in_('service_type_code', service_type_codes[service_type])

        # Apply filters
        if vendor_id:
            query = query.eq('vendor_id', vendor_id)

        if month:
            # Filter by month (YYYY-MM format)
            year, mon = month.split('-')
            start_date = f"{year}-{mon}-01"
            # Calculate end of month
            import calendar
            last_day = calendar.monthrange(int(year), int(mon))[1]
            end_date = f"{year}-{mon}-{last_day}"
            query = query.gte('scheduled_date', start_date).lte('scheduled_date', end_date)
        else:
            if from_date:
                query = query.gte('scheduled_date', from_date)
            if to_date:
                query = query.lte('scheduled_date', to_date)

        query = query.order('scheduled_date', desc=True)
        result = query.execute()

        # Filter by customer_id in Python (since it's in the joined table)
        raw_data = result.data
        if customer_id:
            raw_data = [r for r in raw_data if r.get('jobs', {}).get('customer_id') == customer_id]

        if not raw_data:
            raise HTTPException(404, "Không có dữ liệu để xuất")

        # Transform data based on service type
        services = []
        for row in raw_data:
            job = row.get('jobs') or {}
            vendor = row.get('vendors') or {}
            customer = job.get('customers') or {}

            # Parse JSON fields
            service_details = row.get('service_details') or {}
            if isinstance(service_details, str):
                try:
                    service_details = json.loads(service_details)
                except:
                    service_details = {}

            vendor_text = row.get('vendor_text_input') or {}
            if isinstance(vendor_text, str):
                try:
                    vendor_text = json.loads(vendor_text)
                except:
                    vendor_text = {}

            if service_type == "trucking":
                svc = {
                    'job_no': job.get('job_no'),
                    'customer_code': customer.get('customer_code'),
                    'customer_name': customer.get('short_name'),
                    'ngay_thuc_hien': row.get('scheduled_date'),
                    'gio_lay_hang': row.get('scheduled_time'),
                    'diem_di': row.get('origin_address'),
                    'diem_den': row.get('dest_address'),
                    'loai_hang': row.get('cargo_type'),
                    'so_kien': service_details.get('package_quantity') or row.get('package_quantity'),
                    'don_vi': service_details.get('package_unit') or row.get('package_unit'),
                    'trong_luong_kg': row.get('weight_kg'),
                    'invoice': service_details.get('invoice_numbers') or row.get('invoice_numbers'),
                    'bien_so_xe': vendor_text.get('license_plate') if isinstance(vendor_text, dict) else None,
                    'ten_tai_xe': vendor_text.get('driver_name') if isinstance(vendor_text, dict) else None,
                    'sdt_tai_xe': vendor_text.get('driver_phone') if isinstance(vendor_text, dict) else None,
                    'cccd_tai_xe': vendor_text.get('driver_id_card') if isinstance(vendor_text, dict) else None,
                    'nha_van_chuyen': vendor.get('short_name'),
                    'ten_cong_ty_vc': vendor.get('company_name'),
                    'trang_thai': row.get('status_code'),
                    'ngay_tao': row.get('created_at'),
                    'ngay_cap_nhat': row.get('updated_at')
                }
            elif service_type == "warehouse":
                svc = {
                    'job_no': job.get('job_no'),
                    'customer_code': customer.get('customer_code'),
                    'customer_name': customer.get('short_name'),
                    'scheduled_date': row.get('scheduled_date'),
                    'ngay_bat_dau': row.get('storage_start_date'),
                    'ngay_ket_thuc': row.get('storage_end_date'),
                    'loai_hang': row.get('cargo_type'),
                    'so_kien': row.get('package_quantity'),
                    'don_vi': row.get('package_unit'),
                    'trong_luong_kg': row.get('weight_kg'),
                    'dai_cm': row.get('dimension_length_cm'),
                    'rong_cm': row.get('dimension_width_cm'),
                    'cao_cm': row.get('dimension_height_cm'),
                    'yeu_cau_dac_biet': row.get('special_requirements'),
                    'nha_cung_cap': vendor.get('short_name'),
                    'trang_thai': row.get('status_code'),
                    'ngay_tao': row.get('created_at')
                }
            elif service_type == "customs":
                svc = {
                    'job_no': job.get('job_no'),
                    'customer_code': customer.get('customer_code'),
                    'customer_name': customer.get('short_name'),
                    'scheduled_date': row.get('scheduled_date'),
                    'so_to_khai': row.get('cd_no'),
                    'ngay_to_khai': row.get('declaration_datetime'),
                    'loai_hinh': row.get('loai_hinh'),
                    'loai_hai_quan': row.get('customs_type'),
                    'cua_khau': row.get('customs_port'),
                    'nguoi_mua': row.get('buyer_name'),
                    'nguoi_ban': row.get('seller_name'),
                    'hs_code': row.get('hs_code'),
                    'so_van_don': row.get('bl_awb_no'),
                    'so_co': row.get('co_no'),
                    'nha_cung_cap': vendor.get('short_name'),
                    'trang_thai': row.get('status_code'),
                    'ngay_tao': row.get('created_at')
                }
            else:  # packing
                svc = {
                    'job_no': job.get('job_no'),
                    'customer_code': customer.get('customer_code'),
                    'customer_name': customer.get('short_name'),
                    'scheduled_date': row.get('scheduled_date'),
                    'loai_dong_goi': row.get('packing_type'),
                    'so_mat_hang': row.get('items_count'),
                    'so_kien_xuat': row.get('packages_output'),
                    'dai_truoc_cm': row.get('before_length_cm'),
                    'rong_truoc_cm': row.get('before_width_cm'),
                    'cao_truoc_cm': row.get('before_height_cm'),
                    'dai_sau_cm': row.get('after_length_cm'),
                    'rong_sau_cm': row.get('after_width_cm'),
                    'cao_sau_cm': row.get('after_height_cm'),
                    'co_shrink_wrap': row.get('shrink_wrap'),
                    'co_hut_chan_khong': row.get('vacuum_pack'),
                    'co_chong_buoc': row.get('lashing'),
                    'co_xong_khu_trung': row.get('fumigation'),
                    'nha_cung_cap': vendor.get('short_name'),
                    'trang_thai': row.get('status_code'),
                    'ngay_tao': row.get('created_at')
                }

            services.append(svc)

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{service_type.capitalize()} Export"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Get column names from first row
        columns = list(services[0].keys()) if services else []

        # Vietnamese header mapping for better readability
        vn_headers = {
            'job_no': 'Mã Job',
            'customer_code': 'Mã KH',
            'customer_name': 'Tên KH',
            'ngay_thuc_hien': 'Ngày TH',
            'gio_lay_hang': 'Giờ lấy',
            'diem_di': 'Điểm đi',
            'diem_den': 'Điểm đến',
            'loai_hang': 'Loại hàng',
            'so_kien': 'Số kiện',
            'don_vi': 'Đơn vị',
            'trong_luong_kg': 'KL (kg)',
            'invoice': 'Invoice',
            'bien_so_xe': 'Biển số xe',
            'ten_tai_xe': 'Tài xế',
            'sdt_tai_xe': 'SĐT tài xế',
            'cccd_tai_xe': 'CCCD',
            'nha_van_chuyen': 'NVC',
            'ten_cong_ty_vc': 'Công ty VC',
            'trang_thai': 'Trạng thái',
            'ngay_tao': 'Ngày tạo',
            'ngay_cap_nhat': 'Cập nhật'
        }

        # Write headers
        for col_idx, col_name in enumerate(columns, 1):
            header_text = vn_headers.get(col_name, col_name.replace('_', ' ').title())
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Write data
        for row_idx, row in enumerate(services, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = row.get(col_name)
                # Convert special types
                if isinstance(value, (date, datetime)):
                    value = value.strftime('%Y-%m-%d') if hasattr(value, 'strftime') else str(value)
                elif isinstance(value, time):
                    value = value.strftime('%H:%M') if hasattr(value, 'strftime') else str(value)
                # Parse JSON array for invoice_numbers
                elif col_name == 'invoice' and value:
                    if isinstance(value, list):
                        value = ', '.join(str(i) for i in value)
                    elif isinstance(value, str) and value.startswith('['):
                        try:
                            inv_list = json.loads(value)
                            value = ', '.join(str(i) for i in inv_list)
                        except:
                            pass

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Auto-adjust column widths
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(vn_headers.get(col_name, col_name))
            for row in services[:50]:  # Check first 50 rows
                val = row.get(col_name)
                if val:
                    max_length = max(max_length, min(len(str(val)), 40))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Save to temp file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()

        # Generate filename
        filename = f"{service_type}_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return FileResponse(
            path=temp_file.name,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            background=None  # Cleanup handled by FileResponse
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting {service_type} services: {e}")
        logger.error(traceback.format_exc())
        raise HTTPException(500, f"Export error: {str(e)}")


# ========================================
# Job Status Management
# ========================================

class StatusUpdateRequest(BaseModel):
    status_code: str


class JobNumberUpdateRequest(BaseModel):
    job_no: str


@router.put("/{job_id}/job-number")
async def update_job_number(job_id: int, request: JobNumberUpdateRequest):
    """
    Update job_no for a job. No admin auth required (operators can use this).
    Validates format loosely: must be non-empty string.
    """
    try:
        client = get_supabase()

        new_job_no = request.job_no.strip()
        if not new_job_no:
            return {"success": False, "message": "job_no không được để trống"}

        # Check job exists
        job_result = client.table('jobs').select(
            'job_id, job_no'
        ).eq('job_id', job_id).limit(1).execute()

        if not job_result.data:
            raise HTTPException(404, f"Job {job_id} not found")

        old_job_no = job_result.data[0]['job_no']

        # Check uniqueness (avoid duplicate job_no)
        dup_result = client.table('jobs').select('job_id').eq('job_no', new_job_no).neq('job_id', job_id).limit(1).execute()
        if dup_result.data:
            return {"success": False, "message": f"Mã job '{new_job_no}' đã tồn tại"}

        client.table('jobs').update({
            'job_no': new_job_no,
            'updated_at': datetime.now().isoformat()
        }).eq('job_id', job_id).execute()

        logger.info(f"Job {job_id} job_no updated: {old_job_no} -> {new_job_no}")

        return {
            "success": True,
            "job_id": job_id,
            "job_no": new_job_no,
            "message": f"Đã cập nhật mã job thành {new_job_no}"
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating job_no: {e}")
        return {"success": False, "message": str(e)}


@router.put("/{job_id}/status")
async def update_job_status(job_id: int, request: StatusUpdateRequest):
    """
    Update job status and sync all services.
    """
    try:
        client = get_supabase()

        # Get current job
        job_result = client.table('jobs').select(
            'job_id, job_no, status_code'
        ).eq('job_id', job_id).limit(1).execute()

        if not job_result.data:
            raise HTTPException(404, f"Job {job_id} not found")

        job = job_result.data[0]
        new_status = request.status_code.upper()

        # Valid statuses
        valid_statuses = ['PENDING', 'CONFIRMED', 'DISPATCHED', 'IN_TRANSIT', 'ASSIGNED', 'COMPLETED', 'CANCELLED']
        if new_status not in valid_statuses:
            return {"success": False, "message": f"Trạng thái '{new_status}' không hợp lệ"}

        # Update job status
        client.table('jobs').update({
            'status_code': new_status,
            'updated_at': datetime.now().isoformat()
        }).eq('job_id', job_id).execute()

        # Sync services status
        client.table('job_services').update({
            'status_code': new_status,
            'updated_at': datetime.now().isoformat()
        }).eq('job_id', job_id).execute()

        logger.info(f"Job {job['job_no']} status updated: {job['status_code']} -> {new_status}")

        return {
            "success": True,
            "job_id": job_id,
            "status_code": new_status,
            "message": f"Đã cập nhật trạng thái thành {new_status}"
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating status: {e}")
        return {"success": False, "message": str(e)}


@router.delete("/{job_id}/cancel")
async def cancel_job(job_id: int):
    """
    Cancel a job and all its services.
    """
    try:
        client = get_supabase()

        # Get current job
        job_result = client.table('jobs').select(
            'job_id, job_no, status_code'
        ).eq('job_id', job_id).limit(1).execute()

        if not job_result.data:
            raise HTTPException(404, f"Job {job_id} not found")

        job = job_result.data[0]

        # Prevent cancelling completed jobs
        if job['status_code'] == 'COMPLETED':
            return {"success": False, "message": "Không thể hủy job đã hoàn thành"}

        # Update job status to CANCELLED
        client.table('jobs').update({
            'status_code': 'CANCELLED',
            'updated_at': datetime.now().isoformat()
        }).eq('job_id', job_id).execute()

        # Cancel all services
        client.table('job_services').update({
            'status_code': 'CANCELLED',
            'updated_at': datetime.now().isoformat()
        }).eq('job_id', job_id).execute()

        logger.info(f"Job {job['job_no']} cancelled")

        return {
            "success": True,
            "job_id": job_id,
            "message": f"Đã hủy job {job['job_no']}"
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error cancelling job: {e}")
        return {"success": False, "message": str(e)}


# ========================================
# Customer and Service Management
# ========================================

class CustomerChangeRequest(BaseModel):
    customer_id: int
    customer_code: str
    reason: Optional[str] = None


class AddServiceRequest(BaseModel):
    service_type_code: str
    scheduled_date: Optional[str] = None
    scheduled_time: Optional[str] = None
    origin_address: Optional[str] = None
    dest_address: Optional[str] = None
    cargo_type: Optional[str] = None
    package_quantity: Optional[int] = None
    package_unit: Optional[str] = None
    special_requirements: Optional[str] = None
    service_details: Optional[Dict[str, Any]] = None  # For fee_amount, notes, etc.
    # Vendor and vehicle fields
    vendor_id: Optional[int] = None
    license_plate: Optional[str] = None
    driver_name: Optional[str] = None
    driver_phone: Optional[str] = None
    notes: Optional[str] = None
    # Customs declaration code — required by validator when service_type_code is CUS/CUS_IMPORT/CUS_EXPORT.
    loai_hinh: Optional[str] = None
    cd_no: Optional[str] = None  # Tờ khai number


@router.put("/{job_id}/customer")
async def change_job_customer(job_id: int, request: CustomerChangeRequest):
    """
    Change customer for a job.
    Requires confirmation - logs change for audit.
    """
    try:
        client = get_supabase()

        # Get current job
        job_result = client.table('jobs').select(
            'job_id, job_no, customer_id, status_code'
        ).eq('job_id', job_id).limit(1).execute()

        if not job_result.data:
            raise HTTPException(404, f"Job {job_id} not found")

        job = job_result.data[0]
        old_customer_id = job['customer_id']

        # Prevent change if job is completed/cancelled
        if job['status_code'] in ['COMPLETED', 'CANCELLED']:
            return {
                "success": False,
                "message": f"Không thể đổi KH cho job đã {job['status_code']}"
            }

        # Update job
        client.table('jobs').update({
            'customer_id': request.customer_id,
            'updated_at': datetime.now().isoformat()
        }).eq('job_id', job_id).execute()

        # Log change for audit
        logger.info(
            f"Job {job['job_no']} customer changed: "
            f"{old_customer_id} -> {request.customer_id} "
            f"(reason: {request.reason or 'not specified'})"
        )

        return {
            "success": True,
            "job_id": job_id,
            "message": "Đã đổi khách hàng thành công"
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error changing customer: {e}")
        return {"success": False, "message": str(e)}


@router.post("/{job_id}/services")
async def add_job_service(job_id: int, request: AddServiceRequest, req: Request):
    """
    Add new service to existing job.
    Requires authenticated user so created_by audit trail is populated.
    """
    try:
        client = get_supabase()

        # Auth — required so we can set created_by. Frontend authFetch injects the
        # Bearer token automatically; unsigned requests get rejected.
        current_user = await get_current_user_optional(req)
        if not current_user:
            raise HTTPException(status_code=401, detail="Yêu cầu đăng nhập để thêm dịch vụ.")
        user_id = current_user['user_id']

        # --- VALIDATOR: customs services MUST have a valid loai_hinh ---
        # Uses shared module so endpoint + service layer + DB constraint stay in sync.
        loai_hinh_error = customs_validator.validate_loai_hinh_for_service(
            request.service_type_code, request.loai_hinh
        )
        if loai_hinh_error is not None:
            return {
                "success": False,
                "message": loai_hinh_error["message"],
                "error": loai_hinh_error["error"],
                "suggestions": loai_hinh_error["suggestions"],
            }

        # Verify job exists
        job_result = client.table('jobs').select(
            'job_id, job_no, status_code'
        ).eq('job_id', job_id).limit(1).execute()

        if not job_result.data:
            raise HTTPException(404, f"Job {job_id} not found")

        job = job_result.data[0]

        # Prevent adding to completed/cancelled jobs
        if job['status_code'] in ['COMPLETED', 'CANCELLED']:
            return {
                "success": False,
                "message": f"Không thể thêm dịch vụ cho job đã {job['status_code']}"
            }

        # Parse dates
        scheduled_date = None
        if request.scheduled_date:
            try:
                scheduled_date = parse_date(request.scheduled_date)
            except:
                scheduled_date = request.scheduled_date

        scheduled_time = None
        if request.scheduled_time:
            try:
                scheduled_time = parse_time(request.scheduled_time)
            except:
                scheduled_time = request.scheduled_time

        # Insert new service
        service_data = {
            'job_id': job_id,
            'service_type_code': normalize_service_code(request.service_type_code),
            'origin_address': request.origin_address,
            'dest_address': request.dest_address,
            'cargo_type': request.cargo_type,
            'package_quantity': request.package_quantity,
            'package_unit': request.package_unit,
            'special_requirements': request.special_requirements,
            'status_code': 'PENDING',
            # Audit — who added this service
            'created_by': user_id,
            'updated_by': user_id,
            # Customs declaration fields (normalized uppercase, null when empty)
            'loai_hinh': customs_validator.normalize_loai_hinh(request.loai_hinh) or None,
            'cd_no': (request.cd_no or '').strip() or None,
        }

        # Add vendor if provided
        if request.vendor_id:
            service_data['vendor_id'] = request.vendor_id

        # Add vehicle/driver info as JSON in vendor_text_input
        if request.license_plate or request.driver_name or request.driver_phone:
            vehicle_info = {
                'license_plate': request.license_plate,
                'driver_name': request.driver_name,
                'driver_phone': request.driver_phone
            }
            service_data['vendor_text_input'] = json.dumps(vehicle_info)

        # Add notes to special_requirements if provided
        if request.notes:
            if service_data.get('special_requirements'):
                service_data['special_requirements'] += f"\n{request.notes}"
            else:
                service_data['special_requirements'] = request.notes

        # Add service_details for fee services
        if request.service_details:
            service_data['service_details'] = request.service_details

        if scheduled_date:
            service_data['scheduled_date'] = scheduled_date.isoformat() if hasattr(scheduled_date, 'isoformat') else str(scheduled_date)
        if scheduled_time:
            service_data['scheduled_time'] = str(scheduled_time)

        service_result = client.table('job_services').insert(service_data).execute()

        if not service_result.data:
            return {"success": False, "message": "Không thể thêm dịch vụ"}

        new_service = service_result.data[0]

        logger.info(f"Added service {new_service['svc_id']} to job {job['job_no']}")

        return {
            "success": True,
            "svc_id": new_service['svc_id'],
            "message": f"Đã thêm dịch vụ {request.service_type_code}"
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error adding service: {e}")
        return {"success": False, "message": str(e)}


# ═══════════════════════════════════════════════════════════════════════════════
# QUOTATION ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/quotations/search")
async def search_quotations(
    type: str = "buying",
    service_type: Optional[str] = None,
    customer_id: Optional[int] = None,
    vendor_id: Optional[int] = None
):
    """
    Search for matching quotations based on service parameters.
    type=buying returns vendor_rates, type=selling returns customer_rates.
    Client-side search handles text filtering with diacritic normalization.
    """
    try:
        client = get_supabase()

        if type == "buying":
            # Search vendor_rates (exclude 0-price rates)
            query = client.table('vendor_rates').select(
                '*, vendors(short_name, company_name)'
            ).eq('is_active', True).gt('price', 0)

            if service_type:
                query = query.or_(f'service_type_code.eq.{service_type},service_type_code.is.null')
            if vendor_id:
                query = query.eq('vendor_id', vendor_id)

            result = query.order('price').limit(200).execute()

            rates = []
            for r in result.data:
                vendor = r.pop('vendors', {}) or {}
                rates.append({
                    'rate_id': r.get('id'),  # PK is 'id' in production DB
                    'vendor_id': r.get('vendor_id'),
                    'vendor_name': vendor.get('short_name') or vendor.get('company_name'),
                    'price': r.get('price'),
                    'vehicle_type': r.get('vehicle_type'),
                    'origin': r.get('origin_province'),
                    'destination': r.get('destination_province'),
                    'unit': r.get('unit', 'TRIP'),
                    'service_type_code': r.get('service_type_code'),
                    'notes': r.get('notes'),
                })

        else:
            # Search customer_rates (exclude 0-price rates)
            query = client.table('customer_rates').select(
                '*, customers(short_name, customer_code)'
            ).eq('is_active', True).gt('price', 0)

            if customer_id:
                query = query.eq('customer_id', customer_id)
            if service_type:
                query = query.or_(f'service_type_code.eq.{service_type},service_type_code.is.null')

            result = query.order('price', desc=True).limit(200).execute()

            rates = []
            for r in result.data:
                customer = r.pop('customers', {}) or {}
                rates.append({
                    'rate_id': r.get('id'),  # PK is 'id' in production DB
                    'customer_id': r.get('customer_id'),
                    'customer_name': customer.get('short_name') or customer.get('customer_code'),
                    'price': r.get('price'),
                    'vehicle_type': r.get('vehicle_type'),
                    'origin': r.get('origin_province'),
                    'destination': r.get('destination_province'),
                    'unit': r.get('unit', 'TRIP'),
                    'service_type_code': r.get('service_type_code'),
                })

        return {"rates": rates}

    except Exception as e:
        logger.error(f"Error searching quotations: {e}")
        return {"rates": [], "error": str(e)}


class ServiceQuotationRequest(BaseModel):
    buying_rate_id: Optional[int] = None
    buying_price: Optional[float] = None
    selling_rate_id: Optional[int] = None
    selling_price: Optional[float] = None
    extra_costs: Optional[list] = None  # [{name, amount, vendor?, unit_price?, quantity?, unit?, currency?, exchange_rate?}, ...]
    extra_revenues: Optional[list] = None  # [{name, amount, vendor?, unit_price?, quantity?, unit?, currency?, exchange_rate?}, ...]


def _sync_quotation_to_job_costs(client, job_id: int, svc_id: int, request):
    """Sync service quotation data → job_costs table.
    Deletes old costs for this svc_id and re-inserts from quotation data.
    This ensures jobs.total_revenue/total_cost (generated from job_costs) stay accurate.
    """
    try:
        # Delete existing costs linked to this service
        client.table('job_costs').delete().eq('job_id', job_id).eq('svc_id', svc_id).execute()

        costs_to_insert = []

        # Base selling price → "Doanh thu cơ bản"
        if request.selling_price and request.selling_price > 0:
            costs_to_insert.append({
                'job_id': job_id, 'svc_id': svc_id,
                'cost_name': 'Doanh thu cơ bản',
                'quantity': 1, 'unit': 'lô',
                'buying_rate': 0, 'selling_rate': request.selling_price,
                'vat_rate': 8,
            })

        # Base buying price → "Chi phí cơ bản"
        if request.buying_price and request.buying_price > 0:
            costs_to_insert.append({
                'job_id': job_id, 'svc_id': svc_id,
                'cost_name': 'Chi phí cơ bản',
                'quantity': 1, 'unit': 'lô',
                'buying_rate': request.buying_price, 'selling_rate': 0,
                'vat_rate': 8,
            })

        # Extra costs
        for c in (request.extra_costs or []):
            amt = float(c.get('amount') or 0)
            if amt <= 0:
                continue
            qty = float(c.get('qty') or 1)
            unit_price = float(c.get('unit_price') or (amt / qty if qty else amt))
            costs_to_insert.append({
                'job_id': job_id, 'svc_id': svc_id,
                'cost_name': c.get('name', 'Chi phí khác'),
                'quantity': qty, 'unit': c.get('unit', 'lô'),
                'buying_rate': unit_price, 'selling_rate': 0,
                'vat_rate': 0,
            })

        # Extra revenues
        for r in (request.extra_revenues or []):
            amt = float(r.get('amount') or 0)
            if amt <= 0:
                continue
            qty = float(r.get('qty') or 1)
            unit_price = float(r.get('unit_price') or (amt / qty if qty else amt))
            costs_to_insert.append({
                'job_id': job_id, 'svc_id': svc_id,
                'cost_name': r.get('name', 'Doanh thu khác'),
                'quantity': qty, 'unit': r.get('unit', 'lô'),
                'buying_rate': 0, 'selling_rate': unit_price,
                'vat_rate': 0,
            })

        # Batch insert
        if costs_to_insert:
            client.table('job_costs').insert(costs_to_insert).execute()
            logger.info(f"Synced {len(costs_to_insert)} cost lines to job_costs for job_id={job_id}, svc_id={svc_id}")
    except Exception as e:
        logger.error(f"Failed to sync quotation to job_costs: {e}")


@router.put("/services/{svc_id}/quotations")
async def update_service_quotations(svc_id: int, request: ServiceQuotationRequest):
    """
    Update buying/selling quotations for a service.
    Stores in service_details JSONB and syncs to job_costs table.
    """
    try:
        client = get_supabase()

        # Verify service exists
        svc_result = client.table('job_services').select(
            'svc_id, job_id, service_details'
        ).eq('svc_id', svc_id).limit(1).execute()

        if not svc_result.data:
            raise HTTPException(404, f"Service {svc_id} not found")

        svc = svc_result.data[0]
        current_details = svc.get('service_details') or {}
        if isinstance(current_details, str):
            current_details = json.loads(current_details)

        # Merge quotation data
        current_details['buying_rate_id'] = request.buying_rate_id
        current_details['buying_price'] = request.buying_price
        current_details['selling_rate_id'] = request.selling_rate_id
        current_details['selling_price'] = request.selling_price
        # Normalize field names and ensure required fields
        def normalize_extras(items):
            for item in (items or []):
                if 'description' in item and 'name' not in item:
                    item['name'] = item.pop('description')
                # Ensure unit is always present (default "Lô" if missing)
                if not item.get('unit'):
                    item['unit'] = 'Lô'
                # Ensure qty is present (default 1)
                if not item.get('qty'):
                    item['qty'] = 1
                # Ensure unit_price from amount if missing
                if not item.get('unit_price') and item.get('amount'):
                    item['unit_price'] = item['amount'] / (item.get('qty') or 1)
            return items or []
        
        current_details['extra_costs'] = normalize_extras(request.extra_costs)
        current_details['extra_revenues'] = normalize_extras(request.extra_revenues)

        # Calculate total costs and revenues (including extras)
        total_cost = (request.buying_price or 0) + sum(
            c.get('amount', 0) for c in (request.extra_costs or [])
        )
        total_revenue = (request.selling_price or 0) + sum(
            r.get('amount', 0) for r in (request.extra_revenues or [])
        )

        # Calculate profit
        if total_revenue > 0 or total_cost > 0:
            current_details['total_cost'] = total_cost
            current_details['total_revenue'] = total_revenue
            current_details['profit'] = total_revenue - total_cost
            current_details['margin_pct'] = (
                (total_revenue - total_cost) / total_cost * 100
                if total_cost > 0 else 0
            )
        else:
            current_details['profit'] = None
            current_details['margin_pct'] = None

        # Update service
        client.table('job_services').update({
            'service_details': current_details,
            'updated_at': datetime.now().isoformat()
        }).eq('svc_id', svc_id).execute()

        # Sync to job_costs table (source of truth for jobs.total_revenue/total_cost)
        job_id = svc['job_id']
        _sync_quotation_to_job_costs(client, job_id, svc_id, request)

        logger.info(f"Updated quotations for service {svc_id}: buying={request.buying_price}, selling={request.selling_price}")

        return {
            "success": True,
            "svc_id": svc_id,
            "profit": current_details.get('profit'),
            "margin_pct": current_details.get('margin_pct')
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating quotations: {e}")
        return {"success": False, "message": str(e)}


class ServiceAssignRequest(BaseModel):
    vendor_id: Optional[int] = None
    employee_id: Optional[int] = None
    vehicle_id: Optional[int] = None
    driver_id: Optional[int] = None


@router.put("/services/{svc_id}/assign")
async def assign_service(svc_id: int, request: ServiceAssignRequest):
    """Assign vendor/employee to a service from the UI edit mode."""
    try:
        client = get_supabase()
        update_data = {}
        if request.vendor_id is not None:
            update_data['vendor_id'] = request.vendor_id if request.vendor_id else None
        if request.employee_id is not None:
            update_data['employee_id'] = request.employee_id if request.employee_id else None
        if request.vehicle_id is not None:
            update_data['vehicle_id'] = request.vehicle_id if request.vehicle_id else None
        if request.driver_id is not None:
            update_data['driver_id'] = request.driver_id if request.driver_id else None

        if not update_data:
            return {"success": False, "message": "No fields to update"}

        result = client.table('job_services').update(update_data).eq('svc_id', svc_id).execute()
        if not result.data:
            raise HTTPException(404, f"Service {svc_id} not found")

        return {"success": True, "data": result.data[0]}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error assigning service {svc_id}: {e}")
        raise HTTPException(500, str(e))


class ServiceNotesRequest(BaseModel):
    notes: str


@router.put("/services/{svc_id}/notes")
async def update_service_notes(svc_id: int, request: ServiceNotesRequest):
    """
    Update notes/special_requirements for a service.
    Used by AI chat to add notes to jobs.
    """
    try:
        client = get_supabase()

        # Verify service exists
        svc_result = client.table('job_services').select(
            'svc_id, job_id, special_requirements'
        ).eq('svc_id', svc_id).limit(1).execute()

        if not svc_result.data:
            raise HTTPException(404, f"Service {svc_id} not found")

        # Update notes
        client.table('job_services').update({
            'special_requirements': request.notes,
            'updated_at': datetime.now().isoformat()
        }).eq('svc_id', svc_id).execute()

        logger.info(f"Updated notes for service {svc_id}")

        return {"success": True, "svc_id": svc_id}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating service notes: {e}")
        return {"success": False, "message": str(e)}


@router.put("/services/{svc_id}/details")
async def update_service_details(svc_id: int, request: Request):
    """Update editable service details: cargo, route, date, invoice, package info.
    Requires authenticated user (records updated_by audit).
    """
    try:
        client = get_supabase()
        body = await request.json()

        # Auth required — so updated_by gets set and created_by chain stays intact.
        current_user = await get_current_user_optional(request)
        if not current_user:
            raise HTTPException(status_code=401, detail="Yêu cầu đăng nhập để cập nhật dịch vụ.")
        user_id = current_user['user_id']

        # Whitelist of editable columns — loai_hinh now included (customs services).
        allowed = {
            'cargo_type', 'package_quantity', 'package_unit',
            'origin_address', 'dest_address', 'destination_address',
            'scheduled_date', 'scheduled_time',
            'invoice_numbers', 'employee_id', 'vendor_id', 'driver_id',
            'service_type_code', 'weight_kg', 'volume_cbm',
            'bl_awb_no', 'co_no',
            'route', 'chargeable_weight_kg', 'quotation_no',
            'seller_name', 'buyer_name', 'cd_no', 'customs_status',
            'loai_hinh',
        }
        update_data = {k: v for k, v in body.items() if k in allowed}
        # Map alias: destination_address → dest_address (DB column)
        if 'destination_address' in update_data:
            update_data['dest_address'] = update_data.pop('destination_address')

        # --- VALIDATOR: if the caller is changing service_type_code to customs OR
        # updating loai_hinh, re-check against the whitelist. Need to know the
        # resulting service_type_code + loai_hinh after this update.
        if 'service_type_code' in update_data or 'loai_hinh' in update_data:
            # Fetch current row so we know the resulting state (allowing partial updates)
            current_row = client.table('job_services').select(
                'service_type_code, loai_hinh'
            ).eq('svc_id', svc_id).limit(1).execute()
            cur = current_row.data[0] if current_row.data else {}
            final_svc_code = update_data.get('service_type_code', cur.get('service_type_code'))
            final_loai_hinh = update_data.get('loai_hinh', cur.get('loai_hinh'))
            loai_hinh_error = customs_validator.validate_loai_hinh_for_service(
                final_svc_code, final_loai_hinh
            )
            if loai_hinh_error is not None:
                return {
                    "success": False,
                    "message": loai_hinh_error["message"],
                    "error": loai_hinh_error["error"],
                    "suggestions": loai_hinh_error["suggestions"],
                }
            # Normalize loai_hinh if being set
            if 'loai_hinh' in update_data:
                update_data['loai_hinh'] = (
                    customs_validator.normalize_loai_hinh(update_data['loai_hinh']) or None
                )

        # Store extra_info and/or SEA_DOM fields in service_details JSONB
        jsonb_fields = {k: body[k] for k in SEA_DOM_FIELDS if k in body}
        needs_jsonb = 'extra_info' in body or jsonb_fields
        if needs_jsonb:
            svc_row = client.table('job_services').select('service_details').eq('svc_id', svc_id).limit(1).execute()
            existing = svc_row.data[0].get('service_details') or {} if svc_row.data else {}
            if isinstance(existing, str):
                try:
                    existing = json.loads(existing)
                except Exception:
                    existing = {}
            if 'extra_info' in body:
                existing['extra_info'] = body['extra_info']
            if jsonb_fields:
                existing.update(jsonb_fields)
            update_data['service_details'] = json.dumps(existing, ensure_ascii=False)

        if not update_data:
            return {"success": False, "message": "No valid fields to update"}

        update_data['updated_at'] = datetime.now().isoformat()
        update_data['updated_by'] = user_id  # Audit trail
        client.table('job_services').update(update_data).eq('svc_id', svc_id).execute()
        logger.info(f"Updated details for service {svc_id}: {list(update_data.keys())}")

        # Check for duplicate documents after save
        doc_warnings = []
        doc_check = {k: update_data[k] for k in ('cd_no', 'bl_awb_no', 'co_no', 'invoice_numbers') if update_data.get(k)}
        if doc_check:
            try:
                from app.db.session import get_db_context
                with get_db_context() as db:
                    conditions, params = [], []
                    labels = {'cd_no': 'Tờ khai', 'bl_awb_no': 'BL/AWB', 'co_no': 'CO', 'invoice_numbers': 'Invoice'}
                    for f, v in doc_check.items():
                        val = str(v).strip()
                        if not val:
                            continue
                        if f == 'invoice_numbers':
                            conditions.append(f"js.{f} ILIKE %s")
                            params.append(f"%{val}%")
                        else:
                            conditions.append(f"js.{f} = %s")
                            params.append(val)
                    if conditions:
                        params.append(svc_id)
                        db.execute(f"""
                            SELECT j.job_no, js.cd_no, js.bl_awb_no, js.co_no, js.invoice_numbers
                            FROM job_services js JOIN jobs j ON js.job_id = j.job_id
                            WHERE ({" OR ".join(conditions)}) AND js.svc_id != %s LIMIT 5
                        """, params)
                        for row in db.fetchall():
                            for f, v in doc_check.items():
                                db_val = str(row.get(f) or '').strip()
                                val = str(v).strip()
                                if f == 'invoice_numbers' and val.lower() in db_val.lower():
                                    doc_warnings.append(f"{labels[f]} '{val}' trùng với {row['job_no']}")
                                elif db_val == val:
                                    doc_warnings.append(f"{labels[f]} '{val}' trùng với {row['job_no']}")
            except Exception as e:
                logger.warning(f"Doc duplicate check on update failed: {e}")

        result = {"success": True, "svc_id": svc_id}
        if doc_warnings:
            result["warnings"] = doc_warnings
        return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating service details: {e}")
        return {"success": False, "message": str(e)}


@router.put("/services/{svc_id}/sea-dom-details")
async def update_sea_dom_details(svc_id: int, request: Request):
    """
    Update SEA_DOM specific fields in service_details JSONB.
    Merges the 10 SEA_DOM fields into existing service_details without overwriting other fields.
    """
    try:
        client = get_supabase()
        body = await request.json()

        # Only accept SEA_DOM fields
        sea_dom_updates = {k: v for k, v in body.items() if k in SEA_DOM_FIELDS}

        if not sea_dom_updates:
            return {"success": False, "message": "No valid SEA_DOM fields provided"}

        # Fetch existing service_details to merge
        svc_row = client.table('job_services').select(
            'svc_id, service_type_code, service_details'
        ).eq('svc_id', svc_id).limit(1).execute()

        if not svc_row.data:
            return {"success": False, "message": f"Service {svc_id} not found"}

        svc = svc_row.data[0]

        # Warn if not a SEA_DOM service (but don't block)
        if not is_sea_dom_service(svc.get('service_type_code', '')):
            logger.warning(f"Updating SEA_DOM fields on non-SEA_DOM service {svc_id} ({svc.get('service_type_code')})")

        existing = svc.get('service_details') or {}
        if isinstance(existing, str):
            try:
                existing = json.loads(existing)
            except Exception:
                existing = {}

        # Merge SEA_DOM fields (preserve all other fields)
        existing.update(sea_dom_updates)

        # Validate and log warnings for missing fields
        missing = validate_sea_dom_fields(existing)
        if missing:
            logger.warning(f"SEA_DOM service {svc_id} still missing fields: {missing}")

        client.table('job_services').update({
            'service_details': json.dumps(existing, ensure_ascii=False),
            'updated_at': datetime.now().isoformat()
        }).eq('svc_id', svc_id).execute()

        result = {"success": True, "svc_id": svc_id, "updated_fields": list(sea_dom_updates.keys())}
        if missing:
            result["warnings"] = [f"Thiếu các trường SEA_DOM: {', '.join(missing)}"]
        return result

    except Exception as e:
        logger.error(f"Error updating SEA_DOM details for svc {svc_id}: {e}")
        return {"success": False, "message": str(e)}


# ============================================================
# LIGHTWEIGHT LOOKUP ENDPOINTS (no admin auth required)
# Used by job edit dropdowns — avoids admin dependency
# ============================================================

@router.get("/lookup/vendors")
async def lookup_vendors(search: Optional[str] = Query(None)):
    """List active vendors for dropdown selection (no admin auth)."""
    try:
        client = get_supabase()
        query = client.table('vendors').select(
            'vendor_id, vendor_code, short_name, company_name, country, currency, contact_person'
        ).eq('is_active', True)

        if search:
            query = query.or_(
                f"short_name.ilike.%{search}%,company_name.ilike.%{search}%,vendor_code.ilike.%{search}%"
            )

        result = query.order('vendor_code').limit(200).execute()
        return {"data": result.data or []}
    except Exception as e:
        logger.error(f"Error looking up vendors: {e}")
        return {"data": []}


@router.get("/lookup/customers")
async def lookup_customers(search: Optional[str] = Query(None)):
    """List active customers for dropdown selection (no admin auth)."""
    try:
        client = get_supabase()
        query = client.table('customers').select(
            'customer_id, customer_code, short_name, company_name'
        ).eq('is_active', True)

        if search:
            query = query.or_(
                f"short_name.ilike.%{search}%,company_name.ilike.%{search}%,customer_code.ilike.%{search}%"
            )

        result = query.order('customer_code').limit(200).execute()
        return {"data": result.data or []}
    except Exception as e:
        logger.error(f"Error looking up customers: {e}")
        return {"data": []}
