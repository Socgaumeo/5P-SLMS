"""
AR/AP API — Module Kế toán công nợ (phải thu / phải trả).
- AR: hóa đơn gom N job, track đã thu (v_job_ar_status).
- AP: bảng kê chi phí vendor/employee, track đã trả (v_ap_unbilled_costs).
"""
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import date
import logging, io, smtplib
from email.message import EmailMessage
import httpx

from app.db.supabase_client import get_supabase
from app.core.config import settings

logger = logging.getLogger(__name__)
router = APIRouter()


# ============ MODELS ============
class InvoiceCreate(BaseModel):
    customer_id: int
    job_ids: List[int]
    invoice_no: Optional[str] = None
    issue_date: Optional[date] = None
    due_date: Optional[date] = None
    vat_rate: Optional[float] = 10
    note: Optional[str] = None
    created_by: Optional[int] = None


class PaymentUpdate(BaseModel):
    paid_amount: float
    paid_date: Optional[date] = None


class StatusUpdate(BaseModel):
    payment_status: Optional[str] = None   # unpaid/partial/paid
    paid_amount: Optional[float] = None
    paid_date: Optional[date] = None
    note: Optional[str] = None
    due_date: Optional[date] = None


class BillCreate(BaseModel):
    vendor_id: Optional[int] = None
    employee_id: Optional[int] = None
    cost_ids: List[int]
    bill_no: Optional[str] = None
    due_date: Optional[date] = None
    note: Optional[str] = None
    created_by: Optional[int] = None


# ============ AR — PHẢI THU ============
@router.get("/api/ar/invoices")
def list_invoices(status: Optional[str] = None):
    sb = get_supabase()
    q = sb.table("ar_invoices").select("*").order("created_at", desc=True)
    if status:
        q = q.eq("payment_status", status)
    return {"invoices": q.execute().data}


def _vat_by_job(sb, job_ids: list) -> dict:
    """Tính doanh thu đã gồm VAT cho từng job từ job_costs (mỗi dòng vat_rate riêng).
    Trả {job_id: {"vat_amount": x, "revenue_with_vat": y}}.
    Fallback: nếu job chưa có dòng job_costs nào thì để vat=0 (revenue_with_vat = total_revenue)."""
    out = {}
    if not job_ids:
        return out
    rows = sb.table("job_costs").select(
        "job_id,selling_amount,vat_rate,is_reimbursement"
    ).in_("job_id", list(job_ids)).execute().data or []
    for r in rows:
        jid = r.get("job_id")
        sell = float(r.get("selling_amount") or 0)
        if r.get("is_reimbursement"):
            vr = 0.0  # thu hộ/chi hộ không VAT
        else:
            vr = float(r.get("vat_rate") or 0)
        d = out.setdefault(jid, {"vat_amount": 0.0, "revenue_with_vat": 0.0})
        d["vat_amount"] += sell * vr / 100
        d["revenue_with_vat"] += sell * (1 + vr / 100)
    return out


def _enrich_vat(sb, jobs: list) -> list:
    """Thêm field vat_amount + revenue_with_vat vào mỗi job dict."""
    vmap = _vat_by_job(sb, [j.get("job_id") for j in jobs if j.get("job_id")])
    for j in jobs:
        base = float(j.get("total_revenue") or 0)
        v = vmap.get(j.get("job_id"))
        if v and v["revenue_with_vat"] > 0:
            j["vat_amount"] = round(v["vat_amount"])
            j["revenue_with_vat"] = round(v["revenue_with_vat"])
        else:
            # job chưa breakdown job_costs -> không suy được VAT, giữ nguyên
            j["vat_amount"] = 0
            j["revenue_with_vat"] = round(base)
    return jobs


@router.get("/api/ar/job-status")
def job_ar_status(state: Optional[str] = None, customer_id: Optional[int] = None):
    """Job nào đã xuất HĐ / đã thu / chưa xuất."""
    sb = get_supabase()
    q = sb.table("v_job_ar_status").select("*")
    if state:
        q = q.eq("ar_state", state)
    if customer_id:
        q = q.eq("customer_id", customer_id)
    jobs = q.execute().data
    return {"jobs": _enrich_vat(sb, jobs)}


@router.get("/api/ar/by-customer")
def ar_by_customer(state: Optional[str] = "CHUA_XUAT_HD"):
    """Gom công nợ phải thu theo khách hàng (job chưa xuất HĐ)."""
    sb = get_supabase()
    q = sb.table("v_job_ar_status").select("job_id,customer_id,customer_name,total_revenue,ar_state")
    if state:
        q = q.eq("ar_state", state)
    rows = q.execute().data
    vmap = _vat_by_job(sb, [r.get("job_id") for r in rows if r.get("job_id")])
    agg = {}
    for r in rows:
        cid = r.get("customer_id")
        if cid not in agg:
            agg[cid] = {"customer_id": cid, "customer_name": r.get("customer_name") or "?",
                        "job_count": 0, "total": 0, "total_with_vat": 0}
        base = float(r.get("total_revenue") or 0)
        v = vmap.get(r.get("job_id"))
        with_vat = v["revenue_with_vat"] if (v and v["revenue_with_vat"] > 0) else base
        agg[cid]["job_count"] += 1
        agg[cid]["total"] += base
        agg[cid]["total_with_vat"] += with_vat
    for a in agg.values():
        a["total_with_vat"] = round(a["total_with_vat"])
    return {"customers": sorted(agg.values(), key=lambda x: -x["total_with_vat"])}


@router.post("/api/ar/invoices")
def create_invoice(payload: InvoiceCreate):
    sb = get_supabase()
    # Tính tổng từ total_revenue các job
    jobs = sb.table("jobs").select("job_id,total_revenue,customer_id").in_(
        "job_id", payload.job_ids).execute().data
    if not jobs:
        raise HTTPException(400, "Không tìm thấy job")
    bad = [j["job_id"] for j in jobs if j["customer_id"] != payload.customer_id]
    if bad:
        raise HTTPException(400, f"Job {bad} không thuộc khách hàng này")
    subtotal = sum(float(j["total_revenue"] or 0) for j in jobs)
    vat = round(subtotal * (payload.vat_rate or 0) / 100)
    total = subtotal + vat
    inv = sb.table("ar_invoices").insert({
        "invoice_no": payload.invoice_no,
        "customer_id": payload.customer_id,
        "issue_date": str(payload.issue_date) if payload.issue_date else None,
        "due_date": str(payload.due_date) if payload.due_date else None,
        "subtotal": subtotal, "vat_amount": vat, "total": total,
        "note": payload.note, "created_by": payload.created_by,
    }).execute().data[0]
    for j in jobs:
        sb.table("ar_invoice_jobs").insert({
            "invoice_id": inv["invoice_id"], "job_id": j["job_id"],
            "allocated_amount": float(j["total_revenue"] or 0),
        }).execute()
    return {"invoice": inv, "job_count": len(jobs)}


@router.patch("/api/ar/invoices/{invoice_id}/payment")
def pay_invoice(invoice_id: int, payload: PaymentUpdate):
    sb = get_supabase()
    upd = sb.table("ar_invoices").update({
        "paid_amount": payload.paid_amount,
        "paid_date": str(payload.paid_date) if payload.paid_date else str(date.today()),
    }).eq("invoice_id", invoice_id).execute().data
    if not upd:
        raise HTTPException(404, "Không tìm thấy hóa đơn")
    return {"invoice": upd[0]}


@router.patch("/api/ar/invoices/{invoice_id}")
def edit_invoice(invoice_id: int, payload: StatusUpdate):
    """Sửa tình trạng / ghi chú / hạn HĐ."""
    sb = get_supabase()
    data = {k: (str(v) if isinstance(v, date) else v)
            for k, v in payload.dict().items() if v is not None}
    if not data:
        raise HTTPException(400, "Không có gì để sửa")
    upd = sb.table("ar_invoices").update(data).eq("invoice_id", invoice_id).execute().data
    if not upd:
        raise HTTPException(404, "Không tìm thấy hóa đơn")
    return {"invoice": upd[0]}


@router.delete("/api/ar/invoices/{invoice_id}")
def delete_invoice(invoice_id: int):
    """Xóa HĐ (ar_invoice_jobs tự xóa theo CASCADE) → job quay lại 'chưa xuất HĐ'."""
    sb = get_supabase()
    sb.table("ar_invoices").delete().eq("invoice_id", invoice_id).execute()
    return {"deleted": invoice_id}


# ============ AP — PHẢI TRẢ ============
@router.get("/api/ap/unbilled")
def unbilled_costs(vendor_id: Optional[int] = None, employee_id: Optional[int] = None):
    """Chi phí chưa lên bảng kê. Không filter → gom theo vendor."""
    sb = get_supabase()
    if vendor_id:
        rows = sb.table("v_ap_unbilled_costs").select("*").eq(
            "vendor_id", vendor_id).execute().data
        return {"costs": rows, "total": sum(float(r["amount"] or 0) for r in rows)}
    # summary theo vendor
    rows = sb.table("v_ap_unbilled_costs").select("*").execute().data
    agg = {}
    for r in rows:
        vid = r.get("vendor_id")
        if vid is None:
            continue
        a = agg.setdefault(vid, {"vendor_id": vid, "vendor_name": r.get("vendor_name"), "count": 0, "total": 0})
        a["count"] += 1
        a["total"] += float(r["amount"] or 0)
    return {"vendors": sorted(agg.values(), key=lambda x: -x["total"])}


@router.get("/api/ap/bills")
def list_bills(status: Optional[str] = None):
    sb = get_supabase()
    q = sb.table("ap_bills").select("*").order("created_at", desc=True)
    if status:
        q = q.eq("payment_status", status)
    return {"bills": q.execute().data}


@router.post("/api/ap/bills")
def create_bill(payload: BillCreate):
    sb = get_supabase()
    if not payload.vendor_id and not payload.employee_id:
        raise HTTPException(400, "Cần vendor_id hoặc employee_id")
    costs = sb.table("job_costs").select("cost_id,buying_rate,quantity").in_(
        "cost_id", payload.cost_ids).execute().data
    if not costs:
        raise HTTPException(400, "Không tìm thấy chi phí")
    total = sum(float(c["buying_rate"] or 0) * float(c.get("quantity") or 1) for c in costs)
    bill = sb.table("ap_bills").insert({
        "bill_no": payload.bill_no,
        "vendor_id": payload.vendor_id, "employee_id": payload.employee_id,
        "total_amount": total,
        "due_date": str(payload.due_date) if payload.due_date else None,
        "note": payload.note, "created_by": payload.created_by,
    }).execute().data[0]
    for c in costs:
        amt = float(c["buying_rate"] or 0) * float(c.get("quantity") or 1)
        sb.table("ap_bill_items").insert({
            "bill_id": bill["bill_id"], "cost_id": c["cost_id"], "amount": amt,
        }).execute()
    return {"bill": bill, "item_count": len(costs)}


@router.patch("/api/ap/bills/{bill_id}/payment")
def pay_bill(bill_id: int, payload: PaymentUpdate):
    sb = get_supabase()
    upd = sb.table("ap_bills").update({
        "paid_amount": payload.paid_amount,
        "paid_date": str(payload.paid_date) if payload.paid_date else str(date.today()),
    }).eq("bill_id", bill_id).execute().data
    if not upd:
        raise HTTPException(404, "Không tìm thấy bảng kê")
    return {"bill": upd[0]}


@router.patch("/api/ap/bills/{bill_id}")
def edit_bill(bill_id: int, payload: StatusUpdate):
    """Sửa tình trạng / ghi chú / hạn bảng kê."""
    sb = get_supabase()
    data = {k: (str(v) if isinstance(v, date) else v)
            for k, v in payload.dict().items() if v is not None}
    if not data:
        raise HTTPException(400, "Không có gì để sửa")
    upd = sb.table("ap_bills").update(data).eq("bill_id", bill_id).execute().data
    if not upd:
        raise HTTPException(404, "Không tìm thấy bảng kê")
    return {"bill": upd[0]}


@router.delete("/api/ap/bills/{bill_id}")
def delete_bill(bill_id: int):
    """Xóa bảng kê (ap_bill_items tự xóa theo CASCADE) → chi phí quay lại 'chờ thanh toán'."""
    sb = get_supabase()
    sb.table("ap_bills").delete().eq("bill_id", bill_id).execute()
    return {"deleted": bill_id}


@router.get("/api/ap/bills/{bill_id}/items")
def bill_items(bill_id: int):
    """Chi tiết các dòng chi phí trong 1 bảng kê (kèm thông tin đối chiếu)."""
    sb = get_supabase()
    items = sb.table("ap_bill_items").select("cost_id,amount").eq("bill_id", bill_id).execute().data
    return {"items": items}


# ============ XUẤT EXCEL theo cost_ids đã chọn ============
def _build_statement_xlsx(costs: list, title: str) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "BangKe"
    thin = Side(style='thin'); bd = Border(thin, thin, thin, thin)
    is_cus = any(c.get("is_customs") for c in costs)
    ws.merge_cells('A1:J1'); ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'] = f"Ngày xuất: {date.today().strftime('%d/%m/%Y')}"
    if is_cus:
        hdr = ['STT', 'Job No', 'Ngày', 'Tên phí', 'Số tờ khai', 'Đơn giá', 'SL', 'Thành tiền']
    else:
        hdr = ['STT', 'Job No', 'Ngày', 'Tên phí', 'Biển số', 'Tuyến', 'INV/BL', 'Số HĐ', 'Đơn giá', 'SL', 'Thành tiền']
    ws.append([]); ws.append(hdr); hr = ws.max_row
    for c in range(1, len(hdr) + 1):
        cell = ws.cell(hr, c); cell.font = Font(bold=True); cell.border = bd
        cell.fill = PatternFill('solid', fgColor='DDEBF7'); cell.alignment = Alignment(horizontal='center')
    total = 0
    for i, c in enumerate(costs, 1):
        amt = int(c.get("amount") or 0); total += amt
        nm = c.get("cost_name", "") + (" (chi hộ)" if c.get("is_reimbursement") else "")
        dt = str(c.get("cost_date") or "")
        if is_cus:
            row = [i, c.get("job_no"), dt, nm, c.get("declaration_no") or "", int(c.get("buying_rate") or 0), float(c.get("quantity") or 1), amt]
        else:
            row = [i, c.get("job_no"), dt, nm, c.get("plate_number") or "", c.get("route") or "",
                   c.get("bl_awb_no") or "", c.get("job_invoice_no") or "", int(c.get("buying_rate") or 0), float(c.get("quantity") or 1), amt]
        ws.append(row)
        for cc in range(1, len(hdr) + 1): ws.cell(ws.max_row, cc).border = bd
    trow = [''] * (len(hdr) - 5) + ['TỔNG CỘNG', '', '', total]
    ws.append(trow); tr = ws.max_row
    for c in range(1, len(hdr) + 1):
        ws.cell(tr, c).font = Font(bold=True); ws.cell(tr, c).border = bd
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


class ExportRequest(BaseModel):
    cost_ids: List[int]
    title: Optional[str] = "BẢNG KÊ CHI PHÍ PHẢI TRẢ"


@router.post("/api/ap/export")
def export_selected(payload: ExportRequest):
    """Xuất Excel bảng kê cho các dòng chi phí đã tick chọn."""
    sb = get_supabase()
    costs = sb.table("v_ap_unbilled_costs").select("*").in_("cost_id", payload.cost_ids).execute().data
    if not costs:
        raise HTTPException(400, "Không có chi phí")
    xlsx = _build_statement_xlsx(costs, payload.title)
    return StreamingResponse(io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="bangke_{date.today()}.xlsx"'})


# ============ NOTIFY KẾ TOÁN ============
def _costs_for_notify(sb, cost_ids: list) -> list:
    """Lấy chi phí theo cost_ids (kể cả đã lập bảng kê, không còn trong view unbilled)."""
    if not cost_ids:
        return []
    rows = sb.table("job_costs").select(
        "cost_id,job_id,cost_name,buying_rate,quantity,is_reimbursement,vendor_id,created_at"
    ).in_("cost_id", cost_ids).execute().data
    out = []
    for r in rows:
        jid = r.get("job_id")
        job = sb.table("jobs").select("job_no").eq("job_id", jid).execute().data if jid else []
        vend = sb.table("vendors").select("short_name").eq("vendor_id", r.get("vendor_id")).execute().data if r.get("vendor_id") else []
        amt = float(r.get("buying_rate") or 0) * float(r.get("quantity") or 1)
        out.append({
            "cost_id": r["cost_id"], "job_no": job[0]["job_no"] if job else "",
            "cost_name": r.get("cost_name"), "buying_rate": r.get("buying_rate"),
            "quantity": r.get("quantity"), "amount": amt,
            "is_reimbursement": r.get("is_reimbursement"),
            "vendor_name": vend[0]["short_name"] if vend else "?",
            "cost_date": r.get("created_at", "")[:10], "is_customs": False,
        })
    return out


@router.get("/api/ap/users")
def list_users_for_notify():
    """Danh sách user (từ DB) để chọn người nhận thông báo — kèm email + telegram_id."""
    sb = get_supabase()
    rows = sb.table("users").select("user_id,full_name,email,telegram_id").eq("is_active", True).execute().data
    return {"users": rows}


class NotifyConfig(BaseModel):
    telegram_user_ids: List[int] = []   # user_id được chọn nhận Telegram
    emails: List[str] = []              # danh sách email người nhận (tự do)


@router.get("/api/ap/notify-config")
def get_notify_config():
    sb = get_supabase()
    rows = sb.table("ap_notify_config").select("*").eq("is_active", True).execute().data
    return {"config": rows[0] if rows else None}


@router.post("/api/ap/notify-config")
def set_notify_config(payload: NotifyConfig):
    sb = get_supabase()
    sb.table("ap_notify_config").update({"is_active": False}).eq("is_active", True).execute()
    row = sb.table("ap_notify_config").insert({
        "role": "accountant",
        "telegram_user_ids": payload.telegram_user_ids,
        "emails": payload.emails,
    }).execute().data[0]
    return {"config": row}


class NotifyRequest(BaseModel):
    cost_ids: Optional[List[int]] = None      # gửi từ danh sách chi phí chưa lập bảng kê
    bill_id: Optional[int] = None             # HOẶC gửi từ bảng kê đã lập
    telegram_user_ids: List[int] = []         # người nhận Telegram (inline, chọn ngay lúc gửi)
    emails: List[str] = []                     # email người nhận thêm (inline)
    note: Optional[str] = None
    requested_by: Optional[int] = None         # user_id người bấm gửi (để đặt tên file + ghi)


def _slug(s: str) -> str:
    """Bỏ dấu + ký tự lạ để đặt tên file an toàn."""
    import unicodedata, re
    s = unicodedata.normalize("NFKD", s or "").encode("ascii", "ignore").decode()
    s = re.sub(r"[^A-Za-z0-9]+", "", s)
    return s or "NA"


@router.post("/api/ap/notify")
def notify_accountant(payload: NotifyRequest):
    """Gửi đề nghị chi cho kế toán (Telegram + Email). Người nhận chọn ngay lúc gửi."""
    sb = get_supabase()
    tg_uids = payload.telegram_user_ids or []
    emails = list(payload.emails or [])
    if not tg_uids and not emails:
        raise HTTPException(400, "Chưa chọn người nhận (Telegram hoặc Email)")

    # Lấy chi phí: từ bill_id (bảng kê đã lập) hoặc cost_ids
    if payload.bill_id:
        items = sb.table("ap_bill_items").select("cost_id").eq("bill_id", payload.bill_id).execute().data
        cost_ids = [it["cost_id"] for it in items]
        # bảng kê đã lập → cost đã bị loại khỏi view unbilled → query job_costs trực tiếp
        costs = _costs_for_notify(sb, cost_ids)
    else:
        cost_ids = payload.cost_ids or []
        costs = sb.table("v_ap_unbilled_costs").select("*").in_("cost_id", cost_ids).execute().data
    if not costs:
        raise HTTPException(400, "Không có chi phí")
    total = sum(int(c.get("amount") or 0) for c in costs)
    vendors = sorted(set(c.get("vendor_name") or "?" for c in costs))
    title = f"ĐỀ NGHỊ CHI — {', '.join(vendors)}"
    xlsx = _build_statement_xlsx(costs, title)
    # Người đề nghị (người bấm gửi)
    requester_name = ""
    if payload.requested_by:
        ru = sb.table("users").select("full_name").eq("user_id", payload.requested_by).execute().data
        requester_name = ru[0]["full_name"] if ru else ""
    # Tên file rõ ràng: DeNghiChi_<NCC>_<NguoiDeNghi>_<ngay>.xlsx
    vend_slug = "_".join(_slug(v) for v in vendors)[:40]
    req_slug = _slug(requester_name)
    parts = ["DeNghiChi", vend_slug] + ([req_slug] if requester_name else []) + [str(date.today())]
    fname = "_".join(parts) + ".xlsx"
    summary = (f"💸 <b>Đề nghị thanh toán</b>\n"
               f"NCC: {', '.join(vendors)}\n"
               f"Số khoản: {len(costs)}\n"
               f"Tổng: <b>{total:,.0f}đ</b>".replace(",", "."))
    if requester_name:
        summary += f"\nNgười đề nghị: {requester_name}"
    if payload.note:
        summary += f"\nGhi chú: {payload.note}"
    sent = {"telegram": 0, "email": 0}
    telegram_errors = []

    # Telegram — notify công nợ ưu tiên bot Sen (TELEGRAM_NOTIFY_BOT_TOKEN), fallback bot mặc định
    settings.resolve_telegram_token()
    tok = getattr(settings, "TELEGRAM_NOTIFY_BOT_TOKEN", None) or settings.TELEGRAM_BOT_TOKEN
    if tg_uids:
        urows = sb.table("users").select("user_id,telegram_id,email").in_("user_id", tg_uids).execute().data
        for ur in urows:
            # user được chọn nhận TG thì email của họ cũng vào danh sách nhận mail
            if ur.get("email"):
                emails.append(ur["email"])
            chat = ur.get("telegram_id")
            uid = ur.get("user_id")
            if not chat:
                telegram_errors.append(f"uid={uid}: chưa có telegram_id")
                continue
            if not tok:
                telegram_errors.append(f"uid={uid}: backend thiếu bot token")
                continue
            try:
                with httpx.Client(timeout=30) as cli:
                    r = cli.post(f"https://api.telegram.org/bot{tok}/sendDocument",
                        data={"chat_id": str(chat), "caption": summary, "parse_mode": "HTML"},
                        files={"document": (fname, xlsx,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
                    jr = r.json()
                    if jr.get("ok"):
                        sent["telegram"] += 1
                    else:
                        telegram_errors.append(f"uid={uid} chat={chat}: {jr.get('description','?')}")
                        logger.error(f"Telegram notify fail uid={uid}: {jr}")
            except Exception as e:
                telegram_errors.append(f"uid={uid}: {str(e)[:120]}")
                logger.error(f"Telegram notify fail uid={uid}: {e}")

    # Email qua Gmail API (HTTPS 443 — Railway KHÔNG chặn; gửi TỪ chính Gmail 5pvietnam.tas)
    import base64
    from email.message import EmailMessage
    gcid = getattr(settings, "GMAIL_CLIENT_ID", None)
    gcs = getattr(settings, "GMAIL_CLIENT_SECRET", None)
    grt = getattr(settings, "GMAIL_REFRESH_TOKEN", None)
    email_from = getattr(settings, "EMAIL_FROM", None) or getattr(settings, "GMAIL_SENDER", None) or "5pvietnam.tas@gmail.com"
    emails = sorted(set(e for e in emails if e))
    email_error = None
    email_configured = bool(gcid and gcs and grt)
    if emails and email_configured:
        try:
            with httpx.Client(timeout=20) as cli:
                tr = cli.post("https://oauth2.googleapis.com/token", data={
                    "client_id": gcid, "client_secret": gcs,
                    "refresh_token": grt, "grant_type": "refresh_token"})
                at = tr.json().get("access_token")
                if not at:
                    raise RuntimeError(f"token: {tr.text[:150]}")
                body = summary.replace("<b>", "").replace("</b>", "")
                msg = EmailMessage()
                msg["To"] = ", ".join(emails)
                msg["From"] = email_from
                msg["Subject"] = title
                msg.set_content(body)
                msg.add_attachment(xlsx, maintype="application",
                    subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=fname)
                raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
                sr = cli.post("https://gmail.googleapis.com/gmail/v1/users/me/messages/send",
                    headers={"Authorization": f"Bearer {at}"}, json={"raw": raw})
            if sr.status_code == 200:
                sent["email"] = len(emails)
            else:
                email_error = f"Gmail {sr.status_code}: {sr.text[:200]}"
                logger.error(email_error)
        except Exception as e:
            email_error = str(e)
            logger.error(f"Email notify fail: {e}")

    return {"sent": sent, "total": total, "count": len(costs),
            "email_recipients": emails, "smtp_configured": email_configured,
            "email_error": email_error,
            "telegram_error": "; ".join(telegram_errors) if telegram_errors else None}
