# backend/app/ai/excel/booking_form_parser.py
"""
Specialized parser for trucking booking forms.
Supports various formats: MEIKO/DREAMTECH, HOSIDEN, etc.
Automatically detects customer from filename and content.
"""

import re
import logging
import os
from datetime import datetime, date, time
from typing import Dict, Any, List, Optional, Tuple
from openpyxl import load_workbook

from app.ai.utils.smart_parser import (
    parse_date, format_date_iso,
    parse_time, format_time_str,
    parse_number, normalize_unit, parse_quantity_with_unit
)

logger = logging.getLogger(__name__)

# Known customer keywords for auto-detection
KNOWN_CUSTOMERS = {
    'MEIKO': ['meiko', 'mk'],
    'DREAMTECH': ['dreamtech', 'drt'],
    'HOSIDEN': ['hosiden', 'hsd'],
    'SAMSUNG': ['samsung', 'sds'],
    'ASGL': ['asgl', 'asg logistics'],
}


class BookingFormParser:
    """
    Parse structured booking forms like "Phiếu book xe DREAMTECH"

    Known form layouts:
    - Row 7: Department (Bộ phận book xe)
    - Row 8: Contact person (Người book xe)
    - Row 9: Company name & address
    - Row 10: Recipient info
    - Row 13: Headers (Ngày book xe, Thời gian, Invoice, etc.)
    - Row 14+: Data rows
    """

    # Known cell mappings for trucking booking forms
    FORM_MAPPINGS = {
        # Metadata cells (fixed positions)
        'department': [('A7', 'G7'), 'G7'],  # Label at A7, value at G7
        'contact_person': [('A8', 'G8'), 'G8'],
        'company_and_address': [('A9', 'G9'), 'G9'],
        'recipient_info': [('A10', 'G10'), 'G10'],

        # Header row indicators
        'header_keywords': ['ngày book', 'invoice', 'thời gian', 'số kiện', 'trọng lượng'],

        # Data column mappings (column letters)
        'data_columns': {
            'A': 'stt',
            'B': 'booking_date',
            'C': 'pickup_time',
            'D': 'invoice_number',
            'E': 'available_time',
            'F': 'goods_name',
            'G': 'package_quantity',
            'H': 'weight_kg',
            'I': 'delivery_address',
            'J': 'delivery_time',
            'K': 'return_invoice',
            'L': 'return_goods',
            'M': 'return_quantity',
            'N': 'return_weight',
        }
    }

    def __init__(self, file_path: str = None):
        self.workbook = None
        self.sheet = None
        self.file_path = file_path
        self.filename = os.path.basename(file_path) if file_path else None

    def _detect_customer_from_filename(self) -> tuple[Optional[str], Optional[str]]:
        """
        Detect customer and delivery company from filename.

        For "Phiếu book xe XXX" style filenames:
        - XXX is the DELIVERY company (where goods are going)
        - Our customer is usually the one mentioned in header (pickup location)

        Returns:
            (customer_code, delivery_company) tuple
        """
        if not self.filename:
            return None, None

        filename_lower = self.filename.lower()
        delivery_company = None

        # For booking forms, the company in filename is usually the delivery destination
        if any(kw in filename_lower for kw in ['book xe', 'booking', 'phiếu book', 'phieu book']):
            for company, keywords in KNOWN_CUSTOMERS.items():
                for kw in keywords:
                    if kw in filename_lower:
                        delivery_company = company
                        break
                if delivery_company:
                    break
            # Don't return customer here - detect from content instead
            return None, delivery_company

        # For other files, company in filename might be our customer
        for customer_code, keywords in KNOWN_CUSTOMERS.items():
            for kw in keywords:
                if kw in filename_lower:
                    return customer_code, None

        return None, None

    def _detect_customer_from_content(self) -> Optional[str]:
        """
        Detect customer from form content.

        For booking forms, the customer is typically mentioned in:
        - Header row C13 "xe có mặt tại XXX" (pickup location = our customer)
        - Cell content mentioning pickup company
        """
        if not self.sheet:
            return None

        # Check header row C13 for pickup location hint (highest priority)
        # Example: "Thời gian mong muốn xe có mặt tại Meiko"
        header_c13 = self._get_cell_value('C13')
        if header_c13:
            header_str = str(header_c13).lower()
            # Look for "tại XXX" pattern
            for customer_code, keywords in KNOWN_CUSTOMERS.items():
                for kw in keywords:
                    if kw in header_str:
                        logger.info(f"Detected customer '{customer_code}' from header C13")
                        return customer_code

        # Check other header cells
        for row in range(6, 11):
            for col in ['A', 'B', 'C']:
                val = self._get_cell_value(f'{col}{row}')
                if val:
                    val_str = str(val).lower()
                    # Look for pickup hints
                    if any(hint in val_str for hint in ['lấy hàng', 'pickup', 'có mặt tại']):
                        for customer_code, keywords in KNOWN_CUSTOMERS.items():
                            for kw in keywords:
                                if kw in val_str:
                                    return customer_code

        return None
    
    def parse(self, file_path: str) -> Dict[str, Any]:
        """
        Parse a booking form Excel file.

        Returns:
            Dict containing:
            - metadata: Form-level info (company, contact, etc.)
            - bookings: List of booking entries
            - raw_text: Combined text for AI processing
        """
        try:
            self.file_path = file_path
            self.filename = os.path.basename(file_path)
            self.workbook = load_workbook(file_path, data_only=True)
            self.sheet = self.workbook.active

            result = {
                'metadata': self._extract_metadata(),
                'bookings': self._extract_bookings(),
                'raw_text': '',
                'summary': ''
            }

            # Build combined text for AI
            result['raw_text'] = self._build_raw_text(result)
            result['summary'] = self._build_summary(result)

            return result

        except Exception as e:
            logger.error(f"Error parsing booking form: {e}")
            import traceback
            traceback.print_exc()
            return {
                'metadata': {},
                'bookings': [],
                'raw_text': f"Error parsing file: {str(e)}",
                'summary': 'Parse error'
            }
    
    def _get_cell_value(self, coord: str) -> Any:
        """Get cell value, handling merged cells"""
        try:
            cell = self.sheet[coord]
            return cell.value
        except:
            return None
    
    def _extract_metadata(self) -> Dict[str, Any]:
        """Extract form-level metadata"""
        metadata = {}

        # Department
        dept = self._get_cell_value('G7')
        if dept:
            metadata['department'] = str(dept).strip()

        # Contact person
        contact = self._get_cell_value('G8')
        if contact:
            metadata['contact_person'] = str(contact).strip()
            # Try to extract phone
            phone_match = re.search(r'(\d{10,11})', str(contact))
            if phone_match:
                metadata['contact_phone'] = phone_match.group(1)

        # Company and address in G9 - This is the DELIVERY destination
        company_addr = self._get_cell_value('G9')
        if company_addr:
            lines = str(company_addr).split('\n')
            if lines:
                metadata['delivery_company'] = lines[0].strip()  # e.g., "DREAMTECH", "HOSIDEN"
                if len(lines) > 1:
                    # Parse addresses like "Xưởng 1: Số 2 đường..." or full address
                    delivery_addresses = {}
                    full_address_parts = []
                    for line in lines[1:]:
                        line = line.strip()
                        if line:
                            full_address_parts.append(line)
                            # Check if it's like "Xưởng 1: address"
                            if ':' in line:
                                parts = line.split(':', 1)
                                key = parts[0].strip()
                                addr = parts[1].strip() if len(parts) > 1 else ""
                                delivery_addresses[key] = addr
                            else:
                                delivery_addresses[f"Địa chỉ {len(delivery_addresses)+1}"] = line
                    metadata['delivery_addresses_map'] = delivery_addresses
                    # Full list for reference
                    metadata['delivery_addresses'] = [l.strip() for l in lines[1:] if l.strip()]
                    # Single combined address string for export
                    metadata['delivery_full_address'] = ', '.join(full_address_parts)
                else:
                    metadata['delivery_full_address'] = ''

        # Recipient info
        recipient = self._get_cell_value('G10')
        if recipient:
            metadata['recipient_info'] = str(recipient).strip()

        # Auto-detect customer and delivery company
        # Priority: content-based detection > filename detection
        detected_from_content = self._detect_customer_from_content()
        filename_customer, filename_delivery = self._detect_customer_from_filename()

        # Known customer addresses for quick lookup
        customer_addresses = {
            'MEIKO': 'Kho Meiko - Lô CN9, KCN Thạch Thất - Quốc Oai, Hà Nội',
            'DREAMTECH': 'KCN Vsip Bắc Ninh',
            'HOSIDEN': 'KCN Quế Võ, Bắc Ninh',
            'SAMSUNG': 'KCN Yên Phong, Bắc Ninh',
        }

        # Determine customer (who books through us - pickup location)
        if detected_from_content:
            # Content detection takes priority (e.g., "xe có mặt tại Meiko")
            detected_customer = detected_from_content
        elif filename_customer:
            # Use filename-based customer if content didn't detect one
            detected_customer = filename_customer
        else:
            detected_customer = None

        # If delivery company from filename and different from customer, use it
        if filename_delivery and filename_delivery != detected_customer:
            # Override delivery_company from G9 if filename provides it
            if not metadata.get('delivery_company'):
                metadata['delivery_company'] = filename_delivery

        # Set customer info based on detection
        if detected_customer:
            metadata['customer_code'] = detected_customer
            metadata['pickup_company'] = detected_customer
            metadata['pickup_address'] = f'Kho {detected_customer}'
            metadata['pickup_full_address'] = customer_addresses.get(
                detected_customer, f'Kho {detected_customer}'
            )
            logger.info(f"Detected customer: {detected_customer}")
        else:
            # No customer detected - warn
            metadata['customer_code'] = None
            metadata['pickup_company'] = None
            metadata['pickup_address'] = None
            metadata['customer_warning'] = 'Không xác định được khách hàng từ file. Vui lòng chọn khách hàng.'
            logger.warning(f"Could not detect customer from file: {self.filename}")

        return metadata
    
    def _extract_bookings(self) -> List[Dict[str, Any]]:
        """Extract booking data rows"""
        bookings = []
        
        # Find data rows (usually starting at row 14)
        # Look for rows where column A has a number (STT)
        for row_num in range(14, min(30, self.sheet.max_row + 1)):
            stt = self._get_cell_value(f'A{row_num}')
            
            # Check if this is a data row (STT is a number)
            if stt and isinstance(stt, (int, float)):
                booking = self._extract_booking_row(row_num)
                if booking:
                    bookings.append(booking)
        
        return bookings
    
    def _extract_booking_row(self, row_num: int) -> Optional[Dict[str, Any]]:
        """Extract a single booking row"""
        booking = {}

        # Booking date (B) - use smart parser for flexible format handling
        booking_date_raw = self._get_cell_value(f'B{row_num}')
        if booking_date_raw:
            parsed_date = parse_date(booking_date_raw)
            if parsed_date:
                # Store both display format and ISO format
                booking['booking_date'] = parsed_date.strftime('%d/%m/%Y')
                booking['pickup_date'] = parsed_date.strftime('%d/%m/%Y')
                booking['booking_date_iso'] = parsed_date.isoformat()  # For DB
            else:
                # Fallback to string if parsing fails
                booking['booking_date'] = str(booking_date_raw)

        # Pickup time (C) - use smart parser for flexible format handling
        pickup_time_raw = self._get_cell_value(f'C{row_num}')
        if pickup_time_raw:
            parsed_time = parse_time(pickup_time_raw)
            if parsed_time:
                booking['pickup_time'] = parsed_time.strftime('%H:%M')
                booking['pickup_time_str'] = parsed_time.strftime('%H:%M:%S')  # For DB
            else:
                # Fallback to string if parsing fails
                booking['pickup_time'] = str(pickup_time_raw)
        
        # Invoice number (D)
        invoice = self._get_cell_value(f'D{row_num}')
        if invoice:
            # Handle multi-line invoices
            booking['invoice_number'] = str(invoice).strip().replace('\n', ', ')
        
        # Available time at warehouse (E)
        avail_time = self._get_cell_value(f'E{row_num}')
        if avail_time:
            booking['available_time'] = str(avail_time)
        
        # Goods name (F)
        goods = self._get_cell_value(f'F{row_num}')
        if goods:
            booking['goods_name'] = str(goods).strip()
        
        # Package quantity (G) - may contain "1 pallet 3thùng" format
        pkg_qty = self._get_cell_value(f'G{row_num}')
        if pkg_qty:
            qty_str = str(pkg_qty).strip()
            booking['package_quantity_raw'] = qty_str
            
            # Parse and preserve unit breakdown
            # Include both Vietnamese and English unit names
            unit_pattern = r'(\d+)\s*(pallet|pallets|thùng|kiện|khay|túi|hộp|carton|cartons|box|boxes|ctn|ctns|bag|bags)?'
            matches = re.findall(unit_pattern, qty_str.lower())
            
            # Build detailed breakdown
            unit_breakdown = []
            total_qty = 0
            for num, unit in matches:
                if num:
                    qty = int(num)
                    total_qty += qty
                    if unit:
                        unit_breakdown.append(f"{qty} {unit}")
                    else:
                        unit_breakdown.append(str(qty))
            
            if total_qty > 0:
                booking['package_quantity'] = total_qty
            
            # Keep the breakdown for display
            if unit_breakdown:
                booking['package_breakdown'] = unit_breakdown
            
            # Set display string preserving both units
            booking['package_display'] = qty_str
        
        # Weight (H) - use smart parser for number handling
        weight = self._get_cell_value(f'H{row_num}')
        if weight:
            parsed_weight = parse_number(weight)
            booking['weight_kg'] = parsed_weight if parsed_weight else weight
        
        # Delivery location/notes (I) - specific location like "XƯỞNG 2"
        delivery = self._get_cell_value(f'I{row_num}')
        if delivery:
            delivery_str = str(delivery).strip()
            booking['delivery_notes'] = delivery_str
            booking['delivery_location'] = delivery_str  # Specific location (e.g., XƯỞNG 2)
            # Always treat column I as delivery info
            booking['delivery_address'] = delivery_str
        
        # Delivery time (J)
        delivery_time = self._get_cell_value(f'J{row_num}')
        if delivery_time:
            if isinstance(delivery_time, (time, datetime)):
                booking['delivery_time'] = delivery_time.strftime('%H:%M') if hasattr(delivery_time, 'strftime') else str(delivery_time)
            else:
                booking['delivery_time'] = str(delivery_time)
        
        return booking if booking else None
    
    def _build_raw_text(self, result: Dict) -> str:
        """Build combined text for AI processing"""
        lines = []

        meta = result.get('metadata', {})

        if meta.get('customer_code'):
            lines.append(f"Khách hàng: {meta['customer_code']}")

        if meta.get('contact_person'):
            lines.append(f"Người book: {meta['contact_person']}")

        if meta.get('pickup_full_address'):
            lines.append(f"Địa chỉ lấy hàng: {meta['pickup_full_address']}")

        if meta.get('delivery_company'):
            lines.append(f"Công ty nhận: {meta['delivery_company']}")

        if meta.get('delivery_full_address'):
            lines.append(f"Địa chỉ giao (chung): {meta['delivery_full_address']}")

        if meta.get('recipient_info'):
            lines.append(f"Thông tin người nhận: {meta['recipient_info']}")

        for i, booking in enumerate(result.get('bookings', []), 1):
            lines.append(f"\n--- Booking {i} ---")
            if booking.get('booking_date'):
                lines.append(f"Ngày: {booking['booking_date']}")
            if booking.get('pickup_time'):
                lines.append(f"Giờ lấy hàng: {booking['pickup_time']}")
            if booking.get('invoice_number'):
                lines.append(f"Invoice: {booking['invoice_number']}")
            if booking.get('goods_name'):
                lines.append(f"Hàng hóa: {booking['goods_name']}")
            if booking.get('package_quantity_raw'):
                lines.append(f"Số lượng: {booking['package_quantity_raw']}")
            if booking.get('weight_kg'):
                lines.append(f"Trọng lượng: {booking['weight_kg']} kg")
            # Include delivery location
            if booking.get('delivery_location') or booking.get('delivery_address'):
                delivery = booking.get('delivery_location') or booking.get('delivery_address')
                lines.append(f"Điểm giao cụ thể: {delivery}")
            if booking.get('delivery_time'):
                lines.append(f"Giờ giao: {booking['delivery_time']}")

        return '\n'.join(lines)
    
    def _build_summary(self, result: Dict) -> str:
        """Build a short summary"""
        meta = result.get('metadata', {})
        bookings = result.get('bookings', [])
        
        parts = []
        if meta.get('company_name'):
            parts.append(meta['company_name'])
        
        if bookings:
            b = bookings[0]
            if b.get('booking_date'):
                parts.append(f"ngày {b['booking_date']}")
            if b.get('pickup_time'):
                parts.append(f"lúc {b['pickup_time']}")
            if b.get('package_quantity_raw'):
                parts.append(f"{b['package_quantity_raw']}")
        
        return ', '.join(parts) if parts else 'Booking form'


def is_booking_form(file_path: str) -> bool:
    """
    Detect if a file is a trucking booking form.
    Check for characteristic cells and filename patterns.
    """
    try:
        filename = os.path.basename(file_path).lower()

        # Check filename patterns
        if any(kw in filename for kw in ['book xe', 'booking', 'phiếu book', 'phieu book']):
            return True

        wb = load_workbook(file_path, data_only=True, read_only=True)
        sheet = wb.active

        # Check for booking form indicators
        indicators = [
            sheet['A6'].value and 'BOOKING' in str(sheet['A6'].value).upper(),
            sheet['A7'].value and 'book xe' in str(sheet['A7'].value).lower(),
            sheet['B13'].value and 'ngày' in str(sheet['B13'].value).lower(),
        ]

        wb.close()
        return sum(indicators) >= 2

    except:
        return False
