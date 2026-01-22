# backend/app/ai/excel/excel_reader.py

import pandas as pd
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass
from enum import Enum
import openpyxl
from openpyxl.utils import get_column_letter


class ExcelFormatType(Enum):
    """Loại format của Excel file"""
    TABLE = "table"       # Dạng bảng với header row
    FORM = "form"         # Dạng form key-value
    MIXED = "mixed"       # Dạng mixed (header + table)
    UNKNOWN = "unknown"   # Không xác định


class ExcelReaderError(Exception):
    """Exception raised by ExcelReader"""
    pass


@dataclass
class ExcelFormat:
    """Thông tin format của Excel file"""
    format_type: ExcelFormatType
    header_row: Optional[int]        # Row chứa header (0-indexed), None nếu form
    data_start_row: int              # Row bắt đầu data
    data_end_row: int                # Row kết thúc data
    data_columns: List[str]          # Các column có data (A, B, C...)
    has_merged_cells: bool           # Có merged cells không
    form_pairs: Optional[List[Tuple[str, str]]] = None  # Key-value pairs nếu là form


@dataclass  
class RawExcelData:
    """Raw data đọc từ Excel"""
    sheets: List[str]                # Tên các sheets
    active_sheet: str                # Sheet đang active
    headers: Optional[List[str]]     # Headers nếu có
    rows: List[List[Any]]            # Data rows
    format_info: ExcelFormat         # Format information
    raw_df: pd.DataFrame             # Raw pandas DataFrame


class ExcelReader:
    """
    Đọc và phân tích cấu trúc Excel file
    """
    
    def __init__(self):
        self.workbook = None
        self.active_sheet = None
    
    def read(self, file_path: str, sheet_name: Optional[str] = None) -> RawExcelData:
        """
        Đọc Excel file và trả về raw data
        
        Args:
            file_path: Đường dẫn file Excel
            sheet_name: Tên sheet cần đọc (None = active sheet)
        
        Returns:
            RawExcelData object
        """
        # Load workbook với openpyxl để có thêm metadata
        self.workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        sheets = self.workbook.sheetnames
        self.active_sheet = sheet_name or self.workbook.active.title
        
        # Đọc với pandas
        df = pd.read_excel(file_path, sheet_name=self.active_sheet, header=None)
        
        # Detect format
        format_info = self._detect_format(df)
        
        # Extract headers và rows dựa trên format
        headers = None
        rows = []
        
        if format_info.format_type == ExcelFormatType.TABLE:
            if format_info.header_row is not None:
                headers = df.iloc[format_info.header_row].tolist()
                # Clean headers
                headers = [str(h).strip() if pd.notna(h) else f"Column_{i}" 
                          for i, h in enumerate(headers)]
            
            rows = df.iloc[format_info.data_start_row:format_info.data_end_row + 1].values.tolist()
        
        elif format_info.format_type == ExcelFormatType.FORM:
            # Form format - extract key-value pairs
            format_info.form_pairs = self._extract_form_pairs(df)
            rows = [[k, v] for k, v in format_info.form_pairs]
        
        elif format_info.format_type == ExcelFormatType.MIXED:
            # Mixed format - có cả header info và table
            # TODO: Handle mixed format
            pass
        
        return RawExcelData(
            sheets=sheets,
            active_sheet=self.active_sheet,
            headers=headers,
            rows=rows,
            format_info=format_info,
            raw_df=df
        )
    
    def _detect_format(self, df: pd.DataFrame) -> ExcelFormat:
        """
        Phát hiện format của Excel file
        
        Heuristics:
        - Table: Row đầu có nhiều giá trị text, các row sau có pattern tương tự
        - Form: Column A chứa labels, Column B chứa values
        - Mixed: Có section header + table bên dưới
        """
        if df.empty:
            return ExcelFormat(
                format_type=ExcelFormatType.UNKNOWN,
                header_row=None,
                data_start_row=0,
                data_end_row=0,
                data_columns=[],
                has_merged_cells=False
            )
        
        # Check for merged cells
        ws = self.workbook[self.active_sheet]
        has_merged = len(ws.merged_cells.ranges) > 0
        
        # Analyze structure
        num_rows, num_cols = df.shape
        
        # Check if it's a form (key-value pairs in columns A-B)
        if self._is_form_format(df):
            return ExcelFormat(
                format_type=ExcelFormatType.FORM,
                header_row=None,
                data_start_row=0,
                data_end_row=num_rows - 1,
                data_columns=['A', 'B'],
                has_merged_cells=has_merged
            )
        
        # Find header row for table format
        header_row = self._find_header_row(df)
        
        # Find data boundaries
        data_start, data_end = self._find_data_boundaries(df, header_row)
        
        # Get columns with data
        data_cols = self._get_data_columns(df)
        
        return ExcelFormat(
            format_type=ExcelFormatType.TABLE,
            header_row=header_row,
            data_start_row=data_start,
            data_end_row=data_end,
            data_columns=data_cols,
            has_merged_cells=has_merged
        )
    
    def _is_form_format(self, df: pd.DataFrame) -> bool:
        """
        Kiểm tra xem có phải dạng form key-value không
        
        Criteria:
        - Chủ yếu 2 columns
        - Column A chứa text labels (kết thúc bằng ":" hoặc là keyword)
        - Column B chứa values
        """
        if df.shape[1] < 2:
            return False
        
        # Chỉ check 2 columns đầu
        col_a = df.iloc[:, 0].dropna()
        col_b = df.iloc[:, 1].dropna()
        
        if len(col_a) < 3:
            return False
        
        # Check xem column A có pattern của labels không
        label_patterns = [':', 'date', 'time', 'customer', 'vehicle', 
                         'ngày', 'giờ', 'khách', 'xe', 'hàng']
        
        label_count = 0
        for val in col_a:
            val_str = str(val).lower().strip()
            if any(p in val_str for p in label_patterns):
                label_count += 1
        
        # Nếu >50% rows có label pattern -> form format
        return label_count / len(col_a) > 0.5
    
    def _find_header_row(self, df: pd.DataFrame) -> Optional[int]:
        """
        Tìm row chứa header
        
        Heuristics:
        - Row có nhiều giá trị text
        - Row có các keyword phổ biến (date, time, customer, etc.)
        - Row ngay trước các data rows
        """
        header_keywords = [
            'stt', 'no', 'ngày', 'date', 'giờ', 'time', 
            'khách', 'customer', 'xe', 'vehicle', 'hàng', 'cargo',
            'điểm', 'location', 'ghi chú', 'note', 'sl', 'qty'
        ]
        
        best_row = None
        best_score = 0
        
        # Check first 10 rows
        for i in range(min(10, len(df))):
            row = df.iloc[i]
            score = 0
            
            non_empty = row.dropna()
            if len(non_empty) < 2:
                continue
            
            # Check for keywords
            for val in non_empty:
                val_str = str(val).lower().strip()
                if any(kw in val_str for kw in header_keywords):
                    score += 2
                # Bonus nếu là text ngắn (likely header)
                if len(val_str) < 30 and not val_str.replace('.', '').replace(',', '').isdigit():
                    score += 1
            
            if score > best_score:
                best_score = score
                best_row = i
        
        return best_row if best_score >= 3 else 0
    
    def _find_data_boundaries(self, df: pd.DataFrame, header_row: Optional[int]) -> Tuple[int, int]:
        """Tìm row bắt đầu và kết thúc của data"""
        start_row = (header_row + 1) if header_row is not None else 0
        
        # Tìm row cuối có data
        end_row = len(df) - 1
        for i in range(len(df) - 1, start_row - 1, -1):
            if df.iloc[i].notna().any():
                end_row = i
                break
        
        return start_row, end_row
    
    def _get_data_columns(self, df: pd.DataFrame) -> List[str]:
        """Lấy danh sách columns có data"""
        cols = []
        for i in range(df.shape[1]):
            if df.iloc[:, i].notna().any():
                cols.append(get_column_letter(i + 1))
        return cols
    
    def _extract_form_pairs(self, df: pd.DataFrame) -> List[Tuple[str, str]]:
        """Extract key-value pairs từ form format"""
        pairs = []
        
        for i in range(len(df)):
            key = df.iloc[i, 0] if pd.notna(df.iloc[i, 0]) else None
            value = df.iloc[i, 1] if df.shape[1] > 1 and pd.notna(df.iloc[i, 1]) else None
            
            if key is not None:
                # Clean key
                key_str = str(key).strip().rstrip(':')
                value_str = str(value).strip() if value is not None else ""
                pairs.append((key_str, value_str))
        
        return pairs
    
    def get_preview(self, file_path: str, max_rows: int = 5) -> Dict[str, Any]:
        """
        Lấy preview data để hiển thị cho user xác nhận
        """
        data = self.read(file_path)
        
        preview_rows = data.rows[:max_rows]
        
        return {
            "format_type": data.format_info.format_type.value,
            "headers": data.headers,
            "sample_rows": preview_rows,
            "total_rows": len(data.rows),
            "sheets": data.sheets,
            "active_sheet": data.active_sheet
        }
