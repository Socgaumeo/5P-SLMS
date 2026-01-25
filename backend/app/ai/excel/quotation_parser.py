# backend/app/ai/excel/quotation_parser.py
"""
Parser for quotation/packing service Excel files.
Handles multi-sheet files like ASGL Quotation with Debit/Estimate sheets.
"""

import re
import logging
import os
from datetime import datetime, date, time
from typing import Dict, Any, List, Optional
from openpyxl import load_workbook

from app.ai.utils.smart_parser import (
    parse_date, format_date_iso,
    parse_number, normalize_unit
)

logger = logging.getLogger(__name__)


class QuotationParser:
    """
    Parse quotation Excel files with Debit/Estimate sheets.

    Structure:
    - Debit sheet: Contains customer info, items, quantities, pricing
    - Estimate sheet: Contains dimension details before/after packing
    """

    def __init__(self, file_path: str = None):
        self.workbook = None
        self.file_path = file_path
        self.filename = os.path.basename(file_path) if file_path else None

    def parse(self, file_path: str) -> Dict[str, Any]:
        """
        Parse a quotation Excel file.

        Returns:
            Dict containing parsed data from all sheets
        """
        try:
            self.file_path = file_path
            self.filename = os.path.basename(file_path)
            self.workbook = load_workbook(file_path, data_only=True)

            result = {
                'file_type': 'quotation',
                'service_type': 'PACKING',  # Default service type
                'metadata': {},
                'items': [],
                'debit': {},
                'estimate': {},
                'raw_text': '',
                'summary': ''
            }

            # Parse each relevant sheet
            sheet_names = [s.lower() for s in self.workbook.sheetnames]

            if 'debit' in sheet_names:
                idx = sheet_names.index('debit')
                result['debit'] = self._parse_debit_sheet(self.workbook.worksheets[idx])
                result['metadata'] = result['debit'].get('metadata', {})
                result['items'] = result['debit'].get('items', [])

            if 'estimate' in sheet_names:
                idx = sheet_names.index('estimate')
                result['estimate'] = self._parse_estimate_sheet(self.workbook.worksheets[idx])
                # Merge dimension info into items
                self._merge_estimate_to_items(result)

            # If no debit/estimate, try to parse quotation sheet
            if not result['items'] and 'quotation' in sheet_names:
                idx = sheet_names.index('quotation')
                result['quotation'] = self._parse_quotation_sheet(self.workbook.worksheets[idx])
                result['metadata'] = result['quotation'].get('metadata', {})

            # Build text summary
            result['raw_text'] = self._build_raw_text(result)
            result['summary'] = self._build_summary(result)

            return result

        except Exception as e:
            logger.error(f"Error parsing quotation file: {e}")
            import traceback
            traceback.print_exc()
            return {
                'file_type': 'quotation',
                'metadata': {},
                'items': [],
                'raw_text': f"Error parsing file: {str(e)}",
                'summary': 'Parse error'
            }

    def _get_cell_value(self, sheet, coord: str) -> Any:
        """Get cell value safely"""
        try:
            return sheet[coord].value
        except:
            return None

    def _parse_debit_sheet(self, sheet) -> Dict[str, Any]:
        """Parse Debit sheet"""
        result = {'metadata': {}, 'items': [], 'totals': {}}

        # Extract metadata from header rows
        for row in range(1, 15):
            for col in ['A', 'B']:
                val = self._get_cell_value(sheet, f'{col}{row}')
                if val:
                    val_str = str(val).strip()

                    # Date
                    if val_str.lower().startswith('date:'):
                        result['metadata']['date'] = val_str.replace('Date:', '').strip()

                    # Customer name
                    if "customer's name:" in val_str.lower() or 'customer:' in val_str.lower():
                        result['metadata']['customer_name'] = re.sub(
                            r"(?:customer's name:|customer:)\s*", '', val_str, flags=re.IGNORECASE
                        ).strip()

                    # Address
                    if val_str.lower().startswith('address:'):
                        result['metadata']['address'] = val_str.replace('Address:', '').strip()

                    # Tax code
                    if 'tax code:' in val_str.lower():
                        result['metadata']['tax_code'] = val_str.replace('Tax code:', '').strip()

        # Find header row (contains "No.", "Asset Name", etc.)
        header_row = None
        for row in range(1, 20):
            val_a = self._get_cell_value(sheet, f'A{row}')
            val_b = self._get_cell_value(sheet, f'B{row}')
            if val_a and str(val_a).strip().lower() in ['no.', 'no', 'stt']:
                header_row = row
                break

        if not header_row:
            return result

        # Parse column headers
        columns = {}
        for col_idx in range(1, sheet.max_column + 1):
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col_idx)
            val = self._get_cell_value(sheet, f'{col_letter}{header_row}')
            if val:
                col_name = str(val).strip().lower()
                columns[col_letter] = col_name

        # Parse data rows
        for row in range(header_row + 1, sheet.max_row + 1):
            stt = self._get_cell_value(sheet, f'A{row}')

            # Skip total rows
            if stt and str(stt).lower() in ['total', 'tổng']:
                # Extract totals
                for col, name in columns.items():
                    val = self._get_cell_value(sheet, f'{col}{row}')
                    if val and isinstance(val, (int, float)):
                        result['totals'][name] = val
                continue

            # Skip non-data rows
            if not stt or not isinstance(stt, (int, float)):
                continue

            item = {'stt': int(stt)}

            # Extract item data based on column headers
            for col, name in columns.items():
                val = self._get_cell_value(sheet, f'{col}{row}')
                if val is not None:
                    # Map common column names
                    if 'asset' in name and 'name' in name:
                        item['asset_name'] = str(val)
                    elif name == 'b' or 'asset' in name:
                        item['asset_no'] = str(val)
                    elif name in ['note', 'ghi chú']:
                        item['note'] = str(val)
                    elif name in ['plt', 'pallet']:
                        item['pallet_count'] = val
                    elif name in ['n.w', 'nw', 'net weight']:
                        item['net_weight_kg'] = val
                    elif name in ['g.w', 'gw', 'gross weight']:
                        item['gross_weight_kg'] = val
                    elif name in ['cbm', 'cbm/pallet']:
                        item['cbm'] = val
                    elif name in ['m2', 'm2/pallet']:
                        item['m2'] = val

            if item.get('asset_no') or item.get('asset_name'):
                result['items'].append(item)

        # Extract total amount
        for row in range(header_row + 1, sheet.max_row + 1):
            val_a = self._get_cell_value(sheet, f'A{row}')
            if val_a and 'total amount' in str(val_a).lower():
                # Look for amount value in columns E-J
                for col in ['E', 'F', 'G', 'H', 'I', 'J']:
                    val = self._get_cell_value(sheet, f'{col}{row}')
                    if val and isinstance(val, (int, float)):
                        result['totals']['total_amount'] = val
                        break

        return result

    def _parse_estimate_sheet(self, sheet) -> Dict[str, Any]:
        """Parse Estimate sheet with dimension details"""
        result = {'items': [], 'totals': {}}

        # Find header row
        header_row = None
        for row in range(1, 10):
            val_a = self._get_cell_value(sheet, f'A{row}')
            val_b = self._get_cell_value(sheet, f'B{row}')
            if val_a and str(val_a).strip().lower() in ['stt', 'no.', 'no']:
                header_row = row
                break

        if not header_row:
            # Try row 2 as header
            header_row = 2

        # Parse data rows
        for row in range(header_row + 1, sheet.max_row + 1):
            stt = self._get_cell_value(sheet, f'A{row}')

            # Skip non-data rows
            if not stt or not isinstance(stt, (int, float)):
                continue

            item = {
                'stt': int(stt),
                'asset_no': self._get_cell_value(sheet, f'B{row}'),
                'name': self._get_cell_value(sheet, f'C{row}'),
                # Dimensions before packing
                'before_length_m': self._get_cell_value(sheet, f'D{row}'),
                'before_width_m': self._get_cell_value(sheet, f'E{row}'),
                'before_height_m': self._get_cell_value(sheet, f'F{row}'),
                'weight_kg': self._get_cell_value(sheet, f'G{row}'),
                'quantity': self._get_cell_value(sheet, f'H{row}'),
                # Dimensions after packing
                'after_length_m': self._get_cell_value(sheet, f'L{row}'),
                'after_width_m': self._get_cell_value(sheet, f'M{row}'),
                'after_height_m': self._get_cell_value(sheet, f'N{row}'),
            }

            # Calculate CBM if dimensions available
            if item.get('before_length_m') and item.get('before_width_m') and item.get('before_height_m'):
                item['cbm_before'] = (
                    float(item['before_length_m']) *
                    float(item['before_width_m']) *
                    float(item['before_height_m'])
                )

            result['items'].append(item)

        return result

    def _parse_quotation_sheet(self, sheet) -> Dict[str, Any]:
        """Parse Quotation sheet for pricing info"""
        result = {'metadata': {}, 'services': []}

        # Extract customer info
        for row in range(1, 10):
            val_a = self._get_cell_value(sheet, f'A{row}')
            if val_a:
                val_str = str(val_a).strip()
                if val_str.lower().startswith('customer:'):
                    result['metadata']['customer_name'] = val_str.replace('Customer:', '').strip()
                elif val_str.lower().startswith('address:'):
                    result['metadata']['address'] = val_str.replace('Address:', '').strip()

        return result

    def _merge_estimate_to_items(self, result: Dict):
        """Merge estimate dimensions into debit items"""
        estimate_items = result.get('estimate', {}).get('items', [])
        debit_items = result.get('items', [])

        # Create lookup by asset_no
        estimate_lookup = {}
        for item in estimate_items:
            if item.get('asset_no'):
                estimate_lookup[str(item['asset_no'])] = item

        # Merge dimensions
        for item in debit_items:
            asset_no = str(item.get('asset_no', ''))
            if asset_no in estimate_lookup:
                est = estimate_lookup[asset_no]
                item['before_length_m'] = est.get('before_length_m')
                item['before_width_m'] = est.get('before_width_m')
                item['before_height_m'] = est.get('before_height_m')
                item['after_length_m'] = est.get('after_length_m')
                item['after_width_m'] = est.get('after_width_m')
                item['after_height_m'] = est.get('after_height_m')

    def _build_raw_text(self, result: Dict) -> str:
        """Build raw text for AI processing"""
        lines = []

        lines.append(f"Loại file: Quotation/Packing Service")

        meta = result.get('metadata', {})
        if meta.get('customer_name'):
            lines.append(f"Khách hàng: {meta['customer_name']}")
        if meta.get('date'):
            lines.append(f"Ngày: {meta['date']}")
        if meta.get('address'):
            lines.append(f"Địa chỉ: {meta['address']}")

        items = result.get('items', [])
        if items:
            lines.append(f"\n=== Danh sách hàng hóa ({len(items)} items) ===")
            for item in items:
                item_desc = []
                if item.get('asset_no'):
                    item_desc.append(f"Asset: {item['asset_no']}")
                if item.get('note') or item.get('name'):
                    item_desc.append(f"Tên: {item.get('note') or item.get('name')}")
                if item.get('pallet_count'):
                    item_desc.append(f"Pallet: {item['pallet_count']}")
                if item.get('net_weight_kg'):
                    item_desc.append(f"NW: {item['net_weight_kg']}kg")
                if item.get('cbm'):
                    item_desc.append(f"CBM: {item['cbm']}")
                if item.get('before_length_m'):
                    item_desc.append(f"Kích thước: {item['before_length_m']}x{item.get('before_width_m')}x{item.get('before_height_m')}m")
                lines.append(f"  • {', '.join(item_desc)}")

        totals = result.get('debit', {}).get('totals', {})
        if totals:
            lines.append(f"\n=== Tổng cộng ===")
            if totals.get('total_amount'):
                lines.append(f"Tổng tiền: {totals['total_amount']:,.0f} VND")

        return '\n'.join(lines)

    def _build_summary(self, result: Dict) -> str:
        """Build short summary"""
        meta = result.get('metadata', {})
        items = result.get('items', [])

        parts = []
        if meta.get('customer_name'):
            parts.append(meta['customer_name'])
        if items:
            parts.append(f"{len(items)} items")
        if meta.get('date'):
            parts.append(meta['date'])

        return ', '.join(parts) if parts else 'Quotation'


def is_quotation_file(file_path: str) -> bool:
    """
    Detect if a file is a quotation/packing service file.
    """
    try:
        filename = os.path.basename(file_path).lower()

        # Check filename patterns
        if any(kw in filename for kw in ['quotation', 'debit', 'estimate', 'packing']):
            return True

        # Check sheet names
        wb = load_workbook(file_path, data_only=True, read_only=True)
        sheet_names = [s.lower() for s in wb.sheetnames]
        wb.close()

        # Has Debit or Estimate sheet
        if 'debit' in sheet_names or 'estimate' in sheet_names:
            return True

        return False

    except:
        return False
