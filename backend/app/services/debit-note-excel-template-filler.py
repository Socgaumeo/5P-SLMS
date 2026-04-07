"""
Debit Note Excel Template Filler — write job data into openpyxl workbook cells.

Two fill modes:
1. column-based (DAINESE style): copies reference row, writes per-row formulas,
   updates totals/SUM ranges, handles PurchPurchaseOrder cross-sheet refs.
2. cell-based (simple): direct cell_ref → value mapping.
"""

import copy
import logging
from openpyxl.cell.cell import MergedCell

import importlib
_resolver = importlib.import_module("app.services.debit-note-job-data-resolver")
_format_value = _resolver._format_value
_evaluate_formula = _resolver._evaluate_formula

logger = logging.getLogger(__name__)


def fill_template(wb, field_mapping: dict, data: dict) -> None:
    """
    Dispatch to the correct fill mode based on field_mapping structure.
    column-based: requires 'columns' + 'data_start_row' keys.
    cell-based: direct cell ref keys like 'B2', 'D10'.
    """
    sheet_name = field_mapping.get('sheet')
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    if 'columns' in field_mapping and 'data_start_row' in field_mapping:
        _fill_column_based(ws, field_mapping, data, wb)
    else:
        _fill_cell_based(ws, field_mapping, data)


# ─── Column-based fill (DAINESE style) ────────────────────────────────────────

def _fill_column_based(ws, field_mapping: dict, data: dict, wb=None) -> None:
    """
    Fill template rows from a reference row, preserving all formatting and formulas.

    Steps:
    1. Insert extra rows (if n_jobs > 1) by cloning ref_row style
    2. Write data values into mapped columns for each job row
    3. Write per-row Excel formulas (e.g. H=F*G, V=SUM(K:U))
    4. Update totals row SUM formulas to cover actual data range
    5. Update optional post-totals rows (CO template VAT / grand total)
    6. Update PurchPurchaseOrder cross-sheet refs (VAN_CHUYEN template)
    """
    columns = field_mapping.get('columns', {})
    start_row = field_mapping.get('data_start_row', 2)
    ref_row = field_mapping.get('ref_row', start_row)
    row_formulas = field_mapping.get('row_formulas', {})
    totals_cfg = field_mapping.get('totals_config', {})

    jobs = data.get('jobs_detail', [])
    if not jobs:
        return

    n_jobs = len(jobs)

    # Step 1: clone ref_row for additional jobs
    _insert_data_rows_from_ref(ws, ref_row, start_row, n_jobs)

    # Steps 2-3: write data + row formulas
    for idx, job in enumerate(jobs):
        row = start_row + idx
        _write_job_row(ws, row, idx, job, columns, row_formulas)

    # Step: auto-widen text columns (C, E, etc.) based on content
    _auto_widen_text_columns(ws, columns, start_row, n_jobs)

    # Step 4: update totals SUM formulas
    data_end = start_row + n_jobs - 1
    totals_row = data_end + totals_cfg.get('totals_row_offset', 1)
    _write_totals_row(ws, totals_cfg.get('sum_cols', []), start_row, data_end, totals_row)

    # Step 5: CO-style post-totals (VAT row, grand total row)
    _write_post_totals(ws, totals_cfg.get('post_totals', []), totals_row, start_row, data_end)

    # Step 6: VAN_CHUYEN PurchPurchaseOrder cross-sheet refs
    ppo_cfg = field_mapping.get('purch_purchase_order', {})
    if ppo_cfg and wb is not None:
        ppo_sheet = ppo_cfg.get('sheet', 'PurchPurchaseOrder')
        if ppo_sheet in wb.sheetnames:
            _update_purch_purchase_order(wb[ppo_sheet], ppo_cfg, n_jobs)


def _write_job_row(ws, row: int, idx: int, job: dict, columns: dict, row_formulas: dict) -> None:
    """Write one job's data and formula cells into the given row."""
    # Data columns
    for col_letter, col_cfg in columns.items():
        if not isinstance(col_cfg, dict):
            continue
        field = col_cfg.get('field', '')
        fmt = col_cfg.get('format', 'text')
        value = idx + 1 if field == 'stt' else job.get(field, '')
        formatted = _format_value(value, fmt)
        try:
            cell = ws[f"{col_letter}{row}"]
            if not isinstance(cell, MergedCell):
                cell.value = formatted
                if fmt == 'currency':
                    cell.number_format = '#,##0'
        except Exception as e:
            logger.warning(f"Error writing data {col_letter}{row}: {e}")

    # Per-row Excel formulas
    for col_letter, pattern in row_formulas.items():
        formula = pattern.replace('{row}', str(row))
        try:
            cell = ws[f"{col_letter}{row}"]
            if not isinstance(cell, MergedCell):
                # Ensure formula is written as formula, not text
                cell.data_type = 'f'
                cell.value = formula
        except Exception as e:
            logger.warning(f"Error writing formula {col_letter}{row}: {e}")


def _write_totals_row(ws, sum_cols: list, data_start: int, data_end: int, totals_row: int) -> None:
    """Write SUM formulas into totals row for each listed column."""
    for col in sum_cols:
        try:
            cell = ws[f"{col}{totals_row}"]
            if not isinstance(cell, MergedCell):
                cell.data_type = 'f'
                cell.value = f"=SUM({col}{data_start}:{col}{data_end})"
                cell.number_format = '#,##0'
        except Exception as e:
            logger.warning(f"Error writing totals {col}{totals_row}: {e}")


def _write_post_totals(ws, post_totals: list, totals_row: int, data_start: int, data_end: int) -> None:
    """Write CO-template post-totals rows (VAT, grand total) with correct row references."""
    for pt in post_totals:
        row_num = totals_row + pt.get('row_offset', 0)
        col = pt.get('col', '')
        formula_tpl = pt.get('formula', '')
        if not formula_tpl or not col:
            continue
        formula = (formula_tpl
                   .replace('{totals_row}', str(totals_row))
                   .replace('{totals_row_plus1}', str(totals_row + 1))
                   .replace('{totals_row_plus2}', str(totals_row + 2))
                   .replace('{data_start}', str(data_start))
                   .replace('{data_end}', str(data_end)))
        try:
            cell = ws[f"{col}{row_num}"]
            if not isinstance(cell, MergedCell):
                cell.data_type = 'f'
                cell.value = formula
        except Exception as e:
            logger.warning(f"Error writing post-totals {col}{row_num}: {e}")


def _update_purch_purchase_order(ws_ppo, ppo_cfg: dict, n_jobs: int) -> None:
    """
    Update PurchPurchaseOrder cross-sheet references for VAN_CHUYEN template.

    Template has row 31 as reference (job 1) + row 32 (Price unit).
    For n_jobs jobs: insert rows, write cross-sheet refs, add item descriptions.
    """
    first_ppo_row = ppo_cfg.get('first_data_row', 31)
    row_step = ppo_cfg.get('row_step', 2)  # each job = 2 rows (data + "Price unit")
    hd_row_start = ppo_cfg.get('hd_row_start', 15)
    col_map = ppo_cfg.get('col_map', {})
    hd_sheet = ppo_cfg.get('hd_sheet_name', 'HĐ')

    # Insert extra rows for jobs 2..n (job 1 already has rows 31-32)
    if n_jobs > 1:
        insert_at = first_ppo_row + row_step  # after row 32
        ws_ppo.insert_rows(insert_at, (n_jobs - 1) * row_step)

    # Write all job rows
    for job_idx in range(n_jobs):
        ppo_row = first_ppo_row + (job_idx * row_step)
        hd_row = hd_row_start + job_idx

        # Cross-sheet formulas
        for ppo_col, hd_col in col_map.items():
            try:
                cell = ws_ppo[f"{ppo_col}{ppo_row}"]
                if not isinstance(cell, MergedCell):
                    cell.data_type = 'f'
                    cell.value = f"={hd_sheet}!{hd_col}{hd_row}"
            except Exception as e:
                logger.warning(f"PPO ref error {ppo_col}{ppo_row}: {e}")

        # Standard fields
        try:
            ws_ppo[f"A{ppo_row}"] = job_idx + 1
            ws_ppo[f"I{ppo_row}"] = 1
            ws_ppo[f"L{ppo_row}"] = "Chuyến"
            # Item description from HĐ pickup → delivery
            ws_ppo[f"C{ppo_row}"] = f"=CONCATENATE(\"Cước vận chuyển \",{hd_sheet}!C{hd_row},\" - \",{hd_sheet}!E{hd_row})"
            # Price unit row
            ws_ppo[f"P{ppo_row + 1}"] = "Price unit 1.00"
        except Exception:
            pass


# ─── Column width auto-fit ───────────────────────────────────────────────────

def _auto_widen_text_columns(ws, columns: dict, start_row: int, n_jobs: int) -> None:
    """Widen text columns to fit longest content (min 15, max 45 chars)."""
    for col_letter, col_cfg in columns.items():
        if not isinstance(col_cfg, dict):
            continue
        fmt = col_cfg.get('format', 'text')
        if fmt not in ('text', 'date'):
            continue

        max_len = 12  # minimum width
        for r in range(start_row, start_row + n_jobs):
            cell = ws[f"{col_letter}{r}"]
            if cell.value and not isinstance(cell, MergedCell):
                val_len = len(str(cell.value))
                if val_len > max_len:
                    max_len = val_len

        # Clamp between 15 and 45
        width = min(max(max_len + 2, 15), 45)
        ws.column_dimensions[col_letter].width = width


# ─── Row style cloning helpers ────────────────────────────────────────────────

def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    """Copy font/border/fill/alignment/number_format from src_row to dst_row."""
    for col_idx in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col_idx)
        dst = ws.cell(row=dst_row, column=col_idx)
        if isinstance(src, MergedCell) or isinstance(dst, MergedCell):
            continue
        if src.has_style:
            dst.font = copy.copy(src.font)
            dst.border = copy.copy(src.border)
            dst.fill = copy.copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy.copy(src.protection)
            dst.alignment = copy.copy(src.alignment)


def _insert_data_rows_from_ref(ws, ref_row: int, start_row: int, count: int) -> None:
    """
    If count > 1, insert (count-1) blank rows immediately after start_row,
    then copy ref_row styling to all inserted rows.
    ref_row IS start_row — the original template reference row.
    """
    if count > 1:
        ws.insert_rows(start_row + 1, count - 1)
    max_col = ws.max_column
    for i in range(1, count):
        _copy_row_style(ws, ref_row, start_row + i, max_col)


# ─── Cell-based fill (simple mapping) ─────────────────────────────────────────

def _fill_cell_based(ws, field_mapping: dict, data: dict) -> None:
    """Fill worksheet using direct cell_ref → field/formula mapping."""
    for cell_ref, mapping in field_mapping.items():
        if not isinstance(mapping, dict):
            continue
        field = mapping.get('field', '')
        fmt = mapping.get('format', 'text')
        value = (_evaluate_formula(mapping['formula'], data)
                 if mapping.get('formula') else data.get(field, ''))
        formatted = _format_value(value, fmt, mapping.get('date_format', 'DD/MM/YYYY'))
        try:
            cell = ws[cell_ref]
            if isinstance(cell, MergedCell):
                continue
            cell.value = formatted
            if fmt == 'currency':
                cell.number_format = '#,##0'
            elif fmt == 'percentage':
                cell.number_format = '0.00%'
        except Exception as e:
            logger.warning(f"Error writing cell {cell_ref}: {e}")
