"""
Debit Note Generator Service — load Excel templates, fill data, generate debit notes.

Core engine:
1. Load template from server disk (or download from Telegram file_id)
2. Query job data from DB (jobs + job_services + customer info)
3. Fill cells per field_mapping JSON from debit_templates table
4. Return filled workbook as temp file path

Supports: text, currency, date, percentage formats + table ranges for line items.
"""

import io
import os
import re
import tempfile
import logging
import zipfile
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from app.db.supabase_client import get_supabase

import importlib
telegram_downloader = importlib.import_module("app.services.telegram-file-downloader")

logger = logging.getLogger(__name__)

# Template cache directory
TEMPLATE_CACHE_DIR = os.path.join(tempfile.gettempdir(), '5p_debit_templates')
os.makedirs(TEMPLATE_CACHE_DIR, exist_ok=True)


async def load_template_workbook(template: dict):
    """
    Load template Excel workbook.
    Tries local_file_path first, then downloads from Telegram file_id.
    Returns openpyxl Workbook or None.
    """
    # Try local file path
    if template.get('local_file_path') and os.path.exists(template['local_file_path']):
        return load_workbook(template['local_file_path'])

    # Try Telegram file_id
    if template.get('telegram_file_id'):
        cache_path = os.path.join(TEMPLATE_CACHE_DIR, f"{template['id']}.xlsx")

        # Check cache
        if os.path.exists(cache_path):
            return load_workbook(cache_path)

        # Download from Telegram
        file_data = await telegram_downloader.download_telegram_file(template['telegram_file_id'])
        if file_data:
            file_bytes, _ = file_data
            with open(cache_path, 'wb') as f:
                f.write(file_bytes)
            return load_workbook(cache_path)

    logger.error(f"Cannot load template {template['id']}: no file source")
    return None


def resolve_job_data(job_ids: List[int], customer_id: Optional[int] = None) -> Dict[str, Any]:
    """
    Query DB and build data dict for template filling.
    Returns combined data from jobs, job_services, and customer info.
    """
    client = get_supabase()

    # Get jobs
    jobs_result = client.table('jobs').select(
        '*, customers(customer_id, customer_code, company_name, short_name, tax_code, address)'
    ).in_('job_id', job_ids).execute()
    jobs = jobs_result.data or []

    if not jobs:
        return {}

    # Get services for these jobs
    services_result = client.table('job_services').select(
        '*, vendors(vendor_code, short_name, company_name)'
    ).in_('job_id', job_ids).order('scheduled_date').execute()
    services = services_result.data or []

    # Customer info from first job
    customer = jobs[0].get('customers', {}) or {}

    # Calculate totals
    total_revenue = sum(float(j.get('total_revenue') or 0) for j in jobs)
    total_cost = sum(float(j.get('total_cost') or 0) for j in jobs)

    # Build per-job detail rows (each job = 1 row in the debit note)
    # Group services by job_id
    svc_by_job = {}
    for svc in services:
        jid = svc['job_id']
        if jid not in svc_by_job:
            svc_by_job[jid] = []
        svc_by_job[jid].append(svc)

    jobs_detail = []
    for idx, job in enumerate(jobs, 1):
        jid = job['job_id']
        job_svcs = svc_by_job.get(jid, [])
        first_svc = job_svcs[0] if job_svcs else {}

        revenue = float(job.get('total_revenue') or 0)
        cost = float(job.get('total_cost') or 0)

        jobs_detail.append({
            'stt': idx,
            'job_no': job.get('job_no', ''),
            'service_date': first_svc.get('scheduled_date') or job.get('booking_date', ''),
            'declaration_date': first_svc.get('scheduled_date') or job.get('booking_date', ''),

            # Locations
            'origin': first_svc.get('origin_address', ''),
            'destination': first_svc.get('dest_address', ''),
            'route': f"{first_svc.get('origin_address', '')} → {first_svc.get('dest_address', '')}",
            'pickup_location': first_svc.get('origin_address', ''),
            'pickup_province': (first_svc.get('origin_address', '').split(',')[-1].strip() if first_svc.get('origin_address') else ''),
            'delivery_location': first_svc.get('dest_address', ''),
            'delivery_province': (first_svc.get('dest_address', '').split(',')[-1].strip() if first_svc.get('dest_address') else ''),

            # Vehicle / transport
            'license_plate': first_svc.get('license_plate', ''),
            'vehicle_type': job.get('vehicle_type', ''),
            'transport_type': first_svc.get('service_type_code', ''),

            # Fees (from total_revenue as main amount)
            'transport_fee': revenue,
            'customs_fee': revenue,
            'amount': revenue,
            'total': revenue,
            'unit_price': revenue,
            'quantity': 1,

            # Placeholders for fees not yet in DB
            'surcharge': 0,
            'fuel_surcharge': 0,
            'inspection_fee': 0,
            'handling_fee': 0,
            'expense_amount': 0,

            # Document refs
            'declaration_number': '',
            'commercial_invoice': '',
            'bill_of_lading': '',
            'container_no': '',
            'clearance_channel': '',
            'co_number': '',
            'co_form': '',
            'co_date': '',
            'invoice_number': '',
            'invoice_ref': '',
            'receipt_number': '',
            'expense_description': '',
            'note': '',
            'notes': '',
            'weight_kg': '',
        })

    # Build data dictionary
    now = datetime.now()
    data = {
        # Customer info
        'customer_name': customer.get('short_name') or customer.get('company_name', ''),
        'customer_code': customer.get('customer_code', ''),
        'company_name': customer.get('company_name', ''),
        'tax_code': customer.get('tax_code', ''),
        'customer_address': customer.get('address', ''),

        # Job info (first job or aggregated)
        'job_no': jobs[0].get('job_no', '') if len(jobs) == 1 else f"{len(jobs)} jobs",
        'job_count': len(jobs),
        'booking_date': jobs[0].get('booking_date', ''),

        # Financial
        'total_revenue': total_revenue,
        'total_cost': total_cost,
        'profit': total_revenue - total_cost,
        'vat_rate': 0.08,
        'vat_amount': total_revenue * 0.08,
        'grand_total': total_revenue * 1.08,

        # Date
        'month_year': now.strftime('%m/%Y'),
        'current_date': now.strftime('%d/%m/%Y'),
        'month': now.strftime('%m'),
        'year': now.strftime('%Y'),

        # Service line items
        'services': jobs_detail,

        # All jobs list
        'jobs': [{'job_no': j.get('job_no', ''), 'job_id': j.get('job_id')} for j in jobs],

        # Per-job detail rows for column-based templates
        'jobs_detail': jobs_detail,
    }

    return data


def _format_value(value: Any, fmt: str, date_format: str = 'DD/MM/YYYY') -> Any:
    """Format value according to field mapping format specification."""
    if value is None:
        return ''

    if fmt == 'currency':
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0

    if fmt == 'date':
        if isinstance(value, str) and value:
            try:
                dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
                # Convert format string
                py_fmt = date_format.replace('DD', '%d').replace('MM', '%m').replace('YYYY', '%Y')
                return dt.strftime(py_fmt)
            except Exception:
                return value
        return value

    if fmt == 'percentage':
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0

    # Default: text
    return str(value) if value else ''


def _evaluate_formula(formula: str, data: dict) -> Any:
    """
    Evaluate simple formulas like 'total_revenue * 0.08'.
    Only supports basic arithmetic with data field references.
    """
    try:
        # Replace field names with values
        expr = formula
        for key, val in data.items():
            if isinstance(val, (int, float)):
                expr = expr.replace(key, str(val))
        # Safe eval of simple arithmetic
        if re.match(r'^[\d\.\+\-\*\/\(\)\s]+$', expr):
            return eval(expr)
        return 0
    except Exception:
        return 0


def fill_template(wb, field_mapping: dict, data: dict) -> None:
    """
    Fill workbook cells according to field_mapping.

    Supports two mapping formats:
    1. Column-based table (DAINESE style):
       {sheet, columns: {A: {field, format}, B: ...}, data_start_row, totals}
    2. Cell-ref based (simple):
       {B2: {field, format}, D10: {field, format, formula}}
    """
    # Select sheet
    sheet_name = field_mapping.get('sheet')
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # Detect mapping format
    if 'columns' in field_mapping and 'data_start_row' in field_mapping:
        _fill_column_based(ws, field_mapping, data)
    else:
        _fill_cell_based(ws, field_mapping, data)


def _fill_column_based(ws, field_mapping: dict, data: dict) -> None:
    """
    Fill template using column-based mapping.
    Each job becomes a row, columns map to job/service fields.
    """
    columns = field_mapping.get('columns', {})
    start_row = field_mapping.get('data_start_row', 2)
    totals_config = field_mapping.get('totals', {})

    jobs = data.get('jobs_detail', [])
    if not jobs:
        return

    # Copy formatting from first data row as reference
    ref_row = start_row
    ref_styles = {}
    for col_letter in columns:
        cell = ws[f"{col_letter}{ref_row}"]
        ref_styles[col_letter] = {
            'number_format': cell.number_format,
            'font': cell.font.copy() if cell.font else None,
            'alignment': cell.alignment.copy() if cell.alignment else None,
            'border': cell.border.copy() if cell.border else None,
        }

    # Write each job as a row
    for idx, job in enumerate(jobs):
        row = start_row + idx

        for col_letter, col_cfg in columns.items():
            if not isinstance(col_cfg, dict):
                continue

            field = col_cfg.get('field', '')
            fmt = col_cfg.get('format', 'text')
            cell_ref = f"{col_letter}{row}"

            # Resolve value from job data
            if field == 'stt':
                value = idx + 1
            else:
                value = job.get(field, '')

            # Format value
            formatted = _format_value(value, fmt)

            try:
                cell_obj = ws[cell_ref]
                if isinstance(cell_obj, MergedCell):
                    continue

                cell_obj.value = formatted

                # Apply stored formatting from reference row
                style = ref_styles.get(col_letter, {})
                if style.get('number_format'):
                    cell_obj.number_format = style['number_format']
                if style.get('font'):
                    cell_obj.font = style['font']
                if style.get('alignment'):
                    cell_obj.alignment = style['alignment']
                if style.get('border'):
                    cell_obj.border = style['border']

                # Override number format for currency
                if fmt == 'currency':
                    cell_obj.number_format = '#,##0'
            except Exception as e:
                logger.warning(f"Error writing {cell_ref}: {e}")

    # Write totals row
    if totals_config:
        total_row = start_row + len(jobs)
        for col_letter, agg_type in totals_config.items():
            if not isinstance(agg_type, str):
                continue
            cell_ref = f"{col_letter}{total_row}"
            cell_obj = ws[cell_ref]
            if isinstance(cell_obj, MergedCell):
                continue
            if agg_type == 'sum':
                cell_obj.value = f"=SUM({col_letter}{start_row}:{col_letter}{start_row + len(jobs) - 1})"
                cell_obj.number_format = '#,##0'


def _fill_cell_based(ws, field_mapping: dict, data: dict) -> None:
    """Fill using simple cell_ref → field mapping."""
    for cell_ref, mapping in field_mapping.items():
        if not isinstance(mapping, dict):
            continue

        field = mapping.get('field', '')
        fmt = mapping.get('format', 'text')

        if mapping.get('formula'):
            value = _evaluate_formula(mapping['formula'], data)
        else:
            value = data.get(field, '')

        formatted = _format_value(value, fmt, mapping.get('date_format', 'DD/MM/YYYY'))

        try:
            cell_obj = ws[cell_ref]
            if isinstance(cell_obj, MergedCell):
                continue
            cell_obj.value = formatted
            if fmt == 'currency':
                cell_obj.number_format = '#,##0'
            elif fmt == 'percentage':
                cell_obj.number_format = '0.00%'
        except Exception as e:
            logger.warning(f"Error writing cell {cell_ref}: {e}")


async def generate_single(template_id: str, job_ids: List[int]) -> Optional[str]:
    """
    Generate a single debit note.
    Returns path to temporary Excel file, or None on error.
    """
    client = get_supabase()

    # Get template
    result = client.table('debit_templates').select('*').eq('id', template_id).limit(1).execute()
    if not result.data:
        logger.error(f"Template {template_id} not found")
        return None

    template = result.data[0]
    field_mapping = template.get('field_mapping', {})

    # Load workbook
    wb = await load_template_workbook(template)
    if not wb:
        return None

    # Resolve data
    data = resolve_job_data(job_ids, template.get('customer_id'))
    if not data:
        logger.error(f"No job data found for job_ids: {job_ids}")
        return None

    # Fill template
    fill_template(wb, field_mapping, data)

    # Save to temp file
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name)
    tmp.close()

    return tmp.name


async def generate_batch(template_id: str, customer_id: int, month: str) -> Optional[str]:
    """
    Generate debit notes for all jobs of a customer in a month.
    Returns path to temporary ZIP file containing individual Excel files.
    """
    import calendar

    client = get_supabase()

    # Get template
    tmpl_result = client.table('debit_templates').select('*').eq('id', template_id).limit(1).execute()
    if not tmpl_result.data:
        return None
    template = tmpl_result.data[0]

    # Get jobs for this customer + month
    year, mon = month.split('-')
    last_day = calendar.monthrange(int(year), int(mon))[1]
    start = f"{year}-{mon}-01T00:00:00"
    end = f"{year}-{mon}-{last_day}T23:59:59"

    jobs_result = client.table('jobs').select('job_id, job_no').eq(
        'customer_id', customer_id
    ).gte('created_at', start).lte('created_at', end).execute()

    jobs = jobs_result.data or []

    # Filter by service type if template has service_type mapping
    field_mapping = template.get('field_mapping', {})
    svc_type = field_mapping.get('service_type', '')
    svc_type_map = {
        "CO": ["CUS_SUBMITTED", "CUS_PROCESSING", "CUS_APPROVED"],
        "SEA_AIR_IMPORT": ["SEA_IMP", "AIR_IMP"],
        "DOM_CUSTOMS": ["BORDER_IMP", "CUS_SUBMITTED"],
        "TRUCKING": ["TRUCKING_DOM", "TRUCKING_SHORT", "TRUCKING_LONG"],
        "EXPORT": ["SEA_EXP", "AIR_EXP"],
    }
    svc_codes = svc_type_map.get(svc_type, [])
    if svc_codes and jobs:
        job_ids_all = [j['job_id'] for j in jobs]
        svcs = client.table('job_services').select('job_id').in_(
            'job_id', job_ids_all
        ).in_('service_type_code', svc_codes).execute()
        matching = set(s['job_id'] for s in (svcs.data or []))
        jobs = [j for j in jobs if j['job_id'] in matching]

    if not jobs:
        logger.info(f"No jobs found for customer {customer_id} in {month} (svc_type={svc_type})")
        return None

    # Generate for all jobs together (single debit note with all jobs)
    job_ids = [j['job_id'] for j in jobs]
    excel_path = await generate_single(template_id, job_ids)

    if not excel_path:
        return None

    # Wrap in ZIP
    zip_path = tempfile.NamedTemporaryFile(delete=False, suffix='.zip').name
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        customer_result = client.table('customers').select('customer_code, short_name').eq(
            'customer_id', customer_id
        ).limit(1).execute()
        cust_code = customer_result.data[0]['customer_code'] if customer_result.data else 'unknown'
        filename = f"DEBIT_{cust_code}_{month}.xlsx"
        zf.write(excel_path, filename)

    # Cleanup temp excel
    os.unlink(excel_path)

    return zip_path
