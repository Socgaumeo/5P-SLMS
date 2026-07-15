#!/usr/bin/env python3
"""
AP Settlement — Xuất bảng kê chi phí theo vendor/employee (module Công nợ AR/AP).
Gom mọi job_costs CHƯA nằm trong bảng kê (view v_ap_unbilled_costs) của 1 vendor/người
→ xuất Excel + (tùy chọn) tạo ap_bill để track thanh toán.

Usage:
  python3 ap_settlement.py --list-vendors             # liệt kê vendor có chi phí chờ
  python3 ap_settlement.py --vendor 6                 # bảng kê vendor ASGL (id=6) → Excel
  python3 ap_settlement.py --employee 6               # bảng kê 1 người thực hiện
  python3 ap_settlement.py --vendor 6 --create-bill   # xuất + tạo ap_bill (track trả)
"""
import re, sys, argparse, datetime
import psycopg2
from pathlib import Path

ROOT = Path(__file__).parent.parent

def db():
    url = [l.split('=',1)[1].strip() for l in open(ROOT/'.env') if l.startswith('DATABASE_URL')][0]
    u,pw,h,p,d = re.match(r'postgresql://([^:]+):(.+)@([^:@]+):(\d+)/(\w+)', url).groups()
    return psycopg2.connect(host=h, port=int(p), dbname=d, user=u, password=pw, sslmode='require')

def vnd(n):
    return f"{int(n or 0):,}".replace(',', '.')

def list_vendors(cur):
    cur.execute("""
      SELECT vendor_id, vendor_name, COUNT(*), SUM(amount)
      FROM v_ap_unbilled_costs WHERE vendor_id IS NOT NULL
      GROUP BY vendor_id, vendor_name ORDER BY 4 DESC
    """)
    print("== Vendor có chi phí CHỜ thanh toán ==")
    for vid, name, cnt, tot in cur.fetchall():
        print(f"  [{vid}] {name or '?':30} {cnt} khoản  {vnd(tot)}đ")

def build_statement(cur, vendor_id=None, employee_id=None):
    if vendor_id:
        cur.execute("""SELECT cost_id, job_no, cost_date, cost_name, buying_rate, quantity, amount, is_reimbursement,
            plate_number, route, invoice_numbers, declaration_no, bl_awb_no, job_invoice_no
          FROM v_ap_unbilled_costs WHERE vendor_id=%s ORDER BY cost_date, job_no""", (vendor_id,))
        cur2 = cur.connection.cursor()
        cur2.execute("SELECT short_name, company_name FROM vendors WHERE vendor_id=%s", (vendor_id,))
        who = cur2.fetchone()
        title = f"BẢNG KÊ CHI PHÍ PHẢI TRẢ — {who[1] or who[0]}"
    else:
        # employee: join qua job_costs (nếu cost gắn employee) — fallback: chưa hỗ trợ, cần cột
        cur.execute("""SELECT jc.cost_id, j.job_no, jc.created_at::date, jc.cost_name,
            jc.buying_rate, jc.quantity, (jc.buying_rate*COALESCE(jc.quantity,1)), jc.is_reimbursement
          FROM job_costs jc JOIN jobs j ON j.job_id=jc.job_id
          JOIN job_services s ON s.svc_id=jc.svc_id
          WHERE s.employee_id=%s AND jc.buying_rate>0
            AND NOT EXISTS (SELECT 1 FROM ap_bill_items abi WHERE abi.cost_id=jc.cost_id)
          ORDER BY 3, 2""", (employee_id,))
        cur2 = cur.connection.cursor()
        cur2.execute("SELECT full_name FROM employees WHERE employee_id=%s", (employee_id,))
        e = cur2.fetchone()
        title = f"BẢNG KÊ CHI PHÍ NGƯỜI THỰC HIỆN — {e[0] if e else employee_id}"
    rows = cur.fetchall()
    return title, rows

def to_excel(title, rows, out_path):
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "BangKe"
    thin = Side(style='thin'); border = Border(thin,thin,thin,thin)
    ws.merge_cells('A1:G1'); ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = Alignment('center')
    ws['A2'] = f"Ngày xuất: {datetime.date.today().strftime('%d/%m/%Y')}"
    has_ref = rows and len(rows[0]) >= 14
    if has_ref:
        hdr = ['STT','Job No','Ngày','Tên phí','Biển số','Tuyến','INV','Số TK','B/L-AWB','Số HĐ','Đơn giá','SL','Thành tiền']
        ncol = 13
    else:
        hdr = ['STT','Job No','Ngày','Tên phí','Đơn giá','SL','Thành tiền']
        ncol = 7
    ws.append([]); ws.append(hdr)
    hr = ws.max_row
    for c in range(1,ncol+1):
        cell = ws.cell(hr,c); cell.font = Font(bold=True); cell.border = border
        cell.fill = PatternFill('solid', fgColor='DDEBF7'); cell.alignment = Alignment('center')
    total = 0
    for i,r in enumerate(rows,1):
        cid,jn,dt,name,rate,qty,amt,reim = r[:8]
        nm = name + (' (chi hộ)' if reim else '')
        d = dt.strftime('%d/%m') if dt else ''
        if has_ref:
            plate,route,inv,cd,bl,jinv = r[8:14]
            ws.append([i, jn, d, nm, plate or '', route or '', inv or '', cd or '', bl or '', jinv or '', int(rate or 0), float(qty or 1), int(amt or 0)])
        else:
            ws.append([i, jn, d, nm, int(rate or 0), float(qty or 1), int(amt or 0)])
        total += int(amt or 0)
        for c in range(1,ncol+1): ws.cell(ws.max_row,c).border = border
    trow = ['']*(ncol-4) + ['TỔNG CỘNG'] + ['']*2 + [total] if not has_ref else ['','','','TỔNG CỘNG'] + ['']*8 + [total]
    ws.append(trow)
    tr = ws.max_row
    for c in range(1,ncol+1):
        ws.cell(tr,c).font = Font(bold=True); ws.cell(tr,c).border = border
    widths = [6,18,10,34,12,26,16,16,16,16,14,6,16] if has_ref else [6,18,10,40,14,6,16]
    for idx,w in enumerate(widths):
        ws.column_dimensions[chr(65+idx)].width = w
    wb.save(out_path)
    return total

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--list-vendors', action='store_true')
    ap.add_argument('--vendor', type=int)
    ap.add_argument('--employee', type=int)
    ap.add_argument('--create-bill', action='store_true')
    a = ap.parse_args()
    c = db(); cur = c.cursor()
    if a.list_vendors:
        list_vendors(cur); return
    if not (a.vendor or a.employee):
        list_vendors(cur); print("\n→ Thêm --vendor <id> hoặc --employee <id> để xuất bảng kê"); return

    title, rows = build_statement(cur, a.vendor, a.employee)
    if not rows:
        print("Không có chi phí chờ thanh toán."); return
    tag = f"vendor{a.vendor}" if a.vendor else f"emp{a.employee}"
    out = ROOT / f"tmp/BangKe_ChiPhi_{tag}_{datetime.date.today()}.xlsx"
    out.parent.mkdir(exist_ok=True)
    total = to_excel(title, rows, out)
    print(f"✅ {title}")
    print(f"   {len(rows)} khoản — TỔNG {vnd(total)}đ")
    print(f"   File: {out}")

    if a.create_bill:
        dates = [r[2] for r in rows if r[2]]
        pf, pt = (min(dates), max(dates)) if dates else (None, None)
        bill_no = f"AP-{tag}-{datetime.date.today().strftime('%y%m%d')}"
        cur.execute("""INSERT INTO ap_bills (bill_no,vendor_id,employee_id,period_from,period_to,total_amount,note,created_by)
          VALUES (%s,%s,%s,%s,%s,%s,%s,1) RETURNING bill_id""",
          (bill_no, a.vendor, a.employee, pf, pt, total, title))
        bid = cur.fetchone()[0]
        for r in rows:
            cur.execute("INSERT INTO ap_bill_items (bill_id,cost_id,amount) VALUES (%s,%s,%s)",
                        (bid, r[0], r[6]))
        c.commit()
        print(f"   ✅ Đã tạo ap_bill #{bid} ({bill_no}) — track thanh toán")
    c.close()

if __name__ == '__main__':
    main()
