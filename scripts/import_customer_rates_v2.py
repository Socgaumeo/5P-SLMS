#!/usr/bin/env python3
"""
Enhanced Customer Rate Import with DeepSeek AI.
Features: Rates + Surcharges extraction, File reference tracking.
"""

import os
import json
import shutil
import psycopg2
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
from openai import OpenAI

load_dotenv()

# DeepSeek setup
client = OpenAI(
    api_key=os.getenv('DEEPSEEK_api_key'),
    base_url="https://api.deepseek.com"
)

# Database connection
DB_CONFIG = {
    'host': 'localhost',
    'port': 5432,
    'database': 'slms',
    'user': 'slms_admin',
    'password': os.getenv('POSTGRES_PASSWORD', 'MatKhauManhCua5P_2026!')
}

# File storage directory
STORAGE_DIR = Path("/Users/bear1108/Documents/GitHub/5P-SLMS/stored_files/customer_rates")

def read_pdf_content(file_path: str) -> str:
    """Extract text from PDF."""
    from pypdf import PdfReader
    reader = PdfReader(file_path)
    content = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            content.append(text)
    return "\n".join(content)

def read_excel_content(file_path: str) -> str:
    """Extract text from Excel."""
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)
    content = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        content.append(f"=== SHEET: {sheet} ===")
        for row in ws.iter_rows(max_row=100):
            row_text = " | ".join([str(c.value)[:50] if c.value else "" for c in row])
            if row_text.strip(" |"):
                content.append(row_text)
    return "\n".join(content)

def parse_with_deepseek(content: str, customer_name: str) -> dict:
    """Use DeepSeek to parse ALL rate information including surcharges."""
    
    prompt = f"""Analyze this Vietnamese customer price quotation (báo giá khách hàng).
Customer: {customer_name}

Extract ALL pricing information into a JSON object with TWO sections:

1. "rates" - Main transportation rates:
   - origin: Điểm đầu
   - origin_province: Tỉnh đi
   - destination: Điểm đến
   - destination_province: Tỉnh đến
   - prices: Object with vehicle_type -> price (e.g. {{"1.25T": 200000}})
   - notes: Ghi chú

2. "surcharges" - Additional fees/Phụ phí (chờ giờ, lưu ca, thêm điểm, ngày lễ, etc.):
   - surcharge_code: Code (WAITING_HOUR, OVERNIGHT, EXTRA_STOP, HOLIDAY, etc.)
   - surcharge_name: Tên phụ phí (tiếng Việt)
   - prices: Object with vehicle_type -> price
   - free_allowance: Miễn phí bao nhiêu (e.g. "1 giờ", "30 phút")
   - unit: PER_HOUR, PER_NIGHT, PER_STOP, FLAT, PERCENT

Document content:
```
{content[:15000]}
```

IMPORTANT:
- Return ONLY valid JSON object
- Price = number only (remove commas/dots)
- Include ALL vehicle types and surcharges found
- Common surcharge codes:
  * WAITING_HOUR = Phí chờ giờ
  * OVERNIGHT = Phí lưu ca
  * EXTRA_STOP = Thêm điểm
  * HOLIDAY = Ngày lễ/CN
  * FAILED_PICKUP = Đón trượt

Example output:
{{
  "rates": [
    {{"origin": "KCN Sông Công", "destination": "KCN Yên Bình", "prices": {{"1.25T": 450000, "2.5T": 770000}}, "notes": ""}}
  ],
  "surcharges": [
    {{"surcharge_code": "WAITING_HOUR", "surcharge_name": "Phí chờ giờ", "prices": {{"1.25T": 50000}}, "free_allowance": "1 giờ", "unit": "PER_HOUR"}},
    {{"surcharge_code": "OVERNIGHT", "surcharge_name": "Phí lưu ca", "prices": {{"1.25T": 400000}}, "free_allowance": null, "unit": "PER_NIGHT"}}
  ],
  "validity": {{"from": "2025-08-07", "to": null}},
  "notes": "Giá chưa bao gồm VAT"
}}
"""
    
    response = client.chat.completions.create(
        model="deepseek-chat",
        messages=[
            {"role": "system", "content": "You are a helpful assistant that extracts ALL pricing data from Vietnamese logistics documents. Include both main rates AND surcharges/additional fees. Always return valid JSON."},
            {"role": "user", "content": prompt}
        ],
        temperature=0.1,
        max_tokens=6000
    )
    
    response_text = response.choices[0].message.content.strip()
    
    # Clean up response
    if "```json" in response_text:
        response_text = response_text.split("```json")[1].split("```")[0]
    elif "```" in response_text:
        lines = response_text.split("\n")
        start = next((i for i, l in enumerate(lines) if l.startswith("```")), 0) + 1
        end = next((i for i, l in enumerate(lines[start:]) if l.startswith("```")), len(lines)) + start
        response_text = "\n".join(lines[start:end])
    
    import re
    response_text = re.sub(r',\s*]', ']', response_text)
    response_text = re.sub(r',\s*}', '}', response_text)
    
    try:
        return json.loads(response_text)
    except json.JSONDecodeError as e:
        print(f"⚠️ JSON error: {e}")
        with open("debug_deepseek_response.txt", "w") as f:
            f.write(response_text)
        return {"rates": [], "surcharges": []}

def store_file(file_path: str, customer_code: str) -> str:
    """Copy file to storage directory with timestamp."""
    STORAGE_DIR.mkdir(parents=True, exist_ok=True)
    
    original = Path(file_path)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    new_name = f"{customer_code}_{timestamp}{original.suffix}"
    dest_path = STORAGE_DIR / new_name
    
    shutil.copy2(file_path, dest_path)
    return str(dest_path)

def get_customer_id(cursor, customer_code: str) -> int:
    cursor.execute("""
        SELECT customer_id FROM customers 
        WHERE customer_code ILIKE %s OR short_name ILIKE %s
    """, (f"%{customer_code}%", f"%{customer_code}%"))
    result = cursor.fetchone()
    return result[0] if result else None

def get_or_create_route(cursor, origin: str, destination: str) -> int:
    cursor.execute("""
        SELECT route_id FROM master_routes 
        WHERE origin ILIKE %s AND destination ILIKE %s
    """, (f"%{origin}%", f"%{destination}%"))
    result = cursor.fetchone()
    if result:
        return result[0]
    
    cursor.execute("""
        INSERT INTO master_routes (origin, destination, is_active)
        VALUES (%s, %s, TRUE) RETURNING route_id
    """, (origin, destination))
    return cursor.fetchone()[0]

def create_file_reference(cursor, customer_id: int, file_path: str, stored_path: str) -> int:
    """Create file reference record."""
    original = Path(file_path)
    file_size = original.stat().st_size if original.exists() else 0
    
    cursor.execute("""
        INSERT INTO rate_file_references 
        (entity_type, entity_id, original_filename, stored_path, file_type, file_size_bytes, parsed_by_ai)
        VALUES ('CUSTOMER', %s, %s, %s, %s, %s, 'DEEPSEEK')
        RETURNING id
    """, (customer_id, original.name, stored_path, original.suffix.upper().strip('.'), file_size))
    
    return cursor.fetchone()[0]

def import_rates(cursor, customer_id: int, rates: list, file_ref_id: int) -> int:
    """Import rates to database."""
    imported = 0
    
    for rate in rates:
        try:
            route_id = get_or_create_route(cursor, rate.get('origin', ''), rate.get('destination', ''))
            
            prices = rate.get('prices', {})
            for vehicle_type, price in prices.items():
                cursor.execute("""
                    SELECT id FROM customer_rates 
                    WHERE customer_id = %s AND route_id = %s AND vehicle_type = %s
                    AND effective_date = CURRENT_DATE
                """, (customer_id, route_id, vehicle_type))
                
                if cursor.fetchone():
                    continue
                
                cursor.execute("""
                    INSERT INTO customer_rates 
                    (customer_id, route_id, vehicle_type, price, notes, effective_date, file_reference_id)
                    VALUES (%s, %s, %s, %s, %s, CURRENT_DATE, %s)
                """, (customer_id, route_id, vehicle_type, price, rate.get('notes', ''), file_ref_id))
                imported += 1
                
        except Exception as e:
            print(f"   ⚠️ Rate error: {e}")
    
    return imported

def import_surcharges(cursor, customer_id: int, surcharges: list) -> int:
    """Import surcharges to database."""
    imported = 0
    
    for surcharge in surcharges:
        try:
            prices = surcharge.get('prices', {})
            if not prices:
                continue
            
            # Filter out null prices
            valid_prices = {k: v for k, v in prices.items() if v is not None and v > 0}
            if not valid_prices:
                continue
            
            # Insert surcharge
            cursor.execute("""
                INSERT INTO customer_surcharges 
                (customer_id, surcharge_code, surcharge_name, pricing, free_allowance, unit, effective_date)
                VALUES (%s, %s, %s, %s, %s, %s, CURRENT_DATE)
                RETURNING id
            """, (
                customer_id,
                surcharge.get('surcharge_code', 'OTHER'),
                surcharge.get('surcharge_name', 'Phụ phí'),
                json.dumps(valid_prices),
                surcharge.get('free_allowance'),
                surcharge.get('unit', 'FLAT')
            ))
            surcharge_id = cursor.fetchone()[0]
            
            # Insert normalized prices (only valid ones)
            for vehicle_type, price in valid_prices.items():
                cursor.execute("""
                    INSERT INTO customer_surcharge_prices (surcharge_id, vehicle_type, price)
                    VALUES (%s, %s, %s)
                    ON CONFLICT (surcharge_id, vehicle_type) DO NOTHING
                """, (surcharge_id, vehicle_type, price))
            
            imported += 1
                
        except Exception as e:
            print(f"   ⚠️ Surcharge error: {e}")
    
    return imported

def process_customer_file(file_path: str, customer_code: str):
    """Main function to process customer rate file."""
    print(f"📁 Processing: {file_path}")
    print(f"👤 Customer: {customer_code}")
    
    # Read file
    if file_path.endswith('.pdf'):
        content = read_pdf_content(file_path)
    elif file_path.endswith(('.xlsx', '.xls')):
        content = read_excel_content(file_path)
    else:
        print(f"❌ Unsupported file type")
        return
    
    print(f"📄 Read {len(content)} characters")
    print(f"🤖 Parsing with DeepSeek AI (rates + surcharges)...")
    
    parsed_data = parse_with_deepseek(content, customer_code)
    
    rates = parsed_data.get('rates', [])
    surcharges = parsed_data.get('surcharges', [])
    
    print(f"✅ Extracted: {len(rates)} routes, {len(surcharges)} surcharge types")
    
    # Preview
    if rates:
        print("\n📋 Rate preview:")
        for rate in rates[:2]:
            prices_str = ", ".join([f"{k}: {v:,}" for k, v in rate.get('prices', {}).items()][:3])
            print(f"   {rate.get('origin', 'N/A')} → {rate.get('destination', 'N/A')} | {prices_str}")
    
    if surcharges:
        print("\n💰 Surcharge preview:")
        for s in surcharges[:3]:
            print(f"   {s.get('surcharge_name', 'N/A')} ({s.get('surcharge_code', 'N/A')})")
    
    # Database import
    print(f"\n💾 Importing to database...")
    conn = psycopg2.connect(**DB_CONFIG)
    cursor = conn.cursor()
    
    try:
        customer_id = get_customer_id(cursor, customer_code)
        if not customer_id:
            print(f"❌ Customer '{customer_code}' not found")
            return
        print(f"✅ Customer ID: {customer_id}")
        
        # Store file
        print(f"📂 Storing file...")
        stored_path = store_file(file_path, customer_code)
        file_ref_id = create_file_reference(cursor, customer_id, file_path, stored_path)
        print(f"✅ File stored: {stored_path}")
        
        # Import rates
        rate_count = import_rates(cursor, customer_id, rates, file_ref_id)
        print(f"✅ Imported {rate_count} rates")
        
        # Import surcharges
        surcharge_count = import_surcharges(cursor, customer_id, surcharges)
        print(f"✅ Imported {surcharge_count} surcharges")
        
        # Update file reference
        cursor.execute("""
            UPDATE rate_file_references 
            SET rates_imported = %s, surcharges_imported = %s
            WHERE id = %s
        """, (rate_count, surcharge_count, file_ref_id))
        
        conn.commit()
        print(f"\n🎉 Total imported: {rate_count} rates, {surcharge_count} surcharges")
        print(f"📎 File reference ID: {file_ref_id}")
        
    except Exception as e:
        conn.rollback()
        print(f"❌ Error: {e}")
        raise
    finally:
        cursor.close()
        conn.close()

if __name__ == "__main__":
    import sys
    
    file_path = sys.argv[1] if len(sys.argv) > 1 else "customer_rates/Báo giá NVAF_070825.pdf"
    customer_code = sys.argv[2] if len(sys.argv) > 2 else "NVAF"
    
    process_customer_file(file_path, customer_code)
