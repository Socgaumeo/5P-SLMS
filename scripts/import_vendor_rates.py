#!/usr/bin/env python3
"""
AI-powered Excel parser for vendor rates.
Uses Gemini to intelligently extract and structure data from Excel files.
"""

import os
import json
import openpyxl
import psycopg2
from dotenv import load_dotenv

# Load environment
load_dotenv()

# Gemini setup
import google.generativeai as genai
genai.configure(api_key=os.getenv('GOOGLE_GEMINI_API_KEY'))
model = genai.GenerativeModel('gemini-2.0-flash')

# Database connection
DB_CONFIG = {
    'host': 'localhost',
    'port': 5432,
    'database': 'slms',
    'user': 'slms_admin',
    'password': os.getenv('POSTGRES_PASSWORD', 'MatKhauManhCua5P_2026!')
}

def read_excel_content(file_path: str) -> str:
    """Read Excel file and convert to text for AI processing."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    content = []
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        content.append(f"\n=== SHEET: {sheet_name} ===")
        
        for row in sheet.iter_rows(max_row=50):  # First 50 rows
            row_text = " | ".join([str(cell.value) if cell.value else "" for cell in row])
            if row_text.strip(" |"):
                content.append(row_text)
    
    return "\n".join(content)

def parse_with_ai(excel_content: str) -> list:
    """Use Gemini to parse Excel content into structured vendor rates."""
    
    prompt = f"""Analyze this Excel file containing vendor transportation rates (báo giá vận chuyển).
Extract all pricing information into a JSON array.

Each rate should have:
- origin: Điểm đầu (pickup location)
- destination: Điểm cuối (delivery location)  
- vehicle_type: Loại xe (e.g. "1.25T", "2.5T", "5T", "8T", "15T")
- price: Giá (number only, in VND)
- notes: Ghi chú (any conditions or notes)

Excel Content:
```
{excel_content[:15000]}
```

IMPORTANT:
- Return ONLY a valid JSON array, no markdown or explanation
- Use DOUBLE QUOTES for all strings, not single quotes
- Price should be a number (remove commas, dots for thousands separator)
- If price has multiple values (e.g. range), use the higher value
- Ignore rows that don't contain route/price data
- Limit to maximum 100 most important rates

Example output format:
[
  {{"origin": "KCN Quang Minh", "destination": "Nội thành Hà Nội", "vehicle_type": "1.25T", "price": 850000, "notes": ""}},
  {{"origin": "KCN Quang Minh", "destination": "Cảng Hải Phòng", "vehicle_type": "5T", "price": 4500000, "notes": "Bao cầu xi"}}
]
"""
    
    response = model.generate_content(prompt)
    response_text = response.text.strip()
    
    # Clean up response if wrapped in markdown
    if "```json" in response_text:
        response_text = response_text.split("```json")[1].split("```")[0]
    elif "```" in response_text:
        lines = response_text.split("\n")
        start = next((i for i, l in enumerate(lines) if l.startswith("```")), 0) + 1
        end = next((i for i, l in enumerate(lines[start:]) if l.startswith("```")), len(lines)) + start
        response_text = "\n".join(lines[start:end])
    
    # Replace single quotes with double quotes (common AI mistake)
    import re
    # Fix trailing commas before closing brackets
    response_text = re.sub(r',\s*]', ']', response_text)
    response_text = re.sub(r',\s*}', '}', response_text)
    
    try:
        return json.loads(response_text)
    except json.JSONDecodeError as e:
        print(f"⚠️ JSON parse error: {e}")
        print(f"📄 First 500 chars of response: {response_text[:500]}")
        
        # Try to find and extract just the JSON array
        match = re.search(r'\[[\s\S]*\]', response_text)
        if match:
            try:
                return json.loads(match.group())
            except:
                pass
        
        # Save raw response for debugging
        with open("debug_ai_response.txt", "w") as f:
            f.write(response_text)
        print("💾 Saved raw response to debug_ai_response.txt")
        return []


def get_or_create_route(cursor, origin: str, destination: str) -> int:
    """Get or create route in master_routes table."""
    cursor.execute("""
        SELECT route_id FROM master_routes 
        WHERE origin ILIKE %s AND destination ILIKE %s
    """, (f"%{origin}%", f"%{destination}%"))
    
    result = cursor.fetchone()
    if result:
        return result[0]
    
    # Create new route
    cursor.execute("""
        INSERT INTO master_routes (origin, destination, is_active)
        VALUES (%s, %s, TRUE)
        RETURNING route_id
    """, (origin, destination))
    
    return cursor.fetchone()[0]

def get_vendor_id(cursor, vendor_name: str) -> int:
    """Get vendor ID by name."""
    cursor.execute("""
        SELECT vendor_id FROM vendors 
        WHERE company_name ILIKE %s OR short_name ILIKE %s
    """, (f"%{vendor_name}%", f"%{vendor_name}%"))
    
    result = cursor.fetchone()
    return result[0] if result else None

def import_vendor_rates(file_path: str, vendor_name: str):
    """Main function to import vendor rates from Excel."""
    print(f"📁 Processing: {file_path}")
    
    # Read Excel
    excel_content = read_excel_content(file_path)
    print(f"📄 Read {len(excel_content)} characters from Excel")
    
    # Parse with AI
    print("🤖 Parsing with Gemini AI...")
    rates = parse_with_ai(excel_content)
    print(f"✅ Extracted {len(rates)} rates")
    
    # Preview
    print("\n📋 Preview (first 5 rates):")
    for rate in rates[:5]:
        print(f"   {rate['origin']} → {rate['destination']} | {rate['vehicle_type']} | {rate['price']:,} VND")
    
    # Database import
    print(f"\n💾 Importing to database...")
    conn = psycopg2.connect(**DB_CONFIG)
    cursor = conn.cursor()
    
    try:
        # Get vendor ID
        vendor_id = get_vendor_id(cursor, vendor_name)
        if not vendor_id:
            print(f"❌ Vendor '{vendor_name}' not found in database")
            return
        print(f"✅ Found vendor ID: {vendor_id}")
        
        imported = 0
        skipped = 0
        
        for rate in rates:
            try:
                # Get or create route
                route_id = get_or_create_route(cursor, rate['origin'], rate['destination'])
                
                # Check if rate exists
                cursor.execute("""
                    SELECT id FROM vendor_rates 
                    WHERE vendor_id = %s AND route_id = %s AND vehicle_type = %s
                    AND effective_date = CURRENT_DATE
                """, (vendor_id, route_id, rate['vehicle_type']))
                
                if cursor.fetchone():
                    skipped += 1
                    continue
                
                # Insert rate
                cursor.execute("""
                    INSERT INTO vendor_rates 
                    (vendor_id, route_id, vehicle_type, price, notes, effective_date)
                    VALUES (%s, %s, %s, %s, %s, CURRENT_DATE)
                """, (vendor_id, route_id, rate['vehicle_type'], rate['price'], rate.get('notes', '')))
                
                imported += 1
                
            except Exception as e:
                print(f"   ⚠️ Error importing rate: {e}")
                skipped += 1
        
        conn.commit()
        print(f"\n✅ Import complete: {imported} imported, {skipped} skipped")
        
    except Exception as e:
        conn.rollback()
        print(f"❌ Error: {e}")
        raise
    finally:
        cursor.close()
        conn.close()

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        # Default: process Tam Bao file
        file_path = "vendor_rates/Tam bảo_092025.xlsx"
        vendor_name = "Tam Bảo"
    else:
        file_path = sys.argv[1]
        vendor_name = sys.argv[2] if len(sys.argv) > 2 else "Tam Bảo"
    
    import_vendor_rates(file_path, vendor_name)
