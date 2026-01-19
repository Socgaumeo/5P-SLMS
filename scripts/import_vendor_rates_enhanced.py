#!/usr/bin/env python3
"""
Enhanced AI-powered Excel parser for vendor rates.
Captures: Routes, Rate codes, Surcharges, Refrigerated rates.
"""

import os
import json
import openpyxl
import psycopg2
from dotenv import load_dotenv
import warnings
warnings.filterwarnings("ignore")

# Load environment
load_dotenv()

# Gemini setup
import google.generativeai as genai
genai.configure(api_key=os.getenv('GOOGLE_GEMINI_API_KEY'))
model = genai.GenerativeModel('gemini-2.0-flash')

# Database connection
DB_CONFIG = {
    'host': 'localhost',
    'port': 5432,
    'database': 'slms',
    'user': 'slms_admin',
    'password': os.getenv('POSTGRES_PASSWORD', 'MatKhauManhCua5P_2026!')
}

def read_sheet_content(sheet, max_rows=80) -> str:
    """Read sheet and convert to text."""
    content = []
    for row in sheet.iter_rows(max_row=max_rows):
        row_text = " | ".join([str(cell.value)[:50] if cell.value else "" for cell in row])
        if row_text.strip(" |"):
            content.append(row_text)
    return "\n".join(content)

def parse_rates_with_ai(content: str, rate_type: str) -> list:
    """Use Gemini to parse rate content."""
    
    prompt = f"""Analyze this Excel data containing Vietnamese trucking rates.
Rate type: {rate_type} (STANDARD = xe thường, REFRIGERATED = xe lạnh)

Extract pricing into a JSON array. Each rate should have:
- origin: Điểm đầu (pickup location name only)
- origin_province: Tỉnh đi
- destination: Điểm đến (destination name only)
- destination_province: Tỉnh đến
- rate_code: Mã giá (e.g. "TB18", "TB20")
- temperature_range: Yêu cầu nhiệt độ (for refrigerated only, e.g. "Từ 0 đến 10 độ")
- prices: Object with vehicle_type -> price mapping
  Example: {{"1.25T": 200000, "1.5T": 310000, "2.5T": 420000}}

Data:
```
{content}
```

IMPORTANT:
- Return ONLY valid JSON array
- Price = number (remove commas/dots in thousands)
- Include ALL vehicle types found (1.25T, 1.5T, 2.5T, 3.5T, 5T, 8T, 15T)
- rate_code is usually like TB18, TB20, TB25, etc.
- Skip header rows and notes

Example output:
[
  {{"origin": "Nội Bài", "origin_province": "Hà Nội", "destination": "KCN Quang Minh", "destination_province": "Hà Nội", "rate_code": "TB20", "temperature_range": null, "prices": {{"1.25T": 220000, "1.5T": 330000, "2.5T": 440000}}}}
]
"""
    
    response = model.generate_content(prompt)
    response_text = response.text.strip()
    
    # Clean up response
    if "```json" in response_text:
        response_text = response_text.split("```json")[1].split("```")[0]
    elif "```" in response_text:
        lines = response_text.split("\n")
        start = next((i for i, l in enumerate(lines) if l.startswith("```")), 0) + 1
        end = next((i for i, l in enumerate(lines[start:]) if l.startswith("```")), len(lines)) + start
        response_text = "\n".join(lines[start:end])
    
    import re
    response_text = re.sub(r',\s*]', ']', response_text)
    response_text = re.sub(r',\s*}', '}', response_text)
    
    try:
        return json.loads(response_text)
    except json.JSONDecodeError as e:
        print(f"⚠️ JSON error: {e}")
        # Try regex extraction
        match = re.search(r'\[[\s\S]*\]', response_text)
        if match:
            try:
                return json.loads(match.group())
            except:
                pass
        with open(f"debug_{rate_type}.txt", "w") as f:
            f.write(response_text)
        return []

def get_or_create_route(cursor, origin: str, destination: str) -> int:
    """Get or create route."""
    cursor.execute("""
        SELECT route_id FROM master_routes 
        WHERE origin ILIKE %s AND destination ILIKE %s
    """, (f"%{origin}%", f"%{destination}%"))
    
    result = cursor.fetchone()
    if result:
        return result[0]
    
    cursor.execute("""
        INSERT INTO master_routes (origin, destination, is_active)
        VALUES (%s, %s, TRUE) RETURNING route_id
    """, (origin, destination))
    return cursor.fetchone()[0]

def import_rates(cursor, vendor_id: int, rates: list, rate_type: str):
    """Import rates to database."""
    imported = 0
    
    for rate in rates:
        try:
            route_id = get_or_create_route(cursor, rate['origin'], rate['destination'])
            
            prices = rate.get('prices', {})
            for vehicle_type, price in prices.items():
                # Check existing
                cursor.execute("""
                    SELECT id FROM vendor_rates 
                    WHERE vendor_id = %s AND route_id = %s AND vehicle_type = %s 
                    AND rate_type = %s AND effective_date = CURRENT_DATE
                """, (vendor_id, route_id, vehicle_type, rate_type))
                
                if cursor.fetchone():
                    continue
                
                cursor.execute("""
                    INSERT INTO vendor_rates 
                    (vendor_id, route_id, vehicle_type, price, rate_code, rate_type, 
                     temperature_range, origin_province, destination_province, effective_date)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_DATE)
                """, (
                    vendor_id, route_id, vehicle_type, price,
                    rate.get('rate_code'), rate_type,
                    rate.get('temperature_range'),
                    rate.get('origin_province'),
                    rate.get('destination_province')
                ))
                imported += 1
                
        except Exception as e:
            print(f"   ⚠️ Error: {e}")
    
    return imported

def process_vendor_file(file_path: str, vendor_name: str):
    """Process entire vendor rate file."""
    print(f"📁 Processing: {file_path}")
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print(f"📑 Found sheets: {wb.sheetnames}")
    
    conn = psycopg2.connect(**DB_CONFIG)
    cursor = conn.cursor()
    
    try:
        # Get vendor
        cursor.execute("""
            SELECT vendor_id FROM vendors 
            WHERE company_name ILIKE %s OR short_name ILIKE %s
        """, (f"%{vendor_name}%", f"%{vendor_name}%"))
        result = cursor.fetchone()
        if not result:
            print(f"❌ Vendor '{vendor_name}' not found")
            return
        vendor_id = result[0]
        print(f"✅ Vendor ID: {vendor_id}")
        
        total_imported = 0
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            content = read_sheet_content(sheet)
            
            # Skip surcharge sheets - already handled via SQL
            if 'Phụ Phí' in sheet_name:
                print(f"\n📋 Skipping {sheet_name} (surcharges handled separately)")
                continue
            
            # Determine rate type
            if 'Lạnh' in sheet_name:
                rate_type = 'REFRIGERATED'
            else:
                rate_type = 'STANDARD'
            
            print(f"\n🔄 Processing: {sheet_name} ({rate_type})")
            print(f"🤖 Parsing with AI...")
            
            rates = parse_rates_with_ai(content, rate_type)
            print(f"✅ Extracted {len(rates)} rate entries")
            
            if rates:
                count = import_rates(cursor, vendor_id, rates, rate_type)
                print(f"💾 Imported {count} individual rates")
                total_imported += count
        
        conn.commit()
        print(f"\n✅ Total imported: {total_imported} rates")
        
    except Exception as e:
        conn.rollback()
        print(f"❌ Error: {e}")
        raise
    finally:
        cursor.close()
        conn.close()

if __name__ == "__main__":
    import sys
    
    file_path = sys.argv[1] if len(sys.argv) > 1 else "vendor_rates/Tam bảo_092025.xlsx"
    vendor_name = sys.argv[2] if len(sys.argv) > 2 else "Tam Bảo"
    
    process_vendor_file(file_path, vendor_name)
