"""
Clean DAINESE Excel templates:
- Keep only header rows + 1 reference data row (with formulas intact)
- Delete extra data rows
- Save cleaned templates (overwrite originals)
"""

import os
import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

TEMPLATES_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    'stored_files', 'templates', 'DAINESE'
)


def copy_cell_style(src_cell, dst_cell):
    """Copy all style attributes from src to dst cell."""
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy.copy(src_cell.protection)
        dst_cell.alignment = copy.copy(src_cell.alignment)


def delete_rows_preserve_styles(ws, start_row, end_row):
    """Delete rows from start_row to end_row (inclusive) without breaking sheet structure."""
    if end_row < start_row:
        return
    ws.delete_rows(start_row, end_row - start_row + 1)
    print(f"    Deleted rows {start_row}-{end_row}")


def clean_co_template(path):
    """
    DAINESE_CO_template.xlsx — Sheet "DỊCH VỤ"
    Header: rows 13-14, Data starts row 15 (keep row 15, delete row 16)
    Totals: row 17 =SUM(H15:H16) → after clean becomes =SUM(H15:H15)
    """
    print(f"\nCleaning CO template: {path}")
    wb = load_workbook(path)
    ws = wb['DỊCH VỤ']

    # Delete row 16 (second data row, keep row 15 as reference)
    delete_rows_preserve_styles(ws, 16, 16)

    # Fix totals formulas (row 17 moved to row 16 after deletion)
    # After deleting row 16, the old row 17 becomes row 16
    # SUM(H15:H16) → SUM(H15:H15) since we only have 1 data row
    # But openpyxl doesn't auto-update formulas, so we set them explicitly
    # Row 16 is now the totals row (was row 17)
    totals_row = 16
    ws[f'H{totals_row}'] = '=SUM(H15:H15)'
    ws[f'L{totals_row}'] = '=SUM(L15:L15)'

    # Row 17 (was row 18) — VAT row, row 18 (was row 19) — post-VAT total
    # These rows shift up by 1
    ws[f'H{totals_row + 2}'] = f'=H{totals_row}+H{totals_row + 1}'
    ws[f'L{totals_row + 2}'] = f'=SUM(L15:L15)'
    ws[f'I{totals_row + 3}'] = f'=H{totals_row + 2}+L{totals_row + 2}'

    wb.save(path)
    print(f"    Saved. Sheet rows now: {ws.max_row}")


def clean_nhap_sea_air_template(path):
    """
    DAINESE_NHAP_SEA_AIR_template.xlsx — Sheet "NHẬP"
    Header: rows 11-12, Data: rows 13-28 (keep row 13, delete rows 14-28)
    Totals row 29, VAT row 30, Grand total row 31
    """
    print(f"\nCleaning NHAP SEA AIR template: {path}")
    wb = load_workbook(path)
    ws = wb['NHẬP']

    # Delete rows 14-28 (keep row 13 as reference)
    delete_rows_preserve_styles(ws, 14, 28)

    # After deletion, old rows 29-31 shift up by 15
    # New positions: totals=14, VAT=15, grand_total=16
    totals_row = 14
    vat_row = 15
    grand_total_row = 16

    # Fix totals row formulas to cover only row 13 (1 data row)
    formula_cols_v = ['K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']
    formula_cols_z = ['W', 'X', 'Y']
    summary_cols = ['V', 'Z', 'AA']

    for col in formula_cols_v + formula_cols_z + summary_cols:
        ws[f'{col}{totals_row}'] = f'=SUM({col}13:{col}13)'

    # Grand total = totals + VAT
    for col in formula_cols_v + formula_cols_z + summary_cols:
        ws[f'{col}{grand_total_row}'] = f'=SUM({col}{totals_row}:{col}{vat_row})'

    wb.save(path)
    print(f"    Saved. Sheet rows now: {ws.max_row}")


def clean_xuat_template(path):
    """
    DAINESE_XUAT_template.xlsx — Sheet "XUẤT"
    Header: rows 11-12, Data: rows 13-34 (keep row 13, delete rows 14-34)
    """
    print(f"\nCleaning XUAT template: {path}")
    wb = load_workbook(path)
    ws = wb['XUẤT']

    # Delete rows 14-34 (keep row 13)
    delete_rows_preserve_styles(ws, 14, 34)

    # After deletion, totals row was ~35, now shifts to ~14
    # We need to add a totals row after the data
    totals_row = 14

    # Set totals formulas for key columns
    for col in ['K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC']:
        ws[f'{col}{totals_row}'] = f'=SUM({col}13:{col}13)'

    wb.save(path)
    print(f"    Saved. Sheet rows now: {ws.max_row}")


def clean_tc_nhap_cpn_template(path):
    """
    DAINESE_TC_NHAP_CPN_template.xlsx — Sheet "tc, cpn"
    Header: rows 11-12, Data: rows 13-34 (keep row 13, delete rows 14-34)
    """
    print(f"\nCleaning TC NHAP CPN template: {path}")
    wb = load_workbook(path)
    ws = wb['tc, cpn']

    # Delete rows 14-34 (keep row 13)
    delete_rows_preserve_styles(ws, 14, 34)

    # Totals row after deletion
    totals_row = 14
    for col in ['K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB']:
        ws[f'{col}{totals_row}'] = f'=SUM({col}13:{col}13)'

    wb.save(path)
    print(f"    Saved. Sheet rows now: {ws.max_row}")


def clean_van_chuyen_template(path):
    """
    DAINESE_VAN_CHUYEN_template.xlsx — Sheet "HĐ"
    Header: row 13, sub-header row 14, Data: rows 15-25 (keep row 15, delete rows 16-25)
    Totals row 26 (shifts to row 16 after deletion)
    PurchPurchaseOrder sheet: update cross-sheet refs (rows 31+ reference HĐ rows)
    """
    print(f"\nCleaning VAN CHUYEN template: {path}")
    wb = load_workbook(path)
    ws_hd = wb['HĐ']

    # Check if rows 16-25 have any data
    has_extra_rows = False
    for r in range(16, 26):
        for cell in ws_hd[r]:
            if cell.value is not None:
                has_extra_rows = True
                break

    if has_extra_rows:
        delete_rows_preserve_styles(ws_hd, 16, 25)
        totals_row = 16
    else:
        # Rows already empty, find actual max data row
        totals_row = 16
        print("    No extra rows found, sheet may already be clean")

    # Add totals row formulas at row 16
    for col in ['J', 'K', 'L', 'M', 'N']:
        ws_hd[f'{col}{totals_row}'] = f'=SUM({col}15:{col}15)'

    wb.save(path)
    print(f"    Saved. HĐ sheet rows now: {ws_hd.max_row}")


def verify_templates():
    """Quick verification that templates open and have correct structure."""
    configs = [
        ('DAINESE_CO_template.xlsx', 'DỊCH VỤ', 15),
        ('DAINESE_NHAP_SEA_AIR_template.xlsx', 'NHẬP', 13),
        ('DAINESE_XUAT_template.xlsx', 'XUẤT', 13),
        ('DAINESE_TC_NHAP_CPN_template.xlsx', 'tc, cpn', 13),
        ('DAINESE_VAN_CHUYEN_template.xlsx', 'HĐ', 15),
    ]
    print("\n=== Verification ===")
    for filename, sheet_name, ref_row in configs:
        path = os.path.join(TEMPLATES_DIR, filename)
        try:
            wb = load_workbook(path)
            ws = wb[sheet_name]
            # Check reference row still has some data
            ref_vals = [c.value for c in ws[ref_row] if c.value is not None]
            print(f"  {filename}: sheet='{sheet_name}', ref_row={ref_row}, "
                  f"max_row={ws.max_row}, ref_row_cells={len(ref_vals)}")
        except Exception as e:
            print(f"  ERROR {filename}: {e}")


def main():
    print(f"Templates dir: {TEMPLATES_DIR}")

    clean_co_template(os.path.join(TEMPLATES_DIR, 'DAINESE_CO_template.xlsx'))
    clean_nhap_sea_air_template(os.path.join(TEMPLATES_DIR, 'DAINESE_NHAP_SEA_AIR_template.xlsx'))
    clean_xuat_template(os.path.join(TEMPLATES_DIR, 'DAINESE_XUAT_template.xlsx'))
    clean_tc_nhap_cpn_template(os.path.join(TEMPLATES_DIR, 'DAINESE_TC_NHAP_CPN_template.xlsx'))
    clean_van_chuyen_template(os.path.join(TEMPLATES_DIR, 'DAINESE_VAN_CHUYEN_template.xlsx'))

    verify_templates()
    print("\nDone!")


if __name__ == '__main__':
    main()
