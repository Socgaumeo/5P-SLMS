#!/usr/bin/env python3
"""
Import customer rates using DeepSeek API.
Supports PDF and Excel files.
"""

import os
import json
import psycopg2
from dotenv import load_dotenv
from openai import OpenAI

load_dotenv()

# DeepSeek setup (compatible with OpenAI SDK)
client = OpenAI(
    api_key=os.getenv('DEEPSEEK_api_key'),
    base_url="https://api.deepseek.com"
)

# Database connection
DB_CONFIG = {
    'host': 'localhost',
    'port': 5432,
    'database': 'slms',
    'user': 'slms_admin',
    'password': os.getenv('POSTGRES_PASSWORD', 'MatKhauManhCua5P_2026!')
}

def read_pdf_content(file_path: str) -> str:
    """Extract text from PDF."""
    from pypdf import PdfReader
    
    reader = PdfReader(file_path)
    content = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            content.append(text)
    return "\n".join(content)

def read_excel_content(file_path: str) -> str:
    """Extract text from Excel."""
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)
    content = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        content.append(f"=== SHEET: {sheet} ===")
        for row in ws.iter_rows(max_row=100):
            row_text = " | ".join([str(c.value)[:50] if c.value else "" for c in row])
            if row_text.strip(" |"):
                content.append(row_text)
    return "\n".join(content)

def parse_customer_rates_with_deepseek(content: str, customer_name: str) -> list:
    """Use DeepSeek to parse customer rate content."""
    
    prompt = f"""Analyze this Vietnamese customer price quotation (báo giá khách hàng).
Customer: {customer_name}

Extract all pricing information into a JSON array. Each rate should have:
- origin: Điểm đầu (pickup location name only)
- origin_province: Tỉnh đi
- destination: Điểm đến (destination name only)  
- destination_province: Tỉnh đến
- rate_code: Mã giá (if any)
- prices: Object with vehicle_type -> price mapping
  Example: {{"1.25T": 200000, "1.5T": 310000}}
- notes: Ghi chú, điều kiện

Document content:
```
{content[:12000]}
```

IMPORTANT:
- Return ONLY valid JSON array, no markdown
- Price = number only (remove commas/dots in thousands)
- Include ALL vehicle types found
- If route is not clear, use available location info

Example output:
[
  {{"origin": "KCN VSIP", "origin_province": "Bắc Ninh", "destination": "Nội Bài", "destination_province": "Hà Nội", "rate_code": null, "prices": {{"1.25T": 350000, "2.5T": 500000}}, "notes": ""}}
]
"""
    
    response = client.chat.completions.create(
        model="deepseek-chat",
        messages=[
            {"role": "system", "content": "You are a helpful assistant that extracts structured data from Vietnamese logistics documents. Always return valid JSON."},
            {"role": "user", "content": prompt}
        ],
        temperature=0.1,
        max_tokens=4000
    )
    
    response_text = response.choices[0].message.content.strip()
    
    # Clean up response
    if "```json" in response_text:
        response_text = response_text.split("```json")[1].split("```")[0]
    elif "```" in response_text:
        lines = response_text.split("\n")
        start = next((i for i, l in enumerate(lines) if l.startswith("```")), 0) + 1
        end = next((i for i, l in enumerate(lines[start:]) if l.startswith("```")), len(lines)) + start
        response_text = "\n".join(lines[start:end])
    
    import re
    response_text = re.sub(r',\s*]', ']', response_text)
    response_text = re.sub(r',\s*}', '}', response_text)
    
    try:
        return json.loads(response_text)
    except json.JSONDecodeError as e:
        print(f"⚠️ JSON error: {e}")
        print(f"📄 Response preview: {response_text[:500]}")
        with open("debug_deepseek_response.txt", "w") as f:
            f.write(response_text)
        return []

def get_customer_id(cursor, customer_code: str) -> int:
    """Get customer ID by code."""
    cursor.execute("""
        SELECT customer_id FROM customers 
        WHERE customer_code ILIKE %s OR short_name ILIKE %s
    """, (f"%{customer_code}%", f"%{customer_code}%"))
    result = cursor.fetchone()
    return result[0] if result else None

def get_or_create_route(cursor, origin: str, destination: str) -> int:
    """Get or create route."""
    cursor.execute("""
        SELECT route_id FROM master_routes 
        WHERE origin ILIKE %s AND destination ILIKE %s
    """, (f"%{origin}%", f"%{destination}%"))
    result = cursor.fetchone()
    if result:
        return result[0]
    
    cursor.execute("""
        INSERT INTO master_routes (origin, destination, is_active)
        VALUES (%s, %s, TRUE) RETURNING route_id
    """, (origin, destination))
    return cursor.fetchone()[0]

def import_customer_rates(cursor, customer_id: int, rates: list):
    """Import rates to database."""
    imported = 0
    
    for rate in rates:
        try:
            route_id = get_or_create_route(cursor, rate['origin'], rate['destination'])
            
            prices = rate.get('prices', {})
            for vehicle_type, price in prices.items():
                # Check existing
                cursor.execute("""
                    SELECT id FROM customer_rates 
                    WHERE customer_id = %s AND route_id = %s AND vehicle_type = %s
                    AND effective_date = CURRENT_DATE
                """, (customer_id, route_id, vehicle_type))
                
                if cursor.fetchone():
                    continue
                
                cursor.execute("""
                    INSERT INTO customer_rates 
                    (customer_id, route_id, vehicle_type, price, notes, effective_date)
                    VALUES (%s, %s, %s, %s, %s, CURRENT_DATE)
                """, (customer_id, route_id, vehicle_type, price, rate.get('notes', '')))
                imported += 1
                
        except Exception as e:
            print(f"   ⚠️ Error: {e}")
    
    return imported

def process_customer_file(file_path: str, customer_code: str):
    """Main function to process customer rate file."""
    print(f"📁 Processing: {file_path}")
    print(f"👤 Customer: {customer_code}")
    
    # Read file
    if file_path.endswith('.pdf'):
        content = read_pdf_content(file_path)
    elif file_path.endswith(('.xlsx', '.xls')):
        content = read_excel_content(file_path)
    else:
        print(f"❌ Unsupported file type")
        return
    
    print(f"📄 Read {len(content)} characters")
    print(f"🤖 Parsing with DeepSeek AI...")
    
    rates = parse_customer_rates_with_deepseek(content, customer_code)
    print(f"✅ Extracted {len(rates)} rate entries")
    
    if not rates:
        print("❌ No rates extracted. Check debug_deepseek_response.txt")
        return
    
    # Preview
    print("\n📋 Preview (first 3):")
    for rate in rates[:3]:
        prices_str = ", ".join([f"{k}: {v:,}" for k, v in rate.get('prices', {}).items()][:3])
        print(f"   {rate['origin']} → {rate['destination']} | {prices_str}")
    
    # Database import
    print(f"\n💾 Importing to database...")
    conn = psycopg2.connect(**DB_CONFIG)
    cursor = conn.cursor()
    
    try:
        customer_id = get_customer_id(cursor, customer_code)
        if not customer_id:
            print(f"❌ Customer '{customer_code}' not found")
            return
        print(f"✅ Customer ID: {customer_id}")
        
        count = import_customer_rates(cursor, customer_id, rates)
        conn.commit()
        print(f"\n✅ Imported {count} individual rates")
        
    except Exception as e:
        conn.rollback()
        print(f"❌ Error: {e}")
        raise
    finally:
        cursor.close()
        conn.close()

if __name__ == "__main__":
    import sys
    
    file_path = sys.argv[1] if len(sys.argv) > 1 else "customer_rates/Báo giá NVAF_070825.pdf"
    customer_code = sys.argv[2] if len(sys.argv) > 2 else "NVAF"
    
    process_customer_file(file_path, customer_code)
