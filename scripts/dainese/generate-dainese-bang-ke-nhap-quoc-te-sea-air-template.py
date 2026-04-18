#!/usr/bin/env python3
"""
Generator for DAINESE 'Bảng kê chi tiết hàng nhập quốc tế' (SEA/AIR) Excel.

Reproduces the exact layout, styles, formulas, and logo of the customer's
template file (File 1: Bảng kê nhập tháng 3.2026.sea.air.xlsx).

Usage:
    python this.py <output.xlsx> [logo.png]

Output is compared visually with the original to verify accuracy before
wiring it into the FastAPI export endpoint.
"""

import sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


# ---------- Style constants ----------

FONT_NAME = "Times New Roman"

THIN = Side(style="thin", color="000000")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Number formats
NF_INT = '#,##0'
NF_DEC = '#,##0.00'
NF_ACCOUNTING = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'

# Fills
FILL_DATA = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # soft yellow zebra
FILL_HEADER = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # soft blue
FILL_TOTAL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")


# ---------- Sample data (simulates what comes from DB) ----------

SAMPLE_JOBS = [
    {
        "to_khai": "108021404521", "hd_tm": "2025S0217", "van_don": "2025CA0000102644",
        "ngay_tk": "05/03/2026", "tuyen": "Genoa, Italy - Thái Nguyên", "loai": "SEA",
        "kgs": 5890, "cont": "LCL", "note": "Elantas",
        "phi_mtk": 943830, "phi_kh": 0, "phi_vc": 4561845, "phi_lh": 4089930,
        "phi_psk": 849447, "phi_dnn": 19568742, "cuoc_qt": 60216354, "thc": 20462234,
        "cfs": 0, "do": 1258440, "dly": 0,
        "local": 0, "csht": 94240, "kho": 13013670,
        "hd_traho": "2985, 1148965", "hd_fwd_1": "152", "hd_fwd_2": "183, 184",
    },
    {
        "to_khai": "108016284830", "hd_tm": "00000479", "van_don": "2025CA0000102683",
        "ngay_tk": "03/03/2026", "tuyen": "Genoa, Italy - Thái Nguyên", "loai": "SEA",
        "kgs": 140, "cont": "LCL", "note": "EGI PROJECT",
        "phi_mtk": 629220, "phi_kh": 0, "phi_vc": 2831490, "phi_lh": 629220,
        "phi_psk": 0, "phi_dnn": 0, "cuoc_qt": 5908690, "thc": 2301372,
        "cfs": 0, "do": 943830, "dly": 0,
        "local": 0, "csht": 0, "kho_formula": "=1153600+44550",
        "hd_traho": "2915, 2934", "hd_fwd_1": "149", "hd_fwd_2": "185",
    },
    {
        "to_khai": "108015913130", "hd_tm": "3G/066", "van_don": "DELHPH263669 ",
        "ngay_tk": "03/03/2026", "tuyen": "Mundra, Ấn Độ- Thái Nguyên", "loai": "SEA",
        "kgs": 169, "cont": "LCL", "note": "3 GENERATIONS",
        "phi_mtk": 788670, "phi_kh": 0, "phi_vc": 2497455, "phi_lh": 525780,
        "phi_psk": 0, "phi_dnn": 5731002, "cuoc_qt": 1314450, "thc": 709803,
        "cfs": 0, "do": 788670, "dly": 0,
        "local": 0, "csht": 0, "kho": 685800,
        "hd_traho": "9512", "hd_fwd_1": "148", "hd_fwd_2": "186",
    },
]

CUSTOMER_INFO = {
    "name": "CÔNG TY TNHH DAINESE VIỆT NAM",
    "address": "Lô CN13, Lô CN18, Khu Công nghiệp Yên Bình, Phường Vạn Xuân, Tỉnh Thái Nguyên, Việt Nam",
    "attn": "Ms. Phương Anh - Ms. Dương",
}

COMPANY_INFO = {
    "name": "CÔNG TY TNHH THƯƠNG MẠI VÀ DỊCH VỤ 5P VIỆT NAM",
    "address": "Số nhà 02 Ngõ 1H Phố Trần Quang Diệu, Phường Đống Đa, Thành phố Hà Nội, Việt Nam",
    "mst": "MST: 0110523309",
}

BANK_INFO = [
    "Thông tin chuyển khoản:",
    "Tài khoản: Công Ty TNHH Thương mại và dịch vụ 5P Việt Nam",
    "Số tài khoản: 346886",
    "Tại Ngân hàng: Ngân hàng TMCP Kỹ thương Việt Nam - Techcombank Chi nhánh Đông Đô",
]


# ---------- Builder ----------

def font(size=11, bold=False, italic=False, color=None):
    return Font(name=FONT_NAME, size=size, bold=bold, italic=italic, color=color)


def style_cell(cell, *, f=None, align=None, fill=None, fmt=None, border=True):
    if f is not None:
        cell.font = f
    if align is not None:
        cell.alignment = align
    if fill is not None:
        cell.fill = fill
    if fmt is not None:
        cell.number_format = fmt
    if border:
        cell.border = BORDER_ALL


def build_header(ws, logo_path: str | None, title_month: str):
    # Column widths (sheet NHẬP)
    widths = {
        'A': 5.4,  'B': 13.6, 'C': 19.6, 'D': 19.9, 'E': 16.6, 'F': 26.3, 'G': 11.4,
        'H': 13.5, 'I': 9.5,  'J': 18.5, 'K': 15.3, 'L': 14.5, 'M': 16.4, 'N': 15.2,
        'O': 14.2, 'P': 18.8, 'Q': 17.9, 'R': 13.7, 'S': 12.6, 'T': 13.8, 'U': 11.0,
        'V': 15.7, 'W': 12.5, 'X': 11.5, 'Y': 20.4, 'Z': 15.7, 'AA': 18.5, 'AB': 21.8,
        'AC': 8.2, 'AD': 8.0,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Row heights for header section
    for r in (1, 2, 3, 4):
        ws.row_dimensions[r].height = 23
    ws.row_dimensions[5].height = 67
    ws.row_dimensions[6].height = 38
    for r in (7, 8, 9):
        ws.row_dimensions[r].height = 18
    ws.row_dimensions[10].height = 73
    ws.row_dimensions[11].height = 26
    ws.row_dimensions[12].height = 42

    # Logo (anchor A1, sized to fit B1:B4)
    if logo_path and Path(logo_path).exists():
        img = XLImage(logo_path)
        img.width = 90
        img.height = 90
        ws.add_image(img, 'A1')

    # Company info
    ws['C1'] = COMPANY_INFO['name']
    style_cell(ws['C1'], f=font(12, bold=True), align=ALIGN_LEFT, border=False)
    ws['C2'] = COMPANY_INFO['address']
    style_cell(ws['C2'], f=font(12, bold=True), align=ALIGN_LEFT, border=False)
    ws['C3'] = COMPANY_INFO['mst']
    style_cell(ws['C3'], f=font(12, bold=True), align=ALIGN_LEFT, border=False)

    # Title (merged A5:AD5)
    ws.merge_cells('A5:AD5')
    ws['A5'] = f'BẢNG KÊ CHI TIẾT HÀNG NHẬP QUỐC TẾ THÁNG {title_month}'
    style_cell(ws['A5'], f=font(26, bold=True), align=ALIGN_CENTER_WRAP, border=False)

    # Recipient
    ws['B7'] = f"KÍNH GỬI :  {CUSTOMER_INFO['name']}"
    style_cell(ws['B7'], f=font(14, bold=True), align=ALIGN_LEFT, border=False)
    ws['B8'] = f"Địa chỉ : {CUSTOMER_INFO['address']}"
    style_cell(ws['B8'], f=font(14, bold=True), align=ALIGN_LEFT, border=False)
    ws['B9'] = f"Attn:   {CUSTOMER_INFO['attn']}"
    style_cell(ws['B9'], f=font(14, bold=True), align=ALIGN_LEFT, border=False)


def build_table_header(ws):
    # Row 11 group headers (merged)
    group_headers = [
        ('A11:G11', 'Thông tin chung'),
        ('H11:I11', 'Khối lượng'),
        ('J11:J12', 'Note'),
        ('K11:V11', 'Phí dịch vụ làm hàng  ( VND) '),
        ('W11:Z11', 'Phí trả hộ'),
        ('AA11:AA12', 'Tổng tiền phải trả '),
        ('AB11:AB12', 'Số hóa đơn trả hộ'),
        ('AC11:AD12', 'Số hóa đơn của forwarder'),
    ]
    for rng, label in group_headers:
        ws.merge_cells(rng)
        first = rng.split(':')[0]
        ws[first] = label
        style_cell(ws[first],
                   f=font(12, bold=True, italic=True),
                   align=ALIGN_CENTER_WRAP, fill=FILL_HEADER)
        # Apply border to all cells in merged range
        from openpyxl.utils.cell import range_boundaries
        c1, r1, c2, r2 = range_boundaries(rng)
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = BORDER_ALL
                cell.fill = FILL_HEADER

    # Row 12 column headers
    col_headers = {
        'A12': 'No.',
        'B12': 'Tờ khai',
        'C12': 'Hóa đơn thương mại',
        'D12': 'Vận đơn/Note',
        'E12': 'Ngày tờ khai',
        'F12': 'Tuyến đường',
        'G12': 'Loại hình vận chuyển',
        'H12': 'Kgs',
        'I12': 'No. Cont',
        'K12': 'Phí mở tờ khai hải quan',
        'L12': 'Phí kiểm hóa',
        'M12': 'Phí vận chuyển',
        'N12': 'Phí làm hàng',
        'O12': 'Phí phát sinh khác',
        'P12': 'Phí đầu nước ngoài',
        'Q12': 'Cước vận tải quốc tế',
        'R12': 'Phí xếp dỡ (THC)',
        'S12': 'Phí gom hàng lẻ (CFS)/\nCIC/ LSS',
        'T12': 'Phí lấy lệnh  (DO)',
        'U12': 'Phí  đại lý,',
        'V12': 'Tổng',
        'W12': 'Local charge',
        'X12': 'CSHT, thuế, vé bãi',
        'Y12': 'Lưu kho giao nhận bốc xếp, nâng hạ',
        'Z12': 'Tổng phí trả hộ',
    }
    for coord, val in col_headers.items():
        ws[coord] = val
        style_cell(ws[coord], f=font(10, bold=True), align=ALIGN_CENTER_WRAP, fill=FILL_HEADER)


def build_data_rows(ws, jobs: list, start_row: int = 13) -> int:
    """Write job rows. Returns row index of the row AFTER the last data row."""
    common_font = font(10, bold=True)
    for idx, job in enumerate(jobs):
        r = start_row + idx
        ws.row_dimensions[r].height = 37

        ws.cell(r, 1, idx + 1)
        ws.cell(r, 2, job["to_khai"])
        ws.cell(r, 3, job["hd_tm"])
        ws.cell(r, 4, job["van_don"])
        ws.cell(r, 5, job["ngay_tk"])
        ws.cell(r, 6, job["tuyen"])
        ws.cell(r, 7, job["loai"])
        ws.cell(r, 8, job["kgs"])
        ws.cell(r, 9, job["cont"])
        ws.cell(r, 10, job.get("note", ""))

        # Fees K..U (cols 11..21)
        fee_keys = ["phi_mtk", "phi_kh", "phi_vc", "phi_lh", "phi_psk",
                    "phi_dnn", "cuoc_qt", "thc", "cfs", "do", "dly"]
        for i, k in enumerate(fee_keys):
            v = job.get(k, 0)
            cell = ws.cell(r, 11 + i, v if v else None)
            cell.number_format = NF_DEC

        # V: =SUM(K:U)
        ws.cell(r, 22, f"=SUM(K{r}:U{r})").number_format = NF_ACCOUNTING

        # W..Y: local, csht, kho (cols 23..25)
        for i, k in enumerate(["local", "csht"]):
            v = job.get(k, 0)
            ws.cell(r, 23 + i, v if v else None).number_format = NF_DEC
        # Y can be formula or value
        if "kho_formula" in job:
            ws.cell(r, 25, job["kho_formula"]).number_format = NF_DEC
        else:
            v = job.get("kho", 0)
            ws.cell(r, 25, v if v else None).number_format = NF_DEC

        # Z: =SUM(W:Y)
        ws.cell(r, 26, f"=SUM(W{r}:Y{r})").number_format = NF_ACCOUNTING
        # AA: =V+Z
        ws.cell(r, 27, f"=+V{r}+Z{r}").number_format = NF_ACCOUNTING
        # AB, AC, AD
        ws.cell(r, 28, job.get("hd_traho", ""))
        ws.cell(r, 29, job.get("hd_fwd_1", ""))
        ws.cell(r, 30, job.get("hd_fwd_2", ""))

        # Apply common style
        for c in range(1, 31):
            cell = ws.cell(r, c)
            cell.font = common_font
            cell.alignment = ALIGN_CENTER_WRAP
            cell.border = BORDER_ALL
            cell.fill = FILL_DATA

    return start_row + len(jobs)


def build_totals(ws, data_start: int, data_end_exclusive: int):
    """Build Tổng / VAT / Tổng cộng rows."""
    last_data_row = data_end_exclusive - 1
    # Row 29 = Tổng
    tong_row = data_end_exclusive
    vat_row = tong_row + 1
    tc_row = vat_row + 1

    ws.merge_cells(f'A{tong_row}:J{tong_row}')
    ws[f'A{tong_row}'] = 'Tổng'
    style_cell(ws[f'A{tong_row}'],
               f=font(12, bold=True, italic=True),
               align=ALIGN_CENTER_WRAP, fill=FILL_TOTAL)
    for c in range(1, 11):
        ws.cell(tong_row, c).fill = FILL_TOTAL
        ws.cell(tong_row, c).border = BORDER_ALL

    for c in range(11, 28):  # K..AA
        col = get_column_letter(c)
        cell = ws.cell(tong_row, c, f"=SUM({col}{data_start}:{col}{last_data_row})")
        cell.font = font(12, bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = FILL_TOTAL
        cell.border = BORDER_ALL
        cell.number_format = NF_INT

    ws.merge_cells(f'A{vat_row}:J{vat_row}')
    ws[f'A{vat_row}'] = 'VAT'
    style_cell(ws[f'A{vat_row}'],
               f=font(12, bold=True, italic=True),
               align=ALIGN_CENTER_WRAP, fill=FILL_TOTAL)
    for c in range(1, 11):
        ws.cell(vat_row, c).fill = FILL_TOTAL
        ws.cell(vat_row, c).border = BORDER_ALL
    for c in range(11, 28):
        cell = ws.cell(vat_row, c, 0)
        cell.font = font(12, bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = FILL_TOTAL
        cell.border = BORDER_ALL
        cell.number_format = NF_INT

    ws.merge_cells(f'A{tc_row}:J{tc_row}')
    ws[f'A{tc_row}'] = 'Tổng cộng'
    style_cell(ws[f'A{tc_row}'],
               f=font(12, bold=True, italic=True),
               align=ALIGN_CENTER_WRAP, fill=FILL_TOTAL)
    for c in range(1, 11):
        ws.cell(tc_row, c).fill = FILL_TOTAL
        ws.cell(tc_row, c).border = BORDER_ALL
    for c in range(11, 28):
        col = get_column_letter(c)
        cell = ws.cell(tc_row, c, f"=SUM({col}{tong_row}:{col}{vat_row})")
        cell.font = font(12, bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = FILL_TOTAL
        cell.border = BORDER_ALL
        cell.number_format = NF_INT

    return tc_row


def build_footer(ws, start_row: int):
    for i, line in enumerate(BANK_INFO):
        r = start_row + i
        ws.cell(r, 1, line)
        if i == len(BANK_INFO) - 1:
            ws.merge_cells(f'A{r}:F{r}')
            ws.cell(r, 1).alignment = ALIGN_LEFT_WRAP
        ws.cell(r, 1).font = font(12)


def generate(out_path: str, logo_path: str | None = None, month: str = '03 NĂM 2026'):
    wb = Workbook()
    ws = wb.active
    ws.title = "NHẬP"

    build_header(ws, logo_path, month)
    build_table_header(ws)
    after_data = build_data_rows(ws, SAMPLE_JOBS, start_row=13)
    last_total_row = build_totals(ws, data_start=13, data_end_exclusive=after_data)
    build_footer(ws, last_total_row + 4)

    wb.save(out_path)
    print(f"Wrote: {out_path}")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: generate-...py <output.xlsx> [logo.png] [month_str]")
        sys.exit(1)
    out = sys.argv[1]
    logo = sys.argv[2] if len(sys.argv) > 2 else None
    month = sys.argv[3] if len(sys.argv) > 3 else '03 NĂM 2026'
    generate(out, logo, month)
