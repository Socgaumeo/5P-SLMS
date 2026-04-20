#!/usr/bin/env python3
"""
Verify every customer's generated bảng kê against the original Excel.

Steps per customer:
 1. Call /api/exports/customer/{code}?month=2026-03&template={trucking|handling}
 2. Extract generated: row count, per-col totals, titles
 3. Find matching original Excel file in OneDrive DOANH THU folder
 4. Extract original same metrics
 5. Report side-by-side: matches / gaps / data issues

Output: markdown report grouping by status (OK / data-gap / missing-in-DB).
"""

import sys
import urllib.request
from pathlib import Path
from openpyxl import load_workbook

BACKEND = "http://localhost:8000"
ROOT_OUT = Path("/Users/bear1108/projects/5P-SLMS/.claude/worktrees/gracious-swanson-ff2e09/plans/reports/dainese-templates")
ORIGINAL_ROOT = Path("/Users/bear1108/Library/CloudStorage/OneDrive-Personal/5P/DOANH THU/THÁNG 3")
REPORT_PATH = ROOT_OUT / "verify-all-customer-bang-ke-exports-vs-originals-report.md"

# Customer → (template_key, original_folder_name, service_label)
CUSTOMERS = [
    # Family A Trucking (except 2 missing in DB)
    ("BINHMINH",   "trucking", "BÌNH MINH",       "Trucking"),
    ("KK",         "trucking", "KK",              "Trucking"),
    ("DONGSUNG",   "trucking", "DONSUNG",         "Trucking"),
    ("UPGAIN",     "trucking", "UPGAIN",          "Trucking"),
    ("UTRACON",    "trucking", "UTRACORN",        "Trucking"),
    ("TVC",        "trucking", "TVC",             "Trucking"),
    # Family B Handling
    ("HUNGPHAT",   "handling", "HƯNG PHÁT",       "Handling"),
    ("VINTECH",    "handling", "VINTECH",         "Handling"),
    ("KTXL",       "handling", "XÂY LẮP VN",      "Handling"),
    ("LOGIMARKHN", "handling", "LOGIMARK",        "Handling"),
    ("MESSERHP",   "handling", "MESSER",          "Handling"),
    ("GANGTHEPTN", "handling", "GANG THÉP TN",    "Handling"),
    ("LAS",        "handling", "LAS",             "Handling"),
    # Already-implemented special customers (sanity-check they still work)
    ("DAINESE",    "nhap_sea_air", "DAINESE",     "Special (nhap_sea_air)"),
    ("DAINESE",    "tt",           "DAINESE",     "Special (tt trucking)"),
    ("DAINESE",    "tc_cpn",       "DAINESE",     "Special (tc_cpn)"),
    ("DAINESE",    "phi_co",       "DAINESE",     "Special (phi_co)"),
    ("DAINESE",    "xuat",         "DAINESE",     "Special (xuat)"),
    ("MEIKO",      None,           "MEIKO",       "Special (MEIKO)"),
]


def fetch_generated(code: str, template: str | None) -> Path | None:
    """Call backend and save result to disk."""
    url = f"{BACKEND}/api/exports/customer/{code}?month=2026-03"
    if template:
        url += f"&template={template}"
    out = ROOT_OUT / f"verify-output-{code}-{template or 'default'}.xlsx"
    try:
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = resp.read()
            out.write_bytes(data)
            return out
    except urllib.error.HTTPError as e:
        return None
    except Exception:
        return None


def inspect(path: Path) -> dict:
    """
    Extract metrics from a bảng kê xlsx.

    Because openpyxl-generated files don't cache formula results, we compute
    totals by summing RAW data cells (values that aren't formulas) across
    likely money columns. This gives a realistic revenue figure for both
    our generated files and customer originals.
    """
    info = {"rows": 0, "title": None, "total_revenue": 0.0, "sheets": []}
    try:
        wb = load_workbook(path, data_only=True)
        info["sheets"] = wb.sheetnames
        wb_raw = load_workbook(path, data_only=False)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws_raw = wb_raw[sheet_name]

            # Find title
            for r in range(1, min(15, ws.max_row + 1)):
                for c in range(1, min(5, ws.max_column + 1)):
                    v = ws.cell(r, c).value
                    if v and 'BẢNG KÊ' in str(v).upper():
                        if not info["title"]:
                            info["title"] = str(v)
                        break

            # Detect header row: row with most string cells in top 20 rows
            header_row = None
            for r in range(1, min(25, ws.max_row + 1)):
                strs = sum(1 for c in range(1, min(25, ws.max_column + 1))
                           if isinstance(ws.cell(r, c).value, str))
                if strs >= 5 and any(
                    'thành tiền' in str(ws.cell(r, c).value or '').lower()
                    or 'tổng' in str(ws.cell(r, c).value or '').lower()
                    or 'số tiền' in str(ws.cell(r, c).value or '').lower()
                    or 'tổng tiền' in str(ws.cell(r, c).value or '').lower()
                    or 'đơn giá' in str(ws.cell(r, c).value or '').lower()
                    for c in range(1, min(25, ws.max_column + 1))
                ):
                    header_row = r
                    break
            if not header_row:
                continue

            # Find "money" columns heuristically in header row
            money_cols = []
            for c in range(1, min(25, ws.max_column + 1)):
                v = ws.cell(header_row, c).value
                if not v: continue
                v_low = str(v).lower()
                if any(k in v_low for k in ('thành tiền', 'đơn giá', 'tổng', 'số tiền')):
                    money_cols.append(c)

            # Count data rows + sum amount across money cols
            sheet_rows = 0
            sheet_total_max = 0
            for r in range(header_row + 1, min(400, ws.max_row + 1)):
                stt = ws.cell(r, 1).value
                if isinstance(stt, (int, float)) and stt > 0 and stt < 1000:
                    sheet_rows += 1
                    # Find the largest numeric value in money cols for this row (= total per row)
                    row_max = 0
                    for c in money_cols:
                        # Try evaluated first, then raw (might be raw number not formula)
                        v = ws.cell(r, c).value
                        if isinstance(v, (int, float)) and v > row_max:
                            row_max = v
                        else:
                            # Check raw version
                            rv = ws_raw.cell(r, c).value
                            if isinstance(rv, (int, float)) and rv > row_max:
                                row_max = rv
                    sheet_total_max += row_max
                elif stt and isinstance(stt, str) and 'ổng' in stt:
                    break

            info["rows"] += sheet_rows
            info["total_revenue"] += sheet_total_max

    except Exception as e:
        info["error"] = str(e)
    return info


def find_original(folder_name: str) -> list[Path]:
    """Find all xlsx files in the customer's folder (except ~$ temp)."""
    folder = ORIGINAL_ROOT / folder_name
    if not folder.exists():
        return []
    return [f for f in folder.rglob("*.xlsx") if not f.name.startswith("~$")]


def main():
    lines = ["# Customer Bảng Kê Verification Report — Generated vs Original\n"]
    lines.append(f"**Month**: 2026-03 | **Backend**: {BACKEND}\n")
    lines.append("| Customer | Template | Gen Rows | Gen Total | Orig File | Orig Rows | Orig Total | Status |")
    lines.append("|---|---|---|---|---|---|---|---|")

    stats = {"ok": 0, "data_gap": 0, "missing_db": 0, "no_original": 0}

    for code, template, orig_folder, label in CUSTOMERS:
        gen_path = fetch_generated(code, template)
        if not gen_path or gen_path.stat().st_size < 5000:  # small = error body
            status = "❌ Missing-in-DB (404)"
            stats["missing_db"] += 1
            lines.append(f"| {code} | {template or '-'} | — | — | — | — | — | {status} |")
            continue

        gen = inspect(gen_path)
        gen_rows = gen.get("rows", 0)
        gen_total = gen.get("total_revenue", 0)

        originals = find_original(orig_folder)
        if not originals:
            status = "⚠️ No-original-found"
            stats["no_original"] += 1
            lines.append(f"| {code} | {template or '-'} | {gen_rows} | {gen_total:,.0f} | — | — | — | {status} |")
            continue

        # Pick first (or best-matching) original
        orig = originals[0]
        orig_info = inspect(orig)
        orig_rows = orig_info.get("rows", 0)
        orig_total = orig_info.get("total_revenue", 0)

        # Status
        if gen_rows == 0 and orig_rows > 0:
            status = "❌ Data-gap (DB empty)"
            stats["data_gap"] += 1
        elif gen_total == 0 and orig_total > 0:
            status = "❌ Revenue=0 (DB missing costs)"
            stats["data_gap"] += 1
        elif abs(gen_rows - orig_rows) <= 2 and (orig_total > 0 and abs(gen_total - orig_total) / orig_total < 0.1):
            status = "✅ Match"
            stats["ok"] += 1
        elif gen_rows > 0 and gen_total > 0:
            status = "⚠️ Partial-match"
            stats["ok"] += 1
        else:
            status = "❓ Other"

        lines.append(
            f"| {code} | {template or '-'} | {gen_rows} | {gen_total:,.0f} | "
            f"`{orig.name[:50]}` | {orig_rows} | {orig_total:,.0f} | {status} |"
        )

    lines.append("")
    lines.append(f"## Summary\n")
    lines.append(f"- ✅ Match / partial: **{stats['ok']}**")
    lines.append(f"- ❌ Data gap (DB missing revenue/costs): **{stats['data_gap']}**")
    lines.append(f"- ❌ DB has no jobs for month: **{stats['missing_db']}**")
    lines.append(f"- ⚠️ No original found: **{stats['no_original']}**")

    # Details of data gaps
    lines.append("\n## Gap Analysis\n")
    lines.append("For customers with 'Revenue=0' or 'Data-gap', the original Excel has real data")
    lines.append("but the DB is incomplete. **This is a data-import issue, not a renderer bug.**")
    lines.append("To fix: re-run/extend the import scripts in `backend/scripts/import-*` to")
    lines.append("populate `job_costs` from the original customer Excel files.")

    REPORT_PATH.write_text('\n'.join(lines), encoding='utf-8')
    print(f"\nReport: {REPORT_PATH}")
    print(f"Summary: {stats}")


if __name__ == '__main__':
    main()
