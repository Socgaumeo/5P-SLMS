#!/usr/bin/env python3
"""
Generate a CONCISE markdown summary of a DAINESE Excel file.
Focus on: actual data range, header rows, sample data, formulas, images, styles.
Skips empty trailing columns/rows that openpyxl reports as max_row/col.
"""

import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def find_real_bounds(ws):
    """Find actual data bounds (skip trailing empty cells)."""
    max_r, max_c = 0, 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                if cell.row > max_r:
                    max_r = cell.row
                if cell.column > max_c:
                    max_c = cell.column
    return max_r, max_c


def style_summary(cell):
    parts = []
    if cell.font:
        f = cell.font
        if f.bold:
            parts.append('B')
        if f.italic:
            parts.append('I')
        if f.size and f.size != 11:
            parts.append(f'{f.size}pt')
        if f.name and f.name != 'Calibri':
            parts.append(f.name)
        if f.color and f.color.rgb and str(f.color.rgb) != '00000000':
            parts.append(f'#{str(f.color.rgb)[-6:]}')
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
        rgb = str(cell.fill.fgColor.rgb)
        if rgb != '00000000' and rgb != '00FFFFFF':
            parts.append(f'bg#{rgb[-6:]}')
    if cell.alignment:
        a = cell.alignment
        align_parts = []
        if a.horizontal:
            align_parts.append(a.horizontal[0])
        if a.vertical and a.vertical != 'bottom':
            align_parts.append(a.vertical[0])
        if a.wrap_text:
            align_parts.append('w')
        if align_parts:
            parts.append(''.join(align_parts))
    if cell.number_format and cell.number_format != 'General':
        nf = cell.number_format
        if len(nf) > 20:
            nf = nf[:18] + '..'
        parts.append(f'fmt:{nf}')
    return '|'.join(parts) if parts else ''


def summarize_sheet(ws_formulas, ws_data, out):
    out.append(f"\n## Sheet: `{ws_formulas.title}`\n")
    real_r, real_c = find_real_bounds(ws_formulas)
    out.append(f"- Real data range: **{real_r} rows × {real_c} cols** (openpyxl max: {ws_formulas.max_row} × {ws_formulas.max_column})")
    out.append(f"- Merged cells: `{', '.join(str(r) for r in list(ws_formulas.merged_cells.ranges)[:30])}`")
    out.append(f"- Images: {len(ws_formulas._images) if hasattr(ws_formulas, '_images') else 0}")

    # Column widths
    widths = []
    for c in range(1, real_c + 1):
        letter = get_column_letter(c)
        dim = ws_formulas.column_dimensions.get(letter)
        if dim and dim.width:
            widths.append(f"{letter}={dim.width:.1f}")
    out.append(f"- Column widths: {', '.join(widths[:50])}")

    # Row heights
    heights = []
    for r in range(1, real_r + 1):
        dim = ws_formulas.row_dimensions.get(r)
        if dim and dim.height:
            heights.append(f"r{r}={dim.height:.0f}")
    out.append(f"- Row heights: {', '.join(heights[:30])}")

    # Print first 25 rows of cells
    out.append(f"\n### Cell Map (first 30 rows)\n")
    out.append("```")
    for r in range(1, min(real_r + 1, 31)):
        for c in range(1, real_c + 1):
            cell = ws_formulas.cell(row=r, column=c)
            if cell.value is None and not cell.has_style:
                continue
            coord = cell.coordinate
            val = cell.value
            if val is None:
                continue
            val_str = str(val)
            if len(val_str) > 80:
                val_str = val_str[:77] + '...'
            style = style_summary(cell)
            # If formula, also get computed
            if isinstance(val, str) and val.startswith('='):
                computed = ws_data[coord].value
                comp_str = f" => {computed}"
                out.append(f"  {coord}: FORMULA {val_str}{comp_str}  [{style}]")
            else:
                out.append(f"  {coord}: {val_str!r}  [{style}]")
    out.append("```")

    # Print bottom rows (last 10)
    if real_r > 30:
        out.append(f"\n### Bottom rows ({max(real_r-10, 31)}-{real_r})\n")
        out.append("```")
        for r in range(max(real_r - 10, 31), real_r + 1):
            for c in range(1, real_c + 1):
                cell = ws_formulas.cell(row=r, column=c)
                if cell.value is None:
                    continue
                coord = cell.coordinate
                val = cell.value
                val_str = str(val)
                if len(val_str) > 80:
                    val_str = val_str[:77] + '...'
                style = style_summary(cell)
                if isinstance(val, str) and val.startswith('='):
                    computed = ws_data[coord].value
                    out.append(f"  {coord}: FORMULA {val_str} => {computed}  [{style}]")
                else:
                    out.append(f"  {coord}: {val_str!r}  [{style}]")
        out.append("```")

    # Sample formulas
    formula_count = 0
    formulas = []
    for row in ws_formulas.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formulas.append((cell.coordinate, cell.value))
                formula_count += 1
    if formulas:
        out.append(f"\n### Formulas ({formula_count} total, showing unique patterns)\n")
        # Group by formula pattern (replace numbers with #)
        import re
        patterns = {}
        for coord, f in formulas:
            pat = re.sub(r'\d+', '#', f)
            patterns.setdefault(pat, []).append(coord)
        out.append("```")
        for pat, coords in list(patterns.items())[:20]:
            example_coord = coords[0]
            example_formula = next(f for c, f in formulas if c == example_coord)
            out.append(f"  Pattern: {pat}")
            out.append(f"    Example: {example_coord} = {example_formula}")
            out.append(f"    Used in {len(coords)} cells: {', '.join(coords[:8])}{'...' if len(coords) > 8 else ''}")
        out.append("```")


def main():
    if len(sys.argv) < 2:
        print("Usage: summarize-dainese-excel-template-layout-and-formulas.py <xlsx> [output.md]")
        sys.exit(1)

    filepath = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None

    wb = load_workbook(filepath, data_only=False)
    wb_data = load_workbook(filepath, data_only=True)

    out = []
    out.append(f"# Excel Analysis: `{Path(filepath).name}`\n")
    out.append(f"- Path: `{filepath}`")
    out.append(f"- Sheets: {wb.sheetnames}")

    for sn in wb.sheetnames:
        summarize_sheet(wb[sn], wb_data[sn], out)

    md = '\n'.join(out)
    if output_path:
        Path(output_path).write_text(md, encoding='utf-8')
        print(f"Wrote: {output_path}")
        print(f"Size: {len(md)} bytes")
    else:
        print(md)


if __name__ == '__main__':
    main()
