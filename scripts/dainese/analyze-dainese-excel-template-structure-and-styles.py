#!/usr/bin/env python3
"""
Analyze a DAINESE Excel template file.
Extract: layout, fonts, colors, formulas, merged cells, images, column widths.
Output: detailed JSON + human-readable markdown report.
"""

import sys
import json
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def cell_style(cell):
    """Extract style info from a cell."""
    style = {}
    if cell.font:
        f = cell.font
        style['font'] = {
            'name': f.name,
            'size': f.size,
            'bold': f.bold,
            'italic': f.italic,
            'color': str(f.color.rgb) if f.color and f.color.rgb else None,
        }
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
        style['fill'] = str(cell.fill.fgColor.rgb)
    if cell.alignment:
        a = cell.alignment
        style['align'] = {
            'horizontal': a.horizontal,
            'vertical': a.vertical,
            'wrap_text': a.wrap_text,
        }
    if cell.number_format and cell.number_format != 'General':
        style['number_format'] = cell.number_format
    if cell.border:
        b = cell.border
        borders = {}
        for side_name in ('left', 'right', 'top', 'bottom'):
            side = getattr(b, side_name)
            if side and side.style:
                borders[side_name] = side.style
        if borders:
            style['border'] = borders
    return style


def analyze_sheet(ws):
    """Analyze a single worksheet."""
    info = {
        'name': ws.title,
        'dimensions': ws.dimensions,
        'max_row': ws.max_row,
        'max_col': ws.max_column,
        'merged_ranges': [str(r) for r in ws.merged_cells.ranges],
        'column_widths': {},
        'row_heights': {},
        'images': len(ws._images) if hasattr(ws, '_images') else 0,
        'cells': [],
        'formulas': [],
    }

    # Column widths
    for col_letter, dim in ws.column_dimensions.items():
        if dim.width:
            info['column_widths'][col_letter] = dim.width

    # Row heights
    for row_num, dim in ws.row_dimensions.items():
        if dim.height:
            info['row_heights'][row_num] = dim.height

    # Cells with values, styles, formulas
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None and not cell.has_style:
                continue
            cell_info = {
                'coord': cell.coordinate,
                'row': cell.row,
                'col': cell.column,
            }
            if cell.value is not None:
                # Detect formula
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    cell_info['formula'] = cell.value
                    info['formulas'].append({
                        'coord': cell.coordinate,
                        'formula': cell.value
                    })
                else:
                    cell_info['value'] = str(cell.value) if cell.value is not None else None
            style = cell_style(cell)
            if style:
                cell_info['style'] = style
            info['cells'].append(cell_info)

    return info


def analyze_file(filepath, output_dir):
    """Analyze full Excel file."""
    print(f"\n{'='*70}")
    print(f"Analyzing: {filepath}")
    print(f"{'='*70}")

    # data_only=False to keep formulas
    wb = load_workbook(filepath, data_only=False)
    # also load with data_only=True to see computed values
    wb_data = load_workbook(filepath, data_only=True)

    result = {
        'filepath': str(filepath),
        'filename': Path(filepath).name,
        'sheet_names': wb.sheetnames,
        'sheets': {},
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws_data = wb_data[sheet_name]
        info = analyze_sheet(ws)
        # Get computed values for formula cells
        for f in info['formulas']:
            cell = ws_data[f['coord']]
            f['computed_value'] = str(cell.value) if cell.value is not None else None
        result['sheets'][sheet_name] = info

    # Output
    base = Path(filepath).stem
    safe_base = "".join(c if c.isalnum() or c in '-_' else '_' for c in base)
    json_path = Path(output_dir) / f"{safe_base}.json"

    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2, default=str)

    print(f"JSON output: {json_path}")
    print(f"Sheets: {result['sheet_names']}")
    for name, info in result['sheets'].items():
        print(f"  - {name}: {info['max_row']} rows x {info['max_col']} cols, "
              f"{len(info['merged_ranges'])} merged, "
              f"{len(info['formulas'])} formulas, "
              f"{info['images']} images")

    return result


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: analyze-excel-template.py <xlsx_file> [output_dir]")
        sys.exit(1)

    filepath = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else '.'
    os.makedirs(output_dir, exist_ok=True)

    analyze_file(filepath, output_dir)
