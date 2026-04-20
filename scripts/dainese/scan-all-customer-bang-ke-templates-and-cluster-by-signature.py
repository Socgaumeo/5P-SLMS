#!/usr/bin/env python3
"""
Scan all customer bảng kê (statement) xlsx files in DOANH THU/THÁNG 3,
extract structural fingerprints, then cluster by similarity.

For each file extract:
- sheet names + dimensions
- presence of logo (image)
- title row text + position
- column headers (row identified by content match)
- # data rows
- presence of: VAT row, Tổng row, Reim/Thu hộ section, signature blocks
- formula patterns (per row, sum, etc.)
- column count + column header keywords

Output: markdown report grouping files by template signature.
"""

import sys
import re
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook


ROOT = Path("/Users/bear1108/Library/CloudStorage/OneDrive-Personal/5P/DOANH THU/THÁNG 3")


def find_text_row(ws, keywords: list, max_row: int = 30) -> tuple:
    """Find first row containing any keyword. Returns (row, text) or (None, None)."""
    for r in range(1, min(max_row, ws.max_row + 1)):
        for c in range(1, min(15, ws.max_column + 1)):
            v = ws.cell(r, c).value
            if v and any(k.lower() in str(v).lower() for k in keywords):
                return (r, str(v))
    return (None, None)


def find_real_bounds(ws, max_check_row=300, max_check_col=50):
    """Trim trailing empty rows/cols."""
    last_r, last_c = 0, 0
    for r in range(1, min(max_check_row, ws.max_row + 1)):
        for c in range(1, min(max_check_col, ws.max_column + 1)):
            if ws.cell(r, c).value is not None:
                if r > last_r: last_r = r
                if c > last_c: last_c = c
    return last_r, last_c


def extract_column_headers(ws, header_row: int, max_col: int) -> list:
    """Extract non-empty cell values in the header row."""
    out = []
    for c in range(1, max_col + 1):
        v = ws.cell(header_row, c).value
        if v: out.append(str(v).strip())
    return out


def detect_header_row(ws, max_row=30, max_col=30) -> int:
    """Find row with most columns filled and bold formatting (likely table header)."""
    best_r, best_score = None, 0
    for r in range(1, min(max_row, ws.max_row + 1)):
        score = 0
        for c in range(1, min(max_col, ws.max_column + 1)):
            cell = ws.cell(r, c)
            if cell.value:
                score += 1
                if cell.font and cell.font.bold:
                    score += 1
        if score > best_score:
            best_score = score
            best_r = r
    return best_r


def collect_formula_patterns(ws, max_row=200, max_col=50) -> dict:
    """Count formula patterns (with numbers replaced by #)."""
    patterns = defaultdict(int)
    for r in range(1, min(max_row, ws.max_row + 1)):
        for c in range(1, min(max_col, ws.max_column + 1)):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.startswith('='):
                pat = re.sub(r'\d+', '#', v)
                patterns[pat] += 1
    return dict(patterns)


def has_image(ws) -> bool:
    return hasattr(ws, '_images') and len(ws._images) > 0


def detect_features(ws, last_r, last_c) -> dict:
    """Detect template features: VAT/Tổng/Thu hộ/signature blocks."""
    features = {}
    # Search bottom 60% of sheet for total/vat/thu hộ rows
    bot_start = max(1, int(last_r * 0.4))
    for r in range(bot_start, last_r + 1):
        for c in range(1, min(20, last_c + 1)):
            v = ws.cell(r, c).value
            if v:
                vstr = str(v).lower()
                if 'vat' in vstr and 'vat' not in features:
                    features['vat'] = (r, c, str(v))
                if ('tổng cộng' in vstr or 'tổng tiền' in vstr or 'tổng thanh toán' in vstr) and 'tong_cong' not in features:
                    features['tong_cong'] = (r, c, str(v))
                if vstr.strip() in ('tổng', 'tong') and 'tong' not in features:
                    features['tong'] = (r, c, str(v))
                if ('thu hộ' in vstr or 'chi hộ' in vstr or 'phí trả hộ' in vstr or 'reimburs' in vstr) and 'thu_ho' not in features:
                    features['thu_ho'] = (r, c, str(v))
                if 'bằng chữ' in vstr and 'bang_chu' not in features:
                    features['bang_chu'] = (r, c, str(v)[:80])
                if ('chuyển khoản' in vstr or 'tài khoản' in vstr) and 'bank' not in features:
                    features['bank'] = r
    return features


def signature(info: dict) -> str:
    """Compact signature string for clustering."""
    parts = []
    parts.append(f"cols~{info['col_count']}")
    parts.append("logo" if info['has_image'] else "nologo")
    parts.append(f"sheets={info['sheet_count']}")
    feats = info.get('features', {})
    for k in ('tong_cong', 'tong', 'vat', 'thu_ho', 'bang_chu', 'bank'):
        if k in feats: parts.append(k)
    # Header keywords as fingerprint
    hdrs = info.get('headers', [])
    keys = []
    for h in hdrs:
        h_low = h.lower()
        if 'tờ khai' in h_low or 'cd_no' in h_low: keys.append('TK')
        if 'invoice' in h_low or 'hóa đơn' in h_low or 'hoa don' in h_low: keys.append('INV')
        if 'vận đơn' in h_low or 'bl' in h_low or 'awb' in h_low: keys.append('BL')
        if 'tuyến' in h_low or 'route' in h_low or 'tuyen' in h_low: keys.append('ROUTE')
        if 'điểm' in h_low or 'diem' in h_low: keys.append('PT')
        if 'bks' in h_low or 'biển số' in h_low or 'license' in h_low: keys.append('BKS')
        if 'co' == h_low.strip() or 'số co' in h_low: keys.append('CO')
        if 'cont' in h_low: keys.append('CONT')
    parts.append('hdr=' + ','.join(sorted(set(keys))))
    return '|'.join(parts)


def analyze_file(path: Path) -> dict:
    rel = path.relative_to(ROOT)
    customer = rel.parts[0]
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        return {'path': str(rel), 'customer': customer, 'error': str(e)}

    info = {
        'path': str(rel),
        'customer': customer,
        'filename': path.name,
        'sheet_count': len(wb.sheetnames),
        'sheets': wb.sheetnames,
    }

    # Analyze first non-empty sheet (main one)
    ws = wb[wb.sheetnames[0]]
    last_r, last_c = find_real_bounds(ws)
    info['rows'] = last_r
    info['col_count'] = last_c
    info['has_image'] = has_image(ws)

    # Title + customer name
    title_r, title = find_text_row(ws, ['BẢNG KÊ', 'DEBIT NOTE', 'BANG KE'])
    info['title'] = title
    info['title_row'] = title_r

    cust_r, cust = find_text_row(ws, ['KÍNH GỬI', 'KINH GUI', 'CUSTOMER', 'NGƯỜI MUA'])
    info['recipient'] = cust[:80] if cust else None

    # Header row
    hdr_r = detect_header_row(ws, max_row=min(40, last_r))
    if hdr_r:
        info['header_row'] = hdr_r
        info['headers'] = extract_column_headers(ws, hdr_r, last_c)[:25]

    # Features
    info['features'] = detect_features(ws, last_r, last_c)

    # Formula patterns
    info['formula_patterns'] = collect_formula_patterns(ws, max_row=min(200, last_r), max_col=last_c)

    info['signature'] = signature(info)
    return info


def main():
    out_path = Path("plans/reports/dainese-templates/scan-all-customer-bang-ke-templates-clustering-report.md")
    out_path.parent.mkdir(parents=True, exist_ok=True)

    files = sorted(ROOT.rglob("*.xlsx"))
    files = [f for f in files
             if 'DAINESE' not in str(f) and 'MEIKO' not in str(f)
             and not f.name.startswith('~$')]

    print(f"Scanning {len(files)} files...")
    results = []
    for f in files:
        info = analyze_file(f)
        results.append(info)
        print(f"  {info.get('signature','ERR')[:60]:60} | {info['path']}")

    # Cluster by signature (similarity = same customer count of header keywords + features)
    clusters = defaultdict(list)
    for r in results:
        if 'error' in r:
            clusters['__error__'].append(r)
            continue
        # Cluster key: header keyword set + features set (ignore col count exactness)
        sig = r['signature']
        # Simplify: cluster by header keywords signature only
        m = re.search(r'hdr=([^|]*)', sig)
        hdr_key = m.group(1) if m else ''
        feat_key = '|'.join(sorted(r.get('features', {}).keys()))
        cluster_key = f"hdr=[{hdr_key}] feats=[{feat_key}]"
        clusters[cluster_key].append(r)

    # Write report
    lines = [
        "# Customer Bảng Kê Templates — Clustering Analysis",
        f"\n**Scanned**: {len(files)} files (excl. DAINESE + MEIKO)",
        f"**Customers**: {len(set(r['customer'] for r in results))}",
        f"**Clusters**: {len(clusters)}\n",
    ]

    # Cluster summary
    lines.append("## Cluster Overview\n")
    lines.append("| # | Cluster Signature | Files | Customers |")
    lines.append("|---|---|---|---|")
    for i, (key, members) in enumerate(sorted(clusters.items(), key=lambda x: -len(x[1])), 1):
        custs = sorted(set(m['customer'] for m in members))
        lines.append(f"| C{i} | `{key[:80]}` | {len(members)} | {', '.join(custs)} |")

    # Detail per cluster
    lines.append("\n---\n## Cluster Details\n")
    for i, (key, members) in enumerate(sorted(clusters.items(), key=lambda x: -len(x[1])), 1):
        lines.append(f"\n### Cluster C{i}: `{key}`\n")
        lines.append(f"**Files** ({len(members)}):")
        for m in members:
            lines.append(f"- `{m['path']}`")
            lines.append(f"  - Sheets: {m.get('sheets', [])}")
            lines.append(f"  - Size: {m.get('rows', '?')}r × {m.get('col_count', '?')}c | Image: {m.get('has_image', False)}")
            if m.get('title'):
                lines.append(f"  - Title: `{m['title'][:90]}`")
            if m.get('recipient'):
                lines.append(f"  - Recipient: `{m['recipient']}`")
            if m.get('headers'):
                lines.append(f"  - Headers ({m.get('header_row','?')}): {m['headers'][:15]}")
            feats = m.get('features', {})
            if feats:
                lines.append(f"  - Features: {list(feats.keys())}")
            fps = m.get('formula_patterns', {})
            if fps:
                top_fps = sorted(fps.items(), key=lambda x: -x[1])[:5]
                lines.append(f"  - Top formulas: {[(p,n) for p,n in top_fps]}")

    out_path.write_text('\n'.join(lines), encoding='utf-8')
    print(f"\nWrote: {out_path}")
    print(f"Size: {out_path.stat().st_size} bytes")


if __name__ == '__main__':
    main()
